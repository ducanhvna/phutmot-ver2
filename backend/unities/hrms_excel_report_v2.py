
import math
import sys
import requests
import json
import datetime
import calendar
# from datetime import datetime
import openpyxl
import os
# from dotenv import load_dotenv
# load_dotenv()
from openpyxl.utils import get_column_letter
import re
import pandas as pd
import numpy as np
from collections import Counter
import xmlrpc.client
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors
from apps.home.unities.unities import float_to_hours
from apps.home.unities.reportfile import REPORT_LINK
from apps.employees.models import Employee
import logging
from dateutil.relativedelta import relativedelta
# logging.basicConfig(filename='my_log.txt', format='%(asctime)s - %(levelname)s - %(name)s - %(message)s')
# logger = logging.getLogger('my_program')
# logger.info('This is an info message')

MINIO_ACCESS_KEY = "FLoU4kYrt6EQ8eyWBLjD"
MINIO_SECRET_KEY = "LBa3KybNAxwxHWuFPqKF00ppIi5iOotJXQQzriUa"
MINIO_BUCKET_NAME = "apecerp"
MINIO_ENDPOINT_URL = "http://42.113.122.201:9000/apecerp/"
MINIO_ENDPOINT = "42.113.122.201:9000"

MINIO_PUBLIC_URL = "https://minio.qcloud.asia/apecerp/"

logging.basicConfig(filename='my_log.txt',
                    filemode='a',
                    format='%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S',
                    level=logging.DEBUG)

logging.info("Running HRMS")
logger = logging.getLogger('HrmExcelReport_v2')

url = 'https://hrm.mandalahotel.com.vn'
db = 'apechrm_product_v3'
username = 'admin'
password = 'admin'


class HrmExcelReport_v2():
    def __init__(self, start_date_str, end_date_str, server_url=None, server_db=None, server_username=None, server_password=None,
                 ouput_report_folder='', is_load_df_old=False, is_update_al_cl=False, select_company_id = 0,
                 select_company_name = ''):
        super(HrmExcelReport_v2, self).__init__()
        logger.info('Init class HrmExcelReport_v2')
        self.is_load_df_old = is_load_df_old
        self.select_company_id = select_company_id
        self.select_company_name = select_company_name
        
        # if not is_load_df_old:
        #     logger.warning("Dont load df_old")
        if server_url:
            self.url = server_url
        else:
            self.url = url
        if server_db:
            self.db = server_db
        else:
            self.db = db
        if server_username:
            self.username = server_username
        else:
            self.username = username
        if server_password:
            self.password = server_password
        else:
            self.password = password
        self.is_update_al_cl = is_update_al_cl
        self.output_report_folder = ouput_report_folder
        config_file = 'https://dl.dropboxusercontent.com/scl/fi/tup2bwao3ososd1h7hj04/Config.xlsx?rlkey=h7syiq8iby3l545e7fdmr0doq&dl=0'
        self.df_config_company = pd.read_excel(config_file, engine='openpyxl', index_col=0, sheet_name='COMPANY') 
        isExist = os.path.exists(ouput_report_folder)
        if not isExist:
            os.makedirs(ouput_report_folder)
        print(f'{start_date_str}-{end_date_str}')
        common = xmlrpc.client.ServerProxy(
            '{}/xmlrpc/2/common'.format(self.url))
        self.models = xmlrpc.client.ServerProxy(
            '{}/xmlrpc/2/object'.format(self.url))
        self.uid = common.authenticate(
            self.db, self.username, self.password, {})
        self.user_profile = self.models.execute_kw(self.db, self.uid, self.password, 'res.users', 'read', [
                                                   [self.uid]], {'fields': ['id', 'name', 'company_id']})[0]
        self.company_id = self.user_profile['company_id'][0]
        self.company_name = self.user_profile['company_id'][1]

        self.date_array = pd.date_range(start=start_date_str, end=end_date_str)
        self.download_attendance_report_al_cl_last_month()
        self.df_scheduling_ver_al_cl_last_month.to_excel(os.path.join(self.output_report_folder, 'df_scheduling_ver_al_cl_last_month.xlsx'))
        
        self.load_company_list()
        self.load_hr_weekly_report()
        self.load_kpi_weekly_report_summary()
        self.update_contract_status()
        
        self.download_attendance_report_holiday_last_month()
        self.df_scheduling_ver_holiday_last_month.to_excel(os.path.join(self.output_report_folder, 'df_scheduling_ver_holiday_last_month.xlsx'))
        self.download_employee_info()

        self.merge_download_attendance()
        if self.is_load_df_old:
            self.download_attendance_trans()
        self.merge_employee()
        # self.df_employees.to_excel('df_employees_in_after_merge_attenadant.xlsx')
        self.merge_attendance_data()
        
        self.auto_remove_attendance_restshift()
        # self.df_employees.to_excel('df_employees_in_cl_afteratteen.xlsx')
        self.load_al_cl_report()
        
        # self.df_employees.to_excel('df_employees_in_cl_bf_update.xlsx')
        self.cl_result_update = {}
        self.cl_result_create = {}
        self.al_result_update = {}
        self.al_result_create = {}
        self.load_last_year_al_cl_report()
        if is_update_al_cl:
            self.load_last_al_cl_report()
            # self.update_last_month_data()
        self.merge_shift_info()
        
        self.load_upload_report_data()
    def update_contract_status(self):
        print('update contract status')
        ids_contract = self.models.execute_kw(
            self.db, self.uid, self.password, "hr.contract", "search_read", [[]], {'fields': ['id',
                                                                'minutes_per_day',
                                                                'employee_code', 
                                                                'date_end','date_start', 
                                                                'state','resource_calendar_id']})
        
        self.df_contract = pd.DataFrame.from_dict(ids_contract)
        self.df_contract['date_start'] = pd.to_datetime(
                self.df_contract['date_start'] , format="%Y-%m-%d", errors='coerce')
        self.df_contract['date_end'] = pd.to_datetime(
                self.df_contract['date_end'] , format="%Y-%m-%d", errors='coerce')
        
        # print(len(ids_contract))
        ids_all_close = []
        ids_all_almost = []
        ids_close = []
        ids_almost = []
        ids_open = []
        ids_open_miss_calendar = []
        
        for id in ids_contract:
            
            if id["date_end"]:
                date_end = datetime.datetime.strptime(id["date_end"], "%Y-%m-%d").date()
                
                
                if (date_end < datetime.date.today())  :
                    ids_all_close.append(id["id"])
                    if (id['state'] != 'close'):
                        print(id)
                        ids_close.append(id["id"])
                elif (date_end - datetime.timedelta(days=30) < datetime.date.today()):
                    ids_all_almost.append(id["id"])
                    if (id['state'] != 'almost'):
                        print('almost',id)
                        ids_almost.append(id["id"])

            if id["date_start"] and (not id['id'] in ids_all_almost) and (not id['id'] in ids_all_close):
                date_start = datetime.datetime.strptime(id["date_start"], "%Y-%m-%d").date()
                if (date_start  < datetime.date.today()) and ( (id['state'] != 'open')):
                    if id['resource_calendar_id'] == False:
                        ids_open_miss_calendar.append(id["id"])
                    else:
                        ids_open.append(id["id"])

        if len(ids_almost)>0:
            # print('ids_almost ', ids_almost)
            try:
                self.models.execute_kw(
                    self.db, self.uid, self.password, "hr.contract", "write", [ids_almost, {"state": "almost"}])
            except Exception as ex:
                logger.error(f"almost contract:  {ex}")

        if len(ids_close)>0:
            print('close ', ids_close)
            try:
                self.models.execute_kw(
                    self.db, self.uid, self.password, "hr.contract", "write", [ids_close, {"state": "close"}])
            except Exception as ex:
                logger.error(f"Close contract: {ex}")
        
        if len(ids_open)>0:
            print('ids_open ', ids_open)
            try:
                self.models.execute_kw(
                    self.db, self.uid, self.password, "hr.contract", "write", [ids_open, {"state": "open"}])
            except Exception as ex:
                logger.error(f"open contract: {ex}")

        print('Re download contract status')
        local_index = 0
        LIMIT_SIZE = 500
        ids_contract = []
        while (len(ids_contract) == LIMIT_SIZE) or (local_index == 0):
            offset = local_index * LIMIT_SIZE
            ids_contract = self.models.execute_kw(
                self.db, self.uid, self.password, "hr.contract", "search_read", [[]],
                        {'offset': offset, 'limit': LIMIT_SIZE,
                        'fields': ['id','company_id',
                        'contract_type_id', 'minutes_per_day', 'employee_id',
                        'employee_code', 'date_end','date_start', 'salary_rate',
                        'state','resource_calendar_id']}
            )
            for item in ids_contract:
                try: 
                    item['contract_type_name'] = item['contract_type_id'][1]
                except:
                    item['contract_type_name'] = ''
                try:
                    item['company_sname'] = item['company_id'][1]
                    item['company_sid'] = item['company_id'][0]
                except:
                    item['company_sname'] = ''
                    item['company_sid'] = -1
                try:
                    item['employee_id'] = item['employee_id'][0]
                except:
                    item['employee_id'] = False
            if local_index == 0:  
                self.df_contract = pd.DataFrame.from_dict(ids_contract)  
            else:
                self.df_contract = pd.concat([self.df_contract, pd.DataFrame.from_dict(ids_contract)], ignore_index=True)
            local_index = local_index + 1
        # self.df_contract = pd.DataFrame.from_dict(ids_contract)

        self.df_contract['date_start'] = pd.to_datetime(
                self.df_contract['date_start'] , format="%Y-%m-%d", errors='coerce')
        self.df_contract['date_end'] = pd.to_datetime(
                self.df_contract['date_end'] , format="%Y-%m-%d", errors='coerce')
        # df_contract_ver = self.df_contract.copy()
        # df_contract_ver["date"] = df_contract_ver.apply(
        #         lambda row: pd.date_range(row["date_start"], row["date_end"]),
        #         axis=1
        #     )
        # df_contract_ver = (
        #         df_contract_ver.explode("date", ignore_index=True)
        #         .drop(columns=["date_start", "date_end"])
        #     )
        # df_contract_ver.to_excel(os.path.join(self.output_report_folder, "df_contract_ver.xlsx"), sheet_name='Sheet1')



    def load_company_list(self):
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'res.company', 'search', [[]], {})
        list_company = self.models.execute_kw(self.db, self.uid, self.password, 'res.company', 'read', [ids], {'fields': ['id',
                                            'name', 'is_ho',
                                            'mis_id']})
        
        self.df_all_company = pd.DataFrame.from_dict(list_company)
        self.df_all_company = self.df_all_company.merge(self.df_config_company, 
                    how='left', 
                    left_on=['mis_id'], 
                    right_on=['NAME'])
    def merge_shift_info(self):
        
        local_index = 0
        LIMIT_SIZE = 50
        ids = []
        while (len(ids) == LIMIT_SIZE) or (local_index == 0):
            offset = local_index * LIMIT_SIZE
            if (self.select_company_id >0):
                domain= [('company_id', '=', self.select_company_id)]
            else:
                domain =[]
            ids = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'search', [domain], {'offset': offset, 'limit': LIMIT_SIZE})
            
            list_shifts  = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'read', [ids], {'fields': ['id', 'name', 'start_work_time',
                                                'end_work_time','total_work_time','start_rest_time','end_rest_time', 'company_id',
                                                'rest_shifts', 'fix_rest_time', 'night', 'night_eat','dinner','lunch','breakfast','efficiency_factor', 
                                                'minutes_working_not_reduced']}
                                                )
            # print (self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'fields_get', [], {'attributes': ['string', 'type']}))
            for shift in list_shifts:
                try:
                    company_name = shift['company_id'][1]
                    shift['company_name'] = company_name
                except:
                    shift['company_name'] = False
            if local_index == 0:  
                self.df_shift = pd.DataFrame.from_dict(list_shifts)  
            else:
                self.df_shift = pd.concat([self.df_shift, pd.DataFrame.from_dict(list_shifts)], ignore_index=True)
            
            local_index = local_index + 1
        self.df_shift.to_excel(os.path.join(self.output_report_folder,'self.df_shift.xlsx'))
        self.df_scheduling_ver =  self.df_scheduling_ver.merge(self.df_shift[['id','name', 'total_work_time', \
                    'start_work_time','end_work_time','rest_shifts', 'company_name', \
                    'start_rest_time','end_rest_time', 'fix_rest_time', 'night', \
                    'night_eat','dinner','lunch','breakfast', 'efficiency_factor', 'minutes_working_not_reduced']], \
                    left_on=['shift_name','company_name'], right_on = ['name','company_name'], how='left', suffixes=( '' ,'_shift'))

    def load_upload_report_data(self):
        month = self.date_array[0].month
        year = self.date_array[0].year
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'search', [["&",
                                                                                                           ("month", ">", 0),
                                                                                                           ("year", "=", year)]], {})

        list_upload_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'read', [ids],
                                                    {'fields': ['id','year', 'month', 'template', 'company_id', 'department_id', 'url']})
        for upload_report in list_upload_report:
            try:
                company_id =  upload_report['company_id'][0]

                company_name = upload_report['company_id'][1]
                upload_report['company_name'] = company_name
                upload_report['company_id'] = company_id

            except:
                upload_report['company_name'] = False
            try:
                department_name = upload_report['department_id'][1]
                upload_report['department_name'] = department_name
            except:
                upload_report['department_name'] = ''


        self.df_upload_report = pd.DataFrame.from_dict(list_upload_report)
        if len(self.df_upload_report.index) == 0:
            self.df_upload_report = pd.DataFrame(columns=['id','year', 'month', 'template', 'company_id'])

        self.df_upload_report = self.df_upload_report.merge(self.df_all_company, left_on=['company_id'], \
                        right_on=['id'], how='left')
        # list_upload_ver = []
        # for company_upload_data in REPORT_LINK:
        #     for data_upload in company_upload_data['data']:
        #         try:
        #             item = {'user': company_upload_data['user'],
        #                 'mis-id': company_upload_data['mis-id'],
        #                 'month': data_upload['month'],
        #                 'year': data_upload['year']}
        #             item['type'] = 'report_summary'
        #             item['url'] = data_upload['report_summary']
        #             item['template'] = '1'
        #             list_upload_ver.append(item)
        #         except:
        #             print('report_summary', company_upload_data['mis-id'])

        #         try:
        #             item = {'user': company_upload_data['user'],
        #                 'mis-id': company_upload_data['mis-id'],
        #                 'month': data_upload['month'],
        #                 'year': data_upload['year']}
        #             item['type'] = 'report_detail'
        #             item['url'] = data_upload['report_detail']
        #             item['template'] = '2'
        #             list_upload_ver.append(item)
        #         except:
        #             print('report_detail', company_upload_data['mis-id'])
        #         try:
        #             item = {'user': company_upload_data['user'],
        #                 'mis-id': company_upload_data['mis-id'],
        #                 'month': data_upload['month'],
        #                 'year': data_upload['year']}
        #             item['type'] = 'report_5m'
        #             item['template'] = '3'
        #             item['url'] = data_upload['report_5m']
        #             list_upload_ver.append(item)
        #         except:
        #             print('report_5m', company_upload_data['mis-id'])
        #         try:
        #             item = {'user': company_upload_data['user'],
        #                 'mis-id': company_upload_data['mis-id'],
        #                 'month': data_upload['month'],
        #                 'year': data_upload['year']}
        #             item['template'] = '4'
        #             item['type'] = 'report_night_split'
        #             item['url'] = data_upload['report_night_split']
        #             list_upload_ver.append(item)
        #         except:
        #             print('report_night_split', company_upload_data['mis-id'])
        #         try:
        #             item = {'user': company_upload_data['user'],
        #                 'mis-id': company_upload_data['mis-id'],
        #                 'month': data_upload['month'],
        #                 'year': data_upload['year']}
        #             item['template'] = '5'
        #             item['type'] = 'report_al_cl'
        #             item['url'] = data_upload['report_al_cl']
        #             list_upload_ver.append(item)
        #         except:
        #             print('report_al_cl', company_upload_data['mis-id'])
        #         try:
        #             item = {'user': company_upload_data['user'],
        #                 'mis-id': company_upload_data['mis-id'],
        #                 'month': data_upload['month'],
        #                 'year': data_upload['year']}
        #             item['type'] = 'report_feed'
        #             item['template'] = '7'
        #             item['url'] = data_upload['report_feed']
        #             list_upload_ver.append(item)
        #         except:
        #             print('report_feed', company_upload_data['mis-id'])

        # df_upload_ver = pd.DataFrame.from_dict(list_upload_ver)
        # df_upload_ver = df_upload_ver.merge(self.df_all_company, left_on=['mis-id'], \
        #                 right_on=['mis_id'], how='left')
        # self.df_upload_report = df_upload_ver.merge(self.df_upload_report, right_on=['mis_id', 'year','month','template'], \
        #                 left_on=['mis-id', 'year','month','template'], how='left')
    def load_kpi_weekly_report_summary(self):
        # table kpi.weekly.report.summary
        LIMIT_SIZE = 200
        index_count = 0
        fields = ['employee_code', 'date', 'employee_name', \
            'company_name','department_name', 'employee_level', 'compensation_amount_week_1',
            'compensation_amount_week_2', 'compensation_amount_week_3', 'compensation_amount_week_4',
            'compensation_amount_week_5', 'compensation_status_week_1', 'compensation_status_week_2',
            'compensation_status_week_3' ,'compensation_status_week_4', 'compensation_status_week_5',
            'book_review_compensation_status']
        # for index, company_data in self.df_all_company.iterrows():
        #     company_name = company_data['name']
        #     local_index = 0
            
        #     ids= []
        #     while (len(ids) == LIMIT_SIZE) or (local_index == 0):
        #         offset = local_index * LIMIT_SIZE
        #         ids = self.models.execute_kw(self.db, self.uid, self.password, 'kpi.weekly.report.summary', 'search', 
        #                 [["&",("company_name","=", company_name),("date", "=", self.date_array[0].strftime('%Y-%m-%d'))]], 
        #                 {'offset': offset, 'limit': LIMIT_SIZE})
        #         if len(ids) == 0:
        #             print(f"{index} - No last data content kpi.weekly.report.summary")
                    
        #         else:
        #             print(company_name)
        #             print("found weekly: ",  len(ids))
        #             list_kpi_weekly_report  = self.models.execute_kw(self.db, self.uid, self.password, 'kpi.weekly.report.summary', 'read', [ids], 
        #                                                         {'fields': fields})
        #             if index_count == 0:                                                         
        #                 self.df_kpi_weekly_report = pd.DataFrame.from_dict(list_kpi_weekly_report)
        #                 print("load list_kpi_weekly_report fist")
        #             else:
        #                 self.df_kpi_weekly_report = pd.concat([self.df_kpi_weekly_report, pd.DataFrame.from_dict(list_kpi_weekly_report)], ignore_index=True)

        #             index_count = index_count + 1
        #         local_index = local_index + 1
                
        if index_count == 0:
            self.df_kpi_weekly_report = pd.DataFrame(columns = fields)
        else:
            self.df_kpi_weekly_report['date'] = pd.to_datetime(
                self.df_kpi_weekly_report['date'], format="%Y-%m-%d", errors='coerce')   
                
    def load_hr_weekly_report(self):
        # table: hr.weekly.report
        LIMIT_SIZE = 200
        index_count = 0
        fields = ['employee_code', 'department_id', 'employee_id', \
            'company_id','create_date', 'job_title', 'date',
            'state', 'from_date', 'to_date'
            ]
        # for index, company_data in self.df_all_company.iterrows():
        #     company_sid = company_data['id']
        #     local_index = 0
            
        #     ids= []
        #     print(company_data['name'])
        #     while (len(ids) == LIMIT_SIZE) or (local_index == 0):
        #         offset = local_index * LIMIT_SIZE
        #         ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.weekly.report', 'search', 
        #                 [["&","&",("company_id","=", company_sid), \
        #                 ("date", ">=", (self.date_array[0]-datetime.timedelta(days=7)).strftime('%Y-%m-%d')), \
        #                 ("date", "<=", (self.date_array[-1]+datetime.timedelta(days=7)).strftime('%Y-%m-%d'))]], 
        #                 {'offset': offset, 'limit': LIMIT_SIZE})
        #         if len(ids) == 0:
        #             print(f"{index} - No last data content hr.weekly.report")
        #         else:
        #             print("found weekly report: ",  len(ids))
        #             list_hr_weekly_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.weekly.report', 'read', [ids], 
        #                                                         {'fields': fields})
        #             if index_count == 0:                                                         
        #                 self.hr_weekly_report = pd.DataFrame.from_dict(list_hr_weekly_report)
        #                 print("load list_kpi_weekly_report fist")
        #             else:
        #                 self.hr_weekly_report = pd.concat([self.hr_weekly_report, pd.DataFrame.from_dict(list_hr_weekly_report)], ignore_index=True)

        #             index_count = index_count + 1
        #         local_index = local_index + 1
                
        if index_count == 0:
            self.hr_weekly_report = pd.DataFrame(columns = fields)
            
        else:    
            self.hr_weekly_report['date'] = pd.to_datetime(
                self.hr_weekly_report['date'], format="%Y-%m-%d", errors='coerce')  
            self.hr_weekly_report['from_date'] = pd.to_datetime(
                self.hr_weekly_report['from_date'], format="%Y-%m-%d", errors='coerce') 
            self.hr_weekly_report['to_date'] = pd.to_datetime(
                self.hr_weekly_report['to_date'], format="%Y-%m-%d", errors='coerce') 
            self.hr_weekly_report['create_date'] = pd.to_datetime(
                self.hr_weekly_report['create_date'], format="%Y-%m-%d %H:%M:%S", errors='coerce') 
            
    def load_last_year_al_cl_report(self):
        LIMIT_SIZE = 100
        lastmonth_start_str = (self.date_array[0] + datetime.timedelta(days=-31)).strftime('%Y-%m-%d')
        index_count = 0
        found_last_al_cl = False
        last_date_calculate = self.date_array[0].replace(day=1, month=1).strftime('%Y-%m-%d')
        for index, company_data in self.df_all_company.iterrows():
            
            company_sid = company_data['id']
            if (self.select_company_id>0) and (company_sid != self.select_company_id):
                continue
            # company_data = self.df_all_company[self.df_all_company['id'] == company_sid]
            is_auto_calculate_holiday_increase = company_data['AUTO-CALCULATE-HOLIDAY-WORKTIME-INCREASE']
            is_auto_calculate_al_deduction = company_data['AUTO-CALCULATE-AL-LAST-YEAR-DEDUCTION']
            night_holiday_increase_wage = company_data['NIGHT-HOLIDAY-WAGE']
            local_index = 0
            ids = []
            while (len(ids) == LIMIT_SIZE) or (local_index == 0):
                offset = local_index * LIMIT_SIZE
                
                ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'search', 
                        [["&",("company_id","=", company_sid),("date_calculate_leave", "=", last_date_calculate)]], 
                        {'offset': offset, 'limit': LIMIT_SIZE})
                if len(ids) == 0:
                    print(f"{index} - No last data content al last year")
                else:
                    print("found: ", len(ids))
                    list_al_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'read', [ids], 
                                                                {'fields': ['id','year', 'date_calculate_leave', 'employee_id', 'employee_code','department_id','standard_day',
                                                                    'workingday','date_sign','date_apply_leave','severance_day','seniority_leave', 'job_title', 'company_id',
                                                                    'family_leave','leave_increase_by_seniority_leave','leave_day','leave_year',
                                                                    'remaining_leave','january','february','march','april','may','june','july',
                                                                    'august','september','october','november','december','leave_used','remaining_leave_minute',
                                                                    'remaining_leave_day','note','file','employee_ho','part_time_company_id']})
                    if len(list_al_report ) >0:
                        found_last_al_cl = True
                    for al_report in list_al_report:
                        try:
                            company_name = al_report['company_id'][1]
                            al_report['company_name'] = company_name
                            company_id = company_sid
                            al_report['company_sid'] = company_id
                        except:
                            al_report['company_name'] = False
                            al_report['company_sid'] = False
                        try:
                            employee_name = al_report['employee_id'][1]
                            al_report['employee_name'] = employee_name
                        except:
                            al_report['employee_name'] = False   
                        try:
                            al_report['department_id'] = al_report['department_id'][0]
                        except:
                            al_report['department_id'] = False
                        al_report['NIGHT-HOLIDAY-WAGE'] = night_holiday_increase_wage
                        al_report['AUTO-CALCULATE-HOLIDAY'] = (is_auto_calculate_holiday_increase == True)
                        al_report['AUTO-CALCULATE-AL-LAST-YEAR-DEDUCTION'] = (is_auto_calculate_al_deduction == True)
                    if index_count == 0:                                                         
                        self.df_last_year_al_report = pd.DataFrame.from_dict(list_al_report)
                        print("load fist")
                    else:
                        self.df_last_year_al_report = pd.concat([self.df_last_year_al_report, pd.DataFrame.from_dict(list_al_report)], ignore_index=True)

                    index_count = index_count + 1
                local_index = local_index + 1
        if not found_last_al_cl:
            self.df_last_year_al_report = pd.DataFrame(columns = ['id','year', 'date_calculate_leave', 'employee_id', 'employee_code','department_id','standard_day',
                                                            'workingday','date_sign','date_apply_leave','severance_day','seniority_leave', 'job_title', 
                                                            'family_leave','leave_increase_by_seniority_leave','leave_day','leave_year',
                                                            'remaining_leave','january','february','march','april','may','june','july',
                                                            'august','september','october','november','december','leave_used','remaining_leave_minute',
                                                            'remaining_leave_day','note','file','employee_ho','part_time_company_id',
                                                            'company_name', 'company_sid', 'employee_name',])

        self.df_last_year_al_report['date_calculate_leave'] = pd.to_datetime(self.df_last_year_al_report ['date_calculate_leave'], format="%Y-%m-%d",errors='coerce')
        # self.df_last_al_report = self.df_last_al_report.to_excel('conloncon.xlsx')
        group_max_leave = self.df_cl_report.groupby('company_name').agg({'date_calculate': ['max'], 'company_sid':['max']})
        index_count = 0
        found_last_al_cl = False
        
        for index, company_data in self.df_all_company.iterrows():
            company_sid = company_data['id']
            if (self.select_company_id>0) and (company_sid != self.select_company_id):
                continue
            # company_data = self.df_all_company[self.df_all_company['id'] == company_sid]
            
            is_auto_calculate_holiday_increase = company_data['AUTO-CALCULATE-HOLIDAY-WORKTIME-INCREASE']
            night_holiday_increase_wage = company_data['NIGHT-HOLIDAY-WAGE']
            # last_date_calculate = last_date_calculate
            print(f"last year CLLLLLLLLLLL{index}-{company_sid}-LLLLLLLLLLLLLLLLLLLLLLL{last_date_calculate}")
            local_index = 0
            ids = []
            while (len(ids) == LIMIT_SIZE) or (local_index == 0):
                offset = local_index * LIMIT_SIZE
                ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'search', 
                        [["&", ("company_id","=", company_sid), ("date_calculate", "=", last_date_calculate)]], 
                        {'offset': offset, 'limit': LIMIT_SIZE})
                if len(ids) == 0:
                    print(f"{index}   No last data content")
                else:
                    print('found last : ', len(ids))
                    list_cl_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'read', [ids], 
                                                                {'fields': ['id','year', 'date_calculate', 'employee_id', 'tier','employee_code','department_id','job_title',
                                                                    'workingday','date_sign','contract_type_id','severance_day', 'company_id',
                                                                    'increase_probationary_1','increase_official_1','used_probationary_1','used_official_1','overtime_probationary_1', 'overtime_official_1',
                                                                    'increase_probationary_2','increase_official_2','used_probationary_2','used_official_2','overtime_probationary_2', 'overtime_official_2',
                                                                    'increase_probationary_3','increase_official_3','used_probationary_3','used_official_3','overtime_probationary_3', 'overtime_official_3',
                                                                    'increase_probationary_4','increase_official_4','used_probationary_4','used_official_4','overtime_probationary_4', 'overtime_official_4',
                                                                    'increase_probationary_5','increase_official_5','used_probationary_5','used_official_5','overtime_probationary_5', 'overtime_official_5',
                                                                    'increase_probationary_6','increase_official_6','used_probationary_6','used_official_6','overtime_probationary_6', 'overtime_official_6',
                                                                    'increase_probationary_7','increase_official_7','used_probationary_7','used_official_7','overtime_probationary_7', 'overtime_official_7',
                                                                    'increase_probationary_8','increase_official_8','used_probationary_8','used_official_8','overtime_probationary_8', 'overtime_official_8',
                                                                    'increase_probationary_9','increase_official_9','used_probationary_9','used_official_9','overtime_probationary_9', 'overtime_official_9',
                                                                    'increase_probationary_10','increase_official_10','used_probationary_10','used_official_10','overtime_probationary_10', 'overtime_official_10',
                                                                    'increase_probationary_11','increase_official_11','used_probationary_11','used_official_11','overtime_probationary_11', 'overtime_official_11',
                                                                    'increase_probationary_12','increase_official_12','used_probationary_12','used_official_12','overtime_probationary_12', 'overtime_official_12',
                                                                    'remaining_total_day', 'remaining_probationary_minute', 'remaining_official_minute', 'remaining_total_minute'
                                                                    # 'total_increase_probationary','total_increase_official','total_increase_tv_ovt',
                                                                    # 'total_increase_ct_ovt','total_used_probationary'
                                                                    ]})
                    if len(list_cl_report) >0:
                        found_last_al_cl = True                                             
                    for cl_report in list_cl_report:
                        try:
                            company_name = cl_report['company_id'][1]
                            cl_report['company_name'] = company_name
                            company_id = company_sid
                            cl_report['company_sid'] = company_id
                        except:
                            cl_report['company_name'] = False
                            cl_report['company_sid'] = False
                        try:
                            employee_name = cl_report['employee_id'][1]
                            cl_report['employee_name'] = employee_name
                        except:
                            cl_report['employee_name'] = False
                        try:
                            cl_report['department_id'] = cl_report['department_id'][0]
                        except:
                            cl_report['department_id'] = False
                        cl_report['NIGHT-HOLIDAY-WAGE'] = night_holiday_increase_wage
                        cl_report['AUTO-CALCULATE-HOLIDAY'] = (is_auto_calculate_holiday_increase == True)
                    if index_count == 0:
                        self.df_last_year_cl_report = pd.DataFrame.from_dict(list_cl_report)
                    else:
                        self.df_last_year_cl_report = pd.concat([self.df_last_year_cl_report, pd.DataFrame.from_dict(list_cl_report)], ignore_index=True)
                    index_count = index_count + 1
                local_index = 1 + local_index
            # self.df_last_year_cl_report.to_excel(f'dayladau{index_count}.xlsx')
        
        if not found_last_al_cl:
            self.df_last_year_cl_report = pd.DataFrame(columns = ['id','year', 'date_calculate', 'employee_id', 'tier','employee_code','department_id','job_title',
                                                            'workingday','date_sign','contract_type_id','severance_day', 'company_id',
                                                            'increase_probationary_1','increase_official_1','used_probationary_1','used_official_1','overtime_probationary_1', 'overtime_official_1',
                                                            'increase_probationary_2','increase_official_2','used_probationary_2','used_official_2','overtime_probationary_2', 'overtime_official_2',
                                                            'increase_probationary_3','increase_official_3','used_probationary_3','used_official_3','overtime_probationary_3', 'overtime_official_3',
                                                            'increase_probationary_4','increase_official_4','used_probationary_4','used_official_4','overtime_probationary_4', 'overtime_official_4',
                                                            'increase_probationary_5','increase_official_5','used_probationary_5','used_official_5','overtime_probationary_5', 'overtime_official_5',
                                                            'increase_probationary_6','increase_official_6','used_probationary_6','used_official_6','overtime_probationary_6', 'overtime_official_6',
                                                            'increase_probationary_7','increase_official_7','used_probationary_7','used_official_7','overtime_probationary_7', 'overtime_official_7',
                                                            'increase_probationary_8','increase_official_8','used_probationary_8','used_official_8','overtime_probationary_8', 'overtime_official_8',
                                                            'increase_probationary_9','increase_official_9','used_probationary_9','used_official_9','overtime_probationary_9', 'overtime_official_9',
                                                            'increase_probationary_10','increase_official_10','used_probationary_10','used_official_10','overtime_probationary_10', 'overtime_official_10',
                                                            'increase_probationary_11','increase_official_11','used_probationary_11','used_official_11','overtime_probationary_11', 'overtime_official_11',
                                                            'increase_probationary_12','increase_official_12','used_probationary_12','used_official_12','overtime_probationary_12', 'overtime_official_12',
                                                            'company_name', 'company_sid', 'employee_name', 'AUTO-CALCULATE-HOLIDAY', 'NIGHT-HOLIDAY-WAGE', 'remaining_probationary_minute', 'remaining_official_minute'
                                                            # 'total_increase_probationary','total_increase_official','total_increase_tv_ovt',
                                                            # 'total_increase_ct_ovt','total_used_probationary'
                                                            ])
        self.df_last_year_cl_report['date_calculate'] = pd.to_datetime(self.df_last_year_cl_report ['date_calculate'], format="%Y-%m-%d",errors='coerce')


    def load_last_al_cl_report(self):
        group_max_leave = self.df_al_report.groupby('company_name').agg({'date_calculate_leave': ['max'], 'company_sid':['max']})
        lastmonth_start_str = (self.date_array[0] + datetime.timedelta(days=-31)).strftime('%Y-%m-%d')
        index_count = 0
        found_last_al_cl = False
        LIMIT_SIZE = 200
        for index, data in group_max_leave.iterrows():
            date_calculate_leave_max = data[('date_calculate_leave', 'max')]
            company_sid = data[('company_sid', 'max')]
            if (self.select_company_id>0) and (company_sid != self.select_company_id):
                continue
            date_calculate_ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'search_read', 
                [[("company_id","=", company_sid), ("date_calculate_leave","<", date_calculate_leave_max.strftime('%Y-%m-%d'))]], 
                            {'order': 'date_calculate_leave desc', 'limit':1})
            if len(date_calculate_ids) == 0:
                print(f"{index} - No last data")
                continue
            last_date_calculate = date_calculate_ids[0]['date_calculate_leave']
            ids = []
            local_index = 0
            while (len(ids) == LIMIT_SIZE) or (local_index == 0):
                offset = local_index * LIMIT_SIZE
                
                ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'search', 
                        [["&",("company_id","=", company_sid),("date_calculate_leave", "=", last_date_calculate)]], 
                        {'offset': offset, 'limit': LIMIT_SIZE})
                
                if len(ids) == 0:
                    print(f"{index} - No last data content")
                    
                else:
                    list_al_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'read', [ids], 
                                                                {'fields': ['id','year', 'date_calculate_leave', 'employee_id', 'employee_code','department_id','standard_day',
                                                                    'workingday','date_sign','date_apply_leave','severance_day','seniority_leave', 'job_title', 'company_id',
                                                                    'family_leave','leave_increase_by_seniority_leave','leave_day','leave_year',
                                                                    'remaining_leave','january','february','march','april','may','june','july',
                                                                    'august','september','october','november','december','leave_used','remaining_leave_minute',
                                                                    'remaining_leave_day','note','file','employee_ho','part_time_company_id']})
                    if len(list_al_report ) >0:
                        found_last_al_cl = True
                    for al_report in list_al_report:
                        try:
                            company_name = al_report['company_id'][1]
                            al_report['company_name'] = company_name
                            company_id = company_sid
                            al_report['company_sid'] = company_id
                        except:
                            al_report['company_name'] = False
                            al_report['company_sid'] = False
                        try:
                            employee_name = al_report['employee_id'][1]
                            al_report['employee_name'] = employee_name
                        except:
                            al_report['employee_name'] = False   
                        try:
                            al_report['department_id'] = al_report['department_id'][0]
                        except:
                            al_report['department_id'] = False
                    if index_count == 0:                                                         
                        self.df_last_al_report = pd.DataFrame.from_dict(list_al_report)
                        print("load fist")
                    else:
                        self.df_last_al_report = pd.concat([self.df_last_al_report, pd.DataFrame.from_dict(list_al_report)], ignore_index=True)

                    index_count = index_count + 1
                local_index = local_index + 1
        if not found_last_al_cl:
            self.df_last_al_report = pd.DataFrame(columns = ['id','year', 'date_calculate_leave', 'employee_id', 'employee_code','department_id','standard_day',
                                                            'workingday','date_sign','date_apply_leave','severance_day','seniority_leave', 'job_title', 
                                                            'family_leave','leave_increase_by_seniority_leave','leave_day','leave_year',
                                                            'remaining_leave','january','february','march','april','may','june','july',
                                                            'august','september','october','november','december','leave_used','remaining_leave_minute',
                                                            'remaining_leave_day','note','file','employee_ho','part_time_company_id',
                                                            'company_name', 'company_sid', 'employee_name',])

        self.df_last_al_report['date_calculate_leave'] = pd.to_datetime(self.df_last_al_report ['date_calculate_leave'], format="%Y-%m-%d",errors='coerce')
        # self.df_last_al_report = self.df_last_al_report.to_excel('conloncon.xlsx')
        group_max_leave = self.df_cl_report.groupby('company_name').agg({'date_calculate': ['max'], 'company_sid':['max']})
        index_count = 0
        found_last_al_cl = False
        for index, data in group_max_leave.iterrows():
            # print('Gooooooooooooooo:, ', index)
            date_calculate_max = data[('date_calculate', 'max')]
            company_sid = data[('company_sid', 'max')]
            if (self.select_company_id>0) and (company_sid != self.select_company_id):
                continue
            date_calculate_ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'search_read', 
                [[("company_id","=", company_sid), ('date_calculate','<',date_calculate_max.strftime('%Y-%m-%d'))]], 
                        {'order': 'date_calculate desc', 'limit':1})
            if len(date_calculate_ids) == 0:
                # print(f"{index} - No last data")
                continue
            
            last_date_calculate = date_calculate_ids[0]['date_calculate']
            print(f"CLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLLL{index}-{company_sid}-LLLLLLLLLLLLLLLLLLLLLLL{last_date_calculate}")
            local_index = 0
            ids = []
            while (len(ids) == LIMIT_SIZE) or (local_index == 0):
                offset = local_index * LIMIT_SIZE
                ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'search', 
                        [["&","&", ("company_id","=", company_sid),("date_calculate", "=", last_date_calculate), "|",
                            ("severance_day",">",lastmonth_start_str),(("severance_day","=",False))]], 
                        {'offset': offset, 'limit': LIMIT_SIZE})
                if len(ids) == 0:
                    print(f"CL last - No last data content")
                else:
                    list_cl_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'read', [ids], 
                                                                {'fields': ['id','year', 'date_calculate', 'employee_id', 'tier','employee_code','department_id','job_title',
                                                                    'workingday','date_sign','contract_type_id','severance_day', 'company_id',
                                                                    'increase_probationary_1','increase_official_1','used_probationary_1','used_official_1','overtime_probationary_1', 'overtime_official_1',
                                                                    'increase_probationary_2','increase_official_2','used_probationary_2','used_official_2','overtime_probationary_2', 'overtime_official_2',
                                                                    'increase_probationary_3','increase_official_3','used_probationary_3','used_official_3','overtime_probationary_3', 'overtime_official_3',
                                                                    'increase_probationary_4','increase_official_4','used_probationary_4','used_official_4','overtime_probationary_4', 'overtime_official_4',
                                                                    'increase_probationary_5','increase_official_5','used_probationary_5','used_official_5','overtime_probationary_5', 'overtime_official_5',
                                                                    'increase_probationary_6','increase_official_6','used_probationary_6','used_official_6','overtime_probationary_6', 'overtime_official_6',
                                                                    'increase_probationary_7','increase_official_7','used_probationary_7','used_official_7','overtime_probationary_7', 'overtime_official_7',
                                                                    'increase_probationary_8','increase_official_8','used_probationary_8','used_official_8','overtime_probationary_8', 'overtime_official_8',
                                                                    'increase_probationary_9','increase_official_9','used_probationary_9','used_official_9','overtime_probationary_9', 'overtime_official_9',
                                                                    'increase_probationary_10','increase_official_10','used_probationary_10','used_official_10','overtime_probationary_10', 'overtime_official_10',
                                                                    'increase_probationary_11','increase_official_11','used_probationary_11','used_official_11','overtime_probationary_11', 'overtime_official_11',
                                                                    'increase_probationary_12','increase_official_12','used_probationary_12','used_official_12','overtime_probationary_12', 'overtime_official_12',
                                                                    'remaining_total_day', 'remaining_probationary_minute', 'remaining_official_minute', 'remaining_total_minute'
                                                                    # 'total_increase_probationary','total_increase_official','total_increase_tv_ovt',
                                                                    # 'total_increase_ct_ovt','total_used_probationary'
                                                                    ]})
                    if len(list_cl_report) >0:
                        found_last_al_cl = True                                             
                    for cl_report in list_cl_report:
                        try:
                            company_name = cl_report['company_id'][1]
                            cl_report['company_name'] = company_name
                            company_id = company_sid
                            cl_report['company_sid'] = company_id
                        except:
                            cl_report['company_name'] = False
                            cl_report['company_sid'] = False
                        try:
                            employee_name = cl_report['employee_id'][1]
                            cl_report['employee_name'] = employee_name
                        except:
                            cl_report['employee_name'] = False
                        try:
                            cl_report['department_id'] = cl_report['department_id'][0]
                        except:
                            cl_report['department_id'] = False
                    if index_count == 0:
                        self.df_last_cl_report = pd.DataFrame.from_dict(list_cl_report)
                    else:
                        self.df_last_cl_report = pd.concat([self.df_last_cl_report, pd.DataFrame.from_dict(list_cl_report)], ignore_index=True)
                    index_count = index_count + 1
                local_index = local_index + 1
            # self.df_last_cl_report.to_excel(f'dayladau{index_count}.xlsx')
        if not found_last_al_cl:
            self.df_last_cl_report = pd.DataFrame(columns = ['id','year', 'date_calculate', 'employee_id', 'tier','employee_code','department_id','job_title',
                                                            'workingday','date_sign','contract_type_id','severance_day', 'company_id',
                                                            'increase_probationary_1','increase_official_1','used_probationary_1','used_official_1','overtime_probationary_1', 'overtime_official_1',
                                                            'increase_probationary_2','increase_official_2','used_probationary_2','used_official_2','overtime_probationary_2', 'overtime_official_2',
                                                            'increase_probationary_3','increase_official_3','used_probationary_3','used_official_3','overtime_probationary_3', 'overtime_official_3',
                                                            'increase_probationary_4','increase_official_4','used_probationary_4','used_official_4','overtime_probationary_4', 'overtime_official_4',
                                                            'increase_probationary_5','increase_official_5','used_probationary_5','used_official_5','overtime_probationary_5', 'overtime_official_5',
                                                            'increase_probationary_6','increase_official_6','used_probationary_6','used_official_6','overtime_probationary_6', 'overtime_official_6',
                                                            'increase_probationary_7','increase_official_7','used_probationary_7','used_official_7','overtime_probationary_7', 'overtime_official_7',
                                                            'increase_probationary_8','increase_official_8','used_probationary_8','used_official_8','overtime_probationary_8', 'overtime_official_8',
                                                            'increase_probationary_9','increase_official_9','used_probationary_9','used_official_9','overtime_probationary_9', 'overtime_official_9',
                                                            'increase_probationary_10','increase_official_10','used_probationary_10','used_official_10','overtime_probationary_10', 'overtime_official_10',
                                                            'increase_probationary_11','increase_official_11','used_probationary_11','used_official_11','overtime_probationary_11', 'overtime_official_11',
                                                            'increase_probationary_12','increase_official_12','used_probationary_12','used_official_12','overtime_probationary_12', 'overtime_official_12',
                                                            'company_name', 'company_sid', 'employee_name', 'remaining_official_minute',
                                                            # 'total_increase_probationary','total_increase_official','total_increase_tv_ovt',
                                                            # 'total_increase_ct_ovt','total_used_probationary'
                                                            ])
        self.df_last_cl_report['date_calculate'] = pd.to_datetime(self.df_last_cl_report ['date_calculate'], format="%Y-%m-%d",errors='coerce')

    def load_al_cl_report(self):
        date_calculate_leave_max = self.date_array[len(self.date_array)-1]
        lastmonth_start_str = (self.date_array[0] + datetime.timedelta(days=-31)).strftime('%Y-%m-%d')
        index_count = 0
        LIMIT_SIZE = 200
        df_al_report = None
        df_cl_report = None
        df_al_report = pd.DataFrame(columns = ['id','year', 'date_calculate_leave', 'employee_id', 'company_id','employee_code','department_id','standard_day',
                                                            'workingday','date_sign','date_apply_leave','severance_day','seniority_leave', 'job_title', 
                                                            'family_leave','leave_increase_by_seniority_leave','leave_day','leave_year',
                                                            'remaining_leave','january','february','march','april','may','june','july',
                                                            'august','september','october','november','december','leave_used','remaining_leave_minute','remaining_total_minute',
                                                            'remaining_leave_day','note','file','employee_ho','part_time_company_id','company_name', 'company_sid', 'employee_name',
                                                            'AUTO-CALCULATE-HOLIDAY', 'NIGHT-HOLIDAY-WAGE'])
        
        df_cl_report = pd.DataFrame(columns = ['id','year', 'date_calculate', 'employee_id', 'tier','employee_code','department_id','job_title',
                                                            'workingday','date_sign','contract_type_id','severance_day',
                                                            'increase_probationary_1','increase_official_1','used_probationary_1','used_official_1','overtime_probationary_1', 'overtime_official_1',
                                                            'increase_probationary_2','increase_official_2','used_probationary_2','used_official_2','overtime_probationary_2', 'overtime_official_2',
                                                            'increase_probationary_3','increase_official_3','used_probationary_3','used_official_3','overtime_probationary_3', 'overtime_official_3',
                                                            'increase_probationary_4','increase_official_4','used_probationary_4','used_official_4','overtime_probationary_4', 'overtime_official_4',
                                                            'increase_probationary_5','increase_official_5','used_probationary_5','used_official_5','overtime_probationary_5', 'overtime_official_5',
                                                            'increase_probationary_6','increase_official_6','used_probationary_6','used_official_6','overtime_probationary_6', 'overtime_official_6',
                                                            'increase_probationary_7','increase_official_7','used_probationary_7','used_official_7','overtime_probationary_7', 'overtime_official_7',
                                                            'increase_probationary_8','increase_official_8','used_probationary_8','used_official_8','overtime_probationary_8', 'overtime_official_8',
                                                            'increase_probationary_9','increase_official_9','used_probationary_9','used_official_9','overtime_probationary_9', 'overtime_official_9',
                                                            'increase_probationary_10','increase_official_10','used_probationary_10','used_official_10','overtime_probationary_10', 'overtime_official_10',
                                                            'increase_probationary_11','increase_official_11','used_probationary_11','used_official_11','overtime_probationary_11', 'overtime_official_11',
                                                            'increase_probationary_12','increase_official_12','used_probationary_12','used_official_12','overtime_probationary_12', 'overtime_official_12',
                                                            'company_name','company_sid', 'employee_name', 'AUTO-CALCULATE-HOLIDAY', 'NIGHT-HOLIDAY-WAGE', 'remaining_probationary_minute', 'remaining_official_minute'
                                                            # 'total_increase_probationary','total_increase_official','total_increase_tv_ovt',
                                                            # 'total_increase_ct_ovt','total_used_probationary'
                                                            ])
        for index, data in self.df_all_company.iterrows():
            # date_calculate_leave_max = data[('date_calculate_leave', 'max')]
            company_sid = data[('id')]
            if (self.select_company_id>0) and (company_sid != self.select_company_id):
                continue
            night_holiday_wage = data['NIGHT-HOLIDAY-WAGE']
            date_calculate_ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'search_read', 
                [[("company_id","=", company_sid), \
                  ("date_calculate_leave",">", date_calculate_leave_max.strftime('%Y-%m-%d')), \
                  ("date_calculate_leave","<", (self.date_array[0] + relativedelta(months = 2)).strftime('%Y-%m-%d'))  ]], 
                            {'order': 'date_calculate_leave desc', 'limit':1})
            if len(date_calculate_ids) == 0:
                print(f"{index} - No last data")
                continue
            last_date_calculate = date_calculate_ids[0]['date_calculate_leave']
            local_index = 0
            ids = []
            while (len(ids) == LIMIT_SIZE) or (local_index == 0):
                offset = local_index * LIMIT_SIZE
                ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'search', 
                        [["&",("company_id","=", company_sid),("date_calculate_leave", "=", last_date_calculate)]], 
                        {'offset': offset, 'limit': LIMIT_SIZE})
                if len(ids) == 0:
                    print(f"{index} - No last data content")
                else:
                    list_al_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'read', [ids], 
                                                                {'fields': ['id','year', 'date_calculate_leave', 'employee_id', 'company_id','employee_code','department_id','standard_day',
                                                                    'workingday','date_sign','date_apply_leave','severance_day','seniority_leave', 'job_title', 
                                                                    'family_leave','leave_increase_by_seniority_leave','leave_day','leave_year',
                                                                    'remaining_leave','january','february','march','april','may','june','july',
                                                                    'august','september','october','november','december','leave_used','remaining_leave_minute',
                                                                    'remaining_leave_day','note','file','employee_ho','part_time_company_id']})
                    
                    for al_report in list_al_report:
                        try:
                            company_name = data['name']
                            al_report['company_name'] = company_name
                            company_id =  data['id']
                            al_report['company_sid'] = company_id
                        except:
                            al_report['company_name'] = False
                            al_report['company_sid'] = False
                        try:
                            employee_name = al_report['employee_id'][1]
                            al_report['employee_name'] = employee_name
                        except:
                            al_report['employee_name'] = False   
                        try:
                            al_report['department_id'] = al_report['department_id'][0]
                        except:
                            al_report['department_id'] = False   
                        al_report['NIGHT-HOLIDAY-WAGE'] = night_holiday_wage
                        al_report['AUTO-CALCULATE-AL-LAST-YEAR-DEDUCTION'] = \
                                data['AUTO-CALCULATE-AL-LAST-YEAR-DEDUCTION'] == True
                        al_report['AUTO-CALCULATE-HOLIDAY'] = data['AUTO-CALCULATE-HOLIDAY-WORKTIME-INCREASE'] == True
                    if index_count == 0:                                                         
                        df_al_report = pd.DataFrame.from_dict(list_al_report)
                        print("load fist")
                    else:
                        df_al_report = pd.concat([df_al_report, pd.DataFrame.from_dict(list_al_report)], ignore_index=True)

                    index_count = index_count + 1
                local_index = local_index + 1

        # date_calculate_max = self.date_array[0]
        index_count = 0
        for index, data in self.df_all_company.iterrows():
            company_sid = data['id']
            if (self.select_company_id>0) and (company_sid != self.select_company_id):
                continue
            date_calculate_ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'search_read', 
                [[("company_id","=", company_sid), ('date_calculate','>',date_calculate_leave_max.strftime('%Y-%m-%d')), \
                   ('date_calculate', '<', (self.date_array[0] + relativedelta(months = 2)).strftime('%Y-%m-%d'))  ]], 
                        {'order': 'date_calculate desc', 'limit':1})
            if len(date_calculate_ids) == 0:
                print(f"{index} - No last data")
                continue
            
            print(f"CLLLLLLL{data['id']}LLLLLLLL{data['name']}LLLLLLLLLLLLLLLLLLLLLLLLLL")
            last_date_calculate = date_calculate_ids[0]['date_calculate']
            fields = ['id','year', 'date_calculate', 'employee_id', 'tier','employee_code','department_id','job_title',
                                                            'workingday','date_sign','contract_type_id','severance_day',
                                                            'increase_probationary_1','increase_official_1','used_probationary_1','used_official_1','overtime_probationary_1', 'overtime_official_1',
                                                            'increase_probationary_2','increase_official_2','used_probationary_2','used_official_2','overtime_probationary_2', 'overtime_official_2',
                                                            'increase_probationary_3','increase_official_3','used_probationary_3','used_official_3','overtime_probationary_3', 'overtime_official_3',
                                                            'increase_probationary_4','increase_official_4','used_probationary_4','used_official_4','overtime_probationary_4', 'overtime_official_4',
                                                            'increase_probationary_5','increase_official_5','used_probationary_5','used_official_5','overtime_probationary_5', 'overtime_official_5',
                                                            'increase_probationary_6','increase_official_6','used_probationary_6','used_official_6','overtime_probationary_6', 'overtime_official_6',
                                                            'increase_probationary_7','increase_official_7','used_probationary_7','used_official_7','overtime_probationary_7', 'overtime_official_7',
                                                            'increase_probationary_8','increase_official_8','used_probationary_8','used_official_8','overtime_probationary_8', 'overtime_official_8',
                                                            'increase_probationary_9','increase_official_9','used_probationary_9','used_official_9','overtime_probationary_9', 'overtime_official_9',
                                                            'increase_probationary_10','increase_official_10','used_probationary_10','used_official_10','overtime_probationary_10', 'overtime_official_10',
                                                            'increase_probationary_11','increase_official_11','used_probationary_11','used_official_11','overtime_probationary_11', 'overtime_official_11',
                                                            'increase_probationary_12','increase_official_12','used_probationary_12','used_official_12','overtime_probationary_12', 'overtime_official_12',
                                                            'remaining_total_day', 'remaining_probationary_minute', 'remaining_official_minute', 'remaining_total_minute'
                                                            # 'total_increase_probationary','total_increase_official','total_increase_tv_ovt',
                                                            # 'total_increase_ct_ovt','total_used_probationary'
                                                            ]
            
            
            local_index = 0
            ids = []
            while (len(ids) == LIMIT_SIZE) or (local_index == 0):
                offset = local_index * LIMIT_SIZE
                if (self.select_company_id>0) and (company_sid != self.select_company_id):
                    continue
                ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'search', 
                        [["&", ("company_id","=", company_sid),("date_calculate", "=", last_date_calculate)]], \
                        {'offset': offset, 'limit': LIMIT_SIZE})
                if len(ids) == 0:
                    print(f"{index} - No last data content")
                else:
                    list_cl_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'read', [ids], 
                                                                {'fields': fields})
                
                    for cl_report in list_cl_report:
                        try:
                            company_name = data['name']
                            cl_report['company_name'] = company_name
                            company_id =  data['id']
                            cl_report['company_sid'] = company_id
                        except:
                            cl_report['company_name'] = False
                            cl_report['company_sid'] = False
                        try:
                            employee_name = cl_report['employee_id'][1]
                            cl_report['employee_name'] = employee_name
                        except:
                            cl_report['employee_name'] = False
                        try:
                            cl_report['department_id'] = cl_report['department_id'][0]
                        except:
                            cl_report['department_id'] = False   
                        cl_report['NIGHT-HOLIDAY-WAGE'] = data['NIGHT-HOLIDAY-WAGE']
                        cl_report['AUTO-CALCULATE-HOLIDAY'] = data['AUTO-CALCULATE-HOLIDAY-WORKTIME-INCREASE']
                    if index_count == 0:
                        df_cl_report = pd.DataFrame.from_dict(list_cl_report)
                    else:
                        df_cl_report = pd.concat([df_cl_report, pd.DataFrame.from_dict(list_cl_report)], ignore_index=True)
                    df_cl_report['date_calculate'] = pd.to_datetime(df_cl_report ['date_calculate'], format="%Y-%m-%d",errors='coerce')
                    index_count = index_count + 1
                local_index = local_index + 1
   
                                                            
        self.df_al_report = df_al_report.drop_duplicates(
            subset=['date_calculate_leave', 'employee_code', 'company_name'], keep="last")
        
        df_duplicate = df_al_report[~df_al_report.index.isin(self.df_al_report.index)]
        ids = []
        for index, item in df_duplicate.iterrows():
            if not item['id'] in ids: 
                ids.append(int(item['id']))
        try:
            # self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'unlink', [ids])
            print("deleted: ", ids)
        except Exception as ex:
            logger.error(f"delete hr.al.report : {ex}")

        # self.df_al_report['date_calculate_leave'] = pd.to_datetime(self.df_al_report ['date_calculate_leave'], format="%Y-%m-%d",errors='coerce')
        self.df_al_report['date_calculate_leave'] = pd.to_datetime(self.df_al_report ['date_calculate_leave'], format="%Y-%m-%d",errors='coerce')
        self.df_al_report = self.df_al_report.merge(self.df_employees, left_on=['company_sid','employee_code'], right_on=['company_sid', 'code'], 
                                                            how='left', suffixes=('','_employee'))
        # self.df_al_report.to_excel('self.df_al_report2.xlsx')
        self.df_al_report_copy = self.df_al_report.copy()
        # self.df_al_report.to_excel('conloncon.xlsx')
        # df_duplicate = self.df_cl_report[(self.df_cl_report.groupby(['date_calculate', 'employee_code']).transform('count')>1).values]
        self.df_cl_report = df_cl_report.drop_duplicates(
            subset=['date_calculate', 'employee_code', 'company_name'], keep="last")

        df_duplicate = df_cl_report[~df_cl_report.index.isin(self.df_cl_report.index)]
        ids = []
        for index, item in df_duplicate.iterrows():
            if not item['id'] in ids: 
                # self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'unlink', [[item['id']]])
                print(item['id'])
                ids.append(int(item['id']))
        try:
            # self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'unlink', [ids])
            print("deleted: ", ids)
        except Exception as ex:
            logger.error(f"unlink hr.cl.report: {ex}")

        self.df_cl_report['date_calculate'] = pd.to_datetime(self.df_cl_report ['date_calculate'], format="%Y-%m-%d",errors='coerce')
        self.df_cl_report = self.df_cl_report.merge(self.df_employees, left_on=['company_sid','employee_code'], right_on=['company_sid', 'code'], 
                                                            how='left', suffixes=('','_employee'))

        # self.df_cl_report.to_excel('self.df_cl_report2.xlsx')
        self.df_cl_report_copy = self.df_cl_report.copy()
        # self.df_cl_report.to_excel('list_cl_report.xlsx')
        

   
    def export_data_to_files(self):
        self.hr_weekly_report.to_excel(os.path.join(self.output_report_folder, "hr_weekly_report.xlsx"), sheet_name='Sheet1')
        self.df_kpi_weekly_report.to_excel(os.path.join(self.output_report_folder, "df_kpi_weekly_report.xlsx"), sheet_name='Sheet1')
        self.df_scheduling_ver[(self.df_scheduling_ver['date']>=self.date_array[0]) &
                                (self.df_scheduling_ver['date']<=self.date_array[len(self.date_array)-1])].to_excel(os.path.join(self.output_report_folder, 'attendance_report.xlsx'),sheet_name='Sheet1')
        self.df_hr_leave.to_excel(os.path.join(self.output_report_folder, "hr_leave.xlsx"), sheet_name='Sheet1')
        self.df_explanation_data.to_excel(os.path.join(self.output_report_folder, "explanation_data.xlsx"), sheet_name='Sheet1')
        self.df_al_report_copy[(self.df_al_report_copy['severance_day_employee'].isnull()) |
                                (self.df_al_report_copy['severance_day_employee']== False) | \
                                (self.df_al_report_copy['severance_day_employee'] >= self.date_array[0])] \
                                .to_excel(os.path.join(self.output_report_folder, "al_report.xlsx"), sheet_name='Sheet1')
                                
        self.df_al_report_copy[(self.df_al_report_copy['severance_day_employee'].notnull()) &
                                (self.df_al_report_copy['severance_day_employee']!= False)] \
                                .to_excel(os.path.join(self.output_report_folder, "al_report_severance.xlsx"), sheet_name='Sheet1')
                                
        self.df_cl_report_copy[(self.df_cl_report_copy['severance_day_employee'].isnull()) |
                                (self.df_cl_report_copy['severance_day_employee']== False) | \
                                (self.df_cl_report_copy['severance_day_employee'] >= self.date_array[0])] \
                                .to_excel(os.path.join(self.output_report_folder, "cl_report.xlsx"), sheet_name='Sheet1')
                                
        self.df_cl_report_copy[(self.df_cl_report_copy['severance_day_employee'].notnull()) &
                                (self.df_cl_report_copy['severance_day_employee']!= False)] \
                                .to_excel(os.path.join(self.output_report_folder, "cl_report_severance.xlsx"), sheet_name='Sheet1')
                                
        if self.is_update_al_cl:
            self.df_last_al_report.to_excel(os.path.join(self.output_report_folder, "al_last_report.xlsx"), sheet_name='Sheet1')
            self.df_last_cl_report.to_excel(os.path.join(self.output_report_folder, "cl_last_report.xlsx"), sheet_name='Sheet1')
            self.df_last_year_al_report.to_excel(os.path.join(self.output_report_folder, "al_last_year_report.xlsx"), sheet_name='Sheet1')
            self.df_last_year_cl_report.to_excel(os.path.join(self.output_report_folder, "cl_last_year_report.xlsx"), sheet_name='Sheet1')
        if (self.select_company_id == 0):
            self.df_all_company.to_excel(os.path.join(self.output_report_folder, "company.xlsx"), sheet_name='Sheet1')
        else:
            self.df_all_company[self.df_all_company['id']==self.select_company_id].to_excel \
                (os.path.join(self.output_report_folder, "company.xlsx"), sheet_name='Sheet1')
        self.df_resource_calendar_leaves.to_excel(os.path.join(self.output_report_folder, "resource_calendar_leaves.xlsx"), sheet_name='Sheet1')
        self.df_upload_report.to_excel(os.path.join(self.output_report_folder, "df_upload_report.xlsx"), sheet_name='Sheet1')
        self.df_contract.to_excel(os.path.join(self.output_report_folder, "df_contract.xlsx"), sheet_name='Sheet1')
        if self.is_load_df_old:
            self.df_old.to_excel(os.path.join(self.output_report_folder,'old.xlsx'), sheet_name='Sheet1')
    def download_attendance_trans(self):
        start_str = (self.date_array[0] + datetime.timedelta(days=-1)).strftime('%Y-%m-%d')
        end_str = (self.date_array[len(self.date_array)-1]+ datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        # self.df_old = pd.read_excel(self.input_attendence_file , index_col=None, header=[0,] ,sheet_name='Sheet1')
        # self.df_old['is_from_explanation'] = False
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'search', [["&",("day", ">=", start_str),("day", "<=", end_str)]], {'offset': 0, 'limit': 1000})
        list_attendance_trans  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'read', [ids], {'fields': ['id', 'name', 'day', 'time','in_out']})
        self.df_old = pd.DataFrame.from_dict(list_attendance_trans)
        # self.df_old['shift_name'] = self.df_old['shift_name'].str.strip()
        index = 1
        while len(ids) == 1000:
            ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'search', [["&",("day", ">=", start_str),("day", "<=", end_str)]], {'offset': index * len(ids), 'limit': 1000})
            list_attendance_trans  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'read', [ids], {'fields': ['id', 'name', 'day', 'time','in_out']})
            self.df_old = pd.concat([self.df_old, pd.DataFrame.from_dict(list_attendance_trans)], ignore_index=True)

            index = index + 1
            print(f'load {index}- {len(ids)}')
        # self.df_old = self.df_old[self.df_old['time'].apply(lambda x: type(x) == datetime.datetime)]
        self.df_old ['Gio'] = self.df_old ['time']
        self.df_old ['time'] = pd.to_datetime(self.df_old ['time'], format="%Y-%m-%d %H:%M:%S",errors='coerce')
        # self.df_old ['time'] += pd.Timedelta(hours=7)
        self.df_old = self.df_old.dropna(subset=['time'])
        self.df_old.rename(columns = {'name':'ID', 'time':'Giờ', 'day':'Ngày'}, inplace = True)
        try:
            self.df_old.to_excel('old.xlsx', sheet_name='Sheet1')
        except Exception as ex:
            print(ex)
    def remove_employee_data(self, row):
        print(f"REMOVE {row['code']} - {row['severance_day']}")
        start_str = self.date_array[0].strftime('%Y-%m-%d')

        # self.df_old = pd.read_excel(self.input_attendence_file , index_col=None, header=[0,] ,sheet_name='Sheet1')
        # self.df_old['is_from_explanation'] = False
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [["&", "&",
                                                                                                           ('company', '=', self.company_name),
                                                                                                           ("date", ">=", start_str),
                                                                                                           # ('active','=',True),
                                                                                                           ("employee_code", "=", row['code'])]], {})
        print('IDS: ', ids)
        # if len(ids) > 0:
        #     try:
        #         self.models.execute_kw(
        #             self.db, self.uid, self.password, 'hr.apec.attendance.report', 'unlink', [ids])
        #     except Exception as ex:
        #         logger.error(f"Unlink attence report: {ex}")

    def remove_employee_finished(self, employee_df):
        employee_df.apply(lambda row: self.remove_employee_data(row), axis=1)

    def update_report(self, data):
        month = data['month']
        year = data['year']
        # self.company_id
        summary_report_url = data['report_summary'].replace(
            'https://www.dropbox.com', 'https://dl.dropboxusercontent.com')
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'search', [["&", "&", "&", ("template", "=", '1'),
                                                                                                           ("month", "=", month),
                                                                                                           ("year", "=", year),
                                                                                                           ("company_id", "=", self.company_id)]], {})
        if len(ids) > 0:
            update_data = {'url': summary_report_url,
                           'name': f'Du lieu thang {month}-{year}'}
            # self.models.execute_kw(self.db, self.uid, self.password,
            #                        'hr.upload.attendance', 'write', [ids, update_data])
        else:
            update_data = {
                'template': '1',
                'month': month, 'year': year, 'company_id':  self.company_id,
                'url': summary_report_url, 'name': f'Du lieu thang {month}-{year}'
            }
            # try:
            #     # id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'create',
            #     #                                   [update_data])
            # except Exception as ex:
            #     logger.error(f"Create upload attendance: {ex}")

        # detail_report_url: Cham cong dem so lan
        detail_report_url = data['report_detail'].replace(
            'https://www.dropbox.com', 'https://dl.dropboxusercontent.com')
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'search', [["&", "&",
                                                                                                           ("template", "=", '2'),
                                                                                                           ("month", "=", month),
                                                                                                           ("year", "=", year),
                                                                                                           ("company_id", "=", self.company_id)]], {})
        if len(ids) > 0:
            update_data = {'url': detail_report_url,
                           'name': f'{self.company_name} - Cham cong dem so lan {month}-{year}'}
            # try:
            #     self.models.execute_kw(self.db, self.uid, self.password,
            #                        'hr.upload.attendance', 'write', [ids, update_data])
            # except Exception as ex:
            #     logger.error(f"write attendance: {ex}")
        else:
            update_data = {
                'template': '2', 'month': month, 'year': year, 'company_id':  self.company_id,
                'url': detail_report_url, 'name': f'{self.company_name} - Cham cong dem so lan {month}-{year}'
            }
            # try:
            #     id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'create',
            #                                       [update_data])
            # except Exception as ex:
            #     logger.error(f"create attendance: {ex}")

        # Ca dem ca gay
        report_night_split = data['report_night_split'].replace(
            'https://www.dropbox.com', 'https://dl.dropboxusercontent.com')
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'search', [["&", "&", "&",
                                                                                                           ("template", "=", '4'),
                                                                                                           ("month", "=", month),
                                                                                                           ("year", "=", year),
                                                                                                           ("company_id", "=", self.company_id)]], {})
        if len(ids) > 0:
            update_data = {'url': report_night_split,
                           'name': f'{self.company_name} - Ca dem ca gay {month}-{year}'}
            # try:
            #     self.models.execute_kw(self.db, self.uid, self.password,
            #                        'hr.upload.attendance', 'write', [ids, update_data])
            # except Exception as ex:
            #     logger.error(f"write ca dem ca gay attendance: {ex}")
        else:
            update_data = {
                'template': '4', 'month': month, 'year': year, 'company_id':  self.company_id,
                'url': report_night_split, 'name': f'{self.company_name} - Ca dem ca gay {month}-{year}'
            }
            # try:
            #     id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'create',
            #                                       [update_data])
            # except Exception as ex:
            #     logger.error(f"create ca dem ca gay attendance: {ex}")

        # Ca dem ca gay
        report_5m = data['report_5m'].replace(
            'https://www.dropbox.com', 'https://dl.dropboxusercontent.com')
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'search', [["&", "&", "&",
                                                                                                           ("template", "=", '3'),
                                                                                                           ("month", "=", month),
                                                                                                           ("year", "=", year),
                                                                                                           ("company_id", "=", self.company_id)]], {})
        if len(ids) > 0:
            update_data = {
                'url': report_5m, 'name': f'{self.company_name} - Bao cao di som 5phut {month}-{year}'}
            # try:
            #     self.models.execute_kw(self.db, self.uid, self.password,
            #                        'hr.upload.attendance', 'write', [ids, update_data])
            # except Exception as ex:
            #     logger.error(f"write dimuon attendance: {ex}")
        else:
            update_data = {
                'template': '3', 'month': month, 'year': year, 'company_id':  self.company_id,
                'url': report_5m, 'name': f'{self.company_name} - Bao cao di som 5phut {month}-{year}'
            }
            # try:
            #     id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'create',
            #                                       [update_data])
            # except Exception as ex:
            #     logger.error(f"create dimuon attendance: {ex}")
        # al CL REPORT
        report_al_cl = data['report_al_cl'].replace(
            'https://www.dropbox.com', 'https://dl.dropboxusercontent.com')
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'search', [["&", "&", "&",
                                                                                                           ("template", "=", '5'),
                                                                                                           ("month", "=", month),
                                                                                                           ("year", "=", year),
                                                                                                           ("company_id", "=", self.company_id)]], {})
        if len(ids) > 0:
            update_data = {
                'url': report_al_cl, 'name': f'{self.company_name} - BAO CAO CONG PHEP {month}-{year}'}
            # try:
            #     self.models.execute_kw(self.db, self.uid, self.password,
            #                        'hr.upload.attendance', 'write', [ids, update_data])
            # except Exception as ex:
            #     logger.error(f"write BAO CAO CONG PHEP: {ex}")
        else:
            update_data = {
                'template': '5', 'month': month, 'year': year, 'company_id':  self.company_id,
                'url': report_al_cl, 'name': f'{self.company_name} - BAO CAO CONG PHEP {month}-{year}'
            }
            # try:
            #     id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'create',
            #                                       [update_data])
            # except Exception as ex:
            #     logger.error(f"create BAO CAO CONG PHEP: {ex}")

    def merge_department_data(self):
        # Check deparment
        department_ids = self.models.execute_kw(
            self.db, self.uid, self.password, 'hr.department', 'search', [[('active', '=', True)]])
        list_departments = self.models.execute_kw(self.db, self.uid, self.password, 'hr.department', 'read', [department_ids], {
                                                  'fields': ['id', 'name', 'total_employee', 'company_id', 'member_ids', 'time_keeping_count']})
        for department in list_departments:
            department['company_name'] = ''
            try:
                department['company_name'] =  department['company_id'][1]
            except:
                continue
            
        self.df_departments = pd.DataFrame.from_dict(list_departments)
        self.df_employees = self.df_employees.merge(self.df_departments, left_on='d_id',
                                                    right_on='id', how='left', suffixes=('', 'department'))
    def fix_overtime_error(self, row):
        try:
            if (row['request_date_from'] != False) and (row['overtime_from'] != False):
                if (row['request_date_from'].day != row['overtime_from'].day) and pd.notnull(row['overtime_from']):
                    
                    new_day = row['request_date_from'].day
                    new_month = row['request_date_from'].month
                    new_year = row['request_date_from'].year
                    update_data = {"overtime_from": ( row['overtime_from'].replace(day=new_day, month=new_month, year=new_year)).strftime("%Y-%m-%d %H:%M:%S")}
                    ids = [int(row['id'])]
                    # try:
                    #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.leave', 'write', [ids, update_data])
                    #     print('update overtime from: ', row)
                    # except Exception as ex:
                    #     logger.error(f"write update overtime: {ex}")
            if (row['request_date_to'] != False) and (row['overtime_to'] != False):
                if (row['request_date_to'].day != row['overtime_to'].day) and pd.notnull(row['overtime_to']):
                    new_day = row['request_date_to'].day
                    new_month = row['request_date_to'].month
                    new_year = row['request_date_to'].year
                    update_data = {"overtime_to": ( row['overtime_to'].replace(day=new_day, month=new_month, year=new_year)).strftime("%Y-%m-%d %H:%M:%S")}
                    ids = [int(row['id'])]
                    # try:
                    #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.leave', 'write', [ids, update_data])
                    #     print('update overtime request_date_to: ', row)
                    # except Exception as ex:
                    #     logger.error(f"write update overtime leave: {ex}")
        except Exception as ex:
            logger.error(f'update overtime ex: {ex}')

    def merge_attendance_data(self):
        print('sssssssssssssssssssssssssssssssss')
        start_date = self.date_array[0].strftime("%Y-%m-%d")
        end_date = self.date_array[len(self.date_array)-1].strftime("%Y-%m-%d")
        domain = ['&',
                                                    ("invalid_date", ">=", start_date),
                                                    ("invalid_date", "<=", end_date)]
        if (self.select_company_id>0):
            domain = ['&', '&' ,('company_id','=',self.select_company_id),
                                                    ("invalid_date", ">=", start_date),
                                                    ("invalid_date", "<=", end_date)]
        
        self.df_explanation_data = pd.DataFrame(columns=['id', 'employee_id', 'employee_code', 'department', 'company_id', 'company_name',
                                                                'position', 'invalid_date', 'invalid_type', 'shift_from', 'shift_to', 'shift_break', 'real_time_attendance_data',
                                                                'validated', 'reason', 'reason_str', 'remarks', 'validation_data', 'date_str', 'Lý do', 'Invalid Type'])
        local_index = 0
        LIMIT_SIZE = 50
        explanation_ids = []
        while (len(explanation_ids) == LIMIT_SIZE) or (local_index == 0):
            offset = local_index * LIMIT_SIZE
            explanation_ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.invalid.timesheet', 'search',
                                                    [domain], {'offset': offset, 'limit': LIMIT_SIZE})
            list_old_explanation = self.models.execute_kw(self.db, self.uid, self.password, 'hr.invalid.timesheet', 'read', [explanation_ids], 
                                        {'fields': ['id', 'employee_id', 'employee_code', 'department', 'company_id',
                                            'position', 'invalid_date', 'invalid_type', 'shift_from', 'shift_to',
                                            'shift_break', 'real_time_attendance_data', 'attendance_missing_from', 
                                            'attendance_missing_to',
                                            'validated', 'reason', 'remarks', 'validation_data']})
        
            if len(list_old_explanation) > 0:
                for explanation in list_old_explanation:
                    try:
                        company_name = explanation['company_id'][1]
                        explanation['company_name'] = company_name
                    except:
                        explanation['company_name'] = False
                    explanation['reason_str'] = 'work' if explanation['reason'] =='2' else 'personal' if explanation['reason'] == '1' else '_'
    
                if local_index == 0: 
                    self.df_explanation_data = pd.DataFrame.from_dict(
                        list_old_explanation)
                    
                else:
                    self.df_explanation_data = pd.concat([self.df_explanation_data, 
                        pd.DataFrame.from_dict(list_old_explanation)], ignore_index=True)  
                    
            local_index = local_index + 1   
            
        self.df_explanation_data['employee_code'] = self.df_explanation_data['employee_code'].str.strip()
        self.df_explanation_data['attendance_missing_from'] = pd.to_datetime(
            self.df_explanation_data['attendance_missing_from'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_explanation_data['attendance_missing_to'] = pd.to_datetime(
            self.df_explanation_data['attendance_missing_to'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        
        self.df_explanation_data = self.df_scheduling_ver.merge(self.df_explanation_data,
                                                                left_on=['employee_code', 'date_str'], right_on=['employee_code', 'invalid_date'], how='right', suffixes=('', '_explanation'))
        # self.df_explanation_data['reason_str'] = self.df_explanation_data['reason'].apply(lambda x: 'work' if x =='2' else 'personal' if x == '1' else '_')
        # hr leave
        
        domain = ["&", "&", "&",
                ("request_date_to", ">=", self.date_array[0].strftime('%Y-%m-%d 00:00:00')),
                ("request_date_from", "<=", self.date_array[len(
                self.date_array)-1].strftime('%Y-%m-%d 23:59:59')),
                ('active', '=', True),
                ("state", "=", "validate")
                #   ('employee_company_id', '=', self.company_id) 
                ]
        if (self.select_company_id>0):
            domain = ["&", "&", "&", "&",
                ("employee_company_id",'=', self.select_company_id),
                ("request_date_to", ">=", self.date_array[0].strftime('%Y-%m-%d 00:00:00')),
                ("request_date_from", "<=", self.date_array[len(
                self.date_array)-1].strftime('%Y-%m-%d 23:59:59')),
                ('active', '=', True),
                ("state", "=", "validate")
                #   ('employee_company_id', '=', self.company_id) 
                ]
            
            
        hr_leave_ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.leave', 'search', [domain])
        list_hr_leave = self.models.execute_kw(self.db, self.uid, self.password, 'hr.leave', 'read', [hr_leave_ids], {'fields': ['id', 'employee_id', 'employee_code', 'employee_company_id', 'active',
                                                                                                                       'holiday_status_id', 'minutes', 'time', 'state', 'request_date_from', 'request_date_to',
                                                                                                                       'attendance_missing_from', 'attendance_missing_to', 'reasons', 'for_reasons', 'convert_overtime',
                                                                                                                       'employee_company_id','multiplier_work_time', 'overtime_from', 'overtime_to']})
        # return
        if len(list_hr_leave) > 0:
            # print('hrleave toi')
            print(list_hr_leave)
            addition_leave = []
            for hr_leave in list_hr_leave:
                try:
                    company_sid = hr_leave['employee_company_id'][0]
                    hr_leave['company_sid'] = company_sid
                    company_name = hr_leave['employee_company_id'][1]
                    hr_leave['company_name'] = company_name
                except:
                    hr_leave['company_name'] = False
                try:
                    holiday_status_name = hr_leave['holiday_status_id'][1]
                    hr_leave['holiday_status_name'] = holiday_status_name
                except:
                    hr_leave['holiday_status_name'] = False
                kid_mod = False
                kid_mod_stage_1_start = 13
                kid_mod_stage_1_end = 13.5
                kid_mod_stage_2_start = 13.5
                kid_mod_stage_2_end = 14.0

                if 'con nhỏ' in f"{hr_leave['holiday_status_name']}":
                    kid_mod = True
                    if ('11:00:00' in f"{hr_leave['attendance_missing_from']}") and ('11:00:00' in f"{hr_leave['attendance_missing_to']}"):
                        kid_mod_stage_1_start = 11
                        kid_mod_stage_1_end = 11.5
                        kid_mod_stage_2_start = 11.5
                        kid_mod_stage_2_end = 12
                    elif ('12:00:00' in f"{hr_leave['attendance_missing_from']}") and ('12:00:00' in f"{hr_leave['attendance_missing_to']}"):
                        kid_mod_stage_1_start = 11.5
                        kid_mod_stage_1_end = 12
                        kid_mod_stage_2_start = 13
                        kid_mod_stage_2_end = 13.5
                    elif ('13:00:00' in f"{hr_leave['attendance_missing_from']}") and ('13:00:00' in f"{hr_leave['attendance_missing_to']}"):
                        kid_mod_stage_1_start = 13
                        kid_mod_stage_1_end = 13.5
                        kid_mod_stage_2_start = 13.5
                        kid_mod_stage_2_end = 14.0
                    elif ('16:00:00' in f"{hr_leave['attendance_missing_from']}") and ('16:00:00' in f"{hr_leave['attendance_missing_to']}"):
                        kid_mod_stage_1_start = 16
                        kid_mod_stage_1_end = 16.5
                        kid_mod_stage_2_start = 16.5
                        kid_mod_stage_2_end = 17.0
                    elif ('01:00:00' in f"{hr_leave['attendance_missing_from']}") and ('01:00:00' in f"{hr_leave['attendance_missing_to']}"):
                        kid_mod_stage_1_start = 8
                        kid_mod_stage_1_end = 8.5
                        kid_mod_stage_2_start = 8.5
                        kid_mod_stage_2_end = 9.0
                    elif ('06:00:00' in f"{hr_leave['attendance_missing_from']}") and ('23:00:00' in f"{hr_leave['attendance_missing_to']}"):
                        kid_mod_stage_1_start = 8
                        kid_mod_stage_1_end = 8.5
                        kid_mod_stage_2_start = 16.5
                        kid_mod_stage_2_end = 17.0
                hr_leave['kid_mod'] = kid_mod
                hr_leave['kid_mod_stage_1_start'] = kid_mod_stage_1_start
                hr_leave['kid_mod_stage_1_end'] = kid_mod_stage_1_end
                hr_leave['kid_mod_stage_2_start'] = kid_mod_stage_2_start
                hr_leave['kid_mod_stage_2_end'] = kid_mod_stage_2_end

                try:
                    request_date_from_datetime = datetime.datetime.strptime (hr_leave['request_date_from'], "%Y-%m-%d")
                    request_date_to_datetime = datetime.datetime.strptime (hr_leave['request_date_to'], "%Y-%m-%d")
                    request_date_range = pd.date_range(request_date_from_datetime.strftime("%m/%d/%Y"),
                                request_date_to_datetime.strftime("%m/%d/%Y"))
                    hr_leave['date'] = request_date_range[0]
                    hr_leave['date_str'] = hr_leave['date'].strftime('%Y-%m-%d')
                    if len(request_date_range)>1:
                        for item in request_date_range:
                            if item != request_date_range[0]:
                                my_copy = hr_leave.copy()


                                my_copy['date'] = item
                                my_copy['minutes']  = 0
                                my_copy['time'] = 0
                                my_copy['date_str'] = \
                                        my_copy['date'].strftime('%Y-%m-%d')
                                my_copy['kid_mod'] = kid_mod
                                my_copy['kid_mod_stage_1_start'] = kid_mod_stage_1_start
                                my_copy['kid_mod_stage_1_end'] = kid_mod_stage_1_end
                                my_copy['kid_mod_stage_2_start'] = kid_mod_stage_2_start
                                my_copy['kid_mod_stage_2_end'] = kid_mod_stage_2_end
                                addition_leave.append(my_copy)

                except Exception as ex:
                    print(ex)

            self.df_hr_leave = pd.DataFrame.from_dict(list_hr_leave)
            self.df_hr_leave = pd.concat([self.df_hr_leave,
                pd.DataFrame.from_dict(addition_leave)], ignore_index=True)

            self.df_hr_leave['employee_code'] = self.df_hr_leave['employee_code'].str.strip()

        else:
            self.df_hr_leave = pd.DataFrame(columns=['id', 'employee_id', 'employee_code', 'department','employee_company_id','company_name',
                                                     'holidy_status_id', 'minutes', 'validate', 'request_date_from', 'request_date_to','date_str','attendance_missing_from', 'attendance_missing_to',
                                                     'kid_mod','kid_mod_stage_1_start', 'kid_mod_stage_2_start', 'kid_mod_stage_1_end','kid_mod_stage_2_end'])

        self.df_hr_leave = self.df_hr_leave.merge(self.df_employees[['id', 'code', 'department_id', 'name', 'job_title', 'time_keeping_code',
                                                                     'date_sign', 'time_keeping_count','company_name','company_id','severance_day',
                                                                     'probationary_contract_termination_date', 'workingday', 'probationary_salary_rate', 'resource_calendar_id']],
                                                  left_on=['employee_code', 'company_name'], right_on=['code','company_name'], how='left', suffixes=('', '_employee'))
        self.df_hr_leave['date_from_str'] = self.df_hr_leave.apply(
            lambda row: self.split_date(row), axis=1)
        self.df_hr_leave['request_date_from'] = pd.to_datetime(
            self.df_hr_leave['request_date_from'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_hr_leave['overtime_from'] = pd.to_datetime(
            self.df_hr_leave['overtime_from'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_hr_leave['overtime_to'] = pd.to_datetime(
            self.df_hr_leave['overtime_to'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_hr_leave['request_date_to'] = pd.to_datetime(
            self.df_hr_leave['request_date_to'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_hr_leave['attendance_missing_from'] = pd.to_datetime(
            self.df_hr_leave['attendance_missing_from'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_hr_leave['attendance_missing_to'] = pd.to_datetime(
            self.df_hr_leave['attendance_missing_to'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        # df_fix_overtime=self.df_hr_leave[self.df_hr_leave['overtime_to']<self.df_hr_leave['overtime_from']] 
        # self.df_hr_leave.apply(lambda row: self.fix_overtime_error(row), axis=1)
        # df_fix_overtime.to_excel('df_fix_overtime.xlsx')
        self.df_hr_leave['attendance_missing_from'] = pd.to_datetime(
            self.df_hr_leave['attendance_missing_from'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_hr_leave['attendance_missing_to'] = pd.to_datetime(
            self.df_hr_leave['attendance_missing_to'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        # self.df_scheduling_ver.rename(columns = {'id':'employee_sid'}, inplace = True)
        # self.df_hr_leave['is_probationary'] = True
        self.df_hr_leave[['probationary_contract_termination_date', 'is_probationary', 
                    'contract_company','probationary_salary_rate']] = self.df_hr_leave.apply(
            lambda x: self.calculate_is_probationary(x), axis=1, result_type='expand')
                    
        self.df_hr_leave[['is_leave_probationary','holiday_status_name']] = self.df_hr_leave.apply(
            lambda x: self.cal_proba(x), axis=1, result_type='expand')
        # for group_index, group_data in self.df_hr_leave.groupby('holiday_status_name'):
        self.df_hr_leave =  self.df_hr_leave.merge(self.df_scheduling_ver,
                                                             on=['employee_code','date_str','company_name'], how='left', suffixes=( '' ,'_scheduling' ))

        # self.df_hr_leave =  self.df_scheduling_ver.merge(self.df_hr_leave,
        #                                                      left_on=['employee_code'], right_on = ['employee_code'], how='right', suffixes=( '' ,'_explanation' ))

        # self.df_scheduling_ver = self.df_scheduling_ver.merge(self.df_hr_leave,
        #                                                      left_on=['employee_code'], right_on = ['employee_code'], how='left', suffixes=( '' ,'_explanation' ))
        # Check 'resource.calendar.leaves
        _, lastday = calendar.monthrange(self.date_array[0].year, self.date_array[0].month)

        # start_date = last_month.replace(day=1).strftime('%Y-%m-%d 00:00:00')
        # end_date = last_month.replace(day=lastday).strftime('%Y-%m-%d 00:00:00')
        
        start_datetime_str = (self.date_array[0] -datetime.timedelta(days= 40)) .replace(day=1).strftime('%Y-%m-%d 00:00:00')
        end_datetime_str = (self.date_array[0].replace(day=lastday) + relativedelta(months=1)).strftime('%Y-%m-%d 23:59:59')
        
        resource_calendar_leaves_ids = self.models.execute_kw(
            self.db, self.uid, self.password, 'resource.calendar.leaves', 'search', [[('calendar_id','=',False),
                                    ('date_from','<=', end_datetime_str),
                                    ('date_to','>=', start_datetime_str)]])
        resource_calendar_leaves_list = self.models.execute_kw(self.db, self.uid, self.password, 'resource.calendar.leaves', 'read', [resource_calendar_leaves_ids],
                                                               {'fields': ['id', 'name', 'company_id', 'calendar_id', 'date_from', 'date_to', 'resource_id', 'time_type']})
        for calendar_leave in resource_calendar_leaves_list:
            try:
                company_name = calendar_leave['company_id'][1]
                calendar_leave['company_name'] = company_name
            except:
                calendar_leave['company_name'] = ''

        self.df_resource_calendar_leaves = pd.DataFrame.from_dict(
            resource_calendar_leaves_list)
        self.df_resource_calendar_leaves.to_excel('self.df_resource_calendar_leaves.xlsx')
        print('***************CALENDAR************************************')
        try:
            self.df_resource_calendar_leaves[['date_from', 'date_to']] = self.df_resource_calendar_leaves.apply(
                lambda row: self.refact_date_from(row), axis=1, result_type='expand')
        except:
            self.df_resource_calendar_leaves = pd.DataFrame(
                columns=['date_from', 'date_to','company_name'])
        try:
            self.df_hr_leave[['is_holiday', 'holiday_from', 'holiday_to', 'holiday_name']] = \
                self.df_hr_leave.apply(lambda row: self.merge_holiday(
                    row), axis=1, result_type='expand')
        except Exception as ex:
            print(ex)
            self.df_hr_leave[['is_holiday', 'holiday_from', 'holiday_to', 'holiday_name']] = (
                False, None, None, None)

        index = 0
        self.df_scheduling_ver[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                self.df_scheduling_ver.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
                
        for _, holiday in self.df_resource_calendar_leaves.iterrows():
            # df_company = self.df_scheduling_ver[self.df_scheduling_ver['company']== holiday['company_name']]
            df = self.df_scheduling_ver[(self.df_scheduling_ver['shift_end_datetime'] >= pd.Timestamp(holiday['date_from'])) & \
                                           (self.df_scheduling_ver['shift_start_datetime'] < pd.Timestamp(holiday['date_to'])) & \
                                           (self.df_scheduling_ver['company']== holiday['company_name'])]
            # print(df)
            # print(holiday['date_from'].date())
            # input(holiday['company_name'])
            if len(df.index) >0:
                df['holiday_from'] = holiday['date_from']
                df['is_holiday'] = True
                df['holiday_to'] = holiday['date_to']
            
                if index == 0:
                    self.df_scheduling_ver_holiday_current_month = df 
                else:
                    self.df_scheduling_ver_holiday_current_month = pd.concat([self.df_scheduling_ver_holiday_current_month, df], ignore_index=True)
                index = index + 1
        if index== 0:
            self.df_scheduling_ver_holiday_current_month = pd.DataFrame(columns=['holiday_from', 'holiday_to', 'date_from', 'date_to','company_name', 'is_holiday', \
                'shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime', \
                'probationary_contract_termination_date', 'is_probationary', \
                    'contract_company','probationary_salary_rate', 'code', 'employee_code', \
                    'night_holiday_work_time_minute','holiday_work_time', 'night_worktime_minute'])
        # self.df_scheduling_ver_holiday_current_month['date_str'] = self.df_scheduling_ver_holiday_current_month['date']
        # self.df_scheduling_ver_holiday_current_month['date'] = pd.to_datetime(
            # self.df_scheduling_ver_holiday_current_month['date'], format='%Y-%m-%d',  errors='coerce')
        # for i in range(1,16):
        #     self.df_scheduling_ver_holiday_current_month[f'attendance_attempt_{i}'] = pd.to_datetime(
        #         self.df_scheduling_ver_holiday_current_month[f'attendance_attempt_{i}'], format="%Y-%m-%d %H:%M:%S",  errors='coerce')
        # self.df_scheduling_ver_holiday_current_month['last_attendance_attempt'] = pd.to_datetime(
        #         self.df_scheduling_ver_holiday_current_month['last_attendance_attempt'], format="%Y-%m-%d %H:%M:%S",  errors='coerce')
        # self.df_scheduling_ver_holiday_current_month['probation_completion_wage'] = pd.to_datetime(
        #     self.df_scheduling_ver_holiday_current_month['probation_completion_wage'], format='%Y-%m-%d',  errors='coerce')
        # print(self.df_scheduling_ver)
        else:
            self.df_scheduling_ver_holiday_current_month.to_excel('self_df_scheduling_ver_holiday_current_month.xlsx')
            
            self.df_scheduling_ver_holiday_current_month[['probationary_contract_termination_date', 'is_probationary', 
                    'contract_company','probationary_salary_rate']] = self.df_scheduling_ver_holiday_current_month.apply(
                    lambda x: self.calculate_is_probationary(x), axis=1, result_type='expand')
                    
            
            self.df_scheduling_ver_holiday_current_month[['night_holiday_work_time_minute','holiday_work_time', 'night_worktime_minute']] = \
                    self.df_scheduling_ver_holiday_current_month.apply(lambda row: self.calculate_night_holiday_work_time(row),  axis=1, result_type='expand')
            # self.df_scheduling_ver_holiday_current_month.to_excel('self_df_scheduling_ver_holiday_current_month_last.xlsx')
    def update_last_month_data(self):
        last_month = self.date_array[0] - datetime.timedelta(days=10)
        _, lastday = calendar.monthrange(last_month.year, last_month.month)

        start_date = last_month.replace(day=1).strftime('%Y-%m-%d 00:00:00')
        end_date = last_month.replace(day=lastday).strftime('%Y-%m-%d 00:00:00')
        domain = ["&", "&", "&", "&",
                                ("request_date_to", ">=", start_date),
                                ("request_date_from", ">=", start_date),
                                ("request_date_from", "<=", end_date),
                                ('active', '=', True),
                                ("state", "=", "validate") ]
        if (self.select_company_id > 0):
            domain = ["&", "&", "&", "&", "&",
                ("employee_company_id", "=", self.select_company_id),
                                ("request_date_to", ">=", start_date),
                                ("request_date_from", ">=", start_date),
                                ("request_date_from", "<=", end_date),
                                ('active', '=', True),
                                ("state", "=", "validate") ]
        hr_leave_ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.leave', 'search', [domain])
        list_hr_leave = self.models.execute_kw(self.db, self.uid, self.password, 'hr.leave', 'read', [hr_leave_ids], {'fields': ['id', 'employee_id', 'employee_code', 'employee_company_id', 'active',
                                                                                                                       'holiday_status_id', 'minutes', 'time', 'state', 'request_date_from', 'request_date_to',
                                                                                                                       'attendance_missing_from', 'attendance_missing_to', 'reasons', 'for_reasons', 'convert_overtime',
                                                                                                                       'employee_company_id', 'multiplier_work_time']})
        # return
        if len(list_hr_leave) > 0:
            print('hrleave toi')
            print(list_hr_leave)
            addition_leave = []
            for hr_leave in list_hr_leave:
                try:
                    company_name = hr_leave['employee_company_id'][1]
                    hr_leave['company_name'] = company_name
                    company_sid = hr_leave['employee_company_id'][0]
                    hr_leave['company_sid'] = company_sid
                except:
                    hr_leave['company_name'] = False
                    hr_leave['company_sid'] = False
                try:
                    holiday_status_name = hr_leave['holiday_status_id'][1]
                    hr_leave['holiday_status_name'] = holiday_status_name
                except:
                    hr_leave['holiday_status_name'] = False

                try:
                    request_date_from_datetime = datetime.datetime.strptime (hr_leave['request_date_from'], "%Y-%m-%d")
                    request_date_to_datetime = datetime.datetime.strptime (hr_leave['request_date_to'], "%Y-%m-%d")
                    request_date_range = pd.date_range(request_date_from_datetime.strftime("%m/%d/%Y"),
                                request_date_to_datetime.strftime("%m/%d/%Y"))
                    hr_leave['date'] = request_date_range[0]
                    hr_leave['date_str'] = hr_leave['date'].strftime('%Y-%m-%d')
                    if len(request_date_range)>1:
                        for item in request_date_range:
                            if item != request_date_range[0]:
                                my_copy = hr_leave.copy()


                                my_copy['date'] = item
                                my_copy['minutes']  = 0
                                my_copy['time'] = 0
                                my_copy['date_str'] = \
                                        my_copy['date'].strftime('%Y-%m-%d')
                                addition_leave.append(my_copy)

                except Exception as ex:
                    print(ex)

            self.df_last_month_result = pd.DataFrame.from_dict(list_hr_leave)
            self.df_last_month_result = pd.concat([self.df_last_month_result,
                pd.DataFrame.from_dict(addition_leave)], ignore_index=True)
            
            self.df_last_month_result['employee_code'] = self.df_last_month_result['employee_code'].str.strip()

        else:
            self.df_last_month_result = pd.DataFrame(columns=['id', 'employee_id', 'employee_code', 'department','employee_company_id','company_name',
                                                     'holidy_status_id', 'minutes', 'validate', 'request_date_from', 'request_date_to','date_str'])

        self.df_last_month_result = self.df_last_month_result.merge(self.df_employees[['id', 'code', 'department_id', 'name', 'job_title', 'time_keeping_code',
                                                                     'date_sign', 'time_keeping_count','company_name','company_id','severance_day',
                                                                     'probationary_contract_termination_date', 'workingday', 'probationary_salary_rate', 'resource_calendar_id']],
                                                  left_on=['employee_code', 'company_name'], right_on=['code','company_name'], how='left', suffixes=('', '_employee'))
        self.df_last_month_result['date_from_str'] = self.df_last_month_result.apply(
            lambda row: self.split_date(row), axis=1)
        self.df_last_month_result['request_date_from'] = pd.to_datetime(
            self.df_last_month_result['request_date_from'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_last_month_result['request_date_to'] = pd.to_datetime(
            self.df_last_month_result['request_date_to'], format="%Y-%m-%d %H:%M:%S", errors='coerce')

        self.df_last_month_result['attendance_missing_from'] = pd.to_datetime(
            self.df_last_month_result['attendance_missing_from'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_last_month_result['attendance_missing_to'] = pd.to_datetime(
            self.df_last_month_result['attendance_missing_to'], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        # self.df_scheduling_ver.rename(columns = {'id':'employee_sid'}, inplace = True)
        # self.df_last_month_result['is_probationary'] = True
        self.df_last_month_result[['is_leave_probationary','holiday_status_name']] = self.df_last_month_result.apply(
            lambda x: self.cal_proba(x), axis=1, result_type='expand')
        # for group_index, group_data in self.df_last_month_result.groupby('holiday_status_name'):
        self.df_last_month_result =  self.df_last_month_result.merge(self.df_scheduling_ver,
                                                             on=['employee_code','date_str','company_name'], how='left', suffixes=( '' ,'_scheduling' ))
        # Check 'resource.calendar.leaves
        _, lastday = calendar.monthrange(self.date_array[0].year, self.date_array[0].month)

        # start_date = last_month.replace(day=1).strftime('%Y-%m-%d 00:00:00')
        # end_date = last_month.replace(day=lastday).strftime('%Y-%m-%d 00:00:00')
        
        start_datetime_str = (self.date_array[0] - datetime.timedelta(days= 40)).replace(day=1).strftime('%Y-%m-%d 00:00:00')
        end_datetime_str = (self.date_array[0].replace(day=lastday)+ datetime.timedelta(days= 40)).strftime('%Y-%m-%d 23:59:59')
        
        resource_calendar_leaves_ids = self.models.execute_kw(
            self.db, self.uid, self.password, 'resource.calendar.leaves', 'search', [[('calendar_id','=',False),
                                    ('date_from','<=', end_datetime_str),
                                    ('date_to','>=', start_datetime_str)]])
        resource_calendar_leaves_list = self.models.execute_kw(self.db, self.uid, self.password, 'resource.calendar.leaves', 'read', [resource_calendar_leaves_ids],
                                                               {'fields': ['id', 'name', 'company_id', 'calendar_id', 'date_from', 'date_to', 'resource_id', 'time_type']})
        for calendar_leave in resource_calendar_leaves_list:
            try:
                company_name = calendar_leave['company_id'][1]
                calendar_leave['company_name'] = company_name
            except:
                calendar_leave['company_name'] = ''

        self.df_resource_calendar_leaves = pd.DataFrame.from_dict(
            resource_calendar_leaves_list)

        print('***************CALENDAR************************************')
        try:
            self.df_resource_calendar_leaves[['date_from', 'date_to']] = self.df_resource_calendar_leaves.apply(
                lambda row: self.refact_date_from(row), axis=1, result_type='expand')
        except:
            self.df_resource_calendar_leaves = pd.DataFrame(
                columns=['date_from', 'date_to','company_name'])
        try:
            self.df_last_month_result[['is_holiday', 'holiday_from', 'holiday_to', 'holiday_name']] = \
                self.df_last_month_result.apply(lambda row: self.merge_holiday(
                    row), axis=1, result_type='expand')
        except Exception as ex:
            print(ex)
            self.df_last_month_result[['is_holiday', 'holiday_from', 'holiday_to', 'holiday_name']] = (
                False, None, None, None)
        # self.df_last_month_result = df_hr_leave.groupby(['company_name','employee_code', 'is_probationary', 'holiday_status_name'])

        # for group_index, group_data in df_last_month_result:
        #     group_data.to_excel(f"leave-{group_index[1]}.xlsx")
        self.df_last_month_result.to_excel(os.path.join(self.output_report_folder, "self.df_last_month_hr_leave.xlsx"))
        # self.df_employees.to_excel('df_employees_in_cl_before_merge.xlsx')
        
        if self.is_update_al_cl:
            self.current_month_cl = self.df_employees.merge(self.df_cl_report, left_on=['code','company_sid'],
                     right_on=['employee_code','company_sid'], how='left',  suffixes=( '' ,'_cl'))
            duplicates = self.df_employees[~self.df_employees.index.isin(self.current_month_cl.index)]
            # self.df_employees.to_excel('df_employees_in_cl.xlsx')
            duplicates.to_excel('df_employees_not_in_cl.xlsx')
            
            self.current_month_cl[['tangca_probation', 'tangca_chinhthuc', 'cl_probation', 'cl_chinhthuc', \
                'total_dimuon_vesom', 'total_dimuon_vesom_probationary', 'total_dimuon_vesom_official', \
                'increase_probationary_1', 'increase_official_1']] = self.current_month_cl.apply(lambda row: \
                self.calculate_current_month_cl(row), axis=1, result_type='expand')
            self.current_month_cl.to_excel('df_cl_report_merge.xlsx')
            self.df_cl_report_copy = self.df_cl_report_copy.merge(self.current_month_cl[['id_cl',
                        'total_dimuon_vesom', 'total_dimuon_vesom_probationary', 'total_dimuon_vesom_official']], left_on=['id'],
                     right_on=['id_cl'], how='left',  suffixes=( '' ,'_cl'))
            
            df_al_report = self.df_employees.merge(self.df_al_report, left_on=['code','company_sid'],
                        right_on=['employee_code','company_sid'], how='left',  suffixes=( '' ,'_al'))
            
            df_al_report[['result', 'total_dimuon_vesom', 'remaining_leave']] = df_al_report.apply(lambda row: \
                self.calculate_current_month_al(row), axis=1, result_type='expand')
            
            self.df_al_report_copy = self.df_al_report_copy.merge(df_al_report[['id_al', 'total_dimuon_vesom']], left_on=['id'],
                     right_on=['id_al'], how='left',  suffixes=( '' ,'_al'))
            try:
                df_al_report.to_excel('df_al_report_merge.xlsx')
            except:
                print('savefile ex')
            if self.date_array[0].month == 2:
                df_al_report[(df_al_report['AUTO-CALCULATE-AL-LAST-YEAR-DEDUCTION']== True) &
                            df_al_report['id_al'].notnull()] \
                    .apply(lambda row: self.calculate_last_year_al_deduction(row), axis=1)

            # for index_cl_str, update_data in self.cl_result_update.items():
            #     try:
            #         ids = [int(index_cl_str)]
            #         self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'write', [ids, update_data])
            #         print ('update: ', update_data)
            #     except Exception as ex:
            #         logger.error(f'write cl ex --- : {ex}')
            # for employee_code, update_data in self.cl_result_create.items():
            #     try:
            #         id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 
            #             'hr.cl.report', 'create', [update_data])
            #         print ('Create: ', update_data)
            #     except Exception as ex:
            #         logger.error(f'create cl ex --- : {ex}')
                
    def calculate_last_year_al_deduction(self, row):
        other_info_update = False
        
        al_report_last_year = self.df_last_year_al_report[
                    (self.df_last_year_al_report['employee_code']== row['code']) &
                    (self.df_last_year_al_report['company_sid']== row['company_sid'])]
        cl_report = self.current_month_cl[
            (self.current_month_cl['code']== row['code']) &
            (self.current_month_cl['company_sid']== row['company_sid'])
        ]
        try:
            if str(int(row['id_al'])) in self.al_result_update:
                update_data = self.al_result_update[str(int(row['id_al']))]
            else:
                update_data = {}
        except Exception as ex:
            update_data = {}
            logger.error(f'deduction: {ex}')
        # if row['employee_code'] == 'APG220208015':
        #     print("cl: ", al_report_last_year)
        #     input('APG220208015')
            
        if len(al_report_last_year.index)>0:
            if pd.notnull(al_report_last_year.iloc[0]['remaining_leave_minute']):
                if (al_report_last_year.iloc[0]['remaining_leave_minute']<0) \
                        and(len (cl_report.index)>0) \
                        and((cl_report.iloc[0]['increase_official_1'] > 0) or 
                            (cl_report.iloc[0]['increase_probationary_1'] > 0)):
                            
                    remaining_leave = min(al_report_last_year.iloc[0]['remaining_leave_minute'] \
                        + cl_report.iloc[0]['increase_official_1']
                        + cl_report.iloc[0]['increase_probationary_1'], 0)
                    
                    print('remaining_leave: ', remaining_leave)
                    
                    cl_update_value = remaining_leave - al_report_last_year.iloc[0]['remaining_leave_minute']
                    cl_update_value_probationary = 0
                    cl_update_value_offical = 0
                    print('cl_update_value: ',cl_update_value)
                    if pd.notnull(cl_report.iloc[0]['remaining_probationary_minute']) :

                        cl_update_value_probationary = min(cl_update_value, cl_report.iloc[0]['increase_probationary_1'])

                       
                    cl_update_value = cl_update_value - cl_update_value_probationary
                    print('cl_update_value2: ', cl_update_value)
                    if pd.notnull(cl_report.iloc[0]['increase_official_1']) :
                        cl_update_value_offical = min(cl_update_value, cl_report.iloc[0]['increase_official_1'])
                            
                    ids = [int(cl_report.iloc[0]['id_cl'])]
                    update_data_last_year_cl = {
                        'increase_official_1': int(int(cl_report.iloc[0]['increase_official_1']) - cl_update_value_offical),
                        'increase_probationary_1':int(int( cl_report.iloc[0]['increase_probationary_1']) - cl_update_value_probationary )}
                    
                    try:
                        if f"{int(cl_report.iloc[0]['id_cl'])}" in self.cl_result_update:
                            self.cl_result_update[f"{int(cl_report.iloc[0]['id_cl'])}"]['increase_official_1'] = \
                                int(int(cl_report.iloc[0]['increase_official_1']) - cl_update_value_offical)
                            self.cl_result_update[f"{int(cl_report.iloc[0]['id_cl'])}"]['increase_probationary_1'] = \
                                int(int( cl_report.iloc[0]['increase_probationary_1']) - cl_update_value_probationary )
                        elif row['employee_code'] in self.cl_result_create:
                            self.cl_result_create[row['employee_code']]['increase_official_1'] = \
                                int(int(cl_report.iloc[0]['increase_official_1']) - cl_update_value_offical)

                            self.cl_result_create[row['employee_code']]['increase_probationary_1'] = \
                                int(int( cl_report.iloc[0]['increase_probationary_1']) - cl_update_value_probationary )
                        # else:
                        #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'write', [ids, update_data_last_year_cl])
                        
                    except Exception as ex:
                        logger.error(f'write cl excel report ex: {ex}')
                            
                    if int(row['remaining_leave']) != int(remaining_leave):
                        
                        other_info_update = True
                    update_data['remaining_leave'] = int(remaining_leave)
    # print(f"{row['id_al']} -new_date: {new_date} - check_employee_info: {check_employee_info}")
        
        if (other_info_update):
            ids = [int(row['id_al'])]
            try:
                update_data['date_calculate_leave']= (self.date_array[-1]+datetime.timedelta(days=1)).strftime("%Y-%m-%d")
                self.al_result_update[str(int(row['id_al']))] = update_data
                # self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'write', [ids, update_data])
                print('check_employee_info updated: ', update_data)
            except Exception as ex:
                logger.error(f"write hr al report last year employee: {ex}")
        else: 
            print('updated last year al')
            
        # if row['employee_code'] == 'APG220208015':
        #     print("cl: ", cl_report)
        #     input('APG220208015')
            
    def calculate_current_month_al(self, row):
        result = 0
        total_al_from_al_shift = 0
        new_date = False
        total_dimuon_vesom = 0
        total_cl_from_cl_shift_probationary = 0
        total_cl_from_cl_shift_offical = 0
        # total_al = 0
        # total_al_probationary = 0
        # number_cl_date = 0
        # number_al_date = 0
        
        try:
            remaining_leave = int(row['remaining_leave'])
        except:
            remaining_leave = 0
        try:
            df_filter_al= self.df_al_report[self.df_al_report['company_name']== row['company_name']]
            max_date_calculate = df_filter_al['date_calculate_leave'].max()
            
        # if max_date_calculate< self.date_array[0]:
            if (max_date_calculate <= self.date_array[-1]) or (len(df_filter_al.index)== 0):
                max_date_calculate = self.date_array[-1] + datetime.timedelta(days=1)
                new_date = True
        except Exception as ex:
            print('find max al ex: ', ex)
            max_date_calculate = self.date_array[-1] + datetime.timedelta(days=1)
            new_date = True
            
        # hr_leave_employee_current_month = self.df_hr_leave[
        #     (self.df_hr_leave['employee_code']== row['code']) &
        #     (self.df_hr_leave['company_sid']== row['company_sid'])]
        
        hr_leave_employee = self.df_last_month_result[
            (self.df_last_month_result['employee_code']== row['code']) &
            (self.df_last_month_result['company_sid']== row['company_sid'])]
        # if row['code'] == 'APG210924007':
        #     print('APG210924007', hr_leave_employee)
        
        al_report_before = self.df_last_al_report[
            (self.df_last_al_report['employee_code']== row['code']) &
            (self.df_last_al_report['company_sid']== row['company_sid'])]
        
        al_report_last_year = self.df_last_year_al_report[
                    (self.df_last_year_al_report['employee_code']== row['code']) &
                    (self.df_last_year_al_report['company_sid']== row['company_sid'])]
        
        cl_dimuonvesom = self.current_month_cl[
            (self.current_month_cl['code']== row['code']) &
            (self.current_month_cl['company_sid']== row['company_sid'])]
        
        df_scheduling_ver_al_cl_current_month_employee = self.df_scheduling_ver[ \
            (self.df_scheduling_ver['employee_code'] == row['code']) & \
            (self.df_scheduling_ver['company'] == row['company_name']) & \
            (self.df_scheduling_ver['shift_name'] == 'AL')]
        
        print(f"{row['code']} - {row['company_name']} - {hr_leave_employee} - {row['workingday_employee']}")
        
        # for leave_index, leave_item in hr_leave_employee_current_month.iterrows():
        #     if leave_item['holiday_status_name']:
        #         if ( 'nghỉ phép năm' in leave_item['holiday_status_name'].strip().lower()):
        #             print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
        #             total_al = total_al + \
        #                 max(leave_item['minutes'], leave_item['time'])
        #             if leave_item['is_leave_probationary']:
        #                 total_al_probationary = total_al_probationary + \
        #                     max(leave_item['minutes'],
        #                         leave_item['time'])     
                                
        for index, leave_item in hr_leave_employee.iterrows():

            try:
                if ('phép năm' in leave_item['holiday_status_name'].strip().lower()):

                    result = result + max(leave_item['minutes'],
                                leave_item['time'])
            except Exception as ex:
                print('al: ',ex)
                
        df_scheduling_ver_al_cl_last_month_employee = self.df_scheduling_ver_al_cl_last_month[
            (self.df_scheduling_ver_al_cl_last_month['employee_code'] == row['code']) &
            (self.df_scheduling_ver_al_cl_last_month['company'] == row['company_name']) &
            (self.df_scheduling_ver_al_cl_last_month['shift_name'] == 'AL')]
        for index, al_cl_shift in df_scheduling_ver_al_cl_last_month_employee.iterrows():
            find_leave_item = hr_leave_employee[
                (hr_leave_employee['date_str'] == al_cl_shift['date_str']) &
                ((hr_leave_employee['holiday_status_name'] == 'Nghỉ phép năm') |
                (hr_leave_employee['holiday_status_name'] == 'Nghỉ bù'))]
            if len(find_leave_item.index) == 0:
                # print(row)
                # print(row['code'])
                # print(al_cl_shift['date_str'])
                result = result+  480
                # exit(1)
       
        for index, al_cl_shift in df_scheduling_ver_al_cl_current_month_employee.iterrows():
            find_leave_item = self.df_hr_leave[
                (self.df_hr_leave['date_str'] == al_cl_shift['date_str']) &
                ((self.df_hr_leave['holiday_status_name'] == 'Nghỉ phép năm') |
                (self.df_hr_leave['holiday_status_name'] == 'Nghỉ bù'))]
            if len(find_leave_item.index) == 0:
                # print(row)
                # print(row['code'])
                # print(al_cl_shift['date_str'])
                total_al_from_al_shift =total_al_from_al_shift + 480
                # exit(1)
        update_data = {}
        last_month_index = self.date_array[0].month-1 if self.date_array[0].month>1 else 12
        check_employee_info = False
        if row['year'] != self.date_array[0].year:
            check_employee_info = True
        update_data['year'] = self.date_array[0].year
        if row['standard_day'] != 480 :
            check_employee_info = True
        update_data['standard_day'] = 480
        is_found_date_apply_leave = False
        # if row['employee_code'] == 'APG211014003':
        #     print(f"employee_code.date_apply_leave {row['date_apply_leave']}")
        #     input(f"employee_code.1.APG211014003 {row['date_sign_employee']}")
        if pd.notnull(row['date_sign_employee']):
            if row['date_sign_al'] != row['date_sign_employee']:
                
                check_employee_info = True
            print(f"update date_sign {row['date_sign_employee']}")
            update_data['date_sign'] = row['date_sign_employee'].strftime("%Y-%m-%d")
            # if row['employee_code'] == 'APG211014003':
            #     print(f"employee_code.date_apply_leave {row['date_apply_leave']}")
            #     input(f"employee_code.2.APG211014003 {row['date_sign_employee']}")
            print(pd.Timestamp(row['date_sign_employee'].date()) > self.date_array[-1])
            if pd.Timestamp(row['date_sign_employee'].date()) > self.date_array[-1]:
                
                if row['date_apply_leave'] != row['date_sign_employee'].strftime("%Y-%m-%d"):
                    check_employee_info = True
                update_data['date_apply_leave'] = row['date_sign_employee'].strftime("%Y-%m-%d")
                is_found_date_apply_leave = True
        # if row['employee_code'] == 'APG211014003':
        #     print(f"employee_code.date_apply_leave {row['date_apply_leave']}")
        #     input(f"employee_code.3.APG211014003 {row['workingday_employee']}")
        if pd.notnull(row['workingday_employee']):
            if row['workingday_al'] != row['workingday_employee']:
                
                check_employee_info = True
            update_data['workingday'] = row['workingday_employee'].strftime("%Y-%m-%d")
            
            if not is_found_date_apply_leave:
                # if row['employee_code'] == 'APG211014003':
                #     input(f"employee_code.chua found.APG211014003 {row['date_apply_leave']}")
                if pd.isnull(row['date_sign_employee']):
                    next_two_month_from_workingday = row['workingday_employee'] + relativedelta(months=+2)
                    if next_two_month_from_workingday.year < (self.date_array[0]).year :
                        if row['date_apply_leave'] != (self.date_array[0]).replace(day=1, month=1).strftime("%Y-%m-%d"):
                            
                            check_employee_info = True
                        update_data['date_apply_leave'] = (self.date_array[0]). \
                                replace(day=1, month=1).strftime("%Y-%m-%d")
                    else:
                        if row['date_apply_leave'] != next_two_month_from_workingday.strftime("%Y-%m-%d"):
                            check_employee_info = True
                        update_data['date_apply_leave'] =next_two_month_from_workingday.strftime("%Y-%m-%d")
                else:
                    if row['workingday_employee'].year < (self.date_array[0]).year :
                        if row['date_apply_leave'] != (self.date_array[0]).replace(day=1, month=1).strftime("%Y-%m-%d"):
                            
                            check_employee_info = True
                        update_data['date_apply_leave'] = (self.date_array[0]). \
                                replace(day=1, month=1).strftime("%Y-%m-%d")
                    else:
                        if row['date_apply_leave'] != row['workingday_employee'].strftime("%Y-%m-%d"):
                            check_employee_info = True
                        update_data['date_apply_leave'] = row['workingday_employee'].strftime("%Y-%m-%d")
        # if row['employee_code'] == 'APG211014003':
        #     input(f"employee_code.4.APG211014003 {row['date_apply_leave']}")
        
        if pd.notnull(row['severance_day_employee']):
            if row['severance_day_al'] != row['severance_day_employee']:
                check_employee_info = True
            update_data['severance_day'] = row['severance_day_employee'].strftime("%Y-%m-%d")
        month_array_str = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 
                                'october', 'november', 'december']
        select_month_str = month_array_str[last_month_index-1]
        other_info_update = False
        
        if self.date_array[0].month ==2:
            if len(al_report_last_year.index)>0:
                if pd.notnull(al_report_last_year.iloc[0]['remaining_leave_minute']):
                    if al_report_last_year.iloc[0]['remaining_leave_minute'] >= 0:
                        if row['remaining_leave'] != al_report_last_year.iloc[0]['remaining_leave_minute']:
                            other_info_update = True
                        update_data['remaining_leave'] = int(al_report_last_year.iloc[0]['remaining_leave_minute'].item())
                            
                    remaining_leave = int(al_report_last_year.iloc[0]['remaining_leave_minute'])
        else:
            if len(al_report_before.index)>0:
                if pd.notnull(al_report_before.iloc[0]['remaining_leave']):
                    if al_report_before.iloc[0]['remaining_leave'] >=0:
                        if row['remaining_leave'] != al_report_before.iloc[0]['remaining_leave']:
                            
                            other_info_update = True
                        update_data['remaining_leave'] = int(al_report_before.iloc[0]['remaining_leave'].item())
                            
                    remaining_leave = int(al_report_before.iloc[0]['remaining_leave'])
            else:
                if row['remaining_leave'] != 0:
                    other_info_update = True
                update_data['remaining_leave'] = 0
                remaining_leave = 0
                        
        if len(al_report_before.index)>0:
            print('update before')
            month_count = 0
            for month_count in range(0, last_month_index):
                month_str = month_array_str[month_count]
                if month_str != select_month_str:
                    try:
                        update_data[month_str] = int(al_report_before.iloc[0][month_str])
                        # if (row['code'] == 'APG210727002') and ('Feb' in month_str):
                        #     print(month_str)
                        #     print(update_data)
                        #     print(ids)
                        #     input('5: APG210727002')
                    except Exception as ex:
                        update_data[month_str] = 0
                    if row[month_str] != update_data[month_str]:
                        other_info_update = True
            for month_index in range(self.date_array[0].month, len(month_array_str)):
                current_select_month_str = month_array_str[month_index]
                update_data[current_select_month_str] = 0
                # if (row['code'] == 'APG210727002') and ('Feb' in current_select_month_str):
                #     print(current_select_month_str)
                #     print(update_data)
                #     print(ids)
                #     input('4: APG210727002')
                if row[current_select_month_str] != update_data[current_select_month_str]:
                    other_info_update = True
            
                    
            if self.date_array[0].month !=2:
                if pd.notnull(al_report_before.iloc[0]['remaining_leave']):
                    if al_report_before.iloc[0]['remaining_leave'] >= 0:
                        if row['remaining_leave'] != int(al_report_before.iloc[0]['remaining_leave']):
                            
                            other_info_update = True
                        update_data['remaining_leave'] = int(al_report_before.iloc[0]['remaining_leave'].item())
                    remaining_leave = int(al_report_before.iloc[0]['remaining_leave'])
                else:
                    if row['remaining_leave'] != 0:
                        other_info_update = True
                    update_data['remaining_leave'] = 0
                    remaining_leave = 0
            
        valid_id = False
        try:
            ids = int(row['id_al'])
            print(ids)
            valid_id = ids > 0
        except Exception as ex:
            print(ex)
        if len(cl_dimuonvesom.index)>0:
            total_dimuon_vesom = cl_dimuonvesom.iloc[0]['total_dimuon_vesom']
            total_dimuon_vesom_probationary = cl_dimuonvesom.iloc[0]['total_dimuon_vesom_probationary']
            total_dimuon_vesom_official = cl_dimuonvesom.iloc[0]['total_dimuon_vesom_official']

            total_dimuon_vesom = max(0, total_dimuon_vesom - total_dimuon_vesom_probationary
                    - total_dimuon_vesom_official)
            result = result + int(total_dimuon_vesom)
        # important update current al
        if row[month_array_str[last_month_index]] != int(total_al_from_al_shift):
            update_data[month_array_str[last_month_index]] = int(total_al_from_al_shift)
            other_info_update = True

        # print(f"{row['id_al']} -new_date: {new_date} - check_employee_info: {check_employee_info}")
        update_data[select_month_str] = int(result)
        if(pd.notnull(row['id_al'])) and valid_id and (new_date == False):
            # update_data[select_month_str] = int(result)
            # if row['code'] == 'APG210727002':
            #     print(select_month_str)
            #     print(update_data)
            #     print(ids)
            #     input('2: APG210727002')
            # print(f"update {update_data} - {row}")
            if (int(row[select_month_str]) != int(result)) or (other_info_update) or check_employee_info:
                ids = [int(row['id_al'])]
                try:
                    # self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'write', [ids, update_data])
                    self.al_result_update[str(int(row['id_al']))] = update_data
                    print('check_employee_info updated: ', update_data)
                except Exception as ex:
                    logger.error(f"write hr al report employee: {ex}")
                # if row['code'] == 'APG210727002':
                #     print(select_month_str)
                #     print(update_data)
                #     print(ids)
                #     input('APG210727002')
            else: 
                print('updated al')
        else:
            # print((row['id_al']))
            update_data["employee_id"] = row['id']
            update_data["company_id"]= row['company_sid']
            update_data[select_month_str] = int(result)
            # update_data['year'] =  self.date_array[0].year -1 if i==12 else self.date_array[0].year
            update_data['date_calculate_leave']= (self.date_array[-1]+datetime.timedelta(days=1)).strftime("%Y-%m-%d")
            update_data['year'] = self.date_array[0].year
            # try:
            #     update_data['date_calculate_leave']= max_date_calculate.strftime("%Y-%m-%d")
            #     update_data['year'] = self.date_array[0].year
            # except:
            #     update_data['date_calculate_leave'] = self.date_array[0].strftime("%Y-%m-%d")
            #     update_data['year'] =  self.date_array[0].year
            try:
                # id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 
                #         'hr.al.report', 'create', [update_data])
                self.al_result_create[row['code']] = update_data
                print('created', update_data) 
            except Exception as ex:
                logger.error(f"write hr al report employee: {ex}")
        if self.date_array[0].month == 2:
            is_update_last_year = False
            update_data_last_year = {}
            try:
                if al_report_last_year.iloc[0]['year'] != self.date_array[0].year - 1:
                    is_update_last_year = True
                update_data_last_year['year'] = self.date_array[0].year - 1
            except:
                update_data_last_year['year'] = self.date_array[0].year - 1
                is_update_last_year = True
            try:  
                if al_report_last_year.iloc[0]['employee_id'] != row['id']:
                    is_update_last_year = True
                update_data_last_year["employee_id"] = row['id']
            except:
                update_data_last_year["employee_id"] = row['id']
                is_update_last_year = True
            try: 
                if al_report_last_year.iloc[0]['company_id'] != row['company_sid']:
                    is_update_last_year = True
                update_data_last_year["company_id"] = row['company_sid']
            except:
                update_data_last_year["company_id"] = row['company_sid']
                is_update_last_year = True
            if pd.notnull(row['severance_day_employee']):
                try:
                    if pd.Timestamp(row['severance_day_employee'].date()) < self.date_array[0].replace(day=1, month=1):
                        if al_report_last_year.iloc[0]['severance_day'] != row['severance_day_employee']:
                            
                            is_update_last_year = True
                        update_data_last_year['severance_day'] = row['severance_day_employee'].strftime("%Y-%m-%d")
                    else:
                        if al_report_last_year.iloc[0]['severance_day'] != False:
                            
                            is_update_last_year = True
                        update_data_last_year['severance_day'] = False
                except:
                    if pd.Timestamp(row['severance_day_employee'].date()) < self.date_array[0].replace(day=1, month=1):
                        update_data_last_year['severance_day'] = row['severance_day_employee'].strftime("%Y-%m-%d")
                    else:
                        update_data_last_year['severance_day'] = False
                    is_update_last_year = True
            if pd.notnull(row['workingday_employee']):
                try: 
                    if al_report_last_year.iloc[0]['workingday'].strftime("%Y-%m-%d") != update_data['workingday']:
                        
                        is_update_last_year = True
                    update_data_last_year['workingday'] = update_data['workingday']
                except:
                    update_data_last_year['workingday'] = update_data['workingday']
                    is_update_last_year = True
            
            if pd.notnull(row['date_sign_employee']):
                try:   
                    if al_report_last_year.iloc[0]['date_sign'].strftime("%Y-%m-%d") != update_data['date_sign']:
                        
                        is_update_last_year = True
                    update_data_last_year['date_sign'] = update_data['date_sign']
                except:
                    update_data_last_year['date_sign'] = update_data['date_sign']
                    is_update_last_year = True
                
            for month_index in range(0,12):
                select_month_str = month_array_str[month_index]
                try:
                    if al_report_last_year.iloc[0] [select_month_str] != int(al_report_before.iloc[0][select_month_str]):
                        if len(al_report_before.index) >0:
                            update_data_last_year[select_month_str] = int(al_report_before.iloc[0][select_month_str])
                            is_update_last_year = True
                        else:
                            update_data_last_year[select_month_str] = 0
                            is_update_last_year = True
                except:
                    if len(al_report_before.index) >0:
                        update_data_last_year[select_month_str] = int(al_report_before.iloc[0][select_month_str])
                        is_update_last_year = True
                    else:
                        update_data_last_year[select_month_str] = 0
                        is_update_last_year = True
                    
            if (len(al_report_last_year.index)>0) and (is_update_last_year == True):
                update_data_last_year['date_calculate_leave']= (self.date_array[0].replace(day=1, month=1)).strftime("%Y-%m-%d")
                ids = [int(al_report_last_year.iloc[0]['id'])]
                # try:
                #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'write', [ids, update_data_last_year])
                # except Exception as ex:
                #     logger.error(f"write hr al last year report employee: {ex}")
            # else:
            #     try:
            #         update_data_last_year['date_calculate_leave']= (self.date_array[0].replace(month=1, day=1)).strftime("%Y-%m-%d")
            #         id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 
            #                 'hr.al.report', 'create', [update_data_last_year])
            #         print('created', update_data_last_year) 
            #     except Exception as ex:
            #         logger.error(f"write hr al report last year employee: {ex}")
        # if row['code']=='APG210924007':
        #     exit()
        return result, total_dimuon_vesom, remaining_leave
    def calculate_current_month_cl(self, row):
        tangca_probation = 0
        tangca_chinhthuc = 0
        phatsinhtang_chinhthuc = 0
        phatsinhtang_thuviec = 0
        cl_probation = 0
        cl_chinhthuc = 0
        currnet_month_auto_increase_offical = 0
        currnet_month_auto_increase_probationary = 0
        total_dimuon_vesom = 0
        total_dimuon_vesom_probationary = 0
        total_dimuon_vesom_official = 0
        is_update_last_year = False
        total_cl_from_cl_shift_probationary = 0
        total_cl_from_cl_shift_offical = 0
        try:
            increase_official_1 = row['increase_official_1']
        except:
            increase_official_1 = 0
        try:
            increase_probationary_1 = row['increase_probationary_1']
        except:
            increase_probationary_1 = 0
        new_date = False
        try:
            df_filter_cl = self.df_cl_report[self.df_cl_report['company_name']== row['company_name']]
            max_date_calculate = df_filter_cl['date_calculate'].max()

            if (max_date_calculate <= self.date_array[len(self.date_array) - 1]) or (len(df_filter_cl) == 0):
                max_date_calculate = self.date_array[len(self.date_array) - 1] + datetime.timedelta(days=1)
                new_date = True
        except Exception as ex:
            print(ex)
            max_date_calculate = self.date_array[len(self.date_array) - 1] + datetime.timedelta(days=1)
            new_date = True
        # if max_date_calculate < self.date_array[0]:
        #     max_date_calculate = self.date_array[0]
        hr_leave_employee = self.df_last_month_result[
            (self.df_last_month_result['employee_code']== row['code']) &
            (self.df_last_month_result['company_sid']== row['company_sid'])]
        
        cl_report_before = self.df_last_cl_report[
            (self.df_last_cl_report['employee_code']== row['code']) &
            (self.df_last_cl_report['company_sid']== row['company_sid'])]
        
        holiday_report_before =  self.df_scheduling_ver_holiday_last_month \
            [(self.df_scheduling_ver_holiday_last_month['employee_code']== row['code'])& \
             (self.df_scheduling_ver_holiday_last_month['date'] < self.date_array[0])]
            
        df_scheduling_ver_al_cl_current_month_employee = self.df_scheduling_ver[ \
            (self.df_scheduling_ver['employee_code'] == row['code']) & \
            (self.df_scheduling_ver['company'] == row['company_name']) & \
            (self.df_scheduling_ver['shift_name'] == 'CL') & \
            (self.df_scheduling_ver['date'] >= self.date_array[0])]
        
        cl_report_last_year = self.df_last_year_cl_report[
                    (self.df_last_year_cl_report['employee_code']== row['code']) &
                    (self.df_last_year_cl_report['company_sid']== row['company_sid'])]
        # print(f"{row['employee_code']} - {row['company_name']} - {hr_leave_employee}")
        for index, leave_item in hr_leave_employee.iterrows():
            try:
                if ('nghỉ bù' in leave_item['holiday_status_name'].strip().lower()):

                    if leave_item['is_leave_probationary']:
                        cl_probation = cl_probation + \
                            max(leave_item['minutes'],
                                leave_item['time'])
                    else:
                        cl_chinhthuc = cl_chinhthuc + \
                            max(leave_item['minutes'], leave_item['time'])
                elif ('tăng ca' in leave_item['holiday_status_name'].strip().lower()):
                    # print("don tang ca")
                    if 'phát sinh tăng' in leave_item['reasons'].strip().lower():
                        if leave_item['is_leave_probationary']:
                            phatsinhtang_thuviec = phatsinhtang_thuviec + \
                                max(leave_item['minutes'],
                                    leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                        else:
                            phatsinhtang_chinhthuc = phatsinhtang_chinhthuc + \
                                max(leave_item['minutes'],
                                    leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                    else:
                        if leave_item['is_leave_probationary']:
                            tangca_probation = tangca_probation + \
                                max(leave_item['minutes'],
                                    leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                        else:
                            tangca_chinhthuc = tangca_chinhthuc + \
                                max(leave_item['minutes'],
                                    leave_item['time']) * max(1, leave_item['multiplier_work_time'])

            except Exception as ex:
                print(ex)

            try:
                if (('đi muộn' in leave_item['holiday_status_name'].strip().lower()) or ('về sớm' in leave_item['holiday_status_name'].strip().lower())) and \
                            ((leave_item['for_reasons']=='1') or (leave_item['for_reasons']==1)):
                    # print("don di muon ve som duoc duyet voi ly do ca nhan")
                    total_dimuon_vesom = total_dimuon_vesom + \
                        max(leave_item['minutes'], leave_item['time'])
                    if leave_item['is_leave_probationary']:
                        total_dimuon_vesom_probationary = total_dimuon_vesom_probationary + \
                            max(leave_item['minutes'],
                                leave_item['time'])
            except Exception as ex:
                print("đi muộn CL ",ex)
        if pd.notnull(row['AUTO-CALCULATE-HOLIDAY']):
            if row['AUTO-CALCULATE-HOLIDAY'] == True:
                for _, holiday_item in holiday_report_before.iterrows():
                    if holiday_item['is_probationary']== True:
                        tangca_probation = tangca_probation + \
                                int((holiday_item['holiday_work_time'] - holiday_item['night_holiday_work_time_minute']) * 3 + \
                                 holiday_item['night_holiday_work_time_minute'] * row['NIGHT-HOLIDAY-WAGE'])
                    else:
                        tangca_chinhthuc = tangca_chinhthuc + \
                                int((holiday_item['holiday_work_time'] - holiday_item['night_holiday_work_time_minute']) * 3 + \
                                 holiday_item['night_holiday_work_time_minute'] * row['NIGHT-HOLIDAY-WAGE'])
        # if pd.notnull(row['AUTO-CALCULATE-HOLIDAY-WORKTIME-INCREASE']):
        #     if row['AUTO-CALCULATE-HOLIDAY-WORKTIME-INCREASE'] == True:
                holiday_report_current =  self.df_scheduling_ver_holiday_current_month \
                    [(self.df_scheduling_ver_holiday_current_month['employee_code']== str(row['code']))]
                
                for _, holiday_item in holiday_report_current.iterrows():
                    if holiday_item['is_probationary']== True:
                        # input(row['employee_code'])
                        currnet_month_auto_increase_probationary = currnet_month_auto_increase_probationary + \
                            int((holiday_item['holiday_work_time'] - holiday_item['night_holiday_work_time_minute']) * 3 + \
                                (holiday_item['night_holiday_work_time_minute'] * row['NIGHT-HOLIDAY-WAGE']))
                        print('currnet_month_auto_increase_probationary: ', currnet_month_auto_increase_probationary)
                    else:
                        
                        currnet_month_auto_increase_offical = currnet_month_auto_increase_offical + \
                            int((holiday_item['holiday_work_time'] - holiday_item['night_holiday_work_time_minute']) * 3 + \
                                (holiday_item['night_holiday_work_time_minute'] * row['NIGHT-HOLIDAY-WAGE']))
                        # print(currnet_month_auto_increase_offical)      
        # if row['employee_code']=='APG210924004':
        # # if (currnet_month_auto_increase_offical > 0) or (currnet_month_auto_increase_probationary >0):
        #     input(f"ex pro {row['employee_code']}")
        # if row['code']=='APG231102002':
        #     print("tangca_probation: ",tangca_probation)
        #     print("tangca_chinhthuc: ",tangca_chinhthuc)
        #     print(row['AUTO-CALCULATE-HOLIDAY'])
        #     input(f"{row['company_name']} Press enter to continute...")
        df_scheduling_ver_al_cl_last_month_employee = self.df_scheduling_ver_al_cl_last_month[ \
            (self.df_scheduling_ver_al_cl_last_month['employee_code'] == row['code']) & \
            (self.df_scheduling_ver_al_cl_last_month['company'] == row['company_name']) & \
            (self.df_scheduling_ver_al_cl_last_month['shift_name'] == 'CL') & \
            (self.df_scheduling_ver_al_cl_last_month['date'] < self.date_array[0])]
        
        for index, al_cl_shift in df_scheduling_ver_al_cl_last_month_employee.iterrows():
            find_leave_item = hr_leave_employee[
                (hr_leave_employee['date_str'] == al_cl_shift['date_str']) &
                ((hr_leave_employee['holiday_status_name'] == 'Nghỉ phép năm') |
                (hr_leave_employee['holiday_status_name'] == 'Nghỉ bù'))]

            if len(find_leave_item.index) == 0:
                # print(row)
                # print(row['code'])
                # print(al_cl_shift['date_str'])
                if al_cl_shift['is_probationary'] == True:
                    cl_probation = cl_probation + 480
                else:
                    cl_chinhthuc = cl_chinhthuc + 480
                # exit(1)
        last_month = self.date_array[0].month-1 if self.date_array[0].month>1 else 12
        update_data_last_year  = {}
        # convert bu nam truoc vao phat sinh tang thang 1
        check_increase = False
        # if (row['employee_code']== 'APG230929002'):
        #     print(cl_report_last_year.iloc[0])
        #     print(cl_report_last_year.iloc[0]["remaining_total_minute"])
        #     print(cl_report_before.iloc[0]["remaining_total_minute"])
        #     input('72238')
        if last_month == 1:
                
            remaining_leave_minute_last_month = 0
            try:
                if len(cl_report_before.index)>0:
                    remaining_leave_minute_last_month = int(cl_report_before.iloc[0]["remaining_total_minute"]) \
                                    if pd.notnull(cl_report_before.iloc[0]["remaining_total_minute"]) else 0
                # if (cl_report_last_year.iloc[0]['id']== 72238):
                #     for temp in range(1,13):
                #         print(cl_report_last_year.iloc[0])
                #     # input('72238')
            except:
                if len(cl_report_last_year.index)>0:
                    remaining_leave_minute_last_month = int(cl_report_last_year.iloc[0]["remaining_total_minute"]) \
                        if pd.notnull(cl_report_last_year.iloc[0]["remaining_total_minute"]) else 0
            # if (row['employee_code']== 'APG230929002'):
            #     print(cl_report_last_year.iloc[0])
            #     print(cl_report_last_year.iloc[0]["remaining_total_minute"])
            #     print(cl_report_before.iloc[0]["remaining_total_minute"])
            #     input('72238')
            if len(cl_report_last_year.index)>0:
                if pd.notnull(row['date_sign_employee']):
                    
                    if row['date_sign_employee'] >= (self.date_array[0].replace(month=1)):
                        phatsinhtang_thuviec = phatsinhtang_thuviec + remaining_leave_minute_last_month
                    else:
                        phatsinhtang_chinhthuc = phatsinhtang_chinhthuc + remaining_leave_minute_last_month
                else:
                    phatsinhtang_thuviec = phatsinhtang_thuviec + remaining_leave_minute_last_month
                for month_index in range(1,13):
                    try:
                        if len(cl_report_last_year.index) >0:
                            if (len(cl_report_before.index) > 0):
                                if int(cl_report_last_year.iloc[0][f"used_probationary_{month_index}"]) != \
                                        int(cl_report_before.iloc[0][f"used_probationary_{month_index}"]):
                                    is_update_last_year = True
                                update_data_last_year[f"used_probationary_{month_index}"] = \
                                        int(cl_report_before.iloc[0][f"used_probationary_{month_index}"])
                            else:
                                update_data_last_year[f"used_probationary_{month_index}"] = 0
                                is_update_last_year = True

                    except Exception as ex:
                        logger.error(f"used_probationary_{month_index}: {ex}")
                        if int(cl_report_last_year.iloc[0][f"used_probationary_{month_index}"]) != 0:
                            is_update_last_year = True
                        update_data_last_year[f"used_probationary_{month_index}"] = 0
                    try:
                        if len(cl_report_last_year.index) >0:
                            if (len(cl_report_before.index) > 0):
                                if int(cl_report_last_year.iloc[0][f"used_official_{month_index}"]) != \
                                        int(cl_report_before.iloc[0][f"used_official_{month_index}"]):    
                                    is_update_last_year = True
                                update_data_last_year[f"used_official_{month_index}"] = \
                                        int(cl_report_before.iloc[0][f"used_official_{month_index}"])
                            else:
                                update_data_last_year[f"used_official_{month_index}"] = 0
                                is_update_last_year = True   
                            
                    except Exception as ex:
                        logger.error(f"used_official_{month_index}: {ex}")
                        if int(cl_report_last_year.iloc[0][f"used_official_{month_index}"]) != 0:
                            is_update_last_year = True   
                        update_data_last_year[f"used_official_{month_index}"] = 0
                        # f"used_official_{month_index}": cl_chinhthuc,
                    try:
                        if len(cl_report_last_year.index) >0:
                            if (len(cl_report_before.index) > 0):
                                update_data_last_year[f"overtime_probationary_{month_index}"] = \
                                        int(cl_report_before.iloc[0][f"overtime_probationary_{month_index}"])
                                if int(cl_report_last_year.iloc[0][f"overtime_probationary_{month_index}"]) != \
                                        int(cl_report_before.iloc[0][f"overtime_probationary_{month_index}"]):
                                    is_update_last_year = True
                            else:
                                update_data_last_year[f"overtime_probationary_{month_index}"] = 0
                                is_update_last_year = True
                    except Exception as ex:
                        logger.error(f"overtime_probationary_{month_index}: {ex}")
                        
                        if int(cl_report_last_year.iloc[0][f"overtime_probationary_{month_index}"]) != 0:
                            is_update_last_year = True
                        update_data_last_year[f"overtime_probationary_{month_index}"] = 0
                        # f"overtime_probationary_{last_month}":tangca_probation,
                    try:
                        if (len(cl_report_last_year.index) >0):
                            if (len(cl_report_before.index) > 0):
                                update_data_last_year[f"overtime_official_{month_index}"] = \
                                        int(cl_report_before.iloc[0][f"overtime_official_{month_index}"])
                                if int(cl_report_last_year.iloc[0][f"overtime_official_{month_index}"]) != \
                                        int(cl_report_before.iloc[0][f"overtime_official_{month_index}"]):
                                    
                                    is_update_last_year = True
                            else:
                                update_data_last_year[f"overtime_official_{month_index}"] = 0
                                is_update_last_year = True
                    except Exception as ex:
                        logger.error(f"overtime_official_{month_index}: {ex}")
                        if int(cl_report_last_year.iloc[0][f"overtime_official_{month_index}"]) != 0:
                            is_update_last_year = True
                        update_data_last_year[f"overtime_official_{month_index}"] = 0
                        # f"overtime_official_{last_month}": tangca_chinhthuc,
                    try:
                        if len(cl_report_last_year.index) >0:
                            if (len(cl_report_before.index) > 0):
                                if int(cl_report_last_year.iloc[0][f"increase_official_{month_index}"]) != \
                                        int(cl_report_before.iloc[0][f"increase_official_{month_index}"]):

                                    is_update_last_year = True
                                update_data_last_year[f"increase_official_{month_index}"] = \
                                        int(cl_report_before.iloc[0][f"increase_official_{month_index}"])
                            else:
                                update_data_last_year[f"increase_official_{month_index}"] = 0
                                is_update_last_year = True

                    except Exception as ex:
                        logger.error(f"increase_official_{month_index}: {ex}")
                        if int(cl_report_last_year.iloc[0][f"increase_official_{month_index}"]) != 0:
                            
                            is_update_last_year = True
                        update_data_last_year[f"increase_official_{month_index}"] = 0
                        # f'increase_official_{last_month}': phatsinhtang_chinhthuc,
                    try:
                        if len(cl_report_last_year.index) >0:
                            if (len(cl_report_before.index) > 0):
                                if int(cl_report_last_year.iloc[0][f"increase_probationary_{month_index}"]) != \
                                        int(cl_report_before.iloc[0][f"increase_probationary_{month_index}"]):
                                    
                                    is_update_last_year = True
                                update_data_last_year[f"increase_probationary_{month_index}"] = \
                                        int(cl_report_before.iloc[0][f"increase_probationary_{month_index}"])
                            else:
                                update_data_last_year[f"increase_probationary_{month_index}"] = 0
                                is_update_last_year = True

                    except Exception as ex:
                        logger.error(f"increase_probationary_{month_index}: {ex}")
                        if len(cl_report_last_year.index) >0:
                            if int(cl_report_last_year.iloc[0][f"increase_probationary_{month_index}"]) != 0:
                                
                                is_update_last_year = True
                            update_data_last_year[f"increase_probationary_{month_index}"] = 0
                        # f'increase_probationary_{last_month}': phatsinhtang_thuviec}
            
                # except Exception as ex:
                
                #     logger.error(f'[{row['employee_code']}] last year remaining_total_minute: {ex}')
                if (len(cl_report_last_year.index)>0) and (is_update_last_year == True):
                    update_data_last_year['date_calculate']= (self.date_array[0].replace(day=1, month=1)).strftime("%Y-%m-%d")
                    ids = [int(cl_report_last_year.iloc[0]['id'])]
                    # if (ids[0]== 72238):
                    #     print(cl_report_last_year)
                        
                    #     print(cl_report_before)
                    #     print('remain', update_data_last_year)
                    #     input('sdf')
                    # try:
                    #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'write', [ids, update_data_last_year])
                    # except Exception as ex:
                    #     logger.error(f"write hr cl last year report employee: {ex}")
        if row[f'increase_official_{last_month}'] != phatsinhtang_chinhthuc:
            check_increase = True
        if row[f'increase_probationary_{last_month}'] != phatsinhtang_thuviec:
            check_increase = True

        current_cl_remaining_leave_minute_probationary = row['remaining_probationary_minute'] + \
                row[f"overtime_probationary_{last_month}"] \
                
        current_cl_remaining_leave_minute_offical = row['remaining_official_minute'] + \
                row[f"used_official_{last_month}"]
        #         # - row[f"overtime_official_{last_month}"] - row[f'increase_probationary_{last_month}']
        # total_compensation_leave_dimuon_vesom = min(cl_chinhthuc + cl_probation + total_dimuon_vesom, 
        #                 current_cl_remaining_leave_minute)
        if cl_probation <= current_cl_remaining_leave_minute_probationary:
            total_dimuon_vesom_probationary = min(total_dimuon_vesom_probationary, 
                    current_cl_remaining_leave_minute_probationary - cl_probation)
            cl_probation = cl_probation + total_dimuon_vesom_probationary
        if cl_chinhthuc <= current_cl_remaining_leave_minute_offical:
            total_dimuon_vesom_official = min((total_dimuon_vesom - total_dimuon_vesom_probationary),
                    current_cl_remaining_leave_minute_offical - cl_chinhthuc)
            cl_chinhthuc =cl_chinhthuc +  total_dimuon_vesom_official
        update_data = {}
        for temp_month_index in range(1,13):
            if temp_month_index != last_month:
                update_data[f"used_probationary_{temp_month_index}"] = 0
                update_data[f"used_official_{temp_month_index}"] = 0
                update_data[f"overtime_probationary_{temp_month_index}"] = 0
                update_data[f"overtime_official_{temp_month_index}"] = 0
                update_data[f'increase_official_{temp_month_index}'] = 0
                update_data[f'increase_probationary_{temp_month_index}'] = 0

        update_data[f"used_probationary_{last_month}"] = cl_probation
        update_data[f"used_official_{last_month}"] = cl_chinhthuc
        update_data[f"overtime_probationary_{last_month}"] = tangca_probation
        update_data[f"overtime_official_{last_month}"] = tangca_chinhthuc
        update_data[f'increase_official_{last_month}'] = phatsinhtang_chinhthuc
        update_data[f'increase_probationary_{last_month}'] = phatsinhtang_thuviec
        check_employee_info = False      

        # if pd.notnull(row['year']):
        if  row['year'] != self.date_array[0].year:
            check_employee_info = True
        update_data['year'] = self.date_array[0].year
        # if(row['id_cl']==182086):
        #     print('setp 1: ', update_data)
        #     input('setp 1')
        if pd.notnull(row['date_sign_employee']):
            if row['date_sign_cl'] != row['date_sign_employee']:
                check_employee_info = True
            update_data['date_sign'] = row['date_sign_employee'].strftime("%Y-%m-%d")
        # if(row['id_cl']==182086):
        #     print('setp 2: ', update_data)
        #     input('setp 2')
        if pd.notnull(row['workingday_employee']):
            if row['workingday_cl'] != row['workingday_employee']:
                check_employee_info = True
            update_data['workingday'] = row['workingday_employee'].strftime("%Y-%m-%d")
        if pd.notnull(row['severance_day_employee']):
            if row['severance_day_cl'] != row['severance_day_employee']:
                check_employee_info = True
            update_data['severance_day'] = row['severance_day_employee'].strftime("%Y-%m-%d")
              
        if len(cl_report_before.index)>0:
            for month_index in range (1,13):
                if month_index < last_month:
                    try:
                        update_data[f"used_probationary_{month_index}"] = int(cl_report_before.iloc[0][f"used_probationary_{month_index}"]) \
                                if pd.notnull(cl_report_before.iloc[0][f"used_probationary_{month_index}"]) else 0
                        update_data[f"used_official_{month_index}"] = int(cl_report_before.iloc[0][f"used_official_{month_index}"] ) \
                                if pd.notnull(cl_report_before.iloc[0][f"used_official_{month_index}"] ) else 0
                        update_data[f"overtime_probationary_{month_index}"] = int(cl_report_before.iloc[0][f"overtime_probationary_{month_index}"]) \
                                if pd.notnull(cl_report_before.iloc[0][f"overtime_probationary_{month_index}"]) else 0
                        update_data[f"overtime_official_{month_index}"] = int(cl_report_before.iloc[0][f"overtime_official_{month_index}"] ) \
                                if pd.notnull(cl_report_before.iloc[0][f"overtime_official_{month_index}"] ) else 0
                    except Exception as ex:
                        logger.error("last", ex)
                        update_data[f"used_probationary_{month_index}"] =  0
                        update_data[f"used_official_{month_index}"] =  0
                        update_data[f"overtime_probationary_{month_index}"] =  0
                        update_data[f"overtime_official_{month_index}"] =  0
                    if row[f'increase_official_{month_index}'] != int(cl_report_before.iloc[0][f"increase_official_{month_index}"]):
                        check_increase = True 
                    update_data[f'increase_official_{month_index}'] = int(cl_report_before.iloc[0][f"increase_official_{month_index}"])
                    
                    if row[f'increase_probationary_{month_index}'] != int(cl_report_before.iloc[0][f"increase_probationary_{month_index}"]):
                        check_increase = True
                    update_data[f'increase_probationary_{month_index}'] = int(cl_report_before.iloc[0][f"increase_probationary_{month_index}"])
                         
                elif month_index > last_month:
                    
                    try:
                        if int(row[f'used_probationary_{month_index}']) != 0:
                            check_increase = True
                    except:
                        check_increase = True
                    update_data[f"used_probationary_{month_index}"] =  0
                        
                    try:
                        if int(row[f'used_official_{month_index}']) != 0:
                            check_increase = True
                    except:
                        check_increase = True
                    update_data[f"used_official_{month_index}"] =  0
                    
                    try:
                        update_data[f"overtime_probationary_{month_index}"] =  int(row[f'overtime_probationary_{month_index}'])
                    except:
                        update_data[f"overtime_probationary_{month_index}"] = 0
                    try:
                        update_data[f"overtime_official_{month_index}"] =  int(row[f'overtime_official_{month_index}'])
                    except:
                        update_data[f"overtime_official_{month_index}"] = 0
                    
                    if pd.isnull(row['AUTO-CALCULATE-HOLIDAY']) or (row['AUTO-CALCULATE-HOLIDAY'] == False) :
                        try:
                            if int(row[f'overtime_probationary_{month_index}']) != 0:
                                check_increase = True
                        except:
                            check_increase = True
                        try:
                            if int(row[f'overtime_official_{month_index}']) != 0:
                                check_increase = True
                        except:
                            check_increase = True   
                    else:
                        try:
                            if int(row[f'overtime_probationary_{month_index}']) != 0:
                                check_increase = True
                        except:
                            check_increase = True
                        try:
                            if int(row[f'overtime_official_{month_index}']) != 0:
                                check_increase = True
                        except:
                            check_increase = True        
                        
                    update_data[f'increase_official_{month_index}'] = 0    
                    try:
                        if int(row[f'increase_official_{month_index}']) != 0:
                            check_increase = True
                    except:
                        check_increase = True
                    
                    update_data[f'increase_probationary_{month_index}'] = 0
                    try:
                        if int(row[f'increase_probationary_{month_index}']) != 0:
                            
                            check_increase = True
                    except:
                        check_increase = True
        for index, al_cl_shift in df_scheduling_ver_al_cl_current_month_employee.iterrows():
            find_leave_item = self.df_hr_leave[
                (self.df_hr_leave['date_str'] == al_cl_shift['date_str']) &
                ((self.df_hr_leave['holiday_status_name'] == 'Nghỉ phép năm') |
                (self.df_hr_leave['holiday_status_name'] == 'Nghỉ bù'))]
            if len(find_leave_item.index) == 0:
                print(row)
                print(row['code'])
                print(al_cl_shift['date_str'])
                if al_cl_shift['is_probationary'] == True:
                    total_cl_from_cl_shift_probationary = total_cl_from_cl_shift_probationary + 480
                else:
                    total_cl_from_cl_shift_offical = total_cl_from_cl_shift_offical + 480
                # input(row['company_name'])

                # exit(1)
        
        # if pd.notnull(row['AUTO-CALCULATE-HOLIDAY']):
        #     if row['AUTO-CALCULATE-HOLIDAY'] == True:
        if pd.isnull(total_cl_from_cl_shift_offical):
            total_cl_from_cl_shift_offical = 0
        if pd.isnull(total_cl_from_cl_shift_probationary):
            total_cl_from_cl_shift_probationary = 0
        if pd.isnull(currnet_month_auto_increase_offical):
            currnet_month_auto_increase_offical = 0
        if pd.isnull(currnet_month_auto_increase_probationary):
            currnet_month_auto_increase_probationary = 0
        month_index = self.date_array[0].month
        if pd.notnull(row[f'overtime_official_{month_index}']):
            if int(row[f'overtime_official_{month_index}']) != int(currnet_month_auto_increase_offical ):
                check_increase = True
            update_data[f"overtime_official_{month_index}"] = int(currnet_month_auto_increase_offical )
        elif int(currnet_month_auto_increase_offical) != 0:
            update_data[f"overtime_official_{month_index}"] = int( currnet_month_auto_increase_offical)
            check_increase = True
            
        if pd.notnull(row[f'used_official_{month_index}']):
            if int(row[f'used_official_{month_index}']) != int(total_cl_from_cl_shift_offical):
                check_increase = True
            update_data[f"used_official_{month_index}"] = int(total_cl_from_cl_shift_offical)
                
        elif int(total_cl_from_cl_shift_offical) != 0:
            update_data[f"used_official_{month_index}"] = int( total_cl_from_cl_shift_offical)
            check_increase = True

        if pd.notnull(row[f'overtime_probationary_{month_index}']):
            if int(row[f'overtime_probationary_{month_index}']) != int(currnet_month_auto_increase_probationary):
                
                check_increase = True
            update_data[f"overtime_probationary_{month_index}"] = int(currnet_month_auto_increase_probationary)
        elif int(currnet_month_auto_increase_probationary) != 0:
            update_data[f"overtime_probationary_{month_index}"] = int(currnet_month_auto_increase_probationary)
            check_increase = True
            
        if pd.notnull(row[f'used_probationary_{month_index}']):
            if int(row[f'used_probationary_{month_index}']) != int(total_cl_from_cl_shift_probationary):
                
                check_increase = True
            update_data[f"used_probationary_{month_index}"] = int(total_cl_from_cl_shift_probationary)
        elif int(total_cl_from_cl_shift_probationary) != 0:
            update_data[f"used_probationary_{month_index}"] = int(total_cl_from_cl_shift_probationary)
            check_increase = True    
        
        print(currnet_month_auto_increase_probationary)
        print(currnet_month_auto_increase_offical)
        # if row['employee_code']=='APG210924004':
        # # if (currnet_month_auto_increase_offical > 0) or (currnet_month_auto_increase_probationary >0):
        #     input(row['employee_code'])
 
        if(pd.notnull(row['id_cl'])) and (new_date == False):
            # print(f"update', {row['id_cl']}, {update_data} - increase: {check_increase} - info: {check_employee_info}")
            ids = [int(row['id_cl'])]
            if (int(row[f"used_probationary_{last_month}"]) != cl_probation or
                        int(row[f"used_official_{last_month}"]) != cl_chinhthuc or
                        int(row[f"overtime_probationary_{last_month}"]) != tangca_probation or
                        int(row[f"overtime_official_{last_month}"]) != tangca_chinhthuc or
                        int(row[f"increase_official_{last_month}"]) != phatsinhtang_chinhthuc or
                        int(row[f"increase_probationary_{last_month}"]) != phatsinhtang_thuviec
                        ) or check_increase or check_employee_info:
                try:
                    # self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'write', [ids, update_data])
                    self.cl_result_update[f"{int(row['id_cl'])}"]=update_data

                except Exception as ex:
                    logger.error(f'write cl ex: {ex}')
            # else:
            #     print("CL Existed in db")
            # if(row['employee_code']=='APG230407009'):
            #     print(update_data)
            #     print('remaining_leave_minute_last_month aaaaaaaaaaaaaaaaaaa: ', remaining_leave_minute_last_month)
            #     print('cl_report_before: ', cl_report_before)
            #     print('phatsinhtang_thuviec: ', phatsinhtang_thuviec)
            #     print('phatsinhtang_chinhthuc: ', phatsinhtang_chinhthuc)
            #     print(int(row[f"increase_official_{last_month}"]))
            #     print(int(row[f"increase_probationary_{last_month}"]))
            #     print(row['date_sign_employee'])
            #     print(row['date_sign_employee'] > (self.date_array[0].replace(month=1)))
            #     exit(1)
        else:
            # print('create', update_data)
            update_data["employee_id"]= row['id']
            update_data["company_id"]= row['company_id'][0]
            update_data['date_calculate']= (self.date_array[-1] + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
            update_data['year'] = self.date_array[0].year
            # update_data['year'] =  self.date_array[0].year -1 if last_month==12 else self.date_array[0].year
            # try:
            #     update_data['date_calculate']= max_date_calculate.strftime("%Y-%m-%d")
            #     update_data['year'] = self.date_array[0].year
            # except:
            #     update_data['date_calculate']= self.date_array[0].strftime("%Y-%m-%d")
            #     update_data['year'] = self.date_array[0].year
            try:
                # id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 
                #     'hr.cl.report', 'create', [update_data])
                update_data['date_calculate']= (self.date_array[-1]+datetime.timedelta(days=1)).strftime("%Y-%m-%d")
                print('created', update_data)
                self.cl_result_create[row['code']] = update_data
            except Exception as ex:
                logger.error(f"ex in hr.cl.report. create: {ex}")
        # if(row['employee_code']=='APG230407009'):
        #     print(update_data)
        #     print('last_month', last_month)
        #     print('cl_report_before: ', cl_report_before)
        #     print('no update')
        #     print('remaining_leave_minute_last_month: ', remaining_leave_minute_last_month)
        #     print('phatsinhtang_thuviec: ', phatsinhtang_thuviec)
        #     print('phatsinhtang_chinhthuc: ', phatsinhtang_chinhthuc)
        #     print(int(row[f"increase_official_{last_month}"]))
        #     print(int(row[f"increase_probationary_{last_month}"]))
        #     exit(1)  
        # if row['employee_code']=='APG220301012':
        #     cl_report_before.to_excel('cl_report_before_hanhy.xlsx')
        #     hr_leave_employee.to_excel('hr_leave_employee_APG220301012.xlsx')
            # exit()
        # if row['employee_code'] == 'APG221226008':
        #     exit()
        increase_probationary_1 = update_data['increase_probationary_1']
        increase_official_1 = update_data['increase_official_1']
        # if row['employee_code']=='APG220208015':
        #     input(f': {increase_probationary_1}')
        #     input(f': {increase_official_1}')
        return tangca_probation, tangca_chinhthuc, cl_probation, \
                    cl_chinhthuc, total_dimuon_vesom, \
                    total_dimuon_vesom_probationary, total_dimuon_vesom_official, \
                    increase_probationary_1, increase_official_1


    def merge_holiday(self, row):
        is_holiday = False
        holiday_from = None
        holiday_to = None
        holiday_name = ''
        try:
            shift_time = row['request_date_from']
            df_compare = self.df_resource_calendar_leaves[(self.df_resource_calendar_leaves['date_from'] <= shift_time) &
                                                          (self.df_resource_calendar_leaves['date_to'] > shift_time)]

            if len(df_compare.index) > 0:
                fist_row = df_compare.iloc[0]
                is_holiday = True
                holiday_from = fist_row['date_from']
                holiday_to = fist_row['date_to']
                holiday_name = fist_row['name']

        except Exception as ex:
            is_holiday = False

            print('merge error: ', ex)

        return is_holiday, holiday_from, holiday_to, holiday_name

    def refact_date_from(self, row):
        date_from = row['date_from']
        print('date {} from {}'.format(date_from, type(date_from)))
        try:
            date_from = datetime.datetime.strptime(
                row['date_from'], '%Y-%m-%d %H:%M:%S').replace(hour=0, minute=0, second=0)
        except Exception as ex:
            print(ex)
        date_to = row['date_to']
        try:
            date_to = datetime.datetime.strptime(
                row['date_to'], '%Y-%m-%d %H:%M:%S').replace(hour=0, minute=0, second=0)
            date_to = date_to + datetime.timedelta(days=1)
        except Exception as ex:
            print('date_to', ex)
        return date_from, date_to

    def cal_proba(self, x):
        result_probation = True
        result_holiday_status_name = False
        try:
            if (pd.isnull(x['probationary_contract_termination_date'])):
                df_contract = contract_employee_filter = self.df_contract[(self.df_contract['date_start']<=self.date_array[len(self.date_array)-1])&
                                            (self.df_contract['date_end']>=self.date_array[0])&
                                            (self.df_contract['employee_code']==x['employee_code'])].sort_values("date_end", ascending=False)
            # print(x['probationary_contract_termination_date'])
            if (pd.isnull(x['probationary_contract_termination_date'])) or \
                    (x['date'] <= x['probationary_contract_termination_date']):
                result_probation = True
            else:
                result_probation =  False
        except:
            result_probation = False
        try:
            result_holiday_status_name = x['holiday_status_id'][1]
        except:
            result_holiday_status_name = False
        return result_probation, result_holiday_status_name

    def split_date(self, row):
        try:
            return row['request_date_from'].split(' ')[0]
        except:
            return ''
    def download_attendance_report_holiday_last_month(self):
        last_month = self.date_array[0] - datetime.timedelta(days=10)
        _, lastday = calendar.monthrange(last_month.year, last_month.month)

        # start_date = last_month.replace(day=1).strftime('%Y-%m-%d 00:00:00')
        # end_date = last_month.replace(day=lastday).strftime('%Y-%m-%d 00:00:00')
        print(last_month)
        print(lastday)
        start_datetime_str = (last_month.replace(day=1) - relativedelta(months=1) ).strftime('%Y-%m-%d 00:00:00')
        end_datetime_str = (last_month.replace(day=lastday) + relativedelta(months=1)).strftime('%Y-%m-%d 23:59:59')
        print(start_datetime_str)
        print(end_datetime_str)
        resource_calendar_leaves_ids = self.models.execute_kw(
            self.db, self.uid, self.password, 'resource.calendar.leaves', 'search', 
                                [["&", "&" ,('calendar_id','=',False),
                                  ('date_from','<=', end_datetime_str),
                                  ('date_to','>=', start_datetime_str)]])
        resource_calendar_leaves_list = self.models.execute_kw(self.db, self.uid, self.password, 'resource.calendar.leaves', 'read', [resource_calendar_leaves_ids],
                                                               {'fields': ['id', 'name', 'company_id', 'calendar_id', 'date_from', 'date_to', 'resource_id', 'time_type']})
        # print(resource_calendar_leaves_list)
        # input('end_datetime_str')
        for calendar_leave in resource_calendar_leaves_list:
            try:
                company_name = calendar_leave['company_id'][1]
                calendar_leave['company_name'] = company_name
            except:
                calendar_leave['company_name'] = ''

        self.df_resource_calendar_leaves_last_month = pd.DataFrame.from_dict(
            resource_calendar_leaves_list)

        print('*************** last CALENDAR************************************')
        try:
            self.df_resource_calendar_leaves_last_month[['date_from', 'date_to']] = self.df_resource_calendar_leaves_last_month.apply(
                lambda row: self.refact_date_from(row), axis=1, result_type='expand')
        except:
            self.df_resource_calendar_leaves_last_month = pd.DataFrame(
                columns=['date_from', 'date_to','company_name'])
        index = 0
        fields = ['id', 'employee_name', 'date', 'shift_name', 'employee_code', 'company','additional_company',
                    'shift_start', 'shift_end', 'rest_start', 'rest_end', 'rest_shift', 'probation_completion_wage',
                    'total_shift_work_time', 'total_work_time', 'time_keeping_code', 'kid_time',
                    'department', 'attendance_attempt_1', 'attendance_attempt_2', 'minutes_working_reduced',
                    'attendance_attempt_3', 'attendance_attempt_4', 'attendance_attempt_5',
                    'attendance_attempt_6', 'attendance_attempt_7', 'attendance_attempt_8',
                    'attendance_attempt_9', 'attendance_attempt_10', 'attendance_attempt_11',
                    'attendance_attempt_12', 'attendance_attempt_13', 'attendance_attempt_14',
                    'attendance_inout_1','attendance_inout_2','attendance_inout_3',
                    'attendance_inout_4','attendance_inout_5','attendance_inout_6',
                    'attendance_inout_7','attendance_inout_8','attendance_inout_9', 'amount_al_reserve', 'amount_cl_reserve',
                    'attendance_inout_10','attendance_inout_11','attendance_inout_12', 
                    'attendance_inout_13','attendance_inout_14','attendance_inout_15','actual_total_work_time', 'standard_working_day',
                    'attendance_attempt_15', 'last_attendance_attempt', 'night_hours_normal', 'night_hours_holiday', 'probation_wage_rate', 
                    'split_shift', 'missing_checkin_break', 'leave_early', 'attendance_late', 'night_shift', 'minute_worked_day_holiday', 'total_attendance',
                    'ot_holiday', 'ot_normal']
        for _, holiday in self.df_resource_calendar_leaves_last_month.iterrows():
            start_str = (holiday['date_from'] - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            end_str = (holiday['date_to'] + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
            domain = ["&","&", ("date", ">=", start_str),
                                ("date", "<", end_str),
                              ('company', '=', holiday['company_name'])
                        ]
            # if (self.select_company_id)> 0:
            #     domain = ["&","&","&", 
            #           ("company", ">=", self.select_company_name),
            #           ("date", ">=", start_str),
            #           ("date", "<", end_str),
            #           ('company', '=', holiday['company_name'])]
            
            
            
            ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [domain], {})
            if len(ids) == 0:
                continue
            list_scheduling_ver = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'read', [ids],
                                                        {'fields': fields})
            
            if index == 0:
                self.df_scheduling_ver_holiday_last_month = pd.DataFrame.from_dict(list_scheduling_ver)
                self.df_scheduling_ver_holiday_last_month['holiday_from'] = holiday['date_from']
                self.df_scheduling_ver_holiday_last_month['is_holiday'] = True
                self.df_scheduling_ver_holiday_last_month['holiday_to'] = holiday['date_to']
                self.df_scheduling_ver_holiday_last_month.set_index('id')

            else:
                df = pd.DataFrame.from_dict(list_scheduling_ver)
                df['holiday_from'] = holiday['date_from']
                df['is_holiday'] = True
                df['holiday_to'] = holiday['date_to']
                df.set_index('id')
                
                self.df_scheduling_ver_holiday_last_month = pd.concat([self.df_scheduling_ver_holiday_last_month,
                    df], ignore_index=True)
            print(f'load df_scheduling_ver_holiday_last_month {holiday["company_name"]} {start_str} - {end_str}- {len(ids)}')
            
            index = index + 1
        if index == 0:
            self.df_scheduling_ver_holiday_last_month = pd.DataFrame(
                columns=fields)
        else:
            self.df_scheduling_ver_holiday_last_month['employee_code'] = self.df_scheduling_ver_holiday_last_month['employee_code'].str.strip(
            )
            # print(list_scheduling_ver)
        self.df_scheduling_ver_holiday_last_month['date_str'] = self.df_scheduling_ver_holiday_last_month['date']
        self.df_scheduling_ver_holiday_last_month['date'] = pd.to_datetime(
            self.df_scheduling_ver_holiday_last_month['date'], format='%Y-%m-%d',  errors='coerce')
        for i in range(1,16):
            self.df_scheduling_ver_holiday_last_month[f'attendance_attempt_{i}'] = pd.to_datetime(
                self.df_scheduling_ver_holiday_last_month[f'attendance_attempt_{i}'], format="%Y-%m-%d %H:%M:%S",  errors='coerce')
        self.df_scheduling_ver_holiday_last_month['last_attendance_attempt'] = pd.to_datetime(
                self.df_scheduling_ver_holiday_last_month['last_attendance_attempt'], format="%Y-%m-%d %H:%M:%S",  errors='coerce')
        self.df_scheduling_ver_holiday_last_month['probation_completion_wage'] = pd.to_datetime(
            self.df_scheduling_ver_holiday_last_month['probation_completion_wage'], format='%Y-%m-%d',  errors='coerce')
        # print(self.df_scheduling_ver)
        if len(self.df_scheduling_ver_holiday_last_month) > 0:
            self.df_scheduling_ver_holiday_last_month[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                self.df_scheduling_ver_holiday_last_month.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
            self.df_scheduling_ver_holiday_last_month[['probationary_contract_termination_date', 'is_probationary', 
                    'contract_company','probationary_salary_rate']] = self.df_scheduling_ver_holiday_last_month.apply(
                    lambda x: self.calculate_is_probationary(x), axis=1, result_type='expand')
            self.df_scheduling_ver_holiday_last_month[['night_holiday_work_time_minute','holiday_work_time', 'night_worktime_minute']] = \
                    self.df_scheduling_ver_holiday_last_month.apply(lambda row: self.calculate_night_holiday_work_time(row),  axis=1, result_type='expand')
    def calculate_night_holiday_work_time(self, row):
        """
        calculate night holiday work time with real time in, time out , restime in out
        cobine night holiday time
        """
        holiday_work_time = 0
        night_worktime_minute = 0
        # if not row['is_holiday']:
        #     return 0, holiday_work_time, 0
        
        try:
            # rest_shifts = row['rest_shifts']
            total_work_time = row['total_work_time']
            # if rest_shifts or total_work_time == 0:
            #     return 0
            real_time_in = row['attendance_attempt_1'].replace(second=0)
            real_time_out = row['last_attendance_attempt'].replace(second=0)
            for time_item_index in range(1, 16):
                try:
                    if isinstance(row[f'attendance_attempt_{time_item_index}'], datetime.datetime):
                        if real_time_out < row[f'attendance_attempt_{time_item_index}']:
                            real_time_out = row[f'attendance_attempt_{time_item_index}']
                            
                        if real_time_in > row[f'attendance_attempt_{time_item_index}']:
                            real_time_in = row[f'attendance_attempt_{time_item_index}']
                except Exception as ex:
                    print('realtime out min max error: ', ex)
                    
            # if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
            #     return 0, holiday_work_time, 0
            start_work_date_time = real_time_out
            if pd.notnull(row['shift_start_datetime']):
                start_work_date_time = row['shift_start_datetime'].replace(second=0)
            end_work_date_time = real_time_out
            if pd.notnull(row['shift_end_datetime']):
                end_work_date_time = row['shift_end_datetime'].replace(second=0)
          

            # start night datetime from 22h to 6h:00
            start_night_date_time = start_work_date_time.replace(hour=22, minute=0, second=0)
            end_night_date_time = (start_night_date_time + datetime.timedelta(days=1)).replace(hour=6)
            try:
                start_holiday_datetime = row['holiday_from'].replace(hour=0, minute=0, second=0)
                end_holiday_datetime = row['holiday_to'].replace(hour=0, minute=0, second=0)
            except:
                start_holiday_datetime = None
                end_holiday_datetime = None
            start_rest_date_time = row['rest_start_datetime'].replace(second=0)
            end_rest_date_time = row['rest_end_datetime'].replace(second=0)
            # night stage fist
            current_program_night = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time)
            stage_fist_night = max(0, current_program_night.total_seconds()/60.0)
            
            #night holiday
            if not (start_holiday_datetime is None):
                current_program = min(real_time_out, start_rest_date_time, end_holiday_datetime, end_night_date_time) - \
                        max(real_time_in, start_work_date_time, start_night_date_time, start_holiday_datetime)
                stage_fist = max(0, current_program.total_seconds()/60.0)
            
                
                # holiday stage fist
                current_program_holiday = min(real_time_out, start_rest_date_time, end_holiday_datetime) - \
                        max(real_time_in, start_work_date_time, start_holiday_datetime)
                stage_fist_holiday = max(0, current_program_holiday.total_seconds()/60.0)
            else:
                stage_fist = 0
                stage_fist_holiday = 0
            # night stage second
            current_program_night = min(real_time_out, end_work_date_time, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time)
            stage_second_night = max(0, current_program_night.total_seconds()/60.0)
            
            if not (start_holiday_datetime is None):
                # night holiday
                current_program = min(real_time_out, end_work_date_time, end_holiday_datetime, end_night_date_time) - \
                        max(real_time_in, end_rest_date_time, start_night_date_time, start_holiday_datetime)
                stage_second = max(0, current_program.total_seconds()/60.0)
                
                # holiday stage second
                current_program_holiday = min(real_time_out, end_work_date_time, end_holiday_datetime) - \
                        max(real_time_in, end_rest_date_time, start_holiday_datetime)
                stage_second_holiday = max(0, current_program_holiday.total_seconds()/60.0)
            else:
                stage_second = 0
                stage_second_holiday = 0
            # print('{}: night after now - holiday: {} - {} - {} -night - {}- {} - real - {} - {} '.format(row['name_employee'] ,row['name'], \
            #                     stage_fist, stage_second, start_night_date_time, end_night_date_time, row['start_work_date_time'], row['end_work_date_time']))
            
            stage_night_only = int(stage_fist_night + stage_second_night)
            stage_night =  int(stage_fist + stage_second)
            holiday_work_time = int(stage_fist_holiday + stage_second_holiday)
            
            end_night_date_time = start_work_date_time.replace(hour=6, minute=0, second=0)
            start_night_date_time = (end_night_date_time+ datetime.timedelta(days=-1)).replace(hour=22, minute=0, second=0)

            # night stage 1
            current_program_night = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time)
            stage_fist_night = max(0, current_program_night.total_seconds()/60.0)
            if not (start_holiday_datetime is None):
                # night holiday
                current_program = min(real_time_out, start_rest_date_time, end_holiday_datetime, end_night_date_time) - \
                        max(real_time_in, start_work_date_time, start_night_date_time, start_holiday_datetime)
                stage_fist = max(0, current_program.total_seconds()/60.0)
            else:
                stage_fist = 0
            # night stage 2

            current_program_night = min(real_time_out, end_work_date_time, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time)
            stage_second_night = max(0, current_program_night.total_seconds()/60.0)
            if not (start_holiday_datetime is None):
                # night holiday stage 2
                current_program = min(real_time_out, end_work_date_time, end_holiday_datetime, end_night_date_time) - \
                        max(real_time_in, end_rest_date_time, start_night_date_time, start_holiday_datetime)
                stage_second = max(0, current_program.total_seconds()/60.0)
            else:
                stage_second = 0
            # print('{}: night before now - holiday: {} - {} - {} -night - {}- {} - real - {} - {} '.format(row['name_employee'] ,row['name'], \
            #                     stage_fist, stage_second, start_night_date_time, end_night_date_time, row['start_work_date_time'], row['end_work_date_time']))
            
            stage_morning = int(stage_fist + stage_second)
            stage_morning_only = int(stage_fist_night + stage_second_night)
            night_worktime_minute = int(stage_night_only + stage_morning_only)
            # if row['employee_code'] == 'APG230412016':
            #     print(row['date'])
            #     print('stage_night_only: ', stage_night_only)
            #     print('stage_night_only: ', stage_morning_only)
            #     input('press enter....')
            return min(total_work_time, stage_night + stage_morning), min(total_work_time, holiday_work_time), min(total_work_time, night_worktime_minute)

        except Exception as ex:
            print("calcuate night - holiday time err", ex)
            # print('{}: night - holiday: {} - {} - {} -night - {}- {} - real - {} - {}'.format(row['employee_name'] ,row['shift_name'], 0, 0, 0, 0, \
            #                                                                         row['start_work_date_time'], row['end_work_date_time']))
            return 0, min(total_work_time, holiday_work_time), min(total_work_time, night_worktime_minute)
        
    
    def download_attendance_report_al_cl_last_month(self):
        last_month = self.date_array[0] - datetime.timedelta(days=10)
        _, lastday = calendar.monthrange(last_month.year, last_month.month)

        # start_date = last_month.replace(day=1).strftime('%Y-%m-%d 00:00:00')
        # end_date = last_month.replace(day=lastday).strftime('%Y-%m-%d 00:00:00')
        
        start_str = last_month.replace(day=1).strftime('%Y-%m-%d')
        LIMIT_SIZE = 1000
       
        end_str = last_month.replace(day=lastday).strftime('%Y-%m-%d')
        
        # self.df_old = pd.read_excel(self.input_attendence_file , index_col=None, header=[0,] ,sheet_name='Sheet1')
        # self.df_old['is_from_explanation'] = False
        domain = ["&","&", ("date", ">=", start_str),
                                    ("date", "<=", end_str),
                                    '|',
                                    ('shift_name','=','AL'),
                                    ('shift_name','=','CL')
                                      # ('active','=',True),
                                    #   ('company', '=', self.company_name)
                                    ]
        if (self.select_company_id > 0):
            domain = ["&","&", "&", ('company', '=', self.select_company_name), ("date", ">=", start_str),
                                    ("date", "<=", end_str),
                                    '|',
                                    ('shift_name','=','AL'),
                                    ('shift_name','=','CL')
                                      # ('active','=',True),
                                    #   ('company', '=', self.company_name)
                                    ]
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [domain], {'offset': 0, 'limit': LIMIT_SIZE})
        list_scheduling_ver = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'read', [ids],
                                                     {'fields': ['id', 'employee_name', 'date', 'shift_name', 'employee_code', 'company','additional_company',
                                                                 'shift_start', 'shift_end', 'rest_start', 'rest_end', 'rest_shift', 'probation_completion_wage',
                                                                 'total_shift_work_time', 'total_work_time', 'time_keeping_code', 'kid_time',
                                                                 'department', 'attendance_attempt_1', 'attendance_attempt_2','minutes_working_reduced',
                                                                 'attendance_attempt_3', 'attendance_attempt_4', 'attendance_attempt_5',
                                                                 'attendance_attempt_6', 'attendance_attempt_7', 'attendance_attempt_8',
                                                                 'attendance_attempt_9', 'attendance_attempt_10', 'attendance_attempt_11',
                                                                 'attendance_attempt_12', 'attendance_attempt_13', 'attendance_attempt_14',
                                                                 'attendance_inout_1','attendance_inout_2','attendance_inout_3',
                                                                 'attendance_inout_4','attendance_inout_5','attendance_inout_6',
                                                                 'attendance_inout_7','attendance_inout_8','attendance_inout_9', 'amount_al_reserve', 'amount_cl_reserve',
                                                                 'attendance_inout_10','attendance_inout_11','attendance_inout_12', 
                                                                 'attendance_inout_13','attendance_inout_14','attendance_inout_15','actual_total_work_time', 'standard_working_day',
                                                                 'attendance_attempt_15', 'last_attendance_attempt', 'night_hours_normal', 'night_hours_holiday', 'probation_wage_rate', 
                                                                 'split_shift', 'missing_checkin_break', 'leave_early', 'attendance_late', 'night_shift', 'minute_worked_day_holiday', 
                                                                 'total_attendance', 'ot_holiday','ot_normal']})

        self.df_scheduling_ver_al_cl_last_month = pd.DataFrame.from_dict(list_scheduling_ver)

        index = 1
        while len(ids) == LIMIT_SIZE:
            ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [domain],  {'offset': index * LIMIT_SIZE, 'limit': LIMIT_SIZE})
            list_scheduling_ver = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'read', [ids],
                                                     {'fields': ['id', 'employee_name', 'date', 'shift_name', 'employee_code', 'company','additional_company',
                                                                 'shift_start', 'shift_end', 'rest_start', 'rest_end', 'rest_shift', 'probation_completion_wage',
                                                                 'total_shift_work_time', 'total_work_time', 'time_keeping_code', 'kid_time',
                                                                 'department', 'attendance_attempt_1', 'attendance_attempt_2', 'minutes_working_reduced',
                                                                 'attendance_attempt_3', 'attendance_attempt_4', 'attendance_attempt_5',
                                                                 'attendance_attempt_6', 'attendance_attempt_7', 'attendance_attempt_8',
                                                                 'attendance_attempt_9', 'attendance_attempt_10', 'attendance_attempt_11',
                                                                 'attendance_attempt_12', 'attendance_attempt_13', 'attendance_attempt_14',
                                                                 'attendance_inout_1','attendance_inout_2','attendance_inout_3',
                                                                 'attendance_inout_4','attendance_inout_5','attendance_inout_6',
                                                                 'attendance_inout_7','attendance_inout_8','attendance_inout_9', 'amount_al_reserve', 'amount_cl_reserve',
                                                                 'attendance_inout_10','attendance_inout_11','attendance_inout_12', 
                                                                 'attendance_inout_13','attendance_inout_14','attendance_inout_15','actual_total_work_time', 'standard_working_day',
                                                                 'attendance_attempt_15', 'last_attendance_attempt', 'attendance_inout_last', 'night_hours_normal', 'night_hours_holiday', 'probation_wage_rate', 
                                                                 'split_shift', 'missing_checkin_break', 'leave_early', 'attendance_late', 'night_shift', 'minute_worked_day_holiday',
                                                                 'total_attendance', 'ot_holiday', 'ot_normal']})
            self.df_scheduling_ver_al_cl_last_month = pd.concat([self.df_scheduling_ver_al_cl_last_month,
                pd.DataFrame.from_dict(list_scheduling_ver)], ignore_index=True)

            index = index + 1
            print(f'load df_scheduling_ver_al_cl_last_month {index}- {len(ids)}')
        print('complete list scheduling')

        self.df_scheduling_ver_al_cl_last_month['employee_code'] = self.df_scheduling_ver_al_cl_last_month['employee_code'].str.strip(
        )
        # print(list_scheduling_ver)
        self.df_scheduling_ver_al_cl_last_month['date_str'] = self.df_scheduling_ver_al_cl_last_month['date']
        self.df_scheduling_ver_al_cl_last_month['date'] = pd.to_datetime(
            self.df_scheduling_ver_al_cl_last_month['date'], format='%Y-%m-%d',  errors='coerce')
        for i in range(1,16):
            self.df_scheduling_ver_al_cl_last_month[f'attendance_attempt_{i}'] = pd.to_datetime(
                self.df_scheduling_ver_al_cl_last_month[f'attendance_attempt_{i}'], format="%Y-%m-%d %H:%M:%S",  errors='coerce')
        self.df_scheduling_ver_al_cl_last_month['last_attendance_attempt'] = pd.to_datetime(
                self.df_scheduling_ver_al_cl_last_month['last_attendance_attempt'], format="%Y-%m-%d %H:%M:%S",  errors='coerce')
        self.df_scheduling_ver_al_cl_last_month['probation_completion_wage'] = pd.to_datetime(
            self.df_scheduling_ver_al_cl_last_month['probation_completion_wage'], format='%Y-%m-%d',  errors='coerce')
        # print(self.df_scheduling_ver)

    def merge_download_attendance(self):
        start_str = self.date_array[0].strftime('%Y-%m-%d')
        LIMIT_SIZE = 1000
        if self.is_load_df_old :
            start_str = (self.date_array[0] - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        end_str = self.date_array[len(self.date_array)-1].strftime('%Y-%m-%d')
        if self.is_load_df_old :
            end_str = (self.date_array[len(self.date_array)-1] + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        # self.df_old = pd.read_excel(self.input_attendence_file , index_col=None, header=[0,] ,sheet_name='Sheet1')
        # self.df_old['is_from_explanation'] = False
        domain = ["&", ("date", ">=", start_str),("date", "<=", end_str)]
        if (self.select_company_id > 0):
            domain = ["&", "&", ('company', '=', self.select_company_name), ("date", ">=", start_str),("date", "<=", end_str)]
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [domain], {'offset': 0, 'limit': LIMIT_SIZE})
        list_scheduling_ver = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'read', [ids],
                                                     {'fields': ['id', 'employee_name', 'date', 'shift_name', 'employee_code', 'company','additional_company',
                                                                 'shift_start', 'shift_end', 'rest_start', 'rest_end', 'rest_shift', 'probation_completion_wage',
                                                                 'total_shift_work_time', 'total_work_time', 'time_keeping_code', 'kid_time',
                                                                 'department', 'attendance_attempt_1', 'attendance_attempt_2', 'minutes_working_reduced', 
                                                                 'attendance_attempt_3', 'attendance_attempt_4', 'attendance_attempt_5',
                                                                 'attendance_attempt_6', 'attendance_attempt_7', 'attendance_attempt_8',
                                                                 'attendance_attempt_9', 'attendance_attempt_10', 'attendance_attempt_11',
                                                                 'attendance_attempt_12', 'attendance_attempt_13', 'attendance_attempt_14',
                                                                 'attendance_inout_1','attendance_inout_2','attendance_inout_3',
                                                                 'attendance_inout_4','attendance_inout_5','attendance_inout_6',
                                                                 'attendance_inout_7','attendance_inout_8','attendance_inout_9', 'amount_al_reserve', 'amount_cl_reserve',
                                                                 'attendance_inout_10','attendance_inout_11','attendance_inout_12', 
                                                                 'attendance_inout_13','attendance_inout_14','attendance_inout_15','actual_total_work_time', 'standard_working_day',
                                                                 'attendance_attempt_15', 'last_attendance_attempt', 'night_hours_normal', 'night_hours_holiday', 'probation_wage_rate', 
                                                                 'split_shift', 'missing_checkin_break', 'leave_early', 'attendance_late', 'night_shift', 'minute_worked_day_holiday','total_attendance',
                                                                 'ot_holiday', 'ot_normal']})

        self.df_scheduling_ver = pd.DataFrame.from_dict(list_scheduling_ver)

        index = 1
        while len(ids) == LIMIT_SIZE:
            ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [domain],  {'offset': index * LIMIT_SIZE, 'limit': LIMIT_SIZE})
            list_scheduling_ver = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'read', [ids],
                                                     {'fields': ['id', 'employee_name', 'date', 'shift_name', 'employee_code', 'company','additional_company',
                                                                 'shift_start', 'shift_end', 'rest_start', 'rest_end', 'rest_shift', 'probation_completion_wage',
                                                                 'total_shift_work_time', 'total_work_time', 'time_keeping_code', 'kid_time',
                                                                 'department', 'attendance_attempt_1', 'attendance_attempt_2', 'minutes_working_reduced',
                                                                 'attendance_attempt_3', 'attendance_attempt_4', 'attendance_attempt_5',
                                                                 'attendance_attempt_6', 'attendance_attempt_7', 'attendance_attempt_8',
                                                                 'attendance_attempt_9', 'attendance_attempt_10', 'attendance_attempt_11',
                                                                 'attendance_attempt_12', 'attendance_attempt_13', 'attendance_attempt_14',
                                                                 'attendance_inout_1','attendance_inout_2','attendance_inout_3',
                                                                 'attendance_inout_4','attendance_inout_5','attendance_inout_6',
                                                                 'attendance_inout_7','attendance_inout_8','attendance_inout_9', 'amount_al_reserve', 'amount_cl_reserve',
                                                                 'attendance_inout_10','attendance_inout_11','attendance_inout_12', 
                                                                 'attendance_inout_13','attendance_inout_14','attendance_inout_15','actual_total_work_time', 'standard_working_day',
                                                                 'attendance_attempt_15', 'last_attendance_attempt', 'attendance_inout_last', 'night_hours_normal', 'night_hours_holiday', 'probation_wage_rate', 
                                                                 'split_shift', 'missing_checkin_break', 'leave_early', 'attendance_late', 'night_shift', 'minute_worked_day_holiday','total_attendance',
                                                                 'ot_holiday', 'ot_normal']})
            self.df_scheduling_ver = pd.concat([self.df_scheduling_ver,
                pd.DataFrame.from_dict(list_scheduling_ver)], ignore_index=True)

            index = index + 1
            print(f'load df_scheduling_ver {index}- {len(ids)}')
        print('complete list scheduling')

        self.df_scheduling_ver['employee_code'] = self.df_scheduling_ver['employee_code'].str.strip(
        )
        # print(list_scheduling_ver)
        self.df_scheduling_ver['date_str'] = self.df_scheduling_ver['date']
        self.df_scheduling_ver['date'] = pd.to_datetime(
            self.df_scheduling_ver['date'], format='%Y-%m-%d',  errors='coerce')
        for i in range(1,16):
            self.df_scheduling_ver[f'attendance_attempt_{i}'] = pd.to_datetime(
                self.df_scheduling_ver[f'attendance_attempt_{i}'], format="%Y-%m-%d %H:%M:%S",  errors='coerce')
        self.df_scheduling_ver['last_attendance_attempt'] = pd.to_datetime(
                self.df_scheduling_ver['last_attendance_attempt'], format="%Y-%m-%d %H:%M:%S",  errors='coerce')
        self.df_scheduling_ver['probation_completion_wage'] = pd.to_datetime(
            self.df_scheduling_ver['probation_completion_wage'], format='%Y-%m-%d',  errors='coerce')
        # print(self.df_scheduling_ver)
        # self.auto_remove_attendance_restshift()
        
    def download_employee_info(self):
        last_month_fist_day_str = (self.date_array[0] - datetime.timedelta(days=2)).replace(day=1).strftime('%Y-%m-%d')
        domain = ['&',('active', '=', True), '|', ('severance_day','=',False),
                  ('severance_day','>=',last_month_fist_day_str)]
        
        if (self.select_company_id > 0):
            domain =  ['&', '&', ('company_id','=', self.select_company_id ), ('active', '=', True), '|', ('severance_day','=',False),
                  ('severance_day','>=',last_month_fist_day_str)]
            
        employee_Sids = self.models.execute_kw( self.db, self.uid,  self.password, 'hr.employee', 'search', [domain],
                                                                                                {'offset': 0, 'limit': 1000})
        list_employees = self.models.execute_kw( self.db, self.uid,  self.password, 'hr.employee', 'read', [employee_Sids], {'fields': ['id', 'name', 'user_id', 'employee_ho', 'part_time_company_id', 'part_time_department_id',
                                                                                                                            'company_id', 'code', 'department_id', 'time_keeping_code', 'job_title',
                                                                                                                            'probationary_contract_termination_date', 'severance_day', 'workingday',
                                                                                                                            'probationary_salary_rate', 'resource_calendar_id', 'date_sign', 'level']})

        # The above code is iterating through a list of employees and checking if their
        # 'time_keeping_code' attribute is not None. If it is not None, it removes the '.0' string
        # from the end of the attribute value using the replace() method.
        for employee in list_employees:
            try:
                if employee['time_keeping_code'] != None:
                    employee['time_keeping_code'] = int(
                        employee['time_keeping_code'].replace('.00', '').replace('.0', ''))
            except Exception as ex:
                print(ex)
            try:
                department_id = employee['department_id'][0]
                employee['d_id'] = department_id

            except:

                employee['d_id'] = -1
            try:
                part_time_department_name = employee['part_time_department_id'][1]
                employee['part_time_department_name'] = part_time_department_name
            except:

                employee['part_time_department_name'] = '-'
            try:
                employee['department_name'] = employee['department_id'][1]
            except:

                employee['department_name'] = ''
            try:
                company_name = employee['company_id'][1]
                employee['company_name'] = company_name
                company_id = employee['company_id'][0]
                employee['company_sid'] = company_id
            except:
                employee['company_name'] = False
                employee['company_sid'] = False
        self.df_employees = pd.DataFrame.from_dict(list_employees)
        index = 1
        while len(employee_Sids) == 1000:
            employee_Sids = self.models.execute_kw(self.db, self.uid,  self.password, 'hr.employee', 'search', [domain],
                                                                                                 {'offset': index * len(employee_Sids), 'limit': 1000})
            list_employees = self.models.execute_kw(self.db, self.uid, self.password, 'hr.employee', 'read', [employee_Sids], {'fields': ['id', 'name', 'user_id','employee_ho', 'part_time_company_id', 'part_time_department_id',
                                                                                                                            'company_id', 'code', 'department_id', 'time_keeping_code', 'job_title',
                                                                                                                                'probationary_contract_termination_date', 'severance_day', 'workingday',
                                                                                                                                'probationary_salary_rate', 'resource_calendar_id', 'date_sign', 'level']})

            # The above code is iterating through a list of employees and checking if their
            # 'time_keeping_code' attribute is not None. If it is not None, it removes the '.0' string
            # from the end of the attribute value using the replace() method.
            for employee in list_employees:
                try:
                    if employee['time_keeping_code'] != None:
                        employee['time_keeping_code'] = int(
                            employee['time_keeping_code'].replace('.00', '').replace('.0', ''))
                except Exception as ex:
                    print(ex)
                try:
                    department_id = employee['department_id'][0]
                    employee['d_id'] = department_id

                except:

                    employee['d_id'] = -1
                try:
                    part_time_department_name = employee['part_time_department_id'][1]
                    employee['part_time_department_name'] = part_time_department_name
                except:

                    employee['part_time_department_name'] = '-'
                try:
                    employee['department_name'] = employee['department_id'][1]
                except:

                    employee['department_name'] = ''
                try:
                    company_name = employee['company_id'][1]
                    employee['company_name'] = company_name
                    company_id = employee['company_id'][0]
                    employee['company_sid'] = company_id
                except:
                    employee['company_name'] = False
                    employee['company_sid'] = False
            self.df_employees = pd.concat([self.df_employees, pd.DataFrame.from_dict(list_employees)], ignore_index=True)

            index = index + 1
            # print(f'load emplempoyee {index}- {len(employee_Sids)}')
        # self.df_employees.to_excel(os.path.join(self.output_report_folder, 'employees.xlsx'))
        self.merge_department_data()
        self.df_employees['probationary_contract_termination_date'] = pd.to_datetime(self.df_employees['probationary_contract_termination_date'],
                                                                                     format="%Y-%m-%d %H:%M:%S", errors='coerce')
        self.df_employees['severance_day'] = pd.to_datetime(self.df_employees['severance_day'],
                                                            format="%Y-%m-%d", errors='coerce')
        self.df_employees['workingday'] = pd.to_datetime(self.df_employees['workingday'],
                                                         format="%Y-%m-%d", errors='coerce')
        self.df_employees['date_sign'] = pd.to_datetime(self.df_employees['date_sign'],
                                                        format="%Y-%m-%d", errors='coerce')
        self.df_employees.to_excel(os.path.join(self.output_report_folder, 'employees.xlsx'))
        # self.save_employee_to_db()
        # self.remove_employee_finished(self.df_employees[self.df_employees['severance_day'] < self.date_array[0]])
        # self.df_employees.apply(
        #     lambda row: self.remove_redundant_attendance(row), axis=1)
        # selRows = self.df_employees[self.df_employees['severance_day']
        #                             < self.date_array[0]].index
        # self.df_employees = self.df_employees.drop(selRows, axis=0)
    def save_employee_to_db(self):
        # df_records = self.df_employees.to_dict('records')
        # Employee.objects.all().delete()
        model_instances = [Employee(
            id=record['id'],
            sid=record['id'],
            name=record['name'],
            user_id= None if record['user_id'] == False else record['user_id'][0],
            user_name= None if record['user_id'] == False else record['user_id'][1],
            employee_ho=record['employee_ho'],
            code=record['code'],
            part_time_company_id= None if record['part_time_company_id'] == False else record['part_time_company_id'][0],
            part_time_company_name= None if record['part_time_company_id'] == False else record['part_time_company_id'][1],

            part_time_department_id= None if record['part_time_department_id'] == False else record['part_time_department_id'][0],
            part_time_department_name= None if record['part_time_department_id'] == False else record['part_time_department_id'][1],
            company_id=record['company_sid'],
            company_name=record['company_name'],
            department_id= record['d_id'],
            department_name= None if record['department_id'] == False else record['department_id'][1],
            time_keeping_code=record['time_keeping_code'],
            job_title=record['job_title'],
            probationary_contract_termination_date= None if pd.isnull(record['probationary_contract_termination_date']) 
                                else record['probationary_contract_termination_date'],
            severance_day= None if pd.isnull(record['severance_day']) else record['severance_day'],
            workingday= None if pd.isnull(record['workingday']) else record['workingday'],
            probationary_salary_rate= 0 if pd.isnull(record['probationary_salary_rate']) else record['probationary_salary_rate'],
            resource_calendar_id=None if record['resource_calendar_id'] == False else record['resource_calendar_id'][0],
            resource_calendar_name=None if record['resource_calendar_id'] == False else record['resource_calendar_id'][1],
            date_sign= None if pd.isnull(record['date_sign']) else record['date_sign'],
        ) for index, record in self.df_employees.iterrows() if pd.notnull(record['code'])]

        Employee.objects.bulk_create(model_instances,  update_conflicts=True, unique_fields=["id"])

    def remove_redundant_attendance(self, row):
        result = False
        try:
            if pd.notnull(row['date']) and (row['shift_name'] != '-') and (row['shift_name'] != False):
                if pd.notnull(row['severance_day']):
                    if row['severance_day'] < row['date']:
                        result = 'BY severance DAY'
                if pd.notnull(row['workingday']):
                    if row['workingday'] > row['date']:
                        # self.models.execute_kw(
                        #     self.db, self.uid, self.password, 'hr.apec.attendance.report', 'unlink', [ids])
                        result = 'BY WORKING DAY'
        except Exception as ex:
            print(ex)
            result = False

        return result
    def calculate_is_probationary(self, x):
        try:
            contract_company = x['company']
        except:
            contract_company = x['company_name']
        try:
            probationary_contract_termination_date = x['probationary_contract_termination_date']
        except:
            probationary_contract_termination_date = None
        is_probationary = True
        try:
            probationary_salary_rate = x['probationary_salary_rate']
        except:
            probationary_salary_rate = None
        try:
            if (pd.notnull(x['id_employee'])) and (x['id_employee'] != False):
                contract_employee_filter = self.df_contract[(self.df_contract['date_start']<=x['date'])&
                                                    (self.df_contract['date_end']>=x['date'])&
                                                    (self.df_contract['employee_id']==x['id_employee'])].sort_values("date_end", ascending=False)
            else:
                contract_employee_filter = self.df_contract[(self.df_contract['date_start']<=x['date'])&
                                                (self.df_contract['date_end']>=x['date'])&
                                                (self.df_contract['employee_code']==x['employee_code'])].sort_values("date_end", ascending=False)
        except Exception as ex:
            contract_employee_filter = self.df_contract[(self.df_contract['date_start']<=x['date'])&
                                                (self.df_contract['date_end']>=x['date'])&
                                                (self.df_contract['employee_code']==x['employee_code'])].sort_values("date_end", ascending=False)
            
        # if (x['employee_code']== 'APG231012005'):
        #     contract_employee_filter.to_excel(f'contract_employee_filter_APG231012005{x["date"].day}.xlsx')
        for contract_index, contract_item in contract_employee_filter.iterrows():
            if 'thử việc' in contract_item['contract_type_name'].strip().lower():       
                if not (probationary_contract_termination_date is None):
                    if contract_item['date_end'] >= probationary_contract_termination_date:
                        probationary_contract_termination_date = contract_item['date_end']
                        probationary_salary_rate = contract_item['salary_rate']
                else:
                    probationary_contract_termination_date = contract_item['date_end']
                    probationary_salary_rate = contract_item['salary_rate']

        if len(contract_employee_filter.index) > 0:
            contract_company = contract_employee_filter.iloc[0]['company_sname']
        try:
            if pd.notnull(probationary_contract_termination_date):
                is_probationary = (x['date'] <= probationary_contract_termination_date)
                
            else:
                
                if len(contract_employee_filter.index)>0:
                    date_start = contract_employee_filter.iloc[0]['date_start']
                    probationary_contract_termination_date = date_start - datetime.timedelta(days=1)
                    is_probationary = x['date'] <= probationary_contract_termination_date
                    
        except Exception as ex:
            print(ex)

        return probationary_contract_termination_date, is_probationary, contract_company, probationary_salary_rate


            
    def merge_employee(self):
        self.df_scheduling_ver = self.df_scheduling_ver.merge(self.df_employees[['id', 'code', 'department_id','department_name', 'name', 'job_title', 'time_keeping_code',
                                                                                 'date_sign', 'time_keeping_count', 'company_name', 'severance_day', 'd_id', 'employee_ho',
                                                                                 'probationary_contract_termination_date', 'workingday', 'probationary_salary_rate', 'resource_calendar_id']],
                                                              left_on=['employee_code', 'company'], right_on=['code', 'company_name'], how='left', suffixes=('', '_employee'))
        # self.df_scheduling_ver.rename(columns = {'id':'employee_sid'}, inplace = True)
        self.df_scheduling_ver[['probationary_contract_termination_date', 'is_probationary', 
                    'contract_company','probationary_salary_rate']] = self.df_scheduling_ver.apply(
            lambda x: self.calculate_is_probationary(x), axis=1, result_type='expand')
                    
        self.df_scheduling_ver_al_cl_last_month = self.df_scheduling_ver_al_cl_last_month.merge(self.df_employees[['id', 'code', 'department_id','department_name', 'name', 'job_title', 'time_keeping_code',
                                                                                 'date_sign', 'time_keeping_count', 'company_name', 'severance_day', 'd_id', 'employee_ho',
                                                                                 'probationary_contract_termination_date', 'workingday', 'probationary_salary_rate', 'resource_calendar_id']],
                                                              left_on=['employee_code', 'company'], right_on=['code', 'company_name'], how='left', suffixes=('', '_employee'))

        # self.df_scheduling_ver.rename(columns = {'id':'employee_sid'}, inplace = True)
        self.df_scheduling_ver_al_cl_last_month[['probationary_contract_termination_date', 'is_probationary', 
                    'contract_company','probationary_salary_rate']] = self.df_scheduling_ver_al_cl_last_month.apply(
            lambda x: self.calculate_is_probationary(x), axis=1, result_type='expand')
        # self.df_scheduling_ver[['department_name']] = \
        #     self.df_scheduling_ver.apply(
        #         lambda row: self.getdepartment(row), axis=1, result_type='expand')
        self.df_scheduling_ver['REMOVE'] = self.df_scheduling_ver.apply(
            lambda row: self.remove_redundant_attendance(row), axis=1, result_type='expand')
        self.df_scheduling_ver[self.df_scheduling_ver['REMOVE'] != False].to_excel(os.path.join(self.output_report_folder, 'employees_sche_remove_temp.xlsx'))

        remove_ids = []
        update_data = {'shift_name': '-', 'shift_start':12, 'shift_end': 12, 'total_shift_work_time': 0, \
                'total_work_time':0, 'actual_total_work_time':0,'rest_start':12, 'rest_end':12, 'rest_shift':True, 'split_shift':False, 'night_shift':False}
        # for index, row in self.df_scheduling_ver[self.df_scheduling_ver['REMOVE'] != False].iterrows():
        #     remove_ids.append(int(row['id']))
        # if len(remove_ids)>0:
        #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [remove_ids, update_data])
            

    def calculate_stardard_working_time(self):
        number_sun = len([1 for i in calendar.monthcalendar(self.date_array[0].year,
                                                            self.date_array[0].month) if i[6] != 0])
        number_sat = len([1 for i in calendar.monthcalendar(self.date_array[0].year,
                                                            self.date_array[0].month) if i[5] != 0])
        number_day = len(self.date_array)
        # print('number_sun', number_sun)
        # print('number_sat', number_sat)
        return (number_day - number_sun), (number_day-number_sun - number_sat * 0.5)
    def export_sumary_attendance_report(self):
        index = 0

        for group_index, group_data in self.df_scheduling_ver.groupby('company_name'):
            df_company_filter = self.df_company[self.df_company['name']==group_index]
            if len(df_company_filter.index) > 0:
                output_report_folder = os.path.join(self.output_report_folder, f"{df_company_filter.iloc[0]['user']}")
                # print('out')
            else:
                output_report_folder = os.path.join(self.output_report_folder, f"{index}")
            df_hr_leave = self.df_hr_leave[self.df_hr_leave['company_name']==group_index]
            df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
            df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]

            self.export_sumary_attendance_report_company(group_data,
                                                    df_hr_leave, df_explanation_data,
                                                    output_report_folder)

            index = index + 1

    def export_sumary_attendance_report_company(self, df_scheduling_ver,
                    df_hr_leave,
                    df_explanation_data,
                    output_report_folder):
        working_time_48, working_time_44 = self.calculate_stardard_working_time()
        print("Tong hop cham cong")
        ref_workbook = openpyxl.load_workbook(
            'unity/template/Summary time attendant OK.xlsx')

        ws = ref_workbook.worksheets[0]
        start_row = 11
        start_col = 33
        # ws['AH10'].value = f'Employee Type: All - Working Time: All - Pay Cycle: Month - Payment Period: {self.date_array[0].strftime("%Y/%m/%d")} - {self.date_array[len(self.date_array)-1].strftime("%Y/%m/%d")}'
        # ws['AH6']  = ws['AH10'].value
        ws.merge_cells('A6:BK6')
        top_left_cell = ws['A6']
        top_left_cell.value = f'Employee Type: All - Working Time: All - Pay Cycle: Month - Payment Period: {self.date_array[0].strftime("%d/%m/%Y")} - {self.date_array[len(self.date_array)-1].strftime("%d/%m/%Y")}'
        for date_item in self.date_array:
            day_of_month = date_item.day
            ws.cell(row=start_row-1, column=start_col +
                    day_of_month).value = date_item
            ws.cell(row=start_row-2, column=start_col +
                    day_of_month).value = date_item.strftime('%A')[:3]
        if date_item.day < 31:
            for add_item in range(date_item.day + 1, 32):
                ws.cell(row=start_row-1, column=start_col + add_item).value = ''
                ws.cell(row=start_row-2, column=start_col + add_item).value = ''
                ws.cell(row=start_row, column=start_col + add_item).value = ''

        for g, data in df_scheduling_ver[(df_scheduling_ver['department_name'].notnull())
                                              & (df_scheduling_ver['department_name'] != False)
                                              & (df_scheduling_ver['workingday'].notnull())
                                              # & (self.df_scheduling_ver['employee_code']=='APG230316014')
                                              & (df_scheduling_ver['workingday'] != False)].groupby('department_name'):
            start_row = start_row + 1
            count = 1
            # print(g)
            ws.cell(row=start_row, column=1).value = g
            ws.cell(row=start_row, column=17).value = ''
            ws.cell(row=start_row, column=30).value = ''
            ws.cell(row=start_row, column=31).value = ''
            start_row = start_row + 1

            for sub_group_index, sub_group_data in data.groupby(['employee_name', 'employee_code']):
                total_ncl = 0
                total_ncl_probationary = 0
                total_off = 0

                total_ph = 0
                total_al = 0
                total_al_probationary = 0
                total_tc = 0
                total_tc_probationary = 0
                total_tc_holiday = 0
                total_tc_holiday_probationary = 0
                total_normal_working_minutes = 0
                total_normal_working_minutes_probationary = 0

                total_ph_probationary = 0
                total_unpaid_leave = 0
                total_unpaid_leave_probationary = 0

                total_compensation_leave = 0
                total_compensation_probationary = 0

                total_night = 0
                total_night_probationary = 0

                total_night_holiday = 0
                total_night_holiday_probationary = 0

                total_fix_rest_time = 0
                total_fix_rest_time_probationary = 0

                minute_worked_day_holiday = 0
                minute_worked_day_holiday_probationary = 0

                ws.cell(row=start_row, column=2).value = sub_group_index[1]
                ws.cell(row=start_row, column=3).value = sub_group_index[0]

                col_index = 0

                hr_leave_employee = df_hr_leave[df_hr_leave['employee_code']
                                                     == sub_group_index[1]]

                for index, leave_item in hr_leave_employee.iterrows():
                    # if shift_name == 'CL':
                    #     total_compensation_leave = total_compensation_leave + 1
                    #     if probationary_index:
                    #         total_compensation_probationary = total_compensation_probationary + 1
                    # elif ('/CL' in shift_name) or ('CL/' in shift_name):
                    #     total_compensation_leave = total_compensation_leave + 0.5
                    #     if probationary_index:
                    #         total_compensation_probationary = total_compensation_probationary + 0.5

                    # if shift_name == 'AL':
                    #     total_al = total_al + 1
                    #     if probationary_index:
                    #         total_al_probationary = total_al_probationary + 1
                    # elif ('/AL' in shift_name) or ('AL/' in shift_name):
                    #     total_al = total_al + 0.5
                    #     if probationary_index:
                    #         total_al_probationary = total_al_probationary + 0.5
                    # if (leave_item['request_date_from'] <= self.date_array[len(self.date_array)-1]):
                    # Đơn xin nghỉ có tính lương
                    # print("--------------------",
                    #       leave_item['holiday_status_id'])
                    try:
                        if ('có tính lương' in leave_item['holiday_status_id'][1].strip().lower()):
                            # print("aaaaaa'có tính lương'aaaaaaa")
                            total_ncl = total_ncl + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_tc_probationary = total_tc_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)

                    # print("--------------------",
                    #       leave_item['holiday_status_id'])
                    try:
                        if (leave_item['holiday_status_id'][1].strip().lower() == 'nghỉ bù'):
                            # print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                            total_compensation_leave = total_compensation_leave + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_compensation_probationary = total_compensation_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)
                    try:
                        if (leave_item['holiday_status_id'][1].strip().lower() == 'nghỉ phép năm'):
                            # print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                            total_al = total_al + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_al_probationary = total_al_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                        elif ('tăng ca' in leave_item['holiday_status_id'][1].strip().lower()):
                            # print("don tang ca")
                            if leave_item['is_holiday']:
                                total_tc_holiday = total_tc_holiday + \
                                    max(leave_item['minutes'],
                                        leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                                if leave_item['is_leave_probationary']:
                                    total_tc_holiday_probationary = total_tc_holiday_probationary + \
                                        max(leave_item['minutes'],
                                            leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                            else:
                                total_tc = total_tc + \
                                    max(leave_item['minutes'],
                                        leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                                if leave_item['is_leave_probationary']:
                                    total_tc_probationary = total_tc_probationary + \
                                        max(leave_item['minutes'],
                                            leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                    except Exception as ex:
                        print(ex)
                # print('hr_learve_employee: ', hr_leave_employee)
                hr_leave_employee.to_excel('hr_leave_employee.xlsx')
                # for index, row in sub_group_data.iterrows():
                for probationary_index, probationary_data in sub_group_data.groupby('is_probationary'):
                    for index, row in probationary_data.iterrows():
                        # print(f"{row}")
                        date = row['date']
                        shift_name = row['shift_name'] if row['shift_name'] else ''
                        resource_calendar_id = row['resource_calendar_id']
                        actual_work_time =  min (row['total_work_time'] + row['kid_time'], row['total_shift_work_time'] * 60)

                        ws.cell(row=start_row, column=9).value = 0
                        ws.cell(row=start_row, column=10).value = 0
                        ws.cell(row=start_row, column=11).value = 0
                        ws.cell(row=start_row, column=12).value = 0
                        ws.cell(row=start_row, column=13).value = 0

                        ws.cell(row=start_row, column=14).value = 0
                        ws.cell(row=start_row, column=15).value = 0
                        ws.cell(row=start_row, column=16).value = 0
                        ws.cell(row=start_row, column=17).value = 0
                        ws.cell(row=start_row, column=18).value = 0

                        ws.cell(row=start_row, column=20).value = 0
                        ws.cell(row=start_row, column=21).value = 0
                        ws.cell(row=start_row, column=22).value = 0
                        ws.cell(row=start_row, column=23).value = 0
                        ws.cell(row=start_row, column=24).value = 0

                        ws.cell(row=start_row, column=25).value = 0
                        ws.cell(row=start_row, column=26).value = 0
                        ws.cell(row=start_row, column=30).value = 0
                        ws.cell(row=start_row, column=31).value = 0
                        # if resource_calendar_id == False:
                        #     continue
                        # Find explan nation
                        explanation_filter = df_explanation_data[
                            df_explanation_data['date_str'].str.contains(row['date_str']).fillna(False) &
                            df_explanation_data['employee_code'].str.contains(sub_group_index[1]).fillna(False) &
                            df_explanation_data['reason'].str.contains('2').fillna(False) &
                            df_explanation_data['validated']]


                        if len(explanation_filter.index) > 0 and actual_work_time < row['total_shift_work_time']*60:
                            if row['split_shift'] == True:
                                actual_work_time = 480
                            else:
                                actual_work_time = row['total_shift_work_time']*60

                            print('fix: ------------------- buy explanation', explanation_filter.iloc[0])

                        # if len(hr_leave_filter.index)>0 and actual_work_time < 480:
                        #     # try:
                        #     print('fix by cl: -------------------', hr_leave_filter)
                        #     if pd.notnull(hr_leave_filter.iloc[0]['minutes']):

                        #         minute = int(hr_leave_filter.iloc[0]['minutes'])
                        #         actual_work_time = min(row['total_shift_work_time']*60, actual_work_time + minute)
                        #     # except:
                        #     #     pass

                        if (row['missing_checkin_break'] == True) and (row['time_keeping_count'] > 2) and (not ('/OFF' in shift_name)):
                            actual_work_time = min(
                                actual_work_time, row['total_shift_work_time']*30)
                            print(f"Thieu cham cong giua {date}")
                            redFill = PatternFill(
                                patternType='solid', fgColor=colors.Color(rgb='00FDF000'))
                            ws.cell(row=start_row, column=start_col +
                                    date.day).fill = redFill
                        else:
                            # print(f"khong thieu cham cong {row['attendance_attempt_3']} - hop dong {resource_calendar_id}" )
                            blueFill = PatternFill(
                                patternType='solid', fgColor=colors.Color(rgb='0000FF00'))
                            ws.cell(row=start_row, column=start_col +
                                    date.day).fill = blueFill

                        ft = Font(color="0000FF")
                        if shift_name == '-' or shift_name == 'OFF' or row['total_shift_work_time'] == 0:
                            ft = Font(color="0000FF")
                        elif (pd.isnull(row['attendance_attempt_1']) or row['attendance_attempt_1'] == False) \
                                or (pd.isnull(row['last_attendance_attempt']) or row['last_attendance_attempt'] == False):
                            ft = Font(color="FF0000")
                        elif row['total_work_time'] < 480:
                            ft = Font(color="FC1000")
                        else:
                            ft = Font(color="0000FF")

                        ws.cell(row=start_row, column=start_col +
                                date.day).font = ft

                        if actual_work_time > 0 and row['total_shift_work_time'] > 0:
                            shift_name = f"{int(actual_work_time)}_{shift_name}"
                            total_normal_working_minutes = total_normal_working_minutes + actual_work_time
                            if probationary_index:
                                total_normal_working_minutes_probationary = total_normal_working_minutes_probationary + actual_work_time
                        ws.cell(row=start_row, column=start_col +
                                date.day).value = shift_name
                        if col_index == 0:
                            if pd.notnull(row['probationary_contract_termination_date']):
                                ws.cell(row=start_row, column=6).value = row['probationary_contract_termination_date'].strftime(
                                    '%d-%m-%Y')
                            if pd.notnull(row['workingday']):
                                ws.cell(row=start_row, column=5).value = row['workingday'].strftime(
                                    '%d-%m-%Y')
                            if pd.notnull(row['job_title']):
                                ws.cell(row=start_row,
                                        column=4).value = row['job_title']
                            if pd.notnull(row['probationary_salary_rate']):
                                ws.cell(
                                    row=start_row, column=7).value = row['probationary_salary_rate']/100

                            try:

                                print(resource_calendar_id)
                                # Tiêu chuẩn 48 giờ/tuần
                                if '44' in resource_calendar_id[1]:
                                    ws.cell(row=start_row,
                                            column=8).value = working_time_44
                                else:
                                    ws.cell(row=start_row,
                                            column=8).value = working_time_48
                            except Exception as ex:
                                ws.cell(row=start_row,
                                        column=8).value = working_time_48
                                print("gio lam: ", ex)

                        col_index = col_index + 1

                        if shift_name == 'UP':
                            total_unpaid_leave = total_unpaid_leave + 1
                            if probationary_index:
                                total_unpaid_leave_probationary = total_unpaid_leave_probationary + 1
                        elif ('/UP' in shift_name) or ('UP/' in shift_name):
                            total_unpaid_leave = total_unpaid_leave + 0.5
                            if probationary_index:
                                total_unpaid_leave_probationary = total_unpaid_leave_probationary + 0.5
                        # if shift_name == 'CL':
                        #     total_compensation_leave = total_compensation_leave + 1
                        #     if probationary_index:
                        #         total_compensation_probationary = total_compensation_probationary + 1
                        # elif ('/CL' in shift_name) or ('CL/' in shift_name):
                        #     total_compensation_leave = total_compensation_leave + 0.5
                        #     if probationary_index:
                        #         total_compensation_probationary = total_compensation_probationary + 0.5

                        # if shift_name == 'AL':
                        #     total_al = total_al + 1
                        #     if probationary_index:
                        #         total_al_probationary = total_al_probationary + 1
                        # elif ('/AL' in shift_name) or ('AL/' in shift_name):
                        #     total_al = total_al + 0.5
                        #     if probationary_index:
                        #         total_al_probationary = total_al_probationary + 0.5
                        'Ca gay'
                        if row['split_shift']:
                            total_fix_rest_time = total_fix_rest_time + 1
                            if probationary_index:
                                total_fix_rest_time_probationary = total_fix_rest_time_probationary + 1

                        if shift_name == 'OFF':
                            total_off = total_off + 1
                        elif '/OFF' in shift_name:
                            total_off = total_off + 0.5

                        total_night = total_night + row['night_hours_normal']
                        total_night_holiday = total_night_holiday + \
                            row['night_hours_holiday']
                        # total_ph = total_ph + row['minute_worked_day_holiday']
                        if probationary_index:
                            total_night_holiday_probationary = total_night_holiday_probationary + + \
                                row['night_hours_holiday']
                            # total_ph_probationary = total_ph_probationary + row['minute_worked_day_holiday']
                            total_night_probationary = total_night_probationary + \
                                row['night_hours_normal']

                        if shift_name == 'PH':
                            total_ph = total_ph + 1
                            if probationary_index:
                                total_ph_probationary = total_ph_probationary + 1
                        elif '/PH' in shift_name:
                            total_ph = total_ph + 0.5
                            if probationary_index:
                                total_ph_probationary = total_ph_probationary + 0.5

                # summary data

                ws.cell(row=start_row, column=9).value = total_off
                ws.cell(row=start_row, column=10).value = total_unpaid_leave
                ws.cell(row=start_row, column=11).value = round(
                    total_ph_probationary, 2)
                ws.cell(row=start_row, column=12).value = round(
                    total_tc / 60, 2)
                ws.cell(row=start_row, column=13).value = round(
                    total_tc_holiday / 60, 2)

                ws.cell(row=start_row, column=14).value = round(
                    total_night_probationary / 60, 2)
                ws.cell(row=start_row, column=15).value = round(
                    total_night_holiday_probationary / 60, 2)
                ws.cell(row=start_row,
                        column=16).value = total_fix_rest_time_probationary
                ws.cell(row=start_row, column=17).value = round(
                    total_ncl_probationary/480, 2)
                ws.cell(row=start_row, column=17).value = round(
                    total_compensation_probationary/480, 2)
                ws.cell(row=start_row, column=18).value = round(
                    total_normal_working_minutes_probationary / (8 * 60), 2)

                ws.cell(row=start_row, column=20).value = round(
                    total_ph - total_ph_probationary, 2)
                ws.cell(row=start_row, column=21).value = round(
                    (total_tc - total_tc_probationary) / 60, 2)
                ws.cell(row=start_row, column=22).value = round(
                    (total_tc_holiday - total_tc_holiday_probationary) / 60, 2)
                ws.cell(row=start_row, column=23).value = round(
                    (total_night - total_night_probationary) / 60, 2)
                ws.cell(row=start_row, column=24).value = round(
                    (total_night_holiday - total_night_holiday_probationary) / 60, 2)

                ws.cell(row=start_row, column=25).value = total_fix_rest_time - \
                    total_fix_rest_time_probationary
                ws.cell(row=start_row, column=26).value = round(
                    (total_al - total_al_probationary)/480, 2)
                ws.cell(row=start_row, column=29).value = round(
                    (total_ncl - total_ncl_probationary) / 480, 2)
                ws.cell(row=start_row, column=30).value = round(
                    (total_compensation_leave - total_compensation_probationary) / 480, 2)
                ws.cell(row=start_row, column=31).value = round(
                    (total_normal_working_minutes - total_normal_working_minutes_probationary) / (8 * 60), 2)
                start_row = start_row + 1
        isExist = os.path.exists(output_report_folder)
        if not isExist:

        # Create a new directory because it does not exist
            os.makedirs(output_report_folder)
            print("The new output_report_folder is created! ", output_report_folder)
        ref_workbook.save(os.path.join(
            output_report_folder, "Summary time attendant.xlsx"))
        # print(self.df_explanation_data)
        # print(self.df_hr_leave)
        return df_scheduling_ver

    def getdepartment(self, row):
        try:
            return row['department_id'][1]
        except:
            return ''

    def export_log(self, output_file):
        self.df_employees.to_excel(output_file)

    def analyst_log(self, output_file):
        df = pd.read_excel(output_file)
        code_series = df["code"]

        duplicates = code_series[code_series.duplicated(keep=False)]

        print("Các giá trị bị trùng lặp:")
        print(duplicates)

        df_without_duplicates = df.drop_duplicates(
            subset=["code"], keep="first")

        duplicates = df[~df.index.isin(df_without_duplicates.index)]

        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
            duplicates.to_excel(writer, sheet_name="Duplicates", index=False)

        # In ra thông báo khi hoàn thành
        print("Các giá trị bị trùng lặp đã được đưa vào sheet 'Duplicates'")
    def convert_code_to_int(self, row):
        try:
            return int(row)
        except:
            return 0

    def process_report_raw(self, row):
        # print(row)
        shift_start_datetime = None 
        shift_end_datetime = None
        rest_start_datetime = None
        rest_end_datetime = None
        try:
            try:
                start_work_time = row['start_work_time']
            except:
                # if pd.isnull(start_work_time):
                start_work_time = row['shift_start']
            try:
                end_work_time = row['end_work_time']
            except:
                # if pd.isnull(end_work_time):
                end_work_time = row['shift_end']

            try:
                start_rest_time =row['start_rest_time']
            except:
            # if pd.isnull(start_rest_time):
                start_rest_time = row['rest_start']

            try:
                end_rest_time = row['end_rest_time']
            except:
            # if pd.isnull(end_rest_time):
                end_rest_time = row['rest_end']
        except:
            print('No comment')
        try:
            if pd.notnull(start_work_time):
                h, m, s = float_to_hours(start_work_time)
                shift_start_datetime = row['date'].replace(hour=h, minute=m, second=s)

            if pd.notnull(end_work_time):
                h, m, s = float_to_hours(end_work_time)
                shift_end_datetime = row['date'].replace(hour=h, minute=m, second=s)
                shift_end_datetime = shift_end_datetime if end_work_time >= start_work_time else shift_end_datetime + datetime.timedelta(days=1)
            if pd.notnull(start_rest_time):
                h, m, s = float_to_hours(start_rest_time)
                rest_start_datetime = row['date'].replace(hour=h, minute=m, second=s)
                rest_start_datetime = rest_start_datetime if start_rest_time >= start_work_time else rest_start_datetime + datetime.timedelta(days=1)
            if pd.notnull(end_rest_time):
                h, m, s = float_to_hours(end_rest_time)
                rest_end_datetime = row['date'].replace(hour=h, minute=m, second=s)
                rest_end_datetime = rest_end_datetime if end_rest_time >= start_work_time else rest_end_datetime + datetime.timedelta(days=1)
                
        except Exception as ex:
            print(ex)
        return shift_start_datetime, shift_end_datetime, rest_start_datetime, rest_end_datetime 
    def auto_remove_attendance_restshift(self):
        df_scheduling_ver = self.df_scheduling_ver\
                    .sort_values(by=['employee_code','date']) \
                    .copy()
        # df_scheduling_ver['real_shift_end'] = df_scheduling_ver.apply(lambda x: x['shift_end'] if x['shift_end'] >= x['shift_start'] else x['shift_end'] + 24, 
        #                     axis=1)
        
        # df_scheduling_ver.to_excel('df_scheduling_ver_fist_stage.xlsx')
        for group_index, group_data in df_scheduling_ver.groupby('employee_code'):
            try:
                group_data[['id_prev', 'total_shift_work_time_prev', 'total_attendance_prev', 'shift_name_prev','date_prev']]  \
                    = group_data[['id', 'total_shift_work_time','total_attendance','shift_name','date']].shift(1)
                group_data[['id_next', 'total_shift_work_time_next', 'total_attendance_next', 'shift_name_next', 'date_next']]  \
                    = group_data[['id', 'total_shift_work_time','total_attendance','shift_name','date']].shift(-1)
                    
                group_data_filter = group_data[(group_data['total_shift_work_time_next'] == 0) |
                                                        (group_data['total_shift_work_time_prev'] == 0)|
                                                        (group_data['total_shift_work_time'] == 0)] 
                if len(group_data_filter.index)>0:
                    group_data_filter[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                        group_data_filter.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
                    group_data_filter[['cut_to_prev', 'cut_to_next']] = \
                        group_data_filter.apply(lambda row: self.caculate_attendance_restshift(row), axis=1, result_type='expand')
                    group_data_filter['add_to_last'] = group_data_filter['cut_to_prev'].shift(-1)
                    group_data_filter['add_to_prev'] = group_data_filter['cut_to_next'].shift(1)
                    group_data_filter[(group_data_filter['total_shift_work_time']>0) & \
                            (group_data_filter['employee_ho'] == False)].apply(lambda row: self.process_attendance_restshift(row), axis=1, result_type='expand')
            except Exception as ex:
                logger.error(f'update id: {ex}')
        
        
        df_hr_leave = self.df_hr_leave \
                    .sort_values(by=['employee_code','date']) \
                    .copy()
        # df_scheduling_ver['real_shift_end'] = df_scheduling_ver.apply(lambda x: x['shift_end'] if x['shift_end'] >= x['shift_start'] else x['shift_end'] + 24, 
        #                     axis=1)
        
        # df_scheduling_ver.to_excel('df_scheduling_ver_fist_stage.xlsx')
        for group_index, group_data in df_hr_leave.groupby('employee_code'):
            try:
                group_data_filter = self.df_scheduling_ver[self.df_scheduling_ver['employee_code'] == group_index]
                group_data_filter[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                            group_data_filter.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')

                group_data_filter[['id_prev', 'total_shift_work_time_prev', 'total_attendance_prev', 'shift_name_prev','date_prev']]  \
                        = group_data_filter[['id', 'total_shift_work_time','total_attendance','shift_name','date']].shift(1)
                group_data_filter[['id_next', 'total_shift_work_time_next', 'total_attendance_next', 'shift_name_next', 'date_next']]  \
                        = group_data_filter[['id', 'total_shift_work_time','total_attendance','shift_name','date']].shift(-1)
                        
                #     group_data_filter = group_data_filter[(group_data['total_shift_work_time_next'] == 0) |
                #                                             (group_data['total_shift_work_time_prev'] == 0)|
                #                                             (group_data['total_shift_work_time'] == 0)] 
                #     if len(group_data_filter.index)>0:
                    
                group_data[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                        group_data.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
                        
                group_data[['id_prev', 'total_shift_work_time_prev', 'total_attendance_prev', 'shift_name_prev','date_prev']]  \
                    = group_data[['id', 'total_shift_work_time','total_attendance','shift_name','date']].shift(1)
                    
                group_data[['id_next', 'total_shift_work_time_next', 'total_attendance_next', 'shift_name_next', 'date_next']]  \
                    = group_data[['id', 'total_shift_work_time','total_attendance','shift_name','date']].shift(-1)
                
                group_data['shift_end_datetime_prev'] =  group_data['shift_start_datetime'].shift(1)
                group_data['shift_start_datetime_next'] =  group_data['shift_start_datetime'].shift(-1)
                # group_data = group_data[(group_data['total_shift_work_time_next'] == 0) |
                #                                         (group_data['total_shift_work_time_prev'] == 0)|
                #                                         (group_data['total_shift_work_time'] == 0)] 
                
                group_data_filter[['cut_to_prev', 'cut_to_next']] = \
                    group_data_filter.apply(lambda row: self.caculate_attendance_timeoff(row, group_data), axis=1, result_type='expand')
                    
                group_data_filter['add_to_last'] = group_data_filter['cut_to_prev'].shift(-1)
                group_data_filter['add_to_prev'] = group_data_filter['cut_to_next'].shift(1)
                
                group_data[(group_data['total_shift_work_time']>0) & \
                                (group_data['employee_ho'] == False)]. \
                                apply(lambda row: self.process_attendance_timeoff(row, group_data_filter), axis=1, result_type='expand')
                # except Exception as ex:
                #     logger.error(f'update id: {ex}')
            except Exception as ex:
                logger.error(f'update id: {ex}')
                           
        # df_hr_leave.to_excel('df_scheduling_ver_last_stage.xlsx')
    
    def process_attendance_timeoff(self, row, df_scheduling_ver):
        data = []
        io = []
        min_time = None
        max_time = None
        number_out_date = 0
        number_child_date = 0
        hr_leave_employee_date = self.df_hr_leave[(self.df_hr_leave['employee_code']
                                                    == row['employee_code']) & \
                                                    (self.df_hr_leave['date_str'] == row['date_str'])]
        if pd.notnull(row[f'attendance_attempt_1']):
            min_time = row[f'attendance_attempt_1']
        if pd.notnull(row[f'last_attendance_attempt']):
            max_time = row[f'last_attendance_attempt']
        if ('thiếu chấm công' in row['holiday_status_name'].strip().lower()):
            for index, leave_item in hr_leave_employee_date.iterrows():
                if ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()):
                    number_out_date = number_out_date + 1
                        # if pd.notnull(leave_item['request_date_from']):
                        #     out_date_array.append(leave_item)
                        #     if last_out_date > leave_item['request_date_from']:
                        #         last_out_date = leave_item['request_date_from']
                if ('con nhỏ' in leave_item['holiday_status_name'].strip().lower()):
                    number_child_date != 1
            if (number_out_date > 0) or (number_child_date >0) or \
                (row['split_shift'] == True):
                return
            for i in range(1,16):
                if pd.notnull(row[f'attendance_attempt_{i}']):
                    data.append(row[f'attendance_attempt_{i}'])
                    io.append(row[f'attendance_inout_{i}'])
                    last_index = i
            if last_index!= 0:
                if row[f'attendance_attempt_{last_index}'] != row['last_attendance_attempt']:
                    data.append(row['last_attendance_attempt'])
                    io.append(row['attendance_inout_last'])
                    
            if pd.notnull(row['attendance_missing_from']):
                data.insert(0, row['attendance_missing_from'])
                data.insert(0, 'I')
            
            if pd.notnull(row['attendance_missing_from']):
                data.append(row['attendance_missing_from'])
                data.append('O')
            
            if len(data) > 1:
                # input('start')
                try:
                    self.save_calculate_attendance_restshift(row, data, io)
                except Exception as ex:
                    
                    logger.error(f'save_calculate_attendance_restshift {ex}')
            
        else:
            
            for index, leave_item in hr_leave_employee_date.iterrows():
                if ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()):
                    number_out_date = number_out_date + 1
                        # if pd.notnull(leave_item['request_date_from']):
                        #     out_date_array.append(leave_item)
                        #     if last_out_date > leave_item['request_date_from']:
                        #         last_out_date = leave_item['request_date_from']
                if ('con nhỏ' in leave_item['holiday_status_name'].strip().lower()):
                    number_child_date != 1
            if (number_out_date > 0) or (number_child_date >0) or \
                (row['split_shift'] == True):
                return
            
            try:
                data_next_scheduling = []
                io_next_scheduling = []
                data_prev_scheduling = []
                io_prev_scheduling = []
                min_time = None
                max_time = None
                scheduling_prev = df_scheduling_ver[df_scheduling_ver['date']== (row['date']-datetime.timedelta(days=1))]
                scheduling_next = df_scheduling_ver[df_scheduling_ver['date']== (row['date']+datetime.timedelta(days=1))]
                # index = 0
                # for add_to_prev_item in row['add_to_prev']['data']:
                #     data.insert(index, add_to_prev_item)
                #     index = index + 1
                # index = 0
                # for add_to_prev_item in row['add_to_prev']['io']:
                #     io.insert(index, add_to_prev_item)
                #     index = index + 1
                index = 0
                select_index = 0
                for add_to_prev_item in row['cut_to_next']['data']:
                    if (min_time is None) or (add_to_prev_item < min_time):
                        data_next_scheduling.insert(select_index, add_to_prev_item)
                        min_time = add_to_prev_item
                        io_next_scheduling.insert(select_index, row['cut_to_next']['io'][index])
                        select_index = select_index + 1
                    index = index + 1

                
            except Exception as ex:
                print(ex)
            last_index = 0
            for i in range(1,16):
                if pd.notnull(row[f'attendance_attempt_{i}']):
                    data.append(row[f'attendance_attempt_{i}'])
                    io.append(row[f'attendance_inout_{i}'])
                    last_index = i
            if last_index!= 0:
                if row[f'attendance_attempt_{last_index}'] != row['last_attendance_attempt']:
                    data.append(row['last_attendance_attempt'])
                    io.append(row['attendance_inout_last'])
            try:
                index = 0
                for data_prev_scheduling in row['cut_to_prev']['data']:
                    if data_prev_scheduling > max_time:
                        data_prev_scheduling.append(data_prev_scheduling)
                    for add_to_last_item in row['add_to_last']['io']:
                        io_prev_scheduling.append( row['cut_to_prev']['io'][index])
                        max_time = add_to_last_item
                    index = index + 1
            except Exception as ex:
                print(ex)
            # last_index = 0
            
            # for i in range(1,16):
            #     if pd.notnull(row[f'attendance_attempt_{i}']):
            #         data.append(row[f'attendance_attempt_{i}'])
            #         io.append(row[f'attendance_inout_{i}'])
            #         last_index = i
            # if last_index!= 0:
            #     if row[f'attendance_attempt_{last_index}'] != row['last_attendance_attempt']:
            #         data.append(row['last_attendance_attempt'])
            #         io.append(row['attendance_inout_last'])
            # try:
            #     for add_to_last_item in row['add_to_last']['data']:
            #         data.append(add_to_last_item)
            #     for add_to_last_item in row['add_to_last']['io']:
            #         io.append(add_to_last_item)
            # except Exception as ex:
            #     print(ex)
     
            print(row['company'])
            print('io_prev', io_prev_scheduling)
            print('data_prev', data_prev_scheduling)
            print('io_prev', io_next_scheduling)
            print('data_prev', data_next_scheduling)
            # print(row['cut_to_prev'])
            # print(row['cut_to_next'])
            print('current shift: ',row['shift_name'])
            print(row['employee_code'])
            print('current date: ', row['date'])
            # input("cut_to_next")
            ids = [int(row['id'])]
            print(len(data))
            if len(data_prev_scheduling) > 1:
                # input('start')
                try:
                    self.save_calculate_attendance_restshift(row, data_prev_scheduling, io_prev_scheduling)
                except Exception as ex:
                    
                    logger.error(f'save_calculate_attendance_restshift io_prev_scheduling {ex}')
            if len(data_next_scheduling) > 1:
                # input('start')
                try:
                    self.save_calculate_attendance_restshift(row, data_next_scheduling, io_next_scheduling)
                except Exception as ex:
                    
                    logger.error(f'save_calculate_attendance_restshift io_next_scheduling {ex}')
    def process_attendance_restshift(self, row):
        data = []
        io = []
        min_time = None
        max_time = None
        if pd.notnull(row[f'attendance_attempt_1']):
            min_time = row[f'attendance_attempt_1']
        if pd.notnull(row[f'last_attendance_attempt']):
            max_time = row[f'last_attendance_attempt']
        
        hr_leave_employee_date = self.df_hr_leave[(self.df_hr_leave['employee_code']
                                                     == row['employee_code']) & \
                                                    (self.df_hr_leave['date_str'] == row['date_str']) ]
        number_out_date = 0
        number_child_date = 0
        
        for index, leave_item in hr_leave_employee_date.iterrows():
            if ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()):
                number_out_date =  number_out_date + 1
                    # if pd.notnull(leave_item['request_date_from']):
                    #     out_date_array.append(leave_item)
                    #     if last_out_date > leave_item['request_date_from']:
                    #         last_out_date = leave_item['request_date_from']
            if ('con nhỏ' in leave_item['holiday_status_name'].strip().lower()):
                number_child_date != 1
        if (number_out_date > 0) or (number_child_date >0) or \
            (row['split_shift'] == True):
            return
        
        try:
            index = 0
            select_index = 0
            for add_to_prev_item in row['add_to_prev']['data']:
                if (min_time is None) or (add_to_prev_item < min_time):
                    data.insert(select_index, add_to_prev_item)
                    min_time = add_to_prev_item
                    io.insert(select_index, row['add_to_prev']['io'][index])
                    select_index = select_index + 1
                index = index + 1
            # index = 0
            # for add_to_prev_item in row['add_to_prev']['io']:
            #     io.insert(index, add_to_prev_item)
            #     index = index + 1
        except Exception as ex:
            print(ex)
        last_index = 0
        for i in range(1,16):
            if pd.notnull(row[f'attendance_attempt_{i}']):
                data.append(row[f'attendance_attempt_{i}'])
                io.append(row[f'attendance_inout_{i}'])
                last_index = i
        if last_index!= 0:
            if row[f'attendance_attempt_{last_index}'] != row['last_attendance_attempt']:
                data.append(row['last_attendance_attempt'])
                io.append(row['attendance_inout_last'])
        try:
            index = 0
            for add_to_last_item in row['add_to_last']['data']:
                if add_to_last_item > max_time:
                    data.append(add_to_last_item)
                # for add_to_last_item in row['add_to_last']['io']:
                    io.append( row['add_to_last']['io'][index])
                    max_time = add_to_last_item
                index = index + 1
        except Exception as ex:
            print(ex)
        print('shift_name_prev: ', row['shift_name_prev'])
        print('add_to_prev: ', row['add_to_prev'])
        
        print(row['shift_name_next'])
        print('add_to_last: ', row['add_to_last'])
        
        print(row['company'])
        print('io', io)
        print('data', data)
        print(row['cut_to_prev'])
        print(row['cut_to_next'])
        print('current shift: ',row['shift_name'])
        print(row['employee_code'])
        print('current date: ', row['date'])
        # input("cut_to_next")
        ids = [int(row['id'])]
        print(len(data))
        if len(data) > 1:
            # input('start')
            try:
                self.save_calculate_attendance_restshift(row, data, io)
            except Exception as ex:
                
                logger.error(f'save_calculate_attendance_restshift {ex}')
        
        
    def save_calculate_attendance_restshift(self, row, data, io):
        
        holiday_work_time = 0
        night_worktime_minute = 0 
        # if pd.notnull(row['add_to_prev']) :
        #     if (len(row['add_to_prev']['data'])>0) :
        total_work_time = row['total_work_time']
        # if rest_shifts or total_work_time == 0:
        #     return 0
        real_time_in = data[0].replace(second=0)
        real_time_out = data[-1].replace(second=0)
        shift_name = str(row['shift_name'])
        
        start_work_date_time = real_time_out
        if pd.notnull(row['shift_start_datetime']):
            start_work_date_time = row['shift_start_datetime'].replace(second=0)
            
        end_work_date_time = real_time_out
        if pd.notnull(row['shift_end_datetime']):
            end_work_date_time = row['shift_end_datetime'].replace(second=0)
        
        start_rest_date_time = row['rest_start_datetime'].replace(second=0)
        end_rest_date_time = row['rest_end_datetime'].replace(second=0)
        
        current_program = min(real_time_out, start_rest_date_time) - \
                max(real_time_in, start_work_date_time)
        stage_fist_worktime = max(0, current_program.total_seconds()/60.0)

        current_program = min(real_time_out, end_work_date_time) - \
                    max(real_time_in, end_rest_date_time)
        stage_second_worktime = max(0, current_program.total_seconds()/60.0)
        total_work_time = stage_fist_worktime + stage_second_worktime
        
        if ('/OFF' in shift_name ) or ('/PH' in shift_name) \
            or ('PH/' in shift_name) or ('OFF/' in shift_name):
            total_work_time = min(240, total_work_time)
        # start night datetime from 22h to 6h:00
        start_night_date_time = start_work_date_time.replace(hour=22, minute=0, second=0)
        end_night_date_time = (start_night_date_time + datetime.timedelta(days=1)).replace(hour=6)
        try:
            start_holiday_datetime = row['holiday_from'].replace(hour=0, minute=0, second=0)
            end_holiday_datetime = row['holiday_to'].replace(hour=0, minute=0, second=0)
        except:
            start_holiday_datetime = None
            end_holiday_datetime = None
        start_rest_date_time = row['rest_start_datetime'].replace(second=0)
        end_rest_date_time = row['rest_end_datetime'].replace(second=0)
        # night stage fist
        current_program_night = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                max(real_time_in, start_work_date_time, start_night_date_time)
        stage_fist_night = max(0, current_program_night.total_seconds()/60.0)
        if not (start_holiday_datetime is None):
        #night holiday
            current_program = min(real_time_out, start_rest_date_time, end_holiday_datetime, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time, start_holiday_datetime)
            stage_fist = max(0, current_program.total_seconds()/60.0)
            
            # holiday stage fist
            current_program_holiday = min(real_time_out, start_rest_date_time, end_holiday_datetime) - \
                    max(real_time_in, start_work_date_time, start_holiday_datetime)
            stage_fist_holiday = max(0, current_program_holiday.total_seconds()/60.0)
        else:
            stage_fist = 0
            stage_fist_holiday = 0

        # night stage second
        current_program_night = min(real_time_out, end_work_date_time, end_night_date_time) - \
                max(real_time_in, end_rest_date_time, start_night_date_time)
        stage_second_night = max(0, current_program_night.total_seconds()/60.0)
        if not (start_holiday_datetime is None):
            # night holiday
            current_program = min(real_time_out, end_work_date_time, end_holiday_datetime, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time, start_holiday_datetime)
            stage_second = max(0, current_program.total_seconds()/60.0)
            
            # holiday stage second
            current_program_holiday = min(real_time_out, end_work_date_time, end_holiday_datetime) - \
                    max(real_time_in, end_rest_date_time, start_holiday_datetime)
            stage_second_holiday = max(0, current_program_holiday.total_seconds()/60.0)
        # print('{}: night after now - holiday: {} - {} - {} -night - {}- {} - real - {} - {} '.format(row['name_employee'] ,row['name'], \
        #                     stage_fist, stage_second, start_night_date_time, end_night_date_time, row['start_work_date_time'], row['end_work_date_time']))
        else:
            stage_second = 0
            stage_second_holiday = 0

        stage_night_only = int(stage_fist_night + stage_second_night)
        stage_night =  int(stage_fist + stage_second)
        holiday_work_time = int(stage_fist_holiday + stage_second_holiday)
        
        end_night_date_time = start_work_date_time.replace(hour=6, minute=0, second=0)
        start_night_date_time = (end_night_date_time+ datetime.timedelta(days=-1)).replace(hour=22, minute=0, second=0)
        
        # night stage 1
        current_program_night = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                max(real_time_in, start_work_date_time, start_night_date_time)
        stage_fist_night = max(0, current_program_night.total_seconds()/60.0)
        if not (start_holiday_datetime is None):
            # night holiday stage 1
            current_program = min(real_time_out, start_rest_date_time, end_holiday_datetime, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time, start_holiday_datetime)
            stage_fist = max(0, current_program.total_seconds()/60.0)
        else:
            stage_fist = 0
        ##########################################################
        # night stage 2
        current_program_night = min(real_time_out, end_work_date_time, end_night_date_time) - \
                max(real_time_in, end_rest_date_time, start_night_date_time)
        stage_second_night = max(0, current_program_night.total_seconds()/60.0)
        if not (start_holiday_datetime is None):
            # night holiday stage 2
            current_program = min(real_time_out, end_work_date_time, end_holiday_datetime, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time, start_holiday_datetime)
            stage_second = max(0, current_program.total_seconds()/60.0)
            # print('{}: night before now - holiday: {} - {} - {} -night - {}- {} - real - {} - {} '.format(row['name_employee'] ,row['name'], \
            #                     stage_fist, stage_second, start_night_date_time, end_night_date_time, row['start_work_date_time'], row['end_work_date_time']))
        else:
            stage_second = 0

        stage_morning = int(stage_fist + stage_second)
        stage_morning_only = int(stage_second_night + stage_fist_night)
        night_worktime_minute = int(stage_night_only + stage_morning_only)
                # if row['employee_code']=='APG230320013':    
                #     print('night_worktime_minute', night_worktime_minute)
                #     input("Press Enter to continue...")
                # return min(total_work_time, stage_night + stage_morning), min(total_work_time, holiday_work_time), min(total_work_time, night_worktime_minute)
                  
        if row['total_work_time'] < total_work_time:
        #     print('shift_name_prev: ', row['shift_name_prev'])
        #     print('add_to_prev: ', row['add_to_prev'])
            
        #     print(row['shift_name_next'])
        #     print('add_to_last: ', row['add_to_last'])
            
        #     print(row['company'])
        #     print('io', io)
        #     print('data', data)
        #     print('total_work_time old: ', row['total_work_time'])
        #     print('total_work_time new: ', total_work_time)
        #     print(row['cut_to_prev'])
        #     print(row['cut_to_next'])
        #     print('current shift: ',row['shift_name'])
        #     print(row['employee_code'])
        #     print('current date: ', row['date'])
        #     input("cut_to_next")
            
            ids = [int(row['id'])]
            # input("Press Enter to continue...1")
            update_data = {'total_work_time': total_work_time}
            # input("Press Enter to continue...2")
            update_data['missing_checkin_break'] = False
            # input("Press Enter to continue...3")
            # if (int(row['total_attendance']) != len(data)):
            update_data['total_attendance'] = int(len(data))
            # input("Press Enter to continue...4")
            index= 0
            for attendence_item in data:
                if index > 15:
                    continue
                if pd.notnull(attendence_item):
                    update_data[f'attendance_attempt_{index+1}'] = \
                        attendence_item.strftime('%Y-%m-%d %H:%M:%S')
                index = index + 1
            
            try:
                print('update data', update_data)
                # input("Press Enter to continue...")
                # self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                
            except Exception as ex:
                logger.error(f"save_calculate_attendance_restshift write {update_data} error: {ex}")
                # input("Press Enter to continue ex...")
                
    def caculate_attendance_timeoff(self, row, hr_leave):
        cut_to_prev ={ 'data' : [], 'io' :[]}
        cut_to_next = { 'data' : [], 'io' :[]}
        # diff_with_prev = (row['date'] - row['date_prev']).days
        # diff_with_next = (row['date_next'] - row['date']).days
        
        list_data = []
        list_io = []
        list_data_next = []
        list_io_next = []
        print(row['id'])
        print(hr_leave['id_scheduling'])
        
        df_hr_leave_row = hr_leave[hr_leave['id_scheduling']== row['id']]
        if len(df_hr_leave_row.index)>0:
            # input('con cho bi con')
            hr_leave_row = df_hr_leave_row.iloc[0]
            if (pd.notnull(hr_leave_row['overtime_from'])) and \
                (pd.isnull(hr_leave_row['overtime_to'])):
                if (pd.notnull(hr_leave_row['overtime_from']) != False) and \
                    (pd.isnull(hr_leave_row['overtime_to']) != False):
                    for i in range(1,16):
                        if pd.isnull(row[f'attendance_attempt_{i}']):
                            continue
                        
                        if (row[f'attendance_attempt_{i}'] < hr_leave_row['overtime_from']) or \
                                row[f'attendance_attempt_{i}'] > hr_leave_row['overtime_to']:
                            continue
                        offset = 0
                        diff_with_current = math.abs(row[f'attendance_attempt_{i}']-row['shift_start_datetime'])
                        if diff_with_current < \
                            math.abs(row[f'attendance_attempt_{i}']-row['shift_start_datetime_prev']):
                            offset = -1
                            diff_with_current = math.abs(row[f'attendance_attempt_{i}']-row['shift_start_datetime_prev'])
                            
                        if diff_with_current < \
                            math.abs(row[f'attendance_attempt_{i}']-row['shift_start_datetime_next']):
                            offset = 1
                            
                        if (offset == -1): 
                            # if (diff_with_prev == 1):
                                list_data.append(row[f'attendance_attempt_{i}'])
                                list_io.append(row[f'attendance_inout_{i}'])
                        elif (offset == 1):
                            list_data_next.append(row[f'attendance_attempt_{i}'])
                            list_io_next.append(row[f'attendance_inout_{i}'])
                            
            if (len(list_data) > 0):
                cut_to_prev['data'].append(list_data[0])
                cut_to_prev['io'].append(list_io[0])
                
            if (len(list_data) > 1):
                cut_to_prev['data'].append(list_data[-1])
                cut_to_prev['io'].append(list_io[-1])
                
            if (len(list_data_next) > 0):
                cut_to_next['data'].append(list_data_next[0])
                cut_to_next['io'].append(list_io_next[0])
                
            if (len(list_data_next) > 1):
                cut_to_next['data'].append(list_data_next[-1])
                cut_to_next['io'].append(list_io_next[-1])
            
        return cut_to_prev, cut_to_next       
    def caculate_attendance_restshift(self, row):
        cut_to_prev ={ 'data' : [], 'io' :[]}
        cut_to_next = { 'data' : [], 'io' :[]}
        diff_with_prev = (row['date'] - row['date_prev']).days
        diff_with_next = (row['date_next'] - row['date']).days
        list_data = []
        list_io = []
        list_data_next = []
        list_io_next = []
        if row['total_shift_work_time'] == 0:
            for i in range(1,16):
                if pd.isnull(row[f'attendance_attempt_{i}']):
                    continue
                
                if (row[f'attendance_attempt_{i}'] <= row['shift_start_datetime']): 
                    if (diff_with_prev == 1):
                        list_data.append(row[f'attendance_attempt_{i}'])
                        list_io.append(row[f'attendance_inout_{i}'])

                elif (diff_with_next == 1):
                    list_data_next.append(row[f'attendance_attempt_{i}'])
                    list_io_next.append(row[f'attendance_inout_{i}'])
                    
        if (len(list_data) > 0):
            cut_to_prev['data'].append(list_data[0])
            cut_to_prev['io'].append(list_io[0])
            
        if (len(list_data) > 1):
            cut_to_prev['data'].append(list_data[-1])
            cut_to_prev['io'].append(list_io[-1])
            
        if (len(list_data_next) > 0):
            cut_to_next['data'].append(list_data_next[0])
            cut_to_next['io'].append(list_io_next[0])
            
        if (len(list_data_next) > 1):
            cut_to_next['data'].append(list_data_next[-1])
            cut_to_next['io'].append(list_io_next[-1])
        
        # if len( cut_to_prev['data'])>0:
        #     print(row['shift_name'])
        #     print(row['employee_code'])
        #     print(row['date'])
        #     print( cut_to_prev)
        #     print(diff_with_prev)
        #     print(row['date'])
        #     print('date_prev', row['date_prev'])
        #     print('date_next', row['date_next'])
        #     input("cut_to_prev")
        # if len( cut_to_prev['data'])>0:
        #     print(row['shift_name'])
        #     print(row['employee_code'])
        #     print(row['date'])
        #     print(cut_to_next)
        #     input("cut_to_next")
        return cut_to_prev, cut_to_next
    def auto_fill_missing_attendance_trans(self):
        # df_group = self.df_old.groupby(['EID','PCODE'], as_index=False)
        self.df_old['ID'] = self.df_old['ID'].apply(lambda row: self.convert_code_to_int(row))
        # self.df_old['Date'] = pd.to_datetime(self.df_old['Ngày']).dt.date
        # self.df_old['shift_name']=self.df_old['shift_name'].str.strip()
        # self.df_old = self.df_old.merge(self.df_shift[['id','name', 'total_work_time','company_name']], \
        #         left_on=['shift_name','company_name'], right_on = ['name','company_name'], how='left', suffixes=( '' ,'_y')).drop(['name'], axis=1)
        self.df_old = self.df_old.drop_duplicates(subset=["ID", "Giờ"], keep="last") \
                            .sort_values(['Giờ'])  
        self.df_old = self.df_old.sort_values(['ID', 'Giờ']).groupby(['ID','Ngày']).apply(lambda x: [list(x['Giờ']),list(x['in_out'])]).apply(pd.Series).reset_index()
        self.df_old.columns = ['ID','Ngày', 'time','in_out']
        
        self.df_employees['time_keeping_code'] = self.df_employees['time_keeping_code'].apply \
                    (lambda row: self.convert_code_to_int(row))
        self.df_old = self.df_old.merge(self.df_employees[['time_keeping_code', 'department_name' ,'company_name','code', 'name', 'employee_ho']],
                left_on=['ID'], right_on=['time_keeping_code'], how='left', suffixes=( '' ,'_employee'))

        df_hr_leave_group = self.df_hr_leave.groupby(['employee_code', 'date_str']) \
                    .apply(lambda x: [list(x['id']),list(x['holiday_status_name']), list(x['time']), list(x['state']) , \
                        list(x['minutes']), list(x['attendance_missing_from']), list(x['attendance_missing_to'])]) \
                    .apply(pd.Series)
                    
        df_hr_leave_group.columns = ['id_leave','holiday_status_name', 
                            'time', 'state', 'minutes', 'attendance_missing_from_leave','attendance_missing_to_leave']
        self.df_scheduling_ver['time_keeping_code'] = self.df_scheduling_ver['time_keeping_code'].apply \
                    (lambda row: self.convert_code_to_int(row))
                    
        df_explanation_group = self.df_explanation_data.groupby(['employee_code', 'invalid_date']) \
                    .apply(lambda x: [list(x['id']),list(x['invalid_type']), list(x['shift_from']), list(x['shift_to']) , \
                        list(x['validated']), list(x['attendance_missing_from']), list(x['attendance_missing_to'])]) \
                    .apply(pd.Series)
                    
        df_explanation_group.columns = ['id_explanation','explanation_type', 
                            'shift_from', 'shift_to', 'validated', 'attendance_missing_from_explanation','attendance_missing_to_explanation']
        # df_scheduling_ver.to_excel('step1_old_df_explanation_group.xlsx')
        df_scheduling_ver = self.df_scheduling_ver.merge(df_explanation_group, \
            left_on=['employee_code', 'date_str'], right_index = True, how='left', suffixes=( '' ,'_explanation')) \
            .sort_values(['employee_code', 'date']).reset_index(drop=True)
        # print('new')
        # print(df_hr_leave_group)
        df_scheduling_ver = df_scheduling_ver.merge(df_hr_leave_group, \
            left_on=['employee_code', 'date_str'], right_index = True, how='left', suffixes=( '' ,'_leave')) \
            .sort_values(['employee_code', 'date']).reset_index(drop=True)
        df_scheduling_ver['real_shift_end'] = df_scheduling_ver.apply(lambda x: x['shift_end'] if x['shift_end'] >= x['shift_start'] else x['shift_end'] + 24, 
                            axis=1)
        df_scheduling_ver[['shift_name_prev', 'shift_end_prev', 'employee_code_prev']] = df_scheduling_ver[['shift_name', 'real_shift_end', 'employee_code']].shift(1)
        df_scheduling_ver[['shift_name_next', 'shift_start_next', 'employee_code_next']] = df_scheduling_ver[['shift_name', 'shift_start', 'employee_code']].shift(-1)
        df_scheduling_ver[['total_out_date', 'total_tc_date' , 'total_ncl_date' , 'total_compensation_leave_date', 
                    'total_dimuon_vesom_date' , 'total_al_date', 
                    'total_compensation_leave', 'id_cl'] ]= df_scheduling_ver.apply(lambda x: self.calculate_addtion_time_from_hr_leave(x), axis=1, result_type='expand')
        df_scheduling_ver[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                df_scheduling_ver.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
        # df_scheduling_ver.to_excel('step1_old.xlsx')
        # input('self input')
        self.df_old = self.df_old.merge(df_scheduling_ver[['id','time_keeping_code', 'date_str','shift_name','shift_start','shift_end', 'total_work_time', 'total_work_time_shift',
                                                                'last_attendance_attempt', 'shift_name_prev', 'shift_end_prev', 'date', 'total_shift_work_time',
                                                                'shift_name_next', 'shift_start_next', 'real_shift_end', 'time_keeping_code_employee',
                                                                'attendance_attempt_1','attendance_attempt_2','attendance_attempt_3','attendance_attempt_4',
                                                                'attendance_attempt_5','attendance_attempt_6','attendance_attempt_7','attendance_attempt_8',
                                                                'attendance_attempt_9','attendance_attempt_10','attendance_attempt_11', 'fix_rest_time',
                                                                'attendance_attempt_12','attendance_attempt_13','attendance_attempt_14','attendance_attempt_15',
                                                                'attendance_inout_1','attendance_inout_2','attendance_inout_3',
                                                                'attendance_inout_4','attendance_inout_5','attendance_inout_6',
                                                                'attendance_inout_7','attendance_inout_8','attendance_inout_9', 'amount_al_reserve', 'amount_cl_reserve',
                                                                'attendance_inout_10','attendance_inout_11','attendance_inout_12', 
                                                                'attendance_inout_13','attendance_inout_14','attendance_inout_15',
                                                                'attendance_inout_last','shift_start_datetime','shift_end_datetime', 
                                                                'rest_start_datetime', 'rest_end_datetime', 'start_work_time', 'end_work_time',
                                                                'start_rest_time', 'end_rest_time', 'code', 'employee_code',
                                                                'attendance_missing_from_leave', 'attendance_missing_to_leave',
                                                                'attendance_missing_from_explanation','attendance_missing_to_explanation', 
                                                                'rest_start', 'rest_end', 'rest_shift','company_name','employee_code_prev','employee_code_next',
                                                                'ot_holiday', 'ot_normal']],
                left_on=['ID', 'Ngày'], right_on=['time_keeping_code_employee', 'date_str'], how='left', suffixes=( '' ,'_scheduling'))
        self.df_old = self.df_old.drop_duplicates(
            subset=['ID', 'Ngày'], keep="last")
        
        # df_duplicate = df_al_report[~df_al_report.index.isin(self.df_al_report.index)]
        # self.df_old.to_excel('step2_old.xlsx')
        # input('self step 2 input')
        
        self.df_old[['move_to_prev', 'move_to_next', 'outdate_result', 'total_out', 
                    'total_out_morethan_30', 'total_out_morethan_30_noleave']] = self.df_old.apply(lambda x: self.find_nearest(x), axis=1, result_type= 'expand')
        self.df_old['add_from_next'] = self.df_old['move_to_prev'].shift(-1)
        self.df_old['add_from_prev'] = self.df_old['move_to_next'].shift(1)
        # df_filter_attendence_new = self.df_old[((self.df_old['shift_name'].isnull())|(self.df_old['shift_name']=='-'))

        self.df_old= self.df_old[(self.df_old['date']>=self.date_array[0]) & 
                                (self.df_old['date']<=self.date_array[len(self.date_array)-1])]
        
        df_filter_attendence_new = self.df_old[((self.df_old['shift_name']=='-')|
                                                    ((self.df_old['total_work_time']==0)))
                                                & (self.df_old['code'].notnull())]
        # self.df_old.to_excel('step3_old.xlsx')
        # input('self step 3 input')
        # df_filter_attendence_new.apply(lambda row: self.fill_attendance_tran(row), axis=1)
        # print(len(df_filter_attendence_new.index))
        # for attence_index, row in df_filter_attendence_new.iterrows():
            # print(row)
    def fill_attendance_tran(self, row):
        update_data = {}
        is_need_update = False
        # if pd.notnull(row['add_from_prev']):
        try:
            for item in row['add_from_prev']:
                row['time'].insert(0, item['time'])
                row['in_out'].insert(0, item['in_out'])
        except Exception as ex:
            print("add_from_prev ex: ", ex)
        try:
            for item in row['add_from_next']:
                row['time'].append(item['time'])
                row['in_out'].append(item['in_out'])
        except Exception as ex:
            print("add_from_next ex: ", ex)
        index = 1
        # if pd.notnull(row['time']):
        try:
            for item in row['time']:
                time_str =  item.strftime("%Y-%m-%d %H:%M:%S")
                if index <= 15:
                    update_data[f'attendance_attempt_{index}'] = time_str
                    if update_data[f'attendance_attempt_{index}'] != row[f'attendance_attempt_{index}']:
                        is_need_update = True
                    update_data[f'attendance_inout_{index}'] = row['in_out'][index-1]
                    if update_data[f'attendance_inout_{index}'] != row[f'attendance_inout_{index}']:
                        is_need_update = True
                    if not(row['in_out'][index-1] in ['I', 'O']):
                        if  (update_data[f'attendance_inout_{index}'] != False) and  (update_data[f'attendance_inout_{index}'] != ''):
                            is_need_update = True
                        update_data[f'attendance_inout_{index}'] = False
                update_data['last_attendance_attempt'] = time_str
                update_data['attendance_inout_last']= row['in_out'][index-1]
                index= index + 1
        except Exception as ex:
            print("fill_attendance_tran ex: ", ex)
            
        # assign blank IO
        if index>2:
            if not (row['in_out'][0] in ['I', 'O']):
                row['in_out'][0] = 'I'
            if not (row['in_out'][index-2] in ['I', 'O']):
                row['in_out'][index-2] = 'O'

            if index < 17:
                for index_add_blank in range(index, 16):
                    if (row[f'attendance_attempt_{index_add_blank}']!= False) and pd.notnull(row[f'attendance_attempt_{index_add_blank}']):
                        is_need_update = True
                    update_data[f'attendance_attempt_{index_add_blank}'] = False
                    if (row[f'attendance_inout_{index_add_blank}']!= False) and pd.notnull(row[f'attendance_inout_{index_add_blank}']):
                        is_need_update = True
                    update_data[f'attendance_inout_{index_add_blank}'] = False
       

        # if row['employee_ho'] == True:
        #     worktime_collect = self.calculate_total_work_time_from_couples(row)
            # update_data[f'actual_total_work_time'] = worktime_collect['total_work_time']
            # # if (row['code']=='APG231110002' )and (row['date_str']=='2024-01-02'):
            # #     print(worktime_collect)
            # #     exit(1)
            # if update_data[f'actual_total_work_time'] != 0:
            #     is_need_update = True
            # if pd.notnull(row['start_work_time']):
            #     if row['start_work_time'] != row['shift_start']:
            #         update_data[f'shift_start'] = row['start_work_time'] 
            #         is_need_update = True
            # else:
            #     if row['shift_start'] != 12:
            #         update_data[f'shift_start'] = 12 
            #         is_need_update = True

            # if pd.notnull(row['end_work_time']):
            #     if row['end_work_time'] != row['shift_end']:
            #         update_data[f'shift_end'] = row['end_work_time'] 
            #         is_need_update = True
            # else:
            #     if row['shift_end'] != 12:
            #         update_data[f'shift_end'] = 12 
            #         is_need_update = True

            # if pd.notnull(row['end_rest_time']):
            #     if row['end_rest_time'] != row['rest_end']:
            #         update_data[f'rest_end'] = row['end_rest_time'] 
            #         is_need_update = True
            # else:
            #     if row['rest_end'] != 12:
            #         update_data[f'rest_end'] = 12 
            #         is_need_update = True

            # if pd.notnull(row['start_rest_time']):
            #     if row['start_rest_time'] != row['rest_start']:
            #         update_data[f'rest_start'] = row['start_rest_time'] 
            #         is_need_update = True
            # else:
            #     if row['rest_start'] != 12:
            #         update_data[f'rest_start'] = 12 
            #         is_need_update = True
            # if pd.notnull(row['time_keeping_code_employee']):
            #     if int(row['time_keeping_code_scheduling']) != int(row['time_keeping_code_employee']):
            #         update_data[f'time_keeping_code'] = "{:06d}".format(int(row['time_keeping_code_employee']))
            #         is_need_update = True
            # if pd.notnull(row['total_work_time_shift']):
            #     if row['total_shift_work_time'] != row['total_work_time_shift']:
            #         update_data[f'total_shift_work_time'] = row['total_work_time_shift']
            #         is_need_update = True
            # if (row['code']=='APG230321013' )and (row['date_str']=='2024-01-15'):
            #     print(worktime_collect)
            #     print(row['time_keeping_code_employee'])
            #     print(row['time_keeping_code'])
            #     exit(1)
        
        if is_need_update:
            try:
                ids = [int(row['id'])]
                # self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                print(f'updated: {ids}- {update_data}')
            except Exception as ex:
                logger.error('fill attance trans: ', ex)
                
        else:
            print('no need update', update_data)
    def calculate_total_work_time_from_couples(self, row):
        result = {'total_work_time':0,'night_hours_normal':0}
        couples = self.find_couple(row)
        # print('couples: ', couples)
        result = self.calculate_actual_work_time_ho(row, couples)
        return result
    
    def find_couple_out_date(self, row):
        result = []
        time_queue = []
        index = 0
        for attendance_inout in row['in_out']:
            if attendance_inout == 'O':
                time_queue.append(index)
            elif attendance_inout =='I':
                if len(time_queue)>0:
                    real_time_out = row['time'][time_queue[0]]
                    real_time_in = row['time'][index]
                    out_work_time = self.calculate_actual_out_work_time_couple(row, real_time_out, real_time_in)
                    number_leave = 0
                    number_explanation = 0
                    # result.append([time_queue[0], index, out_work_time])
                    leave_index = 0
                    try:
                        for leave_item_from in row['attendance_missing_from_leave']:
                            leave_item_to =  row['attendance_missing_from_leave'][leave_index]
                            if (leave_item_from <= real_time_out) and(leave_item_to >= real_time_in):
                                number_leave= number_leave + 1
                        
                            leave_index = leave_index + 1
                    except Exception as ex:
                        print(ex)
                    
                    explanation_index = 0
                    try:
                        for explanation_item_from in row['attendance_missing_from_explanation']:
                            explanation_item_to =  row['attendance_missing_to_explanation'][explanation_index]
                            if (explanation_item_from <= real_time_out) and(explanation_item_to >= real_time_in):
                                number_explanation= number_explanation + 1
                        
                            leave_index = leave_index + 1
                    except Exception as ex:
                        print(ex)
                    # if (row['code']=='APG231110002' )and (row['date_str']=='2024-01-02'):
                    #     print('worktime_collect: ', result)
                    #     print(row['in_out'])
                    #     print(row['time'])
                    #     exit(1)
                    leave_data = {'leave': number_leave, 'explanation': number_explanation}
                    result.append([time_queue[0], index, out_work_time, leave_data])
                time_queue = []
            index = index + 1
            
        return result
    
    def calculate_actual_out_work_time_couple(self, row, real_time_in, real_time_out):
        """
        calculate actual work time with real time in, time out , restime in out
        """
        # print('aaaaaa')
        result = 0
        fix_rest_time = row['fix_rest_time']
        
        try:
            if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
                result = 0
            else:
                start_work_date_time = row['shift_start_datetime'].replace(second=0)
                end_work_date_time = row['shift_end_datetime'].replace(second=0)

                start_rest_date_time = row['rest_start_datetime'].replace(second=0)
                end_rest_date_time = row['rest_end_datetime'].replace(second=0)
                
                current_program = min(real_time_out.replace(second=0), start_rest_date_time) \
                                            - max(real_time_in.replace(second=0), start_work_date_time)
                stage_fist = max(0, current_program.total_seconds()//60.0)
                
                current_program = min(real_time_out.replace(second=0), end_work_date_time) \
                                            - max(real_time_in.replace(second=0), end_rest_date_time)
                stage_second = max(0, current_program.total_seconds()//60.0)

                result = int(stage_fist + stage_second)
               
                if fix_rest_time:
                    result = min(result, 480)

        except Exception as ex:
            print(real_time_out)
            print(real_time_in)
            print("In cal out actual", ex)
        
        return result

    def find_couple(self, row):
        result = []
        time_stack = []
        index = 0
        for attendance_inout in row['in_out']:
            if attendance_inout == 'I':
                time_stack.insert(0,index)
            elif attendance_inout =='O':
                if len(time_stack)>0:
                    result.append([time_stack[0], index])
                    # if (row['code']=='APG231110002' )and (row['date_str']=='2024-01-02'):
                    #     print('worktime_collect: ', result)
                    #     print(row['in_out'])
                    #     print(row['time'])
                    #     exit(1)
                time_stack = []
            index = index + 1
        return result

    def calculate_actual_work_time_ho(self, row, couples):
        update_data_item = {'total_work_time':0,'night_hours_normal':0}
        update_data = {'total_work_time':0,'night_hours_normal':0}
        for couple in couples:
            in_index = couple[0]
            out_index = couple[1]
            real_time_in = row['time'][in_index]
            real_time_out = row['time'][out_index]
            update_data_item = self.calculate_actual_work_time_couple(row, real_time_in, real_time_out)
            # update_data['total_work_time'] = update_data['total_work_time'] + update_data_item['total_work_time']
            # update_data['night_hours_normal'] = update_data['night_hours_normal'] + update_data_item['night_hours_normal']
        # if row['id']>0:
        #         ids = [int(row['id'])]
        #         self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
        #         print(f'sucess: {update_data}')
        return update_data_item
    def calculate_actual_work_time_couple(self, row, real_time_in, real_time_out):
        """
        calculate actual work time with real time in, time out , restime in out


        """
        # print('aaaaaa')
        result = 0
        night_work_time = 0
        # try:
        rest_shifts = row['rest_shift']
        # print(f"{rest_shifts}-{total_work_time}")
        # if rest_shifts:
        #     # print(f"{rest_shifts}-{total_work_time}")
        #     result = 0
        # else:
            # try:
            # real_time_in = row['attendance_attempt_1'].replace(second=0)
            # real_time_out = row['last_attendance_attempt'].replace(second=0)
            # print(f"{real_time_in}-{real_time_out}")
        if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
            # print(f"{real_time_in}-{real_time_out}")
            result = 0
        else:
            start_work_date_time = row['shift_start_datetime'].replace(second=0)
            round_minute = round(row['shift_end_datetime'].minute, -1)
            end_work_date_time = row['shift_end_datetime'].replace(minute=round_minute, second=0)

            start_rest_date_time = row['rest_start_datetime'].replace(second=0)
            round_minute = round(row['rest_end_datetime'].minute, -1)
            end_rest_date_time = row['rest_end_datetime'].replace(minute=round_minute, second=0)
            
            current_program = min(real_time_out, start_rest_date_time) - max(real_time_in, start_work_date_time)
            stage_fist = max(0, current_program.total_seconds()//60.0)
            
            current_program = min(real_time_out, end_work_date_time) - max(real_time_in, end_rest_date_time)
            stage_second = max(0, current_program.total_seconds()//60.0)

            result = int(stage_fist + stage_second)
            result = min(result, 480)
            night_work_time = 0
            # night_work_time = self.calculate_night_work_time(row)
            if '/OFF' in row['shift_name']:
                result = min(result, 240)
            night_work_time = min(night_work_time, result)
            # except Exception as ex:
            #     print("In cal actual", ex)
        update_data = {'total_work_time': result, 'night_hours_normal': night_work_time}
        # if result> 0:
        #     try:
        #         resource_calendar_id = row['resource_calendar_id'][1]
        #         update_data['resource_calendar'] = resource_calendar_id
        #         shift_name = row['shift_name']
        #         if  (not '/' in shift_name ) and (row['total_shift_work_time']>0) and \
        #             '44' in resource_calendar_id and row['attendance_attempt_3'] == False:
                    
        #             update_data['missing_checkin_break'] = True
        #             print('Thieu cham cong')
        #         else: 
        #             update_data['missing_checkin_break'] = False
        #     except Exception as ex:
        #         update_data['resource_calendar'] = ''
        #         update_data['missing_checkin_break'] = False
        #         print("calculate thieu cham cong err: ", ex)
        # else:
        #     update_data['missing_checkin_break'] = False

        # print(f'night_work_time {night_work_time}')
        # print(f'total_work_time {result}-{real_time_in}-{real_time_out}-{start_rest_date_time}-{end_rest_date_time}--{stage_fist} + {stage_second}')
        
        
        return update_data

        # except Exception as ex:
        #     print("calcuate actual time err", ex)
        #     return 0
    def find_nearest(self, row):
        move_to_prev = []
        move_to_next = []
        outdate_result = []
        index = 0
        for item in row['time']:
            # if pd.isnull(row['shift_end_prev']) or pd.isnull(row['shift_start_next']):
            #     continue

            item_float = item.hour + item.minute/60 + item.second/3600
            shift_start = row['shift_start'] + 24
            shift_end = row['real_shift_end'] + 24
            # shift_end = shift_end if shift_end >= shift_start else shift_end + 24
            item_float_compare_with_previous = item_float + 24    
            current_dif = min(abs(shift_start-item_float_compare_with_previous), abs(shift_end-item_float_compare_with_previous))
            if row['employee_code'] == row['employee_code_prev'] :
                
                # print(f"float {item_float} --- {item.hour} --- {item.minute/60}")
                # print("prev diff ", abs(row['shift_end_prev']-item_float_compare_with_previous))
                # print("current diff: ", current_dif)
                value_prev = (abs(row['shift_end_prev']-item_float_compare_with_previous) < current_dif)
            
                if value_prev==True:
                    # print("ok pre")
                    move_to_prev.append({'time': item, 'in_out':row['in_out'][index]})     

            if row['employee_code'] == row['employee_code_next'] :
                shift_start_next = row['shift_start_next'] + 24
                value_next = (abs(shift_start_next-item_float) < current_dif)
                # print("next diff: ", abs(shift_start_next-item_float))
                if value_next == True:
                    # print('ok next', shift_start_next)
                    move_to_next.append({'time': item, 'in_out':row['in_out'][index]})
            index = index + 1
        # print(f"{row['time']}-----{row['shift_name']}_{row['shift_start']}-----{move_to_prev}-{row['shift_end_prev']}----{[move_to_next]} = {row['shift_start_next'] }_{move_to_next}")
        for item in move_to_prev:
            row['time'].remove(item['time'])
            row['in_out'].remove(item['in_out'])
        for item in move_to_next:
            row['time'].remove(item['time'])
            row['in_out'].remove(item['in_out'])
            
        outdate_result = self.find_couple_out_date(row=row)
        # leave_data = {'leave': number_leave, 'explanation': number_explanation}
        total_out = 0
        total_out_morethan_30 = 0
        total_out_morethan_30_noleave = 0
        for outdate_item in outdate_result:
            explanation_out = outdate_item[3]['explanation']
            leave_out = outdate_item[3]['leave']
            total_out = outdate_item[2] + total_out
            if outdate_item[2] > 30:
                total_out_morethan_30 = total_out_morethan_30 + outdate_item[2]
                if (explanation_out + leave_out) == 0:
                    total_out_morethan_30_noleave = total_out_morethan_30_noleave + outdate_item[2]
        
        return move_to_prev, move_to_next, outdate_result, total_out, total_out_morethan_30, total_out_morethan_30_noleave

    def calculate_addtion_time_from_hr_leave(self, row):
        total_tc_date =0 
        total_ncl_date =0 
        total_compensation_leave_date = 0 
        total_dimuon_vesom_date = 0
        total_al_date = 0
        total_out_date = 0
        id_cl = []
        actual_work_time =0
        total_shift_work_time=0
                                                
        is_update_cl=False
        total_compensation_leave=0
        # actual_work_time += time_late_early_for_work
        # total_late_early_for_work += time_late_early_for_work
        # shift_name = f"{int(actual_work_time)}_{shift_name}"
        #############   start update hr
        
        calculate_actual_work_time = actual_work_time
        # for index, leave_item in hr_leave_employee_date.iterrows():
            
        #     # print("--------------------",
        #     #     leave_item['holiday_status_id'])
        #     try:
        #         if ('có tính lương' in leave_item['holiday_status_id'].strip().lower()):
        #             print("aaaaaa'có tính lương'aaaaaaa")
                    
        #     except Exception as ex:
        #         print(ex)

            # try:
            #     if (('đi muộn' in leave_item['holiday_status_id'].strip().lower()) or ('về sớm' in leave_item['holiday_status_id'].strip().lower())) and \
            #             ((leave_item['for_reasons']=='1') or (leave_item['for_reasons']==1)) and self.is_calculate_late_early_leave:
            #         print("don di muon ve som duoc duyet voi ly do ca nhan trong ngay")
            #         total_dimuon_vesom_date = total_dimuon_vesom_date + \
            #             max(leave_item['minutes'], leave_item['time'])
                
            # except Exception as ex:
            #     print(ex)
        # print(isinstance(row['holiday_status_name']))
        if (hasattr(row['holiday_status_name'], "__len__")):
            index = 0
            # print(row['holiday_status_name'])
            for leave_item in row['holiday_status_name']:
                if ('nghỉ bù' in leave_item.strip().lower()):
                    # print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                    total_compensation_leave_date = max(row['minutes'][index], row['time'][index])
                    id_cl.append(int(row['id_leave'][index]))
                    if is_update_cl:
                        update_minute = calculate_actual_work_time - total_out_date + total_compensation_leave_date + total_dimuon_vesom_date - total_shift_work_time
                        # print("update_minute : ", update_minute)
                        if (update_minute > 0) and (((update_minute<50)  and (update_minute != total_compensation_leave_date)) \
                                                    or (total_dimuon_vesom_date >0)) and (total_compensation_leave_date > update_minute):
                            update_data = {"minutes": total_compensation_leave_date - update_minute,
                                        "old_minutes": total_compensation_leave_date}
                            ids = [int(row['id_leave'][index])]
                            # try:
                            #     # self.models.execute_kw(self.db, self.uid, self.password, 'hr.leave', 'write', [ids, update_data])
                            #     self.env['hr.leave'].browse(ids).write(update_data)
                            #     total_compensation_leave = total_compensation_leave - update_minute
                            # except Exception as ex:
                            #     pass
                            #     # print(f"hr.leave write error: {ex}")
                            
                            # print ('success update minute')

                if ('nghỉ phép năm' in leave_item.strip().lower()):
                    total_al_date += max(row['minutes'][index], row['time'][index])
                    # print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                elif ( 'có tính lương' in leave_item.strip().lower()):
                    total_ncl_date = max(row['minutes'][index], row['time'][index])
                elif ('tăng ca' in leave_item.strip().lower()):
                    total_tc_date = max(row['minutes'][index], row['time'][index])
                index = index + 1 
  
        return total_out_date, total_tc_date , total_ncl_date , total_compensation_leave_date, total_dimuon_vesom_date , total_al_date, total_compensation_leave, id_cl

