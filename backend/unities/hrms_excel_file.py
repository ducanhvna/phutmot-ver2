# Module-level functions for direct import in other files

def process_report_raw(row):
    """Module-level wrapper for HrmExcelFile.process_report_raw."""
    return HrmExcelFile.process_report_raw(HrmExcelFile, row)

def find_couple(row, list_out_leave_for_work=None):
    """Module-level wrapper for HrmExcelFile.find_couple."""
    if list_out_leave_for_work is None:
        list_out_leave_for_work = []
    return HrmExcelFile.find_couple(HrmExcelFile, row, list_out_leave_for_work)

def find_couple_out_in(row):
    """Module-level wrapper for HrmExcelFile.find_couple_out_in."""
    return HrmExcelFile.find_couple_out_in(HrmExcelFile, row)
import http
import sys
import requests
import json
import datetime
import shutil
import calendar
# from datetime import datetime
import openpyxl
import os
# from dotenv import load_dotenv
# load_dotenv()
from openpyxl.utils import get_column_letter
import re
import pandas as pd
import numpy as np
from collections import Counter
import xmlrpc.client
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors
from unities.unities import float_to_hours, init_media_subfoder_report
from unities.reportfile import REPORT_LINK
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import logging
from io import BytesIO
import urllib
import mimetypes
url_summary_ver_ho = 'https://dl.dropboxusercontent.com/scl/fi/n5ps7uufq1to8zguqvzoh/BANG-CHAM-CONG.xlsx?rlkey=m9e683fnv2xaphgiu9qg0qorz&dl=0'
url_summary = 'https://dl.dropboxusercontent.com/scl/fi/stx1auw6wou3rnqkrjed2/Summary-time-attendant-OK.xlsx?rlkey=k2n6ayrp8r0rvk66rruzu032x'
url_attendace_count = 'https://dl.dropboxusercontent.com/scl/fi/q7dluz004pvcrcpijpetk/Ch-m-c-ng-m-s-l-n.xlsx?rlkey=zbp0nd22gcu3ipo0tczgrxnn1'
url_al_cl_tracking = 'https://dl.dropboxusercontent.com/scl/fi/3iyepiaqbwy0uoj2ogq9h/Theo-doi-ngay-phep-bu.xlsx?rlkey=i88y765pmvsns74p09r6abxt2'
url_al_cl_tracking_ho = 'https://dl.dropboxusercontent.com/scl/fi/7cdru32jew0pigfvk8tsh/M-U-Theo-doi-ngay-phep-bu.xlsx?rlkey=8x9ko6byqogvmmjzyrl9b6wyp'
url_night_split_shift = 'https://dl.dropboxusercontent.com/scl/fi/esn5favz2lr6omja6hyl9/ca-m-chi-ti-t-ca-g-y.xlsx?rlkey=8n9id7p01f7i8tnnfkkh0wp76'
url_late_5_minute = 'https://dl.dropboxusercontent.com/scl/fi/vzkt5xihgjeuenqhd1cgq/B-o-c-o-i-mu-n-d-i-5p-v-tr-n-5p.xlsx?rlkey=yhaksz20oo3zjju1hm0fk1u37'
url_feed_report = 'https://dl.dropboxusercontent.com/scl/fi/bnjpdw8nd2jsjrtowji0g/B-o-c-o-t-ng-h-p-n-ca.xlsx?rlkey=0v2ctgffrd0xgt4s73nvxdx13&dl=0'
url_kpi_weekly_report = 'https://dl.dropboxusercontent.com/scl/fi/hcx60iopldjc34u5s1iq2/Danh-s-ch-c-u-h-nh-Kpi-1.xlsx?rlkey=656mlnefq5m4x68ecunilyszf&dl=0'

# logging.basicConfig(filename='my_log.txt', format='%(asctime)s - %(levelname)s - %(name)s - %(message)s')
# logger = logging.getLogger('my_program')
# logger.info('This is an info message')
logging.basicConfig(filename='my_log.txt',
                    filemode='a',
                    format='%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S',
                    level=logging.DEBUG)

logging.info("Running HRMS")
logger = logging.getLogger('HrmExcelFile')

url = 'https://hrm-test.mandalahotel.com.vn'
db = 'apechrm_product_v3'
username = 'admin'
password = 'admin'


class HrmExcelFile():
    def __init__(self, start_date_str, end_date_str, input_folder,
                server_url=None, server_db=None, server_username=None, server_password=None,
                ouput_report_folder='', is_load_df_old = False, is_load_df_scheduling = True, is_calculate_late_early_leave = False,
                is_post_inspection = False, minioClient=None,bucket_name=None, public_url=None,
                is_update_al_cl = False,
                update_cl = {}, create_cl = {}, update_al = {}, create_al = {}):
        super(HrmExcelFile, self).__init__()
        logger.info("Init HrmExcelFile class")
        if server_url:
            self.url = server_url
        else:
            self.url = url
        if server_db:
            self.db = server_db
        else:
            self.db = db
        if server_username:
            self.username = server_username
        else:
            self.username = username
        if server_password:
            self.password = server_password
        else:
            self.password = password
        if minioClient:
            self.minioClient = minioClient
        if bucket_name:
            self.bucket_name = bucket_name
        if public_url: 
            self.public_url = public_url
        self.al_cl_current_month = {}
        self.cl_result_create = create_cl
        self.cl_result_update = update_cl
        self.al_result_create = create_al
        self.al_result_update = update_al
        self.is_update_al_cl = is_update_al_cl
        self.is_post_inspection = is_post_inspection
        self.is_calculate_late_early_leave = is_calculate_late_early_leave
        self.is_load_df_old = is_load_df_old
        self.is_load_df_scheduling = is_load_df_scheduling
        self.progress_callback = None
        self.input_folder = input_folder
        self.date_array = pd.date_range(start=start_date_str, end=end_date_str)
        
        self.output_report_folder = ouput_report_folder
        # company user profile
        
        common = xmlrpc.client.ServerProxy(
            '{}/xmlrpc/2/common'.format(self.url))
        self.models = xmlrpc.client.ServerProxy(
            '{}/xmlrpc/2/object'.format(self.url))
        self.uid = common.authenticate(
            self.db, self.username, self.password, {})
        self.user_profile = self.models.execute_kw(self.db, self.uid, self.password, 'res.users', 'read', [[self.uid] ], {'fields': ['id', 'name', 'company_id', 'company_ids']})[0]
        self.company_id = self.user_profile['company_id'][0]
        self.company_ids = self.user_profile['company_ids']
        self.company_info  = self.models.execute_kw(self.db, self.uid, self.password, 'res.company', 'read', [[self.company_id]], \
                                                          {'fields': ['id','name','is_ho']})[0]
        self.is_ho = self.company_info ['is_ho']
        # end company user profile
        self.load_employee_from_file()
        self.load_data_from_excel()
        self.working_time_48, self.working_time_44, self.working_time_40 = self.calculate_stardard_working_time()
        
    def load_employee_from_file(self):
        self.df_employees = pd.read_excel(os.path.join(self.input_folder, 'employees.xlsx'), engine='openpyxl', index_col=0) 
    
    def load_data_from_excel(self):
        print("Load data from excel")
        # df_contract\
        self.hr_weekly_report = pd.read_excel(os.path.join(self.input_folder, "hr_weekly_report.xlsx"), engine='openpyxl', index_col=0) 
        self.df_kpi_weekly_report = pd.read_excel(os.path.join(self.input_folder, "df_kpi_weekly_report.xlsx"), engine='openpyxl', index_col=0) 
        
        self.df_contract = pd.read_excel(os.path.join(self.input_folder, 'df_contract.xlsx'), engine='openpyxl', index_col=0) 
        if self.is_load_df_scheduling:
            self.df_scheduling_ver = pd.read_excel(os.path.join(self.input_folder, 'attendance_report.xlsx'), engine='openpyxl', index_col=0) 

        # read self.df_hr_leave
        self.df_hr_leave = pd.read_excel(os.path.join(self.input_folder, 'hr_leave.xlsx'), engine='openpyxl', index_col=0) 
        # print explanation data
        self.df_explanation_data = pd.read_excel(os.path.join(self.input_folder, 'explanation_data.xlsx'), engine='openpyxl', index_col=0) 
        # 
        self.df_al_report = pd.read_excel(os.path.join(self.input_folder, 'al_report.xlsx'), engine='openpyxl', index_col=0) 
        self.df_cl_report = pd.read_excel(os.path.join(self.input_folder, 'cl_report.xlsx'), engine='openpyxl', index_col=0) 
        self.df_al_report_severance = pd.read_excel(os.path.join(self.input_folder, 'al_report_severance.xlsx'), engine='openpyxl', index_col=0) 
        self.df_cl_report_severance = pd.read_excel(os.path.join(self.input_folder, 'cl_report_severance.xlsx'), engine='openpyxl', index_col=0) 
        self.df_all_company= pd.read_excel(os.path.join(self.input_folder, 'company.xlsx'), engine='openpyxl', index_col=0) 
        self.df_shift = pd.read_excel(os.path.join(self.input_folder, 'self.df_shift.xlsx'), engine='openpyxl', index_col=0)
        if self.is_load_df_old:
            print("Load df_old")
            self.df_old = pd.read_excel(os.path.join(self.input_folder, 'old.xlsx'), engine='openpyxl', index_col=0)
        print("load resouce df_resource_calendar_leaves")
        self.df_resource_calendar_leaves= pd.read_excel(os.path.join(self.input_folder,  "resource_calendar_leaves.xlsx"), engine='openpyxl', index_col=0)
        self.df_upload_report= pd.read_excel(os.path.join(self.input_folder,  "df_upload_report.xlsx"), engine='openpyxl', index_col=0)

        self.df_scheduling_ver[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                self.df_scheduling_ver.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
    def update_status_report(self):
        for index, row in self.df_upload_report.iterrows():
            if (row['month'] != self.date_array[0].month) or (row['year'] != self.date_array[0].year) or (f"{row['template']}".strip()=='1') or \
                    (f"{row['template']}".strip()=='3') or (f"{row['template']}".strip()=='4') or (f"{row['template']}".strip()=='7') \
                    or (f"{row['template']}".strip()=='2') or (f"{row['template']}".strip()=='5'):
                continue
            try:
                update_data = {
                        'template': f"{row['template']}".strip(),
                        'month': row['month'], 'year': row['year'], 'company_id':  row['id'],
                        'url': f"{row['url']}".replace('www.dropbox','dl.dropboxusercontent'), 'name': f'{row["type"].strip()} - {row["month"]}-{row["year"]}'
                    }
                if pd.isnull(row['id_x']):
                    
                    id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'create',
                                                            [update_data])
                    print('create', update_data)
                else:
                    ids = [int(row['id_x'])]
                    self.models.execute_kw(self.db, self.uid, self.password,
                                    'hr.upload.attendance', 'write', [ids, update_data])
                    print('update', update_data)
            except Exception as ex:
                logger.error(f"hr.upload.attendance write error: {ex}")
    def load_upload_report_data(self, df_minio_link_report, is_next_month = False):
        month = self.date_array[0].month
        year = self.date_array[0].year
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'search', [[('department_id',"=",False),
                                                                                                           ("month", "=", month),
                                                                                                           ("year", "=", year)]], {})

        list_upload_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'read', [ids],
                                                    {'fields': ['id','year', 'month', 'template', 'company_id', 'department_id']})
        for upload_report in list_upload_report:
            try:
                upload_report['department_name'] = upload_report['department_id'][1]
                upload_report['department_id'] = upload_report['department_id'][0]
            except:
                upload_report['department_name'] = ''
                upload_report['department_id'] = False
                
            try:
                company_id =  upload_report['company_id'][0]

                company_name = upload_report['company_id'][1]
                upload_report['company_name'] = company_name
                upload_report['company_id'] = company_id

            except:
                upload_report['company_name'] = False
        if len(list_upload_report) >0:
            
            df_upload_report = pd.DataFrame.from_dict(list_upload_report)
        else:
            df_upload_report = pd.DataFrame(columns =  ['id','year', 'month', 'template', 'company_id', 'company_name', 'department_name' ,'department_id']) 
        # df_upload_report.to_excel('df_upload_report_test.xlsx')
        
        df_minio_link_report.to_excel('df_minio_link_report_company.xlsx')

        # print(df_upload_report)
        df_minio_link_report = df_minio_link_report[df_minio_link_report['company_id'].notnull()] \
                        .merge(df_upload_report, on=['company_name', 'template'], \
                        how='left', suffixes=('','_upload'))
        # print(df_minio_link_report)
        
        for index, row in df_minio_link_report.iterrows():
            update_data = {
                    'template': f"{row['template']}".strip(),
                    'month': row['month'], 'year': row['year'], 'company_id':  row['company_id'], 'department_id': False,
                    'url': f"{row['url']}".replace('www.dropbox','dl.dropboxusercontent'), 'name': f'{row["type"].strip()} - {row["month"]}-{row["year"]}'
                }
           
            if not is_next_month:
                if (row['month'] != self.date_array[0].month) or (row['year'] != self.date_array[0].year):
                    continue
            
            is_update = False
            # if 'id_upload' in df_minio_link_report.columns:
            if pd.isnull(row['id']):
                try:
                    print('start create', update_data)
                    id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'create',
                                                            [update_data])
                    print('create', id_deparment)
                except Exception as ex:
                    print(ex)
            else:
                try:
                    print(int(row['id']))
                    ids = [int(row['id'])]
                    self.models.execute_kw(self.db, self.uid, self.password,
                                    'hr.upload.attendance', 'write', [ids, update_data])
                    print('update', update_data)
                except Exception as ex:
                    print(ex)

    def load_upload_report_data_department(self, df_minio_link_report):
        month = self.date_array[0].month
        year = self.date_array[0].year
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'search', [["&",
                                                                                                           ("month", "=", month),
                                                                                                           ("year", "=", year)]], {})

        list_upload_report  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'read', [ids],
                                                    {'fields': ['id','year', 'month', 'template', 'company_id', 'department_id']})
        for upload_report in list_upload_report:
            try:
                upload_report['department_name'] = upload_report['department_id'][1]
                upload_report['department_id'] = upload_report['department_id'][0]
            except:
                upload_report['department_name'] = ''
                upload_report['department_id'] = False
                
            try:
                company_id =  upload_report['company_id'][0]

                company_name = upload_report['company_id'][1]
                upload_report['company_name'] = company_name
                upload_report['company_id'] = company_id

            except:
                upload_report['company_name'] = False
        if len(list_upload_report) >0:
            
            df_upload_report = pd.DataFrame.from_dict(list_upload_report)
        else:
            df_upload_report = pd.DataFrame(columns =  ['id','year', 'month', 'template', 'company_id', 'company_name', 'department_name' ,'department_id']) 
        # df_upload_report.to_excel('df_upload_report_test.xlsx')
        # department_ids = self.models.execute_kw(
        #     self.db, self.uid, self.password, 'hr.department', 'search', [[('active', '=', True)]])
        # list_departments = self.models.execute_kw(self.db, self.uid, self.password, 'hr.department', 'read', [department_ids], {
        #                                           'fields': ['id', 'name', 'total_employee', 'company_id', 'member_ids', 'time_keeping_count']})
        # for department in list_departments:
        #     department['company_name'] = ''
        #     try:
        #         department['company_name'] =  department['company_id'][1]
        #         department['company_id'] =  department['company_id'][0]
        #     except Exception as ex:
        #         print(ex)
        #         continue
            
        # df_departments = pd.DataFrame.from_dict(list_departments)
        # print('df_departments', df_departments)
        # print('df_minio_link_report', df_minio_link_report)
        # df_minio_link_report = df_minio_link_report.merge(df_departments, left_on=['company_name', 'department_id'], \
        #                 right_on=['company_name', 'id'], \
        #                 how='left', suffixes=('','_department'))
        # df_minio_link_report.to_excel('df_minio_link_report_step1.xlsx')
        # if not('id' in df_minio_link_report.columns):
        #     print('Cant not join department')
        #     return

        # print(df_upload_report)
        df_minio_link_report = df_minio_link_report[df_minio_link_report['company_name'].notnull()] \
                        .merge(df_upload_report, on=['company_name', 'department_id', 'template'], \
                        how='left', suffixes=('','_upload'))
        # print(df_minio_link_report)
        
        for index, row in df_minio_link_report.iterrows():
            
            update_data = {
                    'template': f"{row['template']}".strip(),
                    'month': row['month'], 'year': row['year'], 'company_id':  row['company_id'], 'department_id': int(row['department_id']),
                    'url': f"{row['url']}".replace('www.dropbox','dl.dropboxusercontent'), 'name': f'{row["type"].strip()} - {row["month"]}-{row["year"]}'
                }
            if (row['month'] != self.date_array[0].month) or (row['year'] != self.date_array[0].year):
                continue
            
            is_update = False
            # if 'id_upload' in df_minio_link_report.columns:
            if pd.isnull(row['id']):
                try:
                    print('start create', update_data)
                    id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.upload.attendance', 'create',
                                                            [update_data])
                    print('create', id_deparment)
                except Exception as ex:
                    print('create report err: ', ex)
            else:
                update_data = {'url': f"{row['url']}".replace('www.dropbox','dl.dropboxusercontent')}
                ids = [int(row['id'])]
                self.models.execute_kw(self.db, self.uid, self.password,
                                'hr.upload.attendance', 'write', [ids, update_data])
                print('update', update_data)
        
    def calculate_stardard_working_time(self):
        number_sun = len([1 for i in calendar.monthcalendar(self.date_array[0].year,
                                                            self.date_array[0].month) if i[6] != 0])
        number_sat = len([1 for i in calendar.monthcalendar(self.date_array[0].year,
                                                            self.date_array[0].month) if i[5] != 0])
        number_day = len(self.date_array)
        print('number_sun', number_sun)
        print('number_sat', number_sat)
        return (number_day - number_sun), (number_day-number_sun - number_sat * 0.5), number_day-number_sun - number_sat
    def update_hr_leave(self):
        index = 0
        output_report_folder = ''
        for group_index, group_data in self.df_scheduling_ver.groupby('company'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            found = False
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                print(mis_id)
                for item in REPORT_LINK:
                    if item['mis-id'] == mis_id:
                        output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}") 
                        found= True
                        break
                print('out')
            else:
                print(index)
                output_report_folder = os.path.join(self.output_report_folder, f"{index}")
            try:
                df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index) &
                                               (self.df_hr_leave['state']=='validate')]
                df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                print('out: ', output_report_folder)
                isExist = os.path.exists(output_report_folder)
                if not isExist:

                # Create a new directory because it does not exist
                    os.makedirs(output_report_folder)
                self.update_hr_leave_company(df_hr_leave, 
                                                        output_report_folder)
            except Exception as ex:
                print(ex)
            index = index + 1
    def update_hr_leave_company(self, df_hr_leave, output_report_folder):
        df_hr_leave['update_minute'] = df_hr_leave.apply(lambda row: self.check_leave_fit_work_time(row), axis=1)
        try:
            df_hr_leave[df_hr_leave['update_minute']>0].to_excel(os.path.join(output_report_folder,"log.xlsx"))
        except:
            print(df_hr_leave[df_hr_leave['update_minute']>0])
    def check_leave_fit_work_time(self, row):
        result = 0
        if (row['total_shift_work_time']>0) and (not ('/' in row['shift_name'])) and \
                (('bù' in row['holiday_status_name']) or ('phép năm' in row['holiday_status_name'])):
            total_work_time =  row['total_work_time']
            if row['missing_checkin_break'] and (total_work_time >=240) :
                total_work_time =  240
            result = max(0, row['minutes'] + total_work_time - row['total_shift_work_time'] * 60)
        return result
    def load_workbook_from_url(self, url):
        try:
            file = urllib.request.urlopen(url).read()
            result =  load_workbook(filename = BytesIO(file))
        except (http.client.IncompleteRead) as e:
            path = 'temp_download_url.xlsx'
            response = requests.get(url, stream=True)
            with open(path, 'wb') as out_file:
                shutil.copyfileobj(response.raw, out_file)
                print('The file was saved successfully')
                
            result = load_workbook(path)
            # file = e.partial
        return result
        
    
    def export_al_cl_report_department(self, al_sheet_name='AL', cl_sheet_name='CL', process_callback=None):
        min_io_results=[]
        index = 0
        for group_index, df_al_report in self.df_al_report[(self.df_al_report['department_name'].notnull())&
                                                              (self.df_al_report['department_name']!=False)].groupby(['company_name', 'department_name']):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index[0]]
            if len(df_company_filter.index) > 0:
                department_id = df_al_report.iloc[0]['department_id']
                index = index +1
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                for item in REPORT_LINK:
                    if item['user'] == self.username :
                        continue
                    if item['mis-id'] == mis_id:
                        # try:
                        department_str_array =  [item_dp[0].upper() for item_dp in group_index[1].strip().split(' ')]
                        department_str = ''
                        for item_dp in department_str_array:
                            if item_dp != '/':
                                department_str =department_str + item_dp
                        department_str = f"{mis_id}-{department_str}-{index}"
                        output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m"), department_str) 
    
                        if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                            os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:

                        # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                        # print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                        # print('al_result', al_result)
                        # try:
                        df_cl_report = self.df_cl_report[(self.df_cl_report['company_name']==group_index[0])&
                                                            (self.df_cl_report['department_name']==group_index[1])] 
                        max_calulate_leave = df_cl_report['date_calculate'].max()
                        df_cl_report =  df_cl_report[df_cl_report['date_calculate']==max_calulate_leave]
                        # print('df_cl_report: ', df_cl_report)
                        # print("AL max: ", al_result[('date_calculate_leave','max')][group_index])
                        
                        df_al_report = self.df_al_report[(self.df_al_report['company_name']==group_index )]
                                                        
                        max_calulate_leave =  df_al_report['date_calculate_leave'].max()
                        df_al_report = df_al_report[df_al_report['date_calculate_leave']==max_calulate_leave]
                        
                        # df_cl_report.to_excel(os.path.join(output_report_folder, 'df_cl_report.xlsx'))
                        # df_al_report.to_excel(os.path.join(output_report_folder, 'df_al_report.xlsx'))
                        # print(output_report_folder)
                        # print(df_cl_report)
                        
                        
                        al_sheet_name='AL' 
                        cl_sheet_name='CL'
                        # continue
                        print("export quan ly phep bu")
                        # ref_workbook = openpyxl.load_workbook(
                        #     'unity/template/THEO_DOI_PHEP_BU.xlsx')
                        
                        ref_workbook = self.load_workbook_from_url(url_al_cl_tracking)
                        
                        print(ref_workbook.sheetnames)
                        index = 0
                        cl_index = 0
                        al_index = 0
                        for sheet_name in ref_workbook.sheetnames:
                            print(sheet_name)
                            if al_sheet_name == sheet_name:
                                print('aaaaa', al_sheet_name)
                                al_sheet_name = sheet_name
                                al_index = index
                            if cl_sheet_name == sheet_name:
                                print('bbbbb', cl_sheet_name)
                                cl_sheet_name = sheet_name
                                cl_index = index
                            index = index + 1
                        ws_al = ref_workbook.worksheets[al_index]
                        ws_cl = ref_workbook.worksheets[cl_index]

                        self.export_al_cl_report_company(df_al_report = df_al_report, 
                                                                df_cl_report = df_cl_report, 
                                                                ws_al=ws_al, ws_cl=ws_cl)
                        

                        new_file_name = "Theo doi ngay phep bu.xlsx"
                        old_file_path = os.path.join(
                            output_report_folder, new_file_name)
                        ref_workbook.save(old_file_path)
                        
                        year = self.date_array[0].year
                        month = self.date_array[0].month
                        
                        company_id = df_company_filter.iloc[0]['id']
                        try:
                            min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}/{department_str}',filename='al_cl', temp_file_path=old_file_path)
                            min_io_results.append( {'company_name': group_index[0], \
                                        'month' : self.date_array[0].month, \
                                        'year' : self.date_array[0].year, \
                                        'company_id': company_id, \
                                        'department_id': department_id, \
                                        'template': '5', \
                                        'department_name': group_index[1], 'url': min_io_result['url'], \
                                        'type': f"Báo cáo phep bu Cong ty {mis_id} phong ban {group_index[1]}"} )
                        except Exception as ex:
                            print(ex)
                        break
                        # except Exception as ex:
                        #     print(f"export_al_cl_report {mis_id} - {ex}")
            # except Exception as ex:
            #     print(ex)
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data_department(df_miniio)
            
    def export_attendence_scheduling_report_department(self):
        min_io_results=[]
        index = 0
        for group_index, df_scheduling_ver in self.df_scheduling_ver[(self.df_scheduling_ver['department'].notnull())&
                                                              (self.df_scheduling_ver['department']!=False)].groupby(['company', 'department']):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index[0]]
            found = False
            if len(df_company_filter.index) > 0:
                department_id = df_scheduling_ver.iloc[0]['d_id']
                index =index + 1
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                for item in REPORT_LINK:
                    if item['user'] == self.username :
                        continue
                    if item['mis-id'] == mis_id:
                        department_str_array =  [item_dp[0].upper() for item_dp in group_index[1].strip().split(' ')]
                        department_str = ''
                        for item_dp in department_str_array:
                            if item_dp != '/':
                                department_str =department_str + item_dp
                        department_str = f"{mis_id}-{department_str}-{index}"
                        output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m"), department_str) 
                        found= True
                        df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index[0] )]
                        # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                        df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index[0]]
                        
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:

                        # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                        print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index[0])
                        
                        
                        ref_workbook = self.load_workbook_from_url(url_attendace_count)
                        ws = ref_workbook.worksheets[0]
                        # source = ref_workbook.worksheets[1]
                        self.export_attendence_scheduling_report_company( df_scheduling_ver, df_hr_leave, ws)
                        print("Save file")
                        # ref_workbook.save(os.path.join(
                        #     output_report_folder, "CHAM_CONG_DEM_SO_LAN.xlsx"))
                        new_file_name = "CHAM_CONG_DEM_SO_LAN.xlsx"
                        old_file_path = os.path.join(
                            output_report_folder, new_file_name)
                        ref_workbook.save(old_file_path)
                        year = self.date_array[0].year
                        month = self.date_array[0].month
                        
                        company_id = df_company_filter.iloc[0]['id']
                        try:
                            min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}/{department_str}', filename='attendance_scheduling', temp_file_path=old_file_path)
                            min_io_results.append( {'company_name': group_index[0], \
                                            'month' : self.date_array[0].month, \
                                            'year' : self.date_array[0].year, \
                                            'company_id': company_id, \
                                            'template': '2', \
                                            'department_id': department_id, \
                                            'department_name': group_index[1], 'url': min_io_result['url'], \
                                            'type': f"Báo cáo chấm công đếm số lần Phong ban {mis_id} {group_index[1]}"} )
                        except Exception as ex:
                            print(ex)
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data_department(df_miniio)    

    def export_sumary_attendance_report_department(self):
        # if not self.is_post_inspection:
        #     self.update_status_report()
        is_post_inspection = self.is_post_inspection
        index = 0
        output_report_folder = ''
        self.df_scheduling_ver[['is_holiday', 'holiday_from', 'holiday_to', \
            'holiday_name', 'standard_working_time', \
                'minutes_per_day']] = self.df_scheduling_ver.apply(lambda row: self.merge_holiday_contract(row), result_type = 'expand', axis=1)
        # self.df_scheduling_ver.to_excel('lalaholiday.xlsx')
        min_io_results = []
        self.is_calculate_late_early_leave = True
        for group_index, group_data in self.df_scheduling_ver[(self.df_scheduling_ver['department'].notnull())&
                                                              (self.df_scheduling_ver['department']!=False)].groupby(['company', 'department']):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index[0]]
            found = False
            index = index +1
            if len(df_company_filter.index) > 0:
                department_id = group_data.iloc[0]['d_id']
                print(df_company_filter)
                auto_calculate_holiday = df_company_filter.iloc[0]['AUTO-CALCULATE-HOLIDAY-IN-SUMMARY']
                night_holiday_wage = df_company_filter.iloc[0]['NIGHT-HOLIDAY-WAGE']
                if pd.isnull(auto_calculate_holiday):
                    auto_calculate_holiday = False
                mis_id = df_company_filter.iloc[0]['mis_id']
                for item in REPORT_LINK:
                    if item['user'] == self.username :
                        continue
                    if item['mis-id'] == mis_id:
                    # if (item['mis-id'] == mis_id) and (mis_id =='MHS-BG'):
                        department_str_array =  [item_dp[0].upper() for item_dp in group_index[1].strip().split(' ')]
                        department_str = ''
                        for item_dp in department_str_array:
                            if item_dp != '/':
                                department_str = department_str + item_dp
                        department_str = f"{mis_id}-{department_str}-{index}"
                        output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m"), department_str) 
                        found= True
                
                        df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index[0]) & 
                                                        (self.df_hr_leave['state']=='validate')]
                        # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                        df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index[0]]
                        print('out: ', output_report_folder)
                        if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                            os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                        
                        if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                            os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:
                        # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                        df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index[0]] 
                        max_calulate_leave = df_cl_report['date_calculate'].max()
                        df_cl_report =  df_cl_report[df_cl_report['date_calculate']==max_calulate_leave]
                        # print('df_cl_report: ', df_cl_report)
                        # print("AL max: ", al_result[('date_calculate_leave','max')][group_index])
                        
                        df_al_report = self.df_al_report[(self.df_al_report['company_name']==group_index[0])]
                                                        
                        max_calulate_leave =  df_al_report['date_calculate_leave'].max()
                        df_al_report = df_al_report[df_al_report['date_calculate_leave']==max_calulate_leave]
                        print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index[0])
                        
                        print("Tong hop cham cong")
                        if not is_post_inspection:
                            ref_workbook = self.load_workbook_from_url(url_summary)
                            # ref_workbook = openpyxl.load_workbook(
                            #     'unity/template/Summary time attendant OK.xlsx')
                            
                            ws = ref_workbook.worksheets[0]
                        else:
                            ws = False
                        is_ho = True
                        
                        self.export_sumary_attendance_report_company(group_index[0], group_data, 
                                                                df_hr_leave,
                                                                df_explanation_data, df_cl_report, df_al_report, ws, is_post_inspection, is_ho, 
                                        auto_calculate_holiday, night_holiday_wage)
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:

                        # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                            print("The new output_report_folder is created! ", output_report_folder)
                        # ref_workbook.save(os.path.join(
                        #     output_report_folder, "Summary time attendant.xlsx"))
                        if (not is_post_inspection) and(department_id) and (department_id != False):
                            try:
                                new_file_name = "Summary time attendant.xlsx"
                                old_file_path = os.path.join(
                                    output_report_folder, new_file_name)
                                ref_workbook.save(old_file_path)
                                year = self.date_array[0].year
                                month = self.date_array[0].month
                                # min_io_result = self.save_excel_to_min_io( document_type='ouput_report', subfoder=f'{mis_id}/{year}-{month}/{department_str}',filename='summary_report', temp_file_path=old_file_path)
                                # min_io_results.append( {'company_name': group_index[0], \
                                #                     'month' : self.date_array[0].month, \
                                #                     'year' : self.date_array[0].year, \
                                #                     'template': '1', \
                                #                     'department_id': department_id, \
                                #                     'department_name': group_index[1], 'url': min_io_result['url'], \
                                #                     'type': f"Báo cáo tổng hợp chấm công phòng ban {group_index[1]}"} )
                            except Exception as ex:
                                logger.error(f"Export department: {ex}")
                                continue
                        # print(self.df_explanation_data)
                        # print(self.df_hr_leave)
                        # index = index + 1
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)
            try:
                df_miniio.to_excel('min_io_results.xlsx')
            except:
                print('Save file error')
            
            self.load_upload_report_data_department(df_miniio)
        
    def save_excel_to_min_io(self, document_type, subfoder, filename, temp_file_path):      
        # minioClient, bucket_name, public_url,
        object_name = f"hrm/{self.db}/{document_type}/{subfoder}/{filename}.xlsx"
        mime_type, encoding = mimetypes.guess_type(temp_file_path)
        try:
            self.minioClient.fput_object(
                self.bucket_name,
                object_name,
                temp_file_path,
                content_type=mime_type
            )
            url_excel = self.public_url + object_name
        except Exception as ex:
            print(ex)
        
        return {
                            "url": url_excel,
                            "name": f"{filename}.xlsx",
                        }
    
    def export_sumary_attendance_report(self, upload_to_min_io=False):
        # if not self.is_post_inspection:
            # self.update_status_report()
        is_post_inspection = self.is_post_inspection
        min_io_results = []
        output_report_folder = ''
        self.df_scheduling_ver[['is_holiday', 'holiday_from', 'holiday_to', \
            'holiday_name', 'standard_working_time', 'minutes_per_day']] = self.df_scheduling_ver.apply(lambda row: self.merge_holiday_contract(row), result_type = 'expand', axis=1)
        # self.df_scheduling_ver.to_excel('lalaholiday.xlsx')
        for group_index, group_data in self.df_scheduling_ver.groupby('company'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            found = False
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                company_id = df_company_filter.iloc[0]['id']
                auto_calculate_holiday = df_company_filter.iloc[0]['AUTO-CALCULATE-HOLIDAY-IN-SUMMARY']
                if pd.isnull(auto_calculate_holiday):
                    auto_calculate_holiday = False
                for item in REPORT_LINK:
                    if item['user'] == self.username :
                        continue
                    if item['mis-id'] == mis_id:
                        output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m")) 
                        found= True
                
                        df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index) & 
                                                        (self.df_hr_leave['state']=='validate')]
                        # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                        df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                        print('out: ', output_report_folder)
                        if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                            os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:
                        # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                        df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index] 
                        max_calulate_leave = df_cl_report['date_calculate'].max()
                        df_cl_report =  df_cl_report[df_cl_report['date_calculate']==max_calulate_leave]
                        # print('df_cl_report: ', df_cl_report)
                        # print("AL max: ", al_result[('date_calculate_leave','max')][group_index])
                        
                        df_al_report = self.df_al_report[(self.df_al_report['company_name']==group_index )]
                                                        
                        max_calulate_leave =  df_al_report['date_calculate_leave'].max()
                        df_al_report = df_al_report[df_al_report['date_calculate_leave']==max_calulate_leave]
                        print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                        media_report_folder = init_media_subfoder_report( f"{item['user']}", self.date_array[0].strftime("%Y-%m"))
                        print("Tong hop cham cong")
                        if not is_post_inspection:
                            ref_workbook = self.load_workbook_from_url(url_summary)
                            
                            ws = ref_workbook.worksheets[0]
                            try:
                                ws.cell(row=8, column=2).value = datetime.datetime.now()
                            except Exception as ex:
                                print(ex)
                        else:
                            ws = False
                        is_ho = False
                        df_scheduling_ver_copany = self.export_sumary_attendance_report_company(group_index, group_data, 
                                                                df_hr_leave,
                                                                df_explanation_data, df_cl_report, df_al_report, ws, is_post_inspection, is_ho, auto_calculate_holiday)
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:

                        # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                            print("The new output_report_folder is created! ", output_report_folder)
                        # ref_workbook.save(os.path.join(
                        #     output_report_folder, "Summary time attendant.xlsx"))
                        if not is_post_inspection:
                            new_file_name = "Summary time attendant.xlsx"
                            old_file_path = os.path.join(
                                output_report_folder, new_file_name)
                            ref_workbook.save(old_file_path)
                            # lod_file_path = os.path.join(
                            #     output_report_folder, 'df_scheduling_ver_copany.xlsx')
                            # df_scheduling_ver_copany.to_excel(lod_file_path)
                            
                            # service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
                            # print("Summary time attendantN url: ", service_url)
                            # # if upload_to_min_io == True:
                            # year = self.date_array[0].year
                            # month = self.date_array[0].month
                            # try:
                            #     min_io_result = self.save_excel_to_min_io(document_type='ouput_report', subfoder=f'{mis_id}/{year}-{month}',filename='summary_report', temp_file_path=old_file_path)
                            #     min_io_results.append( {'company_name': group_index, \
                            #                     'month' : self.date_array[0].month, \
                            #                     'year' : self.date_array[0].year, \
                            #                     'company_id': company_id, \
                            #                     'template': '1', \
                            #                     'department_name': '', 'url': min_io_result['url'], \
                            #                     'type': f"Báo cáo tổng hợp chấm công Cong ty {mis_id}"} )
                            # except Exception as ex:
                            #     print(ex)
                        # print(self.df_explanation_data)
                        # print(self.df_hr_leave)
                        # index = index + 1
        if len(min_io_results)>0:                
            df_miniio = pd.DataFrame.from_dict(min_io_results)
            try:
                df_miniio.to_excel('min_io_results.xlsx')
            except:
                print('Save file error')
            
            self.load_upload_report_data(df_miniio)
        
        # for index_cl_str, update_data in self.cl_result_update.items():
        #     try:
        #         ids = [int(index_cl_str)]
        #         self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'write', [ids, update_data])
        #         print ('update: ', update_data)
        #     except Exception as ex:
        #         logger.error(f'{self.uid}-write cl ex {index_cl_str} --- : {ex}')
        # update_datas = []
        # for employee_code, update_data in self.cl_result_create.items():
        #     update_datas.append(update_data)
        #     try:
        #         id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 
        #             'hr.cl.report', 'create', [update_data])
        #         print ('Create: ', update_data)
        #     except Exception as ex:
        #         logger.error(f'create {employee_code} cl ex --- : {ex}')   
        # try:
        #     id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 
        #         'hr.cl.report', 'create', [update_data])
        #     print ('Create: ', update_data)
        # except Exception as ex:
        #     logger.error(f'create cl ex --- : {ex}')   
    
        # for index_al_str, update_data in self.al_result_update.items():
        #     try:
        #         ids = [int(index_al_str)]
        #         self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'write', [ids, update_data])
        #         print ('update al: ', update_data)
        #     except Exception as ex:
        #         logger.error(f'write al ex {index_al_str} --- : {ex}')
        # update_datas = []
        # for employee_code, update_data in self.al_result_create.items():
        #     update_datas.append(update_data)
        #     try:
        #         id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 
        #             'hr.al.report', 'create', [update_data])
        #         print ('Create al: ', update_data)
        #     except Exception as ex:
        #         logger.error(f' {employee_code} create al ex --- : {ex}')
        # try:
        #     id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 
        #         'hr.al.report', 'create', [update_datas])
        #     print ('Create al: ', update_datas)
        # except Exception as ex:
        #     logger.error(f'create al ex --- : {ex}')   
    def export_sumary_attendance_report_ver2(self):
        min_io_results = []
        is_post_inspection = self.is_post_inspection
        index = 0
        output_report_folder = ''
        self.df_scheduling_ver[['is_holiday', 'holiday_from', 'holiday_to', \
            'holiday_name', 'standard_working_time', 'minutes_per_day']] = self.df_scheduling_ver.apply(lambda row: self.merge_holiday_contract(row), result_type = 'expand', axis=1)
        # self.df_scheduling_ver.to_excel('lalaholiday.xlsx')
        self.calculate_late_early_values()
        ref_workbook = self.load_workbook_from_url(url_summary_ver_ho)

        output_report_folder = os.path.join(self.output_report_folder, f"{self.username}", self.date_array[0].strftime("%Y-%m")) 
        media_report_folder = init_media_subfoder_report( f"{self.username}", self.date_array[0].strftime("%Y-%m"))
        list_sheet_name = []
        source = ref_workbook.active
        for group_index, group_data in self.df_scheduling_ver.groupby('company'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]

            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['id']
                print('company_ids: ', self.company_ids)
                if mis_id in self.company_ids:
                    # if item == mis_id:
                    if not is_post_inspection:
                        target = ref_workbook.copy_worksheet(source)
                        ws = target
                        # ws = ref_workbook.worksheets[len(ref_workbook.worksheets)-2]
                        if (pd.notnull( df_company_filter.iloc[0]['mis_id'])) and not (df_company_filter.iloc[0]['mis_id'] in list_sheet_name):
                            try:
                                ws.title = df_company_filter.iloc[0]['mis_id']
                                list_sheet_name.append(df_company_filter.iloc[0]['mis_id'])
                            except:
                                if len(group_index) > 20:
                                    if not (group_index[19:] in list_sheet_name):
                                        ws.title = group_index[19:]
                                        list_sheet_name.append(group_index[19:])
                    else:
                        ws = False

            
                    df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index) & 
                                                    (self.df_hr_leave['state']=='validate')]
                    # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                    df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                    print('out: ', output_report_folder)
                    if not os.path.exists(os.path.join(self.output_report_folder, f"{self.username}")):
                        os.makedirs(os.path.join(self.output_report_folder, f"{self.username}"))
                    isExist = os.path.exists(output_report_folder)
                    if not isExist:
                    # Create a new directory because it does not exist
                        os.makedirs(output_report_folder)
                    df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index] 
                    max_calulate_leave = df_cl_report['date_calculate'].max()
                    df_cl_report =  df_cl_report[df_cl_report['date_calculate']==max_calulate_leave]
                    # print('df_cl_report: ', df_cl_report)
                    # print("AL max: ", al_result[('date_calculate_leave','max')][group_index])
                    
                    df_al_report = self.df_al_report[(self.df_al_report['company_name']==group_index )]
                                                    
                    max_calulate_leave =  df_al_report['date_calculate_leave'].max()
                    df_al_report = df_al_report[df_al_report['date_calculate_leave']==max_calulate_leave]
                    print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                    
                    print("Tong hop cham cong")
                    
                    
                    is_ho = True
                    self.export_sumary_attendance_report_company_ver2(group_index, group_data, 
                                                            df_hr_leave,
                                                            df_explanation_data, df_cl_report, df_al_report, ws, is_post_inspection, is_ho)
        isExist = os.path.exists(output_report_folder)
        if not isExist:

        # Create a new directory because it does not exist
            os.makedirs(output_report_folder)
            print("The new output_report_folder is created! ", output_report_folder)
        # ref_workbook.save(os.path.join(
        #     output_report_folder, "Summary time attendant.xlsx"))
        if not is_post_inspection:
            new_file_name = "BANG_CHAM_CONG.xlsx"
            old_file_path = os.path.join(
                output_report_folder, new_file_name)
            ref_workbook.save(old_file_path)
            service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
            print("BANG CHAM CONG: ", service_url)
                            # if upload_to_min_io == True:
            year = self.date_array[0].year
            month = self.date_array[0].month
            
            is_save_file = (datetime.datetime.now() - self.date_array[0]).days < 50
            if is_save_file == False:
                return
            try:
                min_io_result = self.save_excel_to_min_io(document_type='ouput_report_ver2', subfoder=f'{mis_id}/{year}-{month}',filename='summary_report_ver_ho', temp_file_path=old_file_path)
                min_io_results.append( {'company_name': self.company_info['name'], \
                                'month' : self.date_array[0].month, \
                                'year' : self.date_array[0].year, \
                                'company_id': self.company_id, \
                                'template': '8', \
                                'department_name': '', 'url': min_io_result['url'], \
                                'type': f"Báo cáo tổng hợp HO VER2 {mis_id}"} )
            except Exception as ex:
                print(ex)
            # print(self.df_explanation_data)
            # print(self.df_hr_leave)
            # index = index + 1
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)
            self.load_upload_report_data(df_miniio)
        
    def export_late_in_5_miniutes_report_ho(self):
        min_io_results = []
        is_post_inspection = self.is_post_inspection
        index = 0
        output_report_folder = ''
        self.df_scheduling_ver[['is_holiday', 'holiday_from', 'holiday_to', \
            'holiday_name', 'standard_working_time', 'minutes_per_day']] = self.df_scheduling_ver.apply(lambda row: self.merge_holiday_contract(row), result_type = 'expand', axis=1)
        # self.df_scheduling_ver.to_excel('lalaholiday.xlsx')

        ref_workbook = self.load_workbook_from_url(url_late_5_minute)

        output_report_folder = os.path.join(self.output_report_folder, f"{self.username}", self.date_array[0].strftime("%Y-%m")) 
        media_report_folder = init_media_subfoder_report( f"{self.username}", self.date_array[0].strftime("%Y-%m"))
        list_sheet_name = []
        source = ref_workbook.active
        for group_index, group_data in self.df_scheduling_ver.groupby('company'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                if mis_id in self.company_ids:
                    # if item == mis_id:
                    if not is_post_inspection:
                        target = ref_workbook.copy_worksheet(source)
                        ws = target
                        # ws = ref_workbook.worksheets[len(ref_workbook.worksheets)-2]
                        if (pd.notnull( df_company_filter.iloc[0]['mis_id'])) and not (df_company_filter.iloc[0]['mis_id'] in list_sheet_name):
                            try:
                                ws.title = df_company_filter.iloc[0]['mis_id']
                                list_sheet_name.append(df_company_filter.iloc[0]['mis_id'])
                            except:
                                if len(group_index) > 20:
                                    if not (group_index[19:] in list_sheet_name):
                                        ws.title = group_index[19:]
                                        list_sheet_name.append(group_index[19:])
                    else:
                        ws = False
                        
                    df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index ) &
                                    (self.df_hr_leave['state']=='validate')]
                    # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                    df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                    print('out: ', output_report_folder)
                    # if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                    #     os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                    isExist = os.path.exists(output_report_folder)
                    if not isExist:

                    # Create a new directory because it does not exist
                        os.makedirs(output_report_folder)
                    print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                    # df_hr_leave = self.df_hr_leave[self.df_hr_leave['company_name']==group_index]
                    # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                    # df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index]
                    # df_al_report = self.df_al_report[self.df_al_report['company_name']==group_index]
                    df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                    # media_report_folder = init_media_subfoder_report( f"{item['user']}", self.date_array[0].strftime("%Y-%m"))

                    ref_workbook = self.load_workbook_from_url(url_late_5_minute)
                    ws = ref_workbook.worksheets[0]
                    self.export_late_in_5_miniutes_report_company( group_data, 
                                                            df_hr_leave, df_explanation_data, ws)
                    # ref_workbook.save(os.path.join(
#     output_report_folder, "BAO_CAO_DI_MUON_TREN_5P.xlsx"))
        isExist = os.path.exists(output_report_folder)
        if not isExist:

        # Create a new directory because it does not exist
            os.makedirs(output_report_folder)
            print("The new output_report_folder is created! ", output_report_folder)
        # ref_workbook.save(os.path.join(
        #     output_report_folder, "Summary time attendant.xlsx"))
        if not is_post_inspection:
            new_file_name = "BAO_CAO_DI_MUON_TREN_5P_ho.xlsx"
            # new_file_name = "Summary time attendant.xlsx"
            old_file_path = os.path.join(
                output_report_folder, new_file_name)
            ref_workbook.save(old_file_path)
            service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
            print("Summary time attendantN url: ", service_url)
                            # if upload_to_min_io == True:
            year = self.date_array[0].year
            month = self.date_array[0].month
            
            is_save_file = (datetime.datetime.now() - self.date_array[0]).days < 50
            if is_save_file == False:
                return
            try:
                min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}',filename='5m_late', temp_file_path=old_file_path)
                min_io_results.append( {'company_name': group_index, \
                                'month' : self.date_array[0].month, \
                                'year' : self.date_array[0].year, \
                                'company_id': self.company_id, \
                                'template': '3', \
                                'department_name': '', 'url': min_io_result['url'], \
                                'type': f"Báo cáo di muon 5ph Cong ty {mis_id}"} )
            except Exception as ex:
                print(ex)
            # break
                # except Exception as ex:
                #     print('5minute ex: ', ex)
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data(df_miniio)

    def export_sumary_attendance_report_ho(self):
        min_io_results = []
        is_post_inspection = self.is_post_inspection
        index = 0
        output_report_folder = ''
        self.df_scheduling_ver[['is_holiday', 'holiday_from', 'holiday_to', \
            'holiday_name', 'standard_working_time', 'minutes_per_day']] = self.df_scheduling_ver.apply(lambda row: self.merge_holiday_contract(row), result_type = 'expand', axis=1)
        # self.df_scheduling_ver.to_excel('lalaholiday.xlsx')

        ref_workbook = self.load_workbook_from_url(url_summary)

        output_report_folder = os.path.join(self.output_report_folder, f"{self.username}", self.date_array[0].strftime("%Y-%m")) 
        media_report_folder = init_media_subfoder_report( f"{self.username}", self.date_array[0].strftime("%Y-%m"))
        list_sheet_name = []
        source = ref_workbook.active
        for group_index, group_data in self.df_scheduling_ver.groupby('company'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]

            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['id']
                print('company_ids: ', self.company_ids)
                if mis_id in self.company_ids:
                    # if item == mis_id:
                    if not is_post_inspection:
                        target = ref_workbook.copy_worksheet(source)
                        ws = target
                        # ws = ref_workbook.worksheets[len(ref_workbook.worksheets)-2]
                        if (pd.notnull( df_company_filter.iloc[0]['mis_id'])) and not (df_company_filter.iloc[0]['mis_id'] in list_sheet_name):
                            try:
                                ws.title = df_company_filter.iloc[0]['mis_id']
                                list_sheet_name.append(df_company_filter.iloc[0]['mis_id'])
                            except:
                                if len(group_index) > 20:
                                    if not (group_index[19:] in list_sheet_name):
                                        ws.title = group_index[19:]
                                        list_sheet_name.append(group_index[19:])
                    else:
                        ws = False

            
                    df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index) & 
                                                    (self.df_hr_leave['state']=='validate')]
                    # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                    df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                    print('out: ', output_report_folder)
                    if not os.path.exists(os.path.join(self.output_report_folder, f"{self.username}")):
                        os.makedirs(os.path.join(self.output_report_folder, f"{self.username}"))
                    isExist = os.path.exists(output_report_folder)
                    if not isExist:
                    # Create a new directory because it does not exist
                        os.makedirs(output_report_folder)
                    df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index] 
                    max_calulate_leave = df_cl_report['date_calculate'].max()
                    df_cl_report =  df_cl_report[df_cl_report['date_calculate']==max_calulate_leave]
                    # print('df_cl_report: ', df_cl_report)
                    # print("AL max: ", al_result[('date_calculate_leave','max')][group_index])
                    
                    df_al_report = self.df_al_report[(self.df_al_report['company_name']==group_index )]
                                                    
                    max_calulate_leave =  df_al_report['date_calculate_leave'].max()
                    df_al_report = df_al_report[df_al_report['date_calculate_leave']==max_calulate_leave]
                    print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                    
                    print("Tong hop cham cong")
                    
                    
                    is_ho = True
                    self.export_sumary_attendance_report_company(group_index, group_data, 
                                                            df_hr_leave,
                                                            df_explanation_data, df_cl_report, df_al_report, ws, is_post_inspection, is_ho)
        isExist = os.path.exists(output_report_folder)
        if not isExist:

        # Create a new directory because it does not exist
            os.makedirs(output_report_folder)
            print("The new output_report_folder is created! ", output_report_folder)
        # ref_workbook.save(os.path.join(
        #     output_report_folder, "Summary time attendant.xlsx"))
        if not is_post_inspection:
            new_file_name = "Summary time attendant.xlsx"
            old_file_path = os.path.join(
                output_report_folder, new_file_name)
            ref_workbook.save(old_file_path)
            service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
            print("Summary time attendantN url: ", service_url)
                            # if upload_to_min_io == True:
            year = self.date_array[0].year
            month = self.date_array[0].month
            
            is_save_file = (datetime.datetime.now() - self.date_array[0]).days < 50
            if is_save_file == False:
                return
            try:
                min_io_result = self.save_excel_to_min_io(document_type='ouput_report', subfoder=f'{mis_id}/{year}-{month}',filename='summary_report', temp_file_path=old_file_path)
                min_io_results.append( {'company_name': self.company_info['name'], \
                                'month' : self.date_array[0].month, \
                                'year' : self.date_array[0].year, \
                                'company_id': self.company_id, \
                                'template': '1', \
                                'department_name': '', 'url': min_io_result['url'], \
                                'type': f"Báo cáo tổng hợp HO {mis_id}"} )
            except Exception as ex:
                print(ex)
            # print(self.df_explanation_data)
            # print(self.df_hr_leave)
            # index = index + 1
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)
            self.load_upload_report_data(df_miniio)
        

    def merge_holiday_contract(self, row):
        is_holiday = False
        holiday_from = None
        holiday_to = None
        holiday_name = ''
        standard_working_time = None
        shift_time = row['date']
        minutes_per_day = 480
        try:
            
            df_compare = self.df_resource_calendar_leaves[(self.df_resource_calendar_leaves['date_from'] <= row['shift_end_datetime']) &
                                                          (self.df_resource_calendar_leaves['date_to'] > row['shift_start_datetime'])]

            if len(df_compare.index) > 0:
                fist_row = df_compare.iloc[0]
                is_holiday = True
                holiday_from = fist_row['date_from']
                holiday_to = fist_row['date_to']
                holiday_name = fist_row['name']
            
        except Exception as ex:
            is_holiday = False

            print('merge error: ', ex)
        try:
            employee_code = str(row['employee_code'])
            if pd.notnull(row['id_employee']):
                df_contract = self.df_contract[(self.df_contract['date_start']<= shift_time) & 
                        (self.df_contract['employee_id'] == row['id_employee']) &
                        (self.df_contract['date_end'] >= shift_time)]
            else:
                df_contract = self.df_contract[(self.df_contract['date_start']<= shift_time) & 
                        (self.df_contract['employee_code'] == employee_code) &
                        (self.df_contract['date_end'] >= shift_time)]
                            
            if len(df_contract.index) > 0:
                resource_calendar_id = df_contract.iloc[0]['resource_calendar_id']
                minutes_per_day = df_contract.iloc[0]['minutes_per_day']
                if minutes_per_day == 480:
                    if ('44' in resource_calendar_id):
                        standard_working_time = self.working_time_44
                    elif '40' in resource_calendar_id:
                        standard_working_time = self.working_time_40
                    elif '48' in resource_calendar_id:
                        standard_working_time= self.working_time_48
                        
                elif minutes_per_day == 530:
                    if ('44' in resource_calendar_id):
                        standard_working_time = self.working_time_40
                    elif '40' in resource_calendar_id:
                        standard_working_time = self.working_time_40
                    elif '48' in resource_calendar_id:
                        standard_working_time= self.working_time_44
                # print("gio lam ex: ", ex)
        except Exception as ex:
            print('merge contract error: ', ex)
        
        return is_holiday, holiday_from, holiday_to, holiday_name, standard_working_time, minutes_per_day


    def calculate_actual_work_time_out_work_reason(self, scheduling_ver_row, hr_leave_ver_row):
        start_work_date_time = scheduling_ver_row['shift_start_datetime'].replace(second=0)
        end_work_date_time = scheduling_ver_row['shift_end_datetime'].replace(second=0)

        start_rest_date_time = scheduling_ver_row['rest_start_datetime'].replace(second=0)
        end_rest_date_time = scheduling_ver_row['rest_end_datetime'].replace(second=0)
         # start night datetime from 22h to 6h:00
        start_night_date_time = hr_leave_ver_row['attendance_missing_from'].replace(second=0)
        end_night_date_time = hr_leave_ver_row['attendance_missing_to'].replace(second=0)
        try:
            real_time_out = scheduling_ver_row['last_attendance_attempt'].replace(second=0)
        except:
            real_time_out = end_night_date_time
        try:
            real_time_in = scheduling_ver_row['attendance_attempt_1'].replace(second=0)
        except:
            real_time_out = start_night_date_time
            
        current_program = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time)
        # stage_fist = max(0, current_program.total_seconds()/60.0)
        return max(0, current_program)

    def calculate_night_holiday_work_time(self, row):
        """
        calculate night holiday work time with real time in, time out , restime in out
        cobine night holiday time
        """
        holiday_work_time = 0
        night_worktime_minute = 0
        # if not row['is_holiday']:
        #     return 0, holiday_work_time, 0
        # if row['employee_code']=='APG230320013'    :
        #     input("Press Enter to continue...")
        try:
            # rest_shifts = row['rest_shifts']
            total_work_time = row['total_work_time']
            # if rest_shifts or total_work_time == 0:
            #     return 0
            real_time_in = row['attendance_attempt_1'].replace(second=0)
            real_time_out = row['last_attendance_attempt'].replace(second=0)
            # if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
            #     return 0, holiday_work_time, 0
            start_work_date_time = real_time_out
            if pd.notnull(row['shift_start_datetime']):
                start_work_date_time = row['shift_start_datetime'].replace(second=0)
            end_work_date_time = real_time_out
            if pd.notnull(row['shift_end_datetime']):
                end_work_date_time = row['shift_end_datetime'].replace(second=0)
          

            # start night datetime from 22h to 6h:00
            start_night_date_time = start_work_date_time.replace(hour=22, minute=0, second=0)
            end_night_date_time = (start_night_date_time + datetime.timedelta(days=1)).replace(hour=6)
            try:
                start_holiday_datetime = row['holiday_from'].replace(hour=0, minute=0, second=0)
                end_holiday_datetime = row['holiday_to'].replace(hour=0, minute=0, second=0)
            except:
                start_holiday_datetime = None
                end_holiday_datetime = None

            start_rest_date_time = row['rest_start_datetime'].replace(second=0)
            end_rest_date_time = row['rest_end_datetime'].replace(second=0)
            # night stage fist
            current_program_night = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time)
            stage_fist_night = max(0, current_program_night.total_seconds()/60.0)
            if not (start_holiday_datetime is None):
                #night holiday
                current_program = min(real_time_out, start_rest_date_time, end_holiday_datetime, end_night_date_time) - \
                        max(real_time_in, start_work_date_time, start_night_date_time, start_holiday_datetime)
                stage_fist = max(0, current_program.total_seconds()/60.0)
                
                # holiday stage fist
                current_program_holiday = min(real_time_out, start_rest_date_time, end_holiday_datetime) - \
                        max(real_time_in, start_work_date_time, start_holiday_datetime)
                stage_fist_holiday = max(0, current_program_holiday.total_seconds()/60.0)
            else:
                stage_fist = 0
                stage_fist_holiday = 0

            # night stage second
            current_program_night = min(real_time_out, end_work_date_time, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time)
            stage_second_night = max(0, current_program_night.total_seconds()/60.0)
            if not (start_holiday_datetime is None):
                # night holiday
                current_program = min(real_time_out, end_work_date_time, end_holiday_datetime, end_night_date_time) - \
                        max(real_time_in, end_rest_date_time, start_night_date_time, start_holiday_datetime)
                stage_second = max(0, current_program.total_seconds()/60.0)
                
                # holiday stage second
                current_program_holiday = min(real_time_out, end_work_date_time, end_holiday_datetime) - \
                        max(real_time_in, end_rest_date_time, start_holiday_datetime)
                stage_second_holiday = max(0, current_program_holiday.total_seconds()/60.0)
            # print('{}: night after now - holiday: {} - {} - {} -night - {}- {} - real - {} - {} '.format(row['name_employee'] ,row['name'], \
            #                     stage_fist, stage_second, start_night_date_time, end_night_date_time, row['start_work_date_time'], row['end_work_date_time']))
            else:
                stage_second = 0
                stage_second_holiday = 0

            stage_night_only = int(stage_fist_night + stage_second_night)
            stage_night =  int(stage_fist + stage_second)
            holiday_work_time = int(stage_fist_holiday + stage_second_holiday)
            
            end_night_date_time = start_work_date_time.replace(hour=6, minute=0, second=0)
            start_night_date_time = (end_night_date_time+ datetime.timedelta(days=-1)).replace(hour=22, minute=0, second=0)
            
            # night stage 1
            current_program_night = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time)
            stage_fist_night = max(0, current_program_night.total_seconds()/60.0)
            if not (start_holiday_datetime is None):
                # night holiday stage 1
                current_program = min(real_time_out, start_rest_date_time, end_holiday_datetime, end_night_date_time) - \
                        max(real_time_in, start_work_date_time, start_night_date_time, start_holiday_datetime)
                stage_fist = max(0, current_program.total_seconds()/60.0)
            else:
                stage_fist = 0
            ##########################################################
            # night stage 2
            current_program_night = min(real_time_out, end_work_date_time, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time)
            stage_second_night = max(0, current_program_night.total_seconds()/60.0)
            if not (start_holiday_datetime is None):
                # night holiday stage 2
                current_program = min(real_time_out, end_work_date_time, end_holiday_datetime, end_night_date_time) - \
                        max(real_time_in, end_rest_date_time, start_night_date_time, start_holiday_datetime)
                stage_second = max(0, current_program.total_seconds()/60.0)
                # print('{}: night before now - holiday: {} - {} - {} -night - {}- {} - real - {} - {} '.format(row['name_employee'] ,row['name'], \
                #                     stage_fist, stage_second, start_night_date_time, end_night_date_time, row['start_work_date_time'], row['end_work_date_time']))
            else:
                stage_second = 0

            stage_morning = int(stage_fist + stage_second)
            stage_morning_only = int(stage_second_night + stage_fist_night)
            night_worktime_minute = int(stage_night_only + stage_morning_only)
            # if row['employee_code']=='APG230320013':    
            #     print('night_worktime_minute', night_worktime_minute)
            #     input("Press Enter to continue...")
            return min(total_work_time, stage_night + stage_morning), min(total_work_time, holiday_work_time), min(total_work_time, night_worktime_minute)

        except Exception as ex:
            print("calcuate night - holiday time err", ex)
            # print('{}: night - holiday: {} - {} - {} -night - {}- {} - real - {} - {}'.format(row['employee_name'] ,row['shift_name'], 0, 0, 0, 0, \
            #                                                                         row['start_work_date_time'], row['end_work_date_time']))
            return 0, min(total_work_time, holiday_work_time), min(total_work_time, night_worktime_minute)
    
    def export_sumary_attendance_report_company_ver2(self, company_name, df_scheduling_ver, 
                    df_hr_leave, 
                    df_explanation_data, df_cl_report, df_al_report, ws, 
                    is_post_inspection=False, is_ho = False, 
                    auto_calculate_holiday = False,
                    night_holiday_wage = 3.0):
        week_day_array = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
        start_col = 5
        start_row = 9
        
        for date_item in self.date_array:
            day_of_month = date_item.day
            ws.cell(row=start_row, column=start_col +
                    day_of_month).value = date_item
            ws.cell(row=start_row + 1, column=start_col +
                    day_of_month).value =week_day_array[date_item.weekday()]
        if date_item.day < 31:
            for add_item in range(date_item.day + 1, 32):
                ws.cell(row=start_row, column=start_col + add_item).value = ''
            ws.cell(row=start_row + 1, column=start_col +
                    day_of_month).value = ''                
        data_probationwage_rate_update = []
        data_probation_completion_wage_update = []
        data_actual_total_work_time_update = []
        data_standard_working_day_update = []
        data_contract_company_update = []
        data_total_work_time_update = []
        
        ids_total_work_time_update = {}
        ids_contract_company_update = {}
        ids_probationwage_rate_update = {}
        ids_probation_completion_wage_update = {}
        ids_actual_total_work_time_update = {}
        ids_standard_working_day_update = {}
        ids_all_off_remove = []
        month_array_str = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 
                                'october', 'november', 'december']
        current_month_str = month_array_str[self.date_array[0].month-1]
        try:
            df_scheduling_ver[['night_holiday_work_time_minute','holiday_work_time', 'night_worktime_minute']] = \
                    df_scheduling_ver.apply(lambda row: self.calculate_night_holiday_work_time(row),  axis=1, result_type='expand')
            print("sucesss")
        except Exception as ex:
            df_scheduling_ver['night_holiday_work_time_minute'] =0
            df_scheduling_ver['holiday_work_time'] = 0
            df_scheduling_ver['night_worktime_minute']= 0
            print("calculate night: ", ex) 
        start_row = 11
        for g, data in df_scheduling_ver[(df_scheduling_ver['department_name'].notnull())
                                              & (df_scheduling_ver['department_name'] != False)
                                              & (df_scheduling_ver['workingday'].notnull())
                                              # & (self.df_scheduling_ver['employee_code']=='APG230316014')
                                              & (df_scheduling_ver['workingday'] != False)].groupby('department_name'):
            # start_row = start_row + 1
            count = 0
            print(g)
            if not is_post_inspection:
                ws.cell(row=start_row, column=3).value = g
                ws.cell(row=start_row, column=17).value = ''
                ws.cell(row=start_row, column=31).value = ''
                ws.cell(row=start_row, column=32).value = ''
                # start_row = start_row + 1

            for sub_group_index, sub_group_data in data.groupby(['name', 'employee_code']):
                last_month_al_employee = df_al_report[df_al_report['employee_code']==sub_group_index[1]]
                last_month_cl_employee = df_cl_report[df_cl_report['employee_code']==sub_group_index[1]]
                
                total_attendance_late_less_than_5 = 0
                total_attendance_late_more_than_5 = 0
                total_actual_work_time_date=0
                if (sub_group_data.iloc[0]['severance_day'] !=False) and (pd.notnull (sub_group_data.iloc[0]['severance_day'])) and \
                        sub_group_data.iloc[0]['severance_day'] < self.date_array[0]:
                    continue
                
                # if sub_group_index[1] == 'APG220801006':
                    # input('APG220801006')
                    # print(sub_group_data.iloc[0]['night_holiday_work_time_minute'])
                    # print('total: ', total_night_holiday)
                list_off_ids = []
                ids_off_remove_array = []
                standard_working_time = sub_group_data['standard_working_time'].mean()
                # minutes_per_day = sub_group_data[sub_group_data['minutes_per_day']>0]['minutes_per_day'].mean()
                minutes_per_day = sub_group_data[sub_group_data['minutes_per_day']>0]['minutes_per_day'].max()
                total_ncl = 0
                total_ncl_probationary = 0
                total_off = 0

                total_ph = 0
                total_al = 0
                total_al_probationary = 0

                total_al_dimuon_vesom = 0

                total_dimuon_vesom = 0
                total_dimuon_vesom_probationary = 0

                total_tc = 0
                total_tc_probationary = 0
                phatsinhtang_thuviec = 0
                phatsinhtang_chinhthuc = 0
                total_tc_holiday = 0
                total_tc_holiday_probationary = 0
                total_normal_working_minutes = 0
                total_normal_working_minutes_probationary = 0

                total_ph_probationary = 0
                total_unpaid_leave = 0
                total_unpaid_leave_probationary = 0

                total_compensation_leave = 0
                total_compensation_probationary = 0

                total_compensation_leave_dimuon_vesom = 0
                total_compensation_dimuon_vesom_probationary = 0

                total_night = 0
                total_night_probationary = 0

                total_night_holiday = 0
                total_night_holiday_probationary = 0

                total_fix_rest_time = 0
                total_fix_rest_time_probationary = 0
                total_leave = 0
                minute_worked_day_holiday = 0
                minute_worked_day_holiday_probationary = 0
                total_minutes_working_reduced = 0
                total_late_early_for_work = 0
                if not is_post_inspection:
                    ws.cell(row=start_row, column=1).value = sub_group_index[1]
                    ws.cell(row=start_row, column=2).value = sub_group_index[0]
                col_index = 0
                total_out = 0
                total_out_no_explaination = 0
                

                hr_leave_employee = df_hr_leave[df_hr_leave['employee_code']
                                                     == sub_group_index[1]]

                for index, leave_item in hr_leave_employee.iterrows():
                    # if sub_group_index[1] == 'APG230610001':
                    #     print(leave_item)
                    #     leave_item['employee_code']
                    #     print(total_tc_holiday)
                    #     print(hr_leave_employee)
                    #     for index2, testitem in hr_leave_employee.iterrows():
                    #         try:
                    #             print(testitem['holiday_status_name'])
                    #         except:
                    #             print(testitem['holiday_status_name'][1])
                    #         if ('tăng ca' in testitem['holiday_status_name'].strip().lower()):
                    #             exit(1)
                    # if shift_name == 'CL':
                    #     total_compensation_leave = total_compensation_leave + 1
                    #     if probationary_index:
                    #         total_compensation_probationary = total_compensation_probationary + 1
                    # elif ('/CL' in shift_name) or ('CL/' in shift_name):
                    #     total_compensation_leave = total_compensation_leave + 0.5
                    #     if probationary_index:
                    #         total_compensation_probationary = total_compensation_probationary + 0.5

                    # if shift_name == 'AL':
                    #     total_al = total_al + 1
                    #     if probationary_index:
                    #         total_al_probationary = total_al_probationary + 1
                    # elif ('/AL' in shift_name) or ('AL/' in shift_name):
                    #     total_al = total_al + 0.5
                    #     if probationary_index:
                    #         total_al_probationary = total_al_probationary + 0.5
                    # if (leave_item['request_date_from'] <= self.date_array[len(self.date_array)-1]):
                    # Đơn xin nghỉ có tính lương
                    # print("--------------------",
                    #       leave_item['holiday_status_name'])
                    # print("--------------------",
                    #       leave_item['holiday_status_name'])
                    try:
                        if ('nghỉ bù' in leave_item['holiday_status_name'].lower()):
                            print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                            total_compensation_leave = total_compensation_leave + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_compensation_probationary = total_compensation_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)
                    try:
                        if ( 'nghỉ phép năm' in leave_item['holiday_status_name'].strip().lower()):
                            print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                            total_al = total_al + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_al_probationary = total_al_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)             
                    try:
                        if ( 'có tính lương' in leave_item['holiday_status_name'].strip().lower()):
                            print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                            total_ncl = total_ncl + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_ncl_probationary = total_ncl_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)
                    try:
                        if ('tăng ca' in leave_item['holiday_status_name'].strip().lower()):
                            print("don tang ca")
                            if leave_item['is_holiday']==True:
                                total_tc_holiday = total_tc_holiday + \
                                    max(leave_item['minutes'],
                                        leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                                if leave_item['is_leave_probationary']:
                                    total_tc_holiday_probationary = total_tc_holiday_probationary + \
                                        max(leave_item['minutes'],
                                            leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                            else:
                                total_tc = total_tc + \
                                    max(leave_item['minutes'],
                                        leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                                if leave_item['is_leave_probationary']:
                                    total_tc_probationary = total_tc_probationary + \
                                        max(leave_item['minutes'],
                                            leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                                        
                            if 'phát sinh tăng' in leave_item['reasons'].strip().lower():
                                if leave_item['is_leave_probationary']:
                                    phatsinhtang_thuviec = phatsinhtang_thuviec + \
                                        max(leave_item['minutes'],
                                            leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                                else:
                                    phatsinhtang_chinhthuc = phatsinhtang_chinhthuc + \
                                        max(leave_item['minutes'],
                                            leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                        
                            # if sub_group_index[1] == 'APG230610001':
                            #     print(leave_item)
                            #     print(total_tc_holiday)
                            #     print(real_total_work_time)
                            #     exit(1)
                    except Exception as ex:
                        print(ex)    
                            
                    try:    
                        if (('đi muộn' in leave_item['holiday_status_name'].strip().lower()) or ('về sớm' in leave_item['holiday_status_name'].strip().lower())) and \
                                ((leave_item['for_reasons']=='1') or (leave_item['for_reasons']==1)) and self.is_calculate_late_early_leave:
                            print("don di muon ve som duoc duyet voi ly do ca nhan")
                            total_dimuon_vesom = total_dimuon_vesom + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_dimuon_vesom_probationary = total_dimuon_vesom_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)   
                    try:                
                        if ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()) and \
                                    leave_item['for_reasons']=='1':
                            duration = leave_item['attendance_missing_to'] - leave_item['attendance_missing_from']
                            seconds = duration.total_seconds()
                            hours = seconds // 3600
                            minutes = (seconds % 3600) // 60.0

                            total_out = total_out + minutes + hours*60
                            print("don ra ngoài cá nhân")
                    except Exception as ex:
                        print(ex)
                
                for probationary_index, probationary_data in sub_group_data.sort_values(by='date').groupby('is_probationary'):   
                    for index, row in probationary_data.iterrows():
                        last_out_date= row['last_attendance_attempt']
                        out_date_array = []
                        hr_leave_employee_date = df_hr_leave[(df_hr_leave['employee_code']
                                                     == sub_group_index[1]) & \
                                                    (df_hr_leave['date_str'] == row['date_str']) ]
                        if pd.notnull(row['contract_company']):
                            if row['company'] != row['contract_company']:
                                try:
                                    # is_need_update= True
                                    if f"{row['contract_company']}" in data_contract_company_update:
                                        ids_contract_company_update[f"{row['contract_company']}"].append({"id":int(row['id']), "employee_code": row['employee_code'], 
                                                                                                        "employee_name": row['employee_name'], 
                                                                                                        'company': row["company"], 'contract_company': row["contract_company"]})
                                    else:
                                        data_contract_company_update.append(f"{row['contract_company']}")
                                        ids_contract_company_update[f"{row['contract_company']}"] = [{"id":int(row['id']), "employee_code": row['employee_code'], 
                                                                                                        'company': row["company"], 'contract_company': row["contract_company"]}]
                                except Exception as ex:
                                    print(ex)
                    # if (sub_group_index[1]=='APG220215002'):
                    #     hr_leave_employee_date.to_excel(os.path.join(output_report_folder, f'{ row["date_str"]}-APG220215002.xlsx'))
                        if len(hr_leave_employee_date.index) >0:
                            print("check date: ",  row['date'].strftime('%Y-%m-%d'))
                        time_late_early_for_work = 0
                        time_bussiness_trip = 0
                        is_update_total_worktime = False
                        real_total_work_time = round(row['total_work_time'],0)

                        if len(hr_leave_employee_date.index) >0:
                            for index, leave_item in hr_leave_employee_date.iterrows():
                                
                                # print("--------------------",
                                #     leave_item['holiday_status_name'])
                                
                                # try:
                                # if ('có tính lương' in leave_item['holiday_status_name'].strip().lower()):
                                #     print("aaaaaa'có tính lương'aaaaaaa")
                                if ('công tác' in leave_item['holiday_status_name'].strip().lower()):
                                    time_bussiness_trip = time_bussiness_trip + max(leave_item['minutes'], leave_item['time'])
                                if ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()):
                                    if pd.notnull(leave_item['request_date_from']):
                                        out_date_array.append(leave_item)
                                        if last_out_date > leave_item['request_date_from']:
                                            last_out_date = leave_item['request_date_from']
                                        if (leave_item['for_reasons']=='1'): 
                                            real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row)
                                            try:
                                                calculate_result_update = self.calculate_actual_work_time_couple(row, row[f'attendance_attempt_1'], leave_item['attendance_missing_to'])
                                                real_total_work_time = max(calculate_result_update['total_work_time'], real_total_work_time)
                                            except Exception as ex:
                                                pass
                                            calculate_result = self.calculate_actual_work_time_couple(row, leave_item['attendance_missing_from'], leave_item['attendance_missing_to'])
                                            
                                            print("time out date xfiiiiiiiiiiiiiiiiiistx", calculate_result)
                                            time_out_date = calculate_result['total_work_time']
                                            real_total_work_time = real_total_work_time - time_out_date
                                # elif (('đi muộn' in leave_item['holiday_status_name'].strip().lower()) or ('về sớm' in leave_item['holiday_status_name'].strip().lower())):
                                #     if row['employee_ho']== False:
                                #         real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row) 
                                #         if row['employee_code'] == 'APG220801006':
                                #             print(row)
                                #             print()
                                #             print(real_total_work_time)
                                #             # exit(1)
                                
                                #     if is_ho:
                                #         if(real_total_work_time != row['total_work_time']):
                                #             if int(real_total_work_time) in data_total_work_time_update:
                                #                 ids_total_work_time_update[int(real_total_work_time)].append(int(row['id']))
                                #             else:
                                #                 data_total_work_time_update.append(int(real_total_work_time))
                                #                 ids_total_work_time_update[int(real_total_work_time)] = [int(row['id'])]

                                #     if((leave_item['for_reasons']=='2') or (leave_item['for_reasons']==2)) and self.is_calculate_late_early_leave:
                                #         time_late_early_for_work += max(leave_item['minutes'], leave_item['time'])
                                # except Exception as ex:
                                #     print(ex)
                            
                        # print(f"{row}")
                        date = row['date']
                        shift_name = row['shift_name'] if row['shift_name'] else ''
                        resource_calendar_id = row['resource_calendar_id']
                        if pd.isnull(resource_calendar_id) or resource_calendar_id == False:
                            if pd.notnull(row['id_employee']):
                                contract_employee_filter = self.df_contract[(self.df_contract['date_start']<=date)&
                                        (self.df_contract['date_end']>=date)&
                                        (self.df_contract['employee_id']==row['id_employee'])].sort_values("date_end", ascending=False)
                            else:
                                contract_employee_filter = self.df_contract[(self.df_contract['date_start']<=date)&
                                        (self.df_contract['date_end']>=date)&
                                        (self.df_contract['employee_code']==sub_group_index[1])].sort_values("date_end", ascending=False)
                            if len(contract_employee_filter.index)>0:
                                resource_calendar_id = contract_employee_filter.iloc[0]['resource_calendar_id']
                            
                            
                        if row['fix_rest_time'] == False:
                            total_shift_work_time = round(row['total_shift_work_time'] * 60, -1)
                        else: 
                            total_shift_work_time = 480

                        # if sub_group_index[1] == 'APG220601007':
                        actual_work_time =  max (time_bussiness_trip, min (real_total_work_time + time_bussiness_trip, total_shift_work_time))

                        # else:    
                        #     actual_work_time =  min (row['total_work_time'] + row['kid_time'], total_shift_work_time)
                        # actual_work_time = actual_work_time - total_out
                        
                        
                        # if resource_calendar_id == False:
                        #     continue
                        # Find explan nation
                        # giai trinh di muon ve som ly do cong viec
                        # thieu cham cong va khong dung gio cho full luon
                        explanation_filter = df_explanation_data[
                            df_explanation_data['date_str'].isin([row['date_str']]) &
                            ((df_explanation_data['invalid_type']==3) | (df_explanation_data['invalid_type']=='3') |
                            (df_explanation_data['invalid_type']==4) | (df_explanation_data['invalid_type']=='4')) &
                            df_explanation_data['employee_code'].str.contains(sub_group_index[1]).fillna(False) &
                            ((df_explanation_data['reason']==2) | (df_explanation_data['reason']== '2')) &
                            ((df_explanation_data['validated']==2) | (df_explanation_data['validated'] == '2'))]
                        
                        if len(explanation_filter.index) > 0 and actual_work_time < row['total_shift_work_time']*60:
                            if ('/OFF' in row['shift_name']) or ('OFF/' in row['shift_name']):
                                actual_work_time = 240
                            elif row['fix_rest_time'] == True:
                                actual_work_time = 480
                            else:
                                actual_work_time = round(row['total_shift_work_time']*60, -1)

                            print('fix: ------------------- buy explanation', explanation_filter.iloc[0])

                        # if len(hr_leave_filter.index)>0 and actual_work_time < 480:
                        #     # try:
                        #     print('fix by cl: -------------------', hr_leave_filter)
                        #     if pd.notnull(hr_leave_filter.iloc[0]['minutes']):

                        #         minute = int(hr_leave_filter.iloc[0]['minutes'])
                        #         actual_work_time = min(row['total_shift_work_time']*60, actual_work_time + minute)
                        #     # except:
                        #     #     pass

                        if (row['missing_checkin_break'] == True) and (not ('/OFF' in shift_name)) and (not ('OFF/' in shift_name)):
                            actual_work_time = min(actual_work_time, 240)
                                # actual_work_time, row['total_shift_work_time']*30)
                            # print(f"Thieu cham cong giua {date}")
                            if not is_post_inspection:
                                redFill = PatternFill(
                                    patternType='solid', fgColor=colors.Color(rgb='00FDF000'))
                                ws.cell(row=start_row, column=start_col +
                                        date.day).fill = redFill
                        else:
                            # print(f"khong thieu cham cong {row['attendance_attempt_3']} - hop dong {resource_calendar_id}" )
                            if not is_post_inspection:
                                blueFill = PatternFill(
                                    patternType='solid', fgColor=colors.Color(rgb='0000FF00'))
                                ws.cell(row=start_row, column=start_col +
                                        date.day).fill = blueFill
                        if not is_post_inspection:
                            ft = Font(color="0000FF")
                            if shift_name == '-' or shift_name == 'OFF' or row['total_shift_work_time'] == 0:
                                ft = Font(color="0000FF")
                            elif (pd.isnull(row['attendance_attempt_1']) or row['attendance_attempt_1'] == False) \
                                    or (pd.isnull(row['last_attendance_attempt']) or row['last_attendance_attempt'] == False):
                                ft = Font(color="FF0000")
                            elif real_total_work_time < 480:
                                ft = Font(color="FC1000")
                            else:
                                ft = Font(color="0000FF")

                            ws.cell(row=start_row, column=start_col +
                                    date.day).font = ft
                        time_out_date = 0
                        total_tc_date =0 
                        total_ncl_date =0 
                        total_compensation_leave_date = 0 
                        number_cl_date = 0
                        total_dimuon_vesom_date = 0
                        total_dimuon_vesom_work_date = 0
                        total_al_date = 0 
                        number_al_date = 0
                        late_mid_time = 0
                        out_no_explaination_date = 0
                        time_late =0 
                        calculate_result_update_total = 0
                        # if actual_work_time > 0 and row['total_shift_work_time'] > 0:
                        convert_overtime = False
                        start_date_work_kid_mode = 0
                        end_date_work_kid_mode = 0
                        kid_mode_time = 0
                        for index, leave_item in hr_leave_employee_date.iterrows():
                            
                            # print("--------------------",
                            #     leave_item['holiday_status_name'])
                            try:
                                if ('có tính lương' in leave_item['holiday_status_name'].strip().lower()):
                                    print("aaaaaa'có tính lương'aaaaaaa")
                                    # total_ncl = total_ncl + \
                                    #     max(leave_item['minutes'], leave_item['time'])
                                    # if leave_item['is_leave_probationary']:
                                    #     total_tc_probationary = total_tc_probationary + \
                                    #         max(leave_item['minutes'],
                                    #             leave_item['time'])
                            except Exception as ex:
                                print(ex)

                            try:
                                if (('đi muộn' in leave_item['holiday_status_name'].strip().lower()) or ('về sớm' in leave_item['holiday_status_name'].strip().lower())):
                                    if ((leave_item['for_reasons']=='1') or (leave_item['for_reasons']==1)) and self.is_calculate_late_early_leave:
                                        print("don di muon ve som duoc duyet voi ly do ca nhan trong ngay")
                                        total_dimuon_vesom_date = total_dimuon_vesom_date + \
                                            max(leave_item['minutes'], leave_item['time'])
                                        real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row)

                                    elif ((leave_item['for_reasons']=='2') or (leave_item['for_reasons']==2)) and self.is_calculate_late_early_leave:
                                        total_dimuon_vesom_work_date = total_dimuon_vesom_work_date + \
                                            max(leave_item['minutes'], leave_item['time'])
                                        time_late_early_for_work = time_late_early_for_work + max(leave_item['minutes'], leave_item['time'])

                                        # if leave_item['is_leave_probationary']:
                                            #     total_dimuon_vesom_probationary = total_dimuon_vesom_probationary + \
                                            #         max(leave_item['minutes'],
                                            #             leave_item['time'])
                            except Exception as ex:
                                print(ex)
                            try:
                                if ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()) and \
                                    ((leave_item['for_reasons']=='1') or (leave_item['for_reasons']==1)):
                                    # duration = leave_item['attendance_missing_to'] - leave_item['attendance_missing_from']
                                    calculate_result = self.calculate_actual_work_time_couple(row, leave_item['attendance_missing_from'], leave_item['attendance_missing_to'])
                                    print("time out date xxxxxxxxxxxxxxxx", calculate_result)
                                    calculate_result_update = self.calculate_actual_work_time_couple(row, row[f'attendance_attempt_1'], leave_item['attendance_missing_to'])
                                    calculate_result_update_total = calculate_result_update['total_work_time']
                                    time_out_date = calculate_result['total_work_time']
                                    # if(row['employee_code']=='APG210922003') and (row['date'].day==30):
                                    #     print(row)
                                    #     print(actual_work_time)
                                    #     print(time_out_date)
                                    #     print(real_total_work_time)
                                    #     exit(1)
                                    # seconds = duration.total_seconds()
                                    # hours = seconds // 3600
                                    # minutes = (seconds % 3600) // 60.0
                                    # time_out_date = minutes + hours*60
                                    
                            except Exception as ex:
                                print(ex)
                                
                            convert_overtime = leave_item['convert_overtime']
                            
                            # if ('con nhỏ' in leave_item['holiday_status_name'].strip().lower()):
                            #     real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row) 
                                
                            #     kid_mod_stage_1_start_datetime, kid_mod_stage_1_end_datetime, \
                            #             kid_mod_stage_2_start_datetime, kid_mod_stage_2_end_datetime = \
                            #             self.process_kid_mode(row, leave_item)
                            #     try:
                            #         if ((leave_item['kid_mod_stage_1_start']!=11) and 
                            #             (leave_item['kid_mod_stage_1_start']!=11.5) and
                            #             (leave_item['kid_mod_stage_1_start']!=13)  ):
                            #             kid_mode_time = 60
                            #             try:    
                            #                 if(leave_item['kid_mod_stage_2_end'] == 9):    
                            #                     start_date_work_kid_mode = (kid_mod_stage_2_end_datetime.replace(second=0) - row['attendance_attempt_1'].replace(second=0)) \
                            #                                 .total_seconds()//60.0
                            #                     start_date_work_kid_mode =min(60, max(0, start_date_work_kid_mode))
                                                
                            #                 elif(leave_item['kid_mod_stage_1_end'] == 8.5):       
                            #                     start_date_work_kid_mode = (kid_mod_stage_1_end_datetime.replace(second=0) - row['attendance_attempt_1'].replace(second=0)) \
                            #                                 .total_seconds()//60.0
                            #                     start_date_work_kid_mode =min(30, max(0, start_date_work_kid_mode))
                                                
                            #                 print("(leave_item['kid_mod_stage_1_start']: ", leave_item['kid_mod_stage_1_start'])
                            #                 print("(leave_item['kid_mod_stage_1_end']: ", leave_item['kid_mod_stage_1_end'])
                            #                 print("(leave_item['kid_mod_stage_2_start']: ", leave_item['kid_mod_stage_2_start'])
                            #                 print("(leave_item['kid_mod_stage_2_end']: ", leave_item['kid_mod_stage_2_end'])
                            #                 print("row['shift_start_datetime']: ", row['shift_start_datetime'])
                            #                 print("row['attendance_attempt_1']: ", row['attendance_attempt_1'])
                            #                 print('kid_mod_stage_1_start_datetime: ', kid_mod_stage_1_start_datetime)
                                            
                            #             except:
                            #                 print('start_date_work_kid_mode')
                            #                 start_date_work_kid_mode = 0
                            #             try:
                            #                 if(leave_item['kid_mod_stage_1_start'] == 16):    
                            #                     end_date_work_kid_mode = (row['last_attendance_attempt'].replace(second=0) - kid_mod_stage_1_start_datetime.replace(second=0)) \
                            #                             .total_seconds()//60.0
                            #                     end_date_work_kid_mode =min(60, max(0, end_date_work_kid_mode))
                                                
                            #                 elif(leave_item['kid_mod_stage_2_start'] == 16.5):    
                            #                     end_date_work_kid_mode = (row['last_attendance_attempt'].replace(second=0) - kid_mod_stage_2_start_datetime.replace(second=0)) \
                            #                             .total_seconds()//60.0
                            #                     end_date_work_kid_mode =min(30, max(0, end_date_work_kid_mode))
                            #             except:
                            #                 print('start_date_work_kid_mode')
                            #                 end_date_work_kid_mode = 0
                                        
                            #     except Exception as ex:
                            #         print(ex)
                            #         # input('con nho 1')         
                                    
                            #     if (row['employee_ho']== False) and \
                            #         ((leave_item['kid_mod_stage_1_start']==11) or 
                            #         (leave_item['kid_mod_stage_1_start']==11.5) or
                            #         (leave_item['kid_mod_stage_1_start']==13)  ) :
                                    
                            #         if kid_mod_stage_1_start_datetime is None:
                            #             continue
                            #         mid_time_fist_array = []
                            #         for time_item_index in range(1,16) :
                            #             try:
                            #                 if (row[f"attendance_attempt_{time_item_index}"] < kid_mod_stage_1_start_datetime) and \
                            #                     (row[f"attendance_attempt_{time_item_index}"] <  row['last_attendance_attempt']) and \
                            #                     (row[f"attendance_attempt_{time_item_index}"] >  row['attendance_attempt_1']) and \
                            #                     (row[f"attendance_attempt_{time_item_index}"] > row['shift_start_datetime'])and \
                            #                     (row[f"attendance_attempt_{time_item_index}"] < row['shift_end_datetime']) :
                            #                     mid_time_fist_array.append(row[f"attendance_attempt_{time_item_index}"])
                            #             except Exception as ex:
                            #                 print("con nho ex mid_time_fist_array", ex)
                            #         mid_time_last_array = []
                            #         for time_item_index in range(1,16): 
                            #             try:
                            #                 if (row[f"attendance_attempt_{time_item_index}"] > kid_mod_stage_2_end_datetime) and \
                            #                     (row[f"attendance_attempt_{time_item_index}"] <  row['last_attendance_attempt']) and \
                            #                     (row[f"attendance_attempt_{time_item_index}"] > row['attendance_attempt_1']) and \
                            #                     (row[f"attendance_attempt_{time_item_index}"] > row['shift_start_datetime'])and \
                            #                     (row[f"attendance_attempt_{time_item_index}"] < row['shift_end_datetime']):
                            #                     mid_time_last_array.append(row[f"attendance_attempt_{time_item_index}"])
                            #             except Exception as ex:
                            #                 print("con nho ex mid_time_last_array", ex)
                            #         delete_item_array = []
                            #         for out_date_item in out_date_array:
                            #             for mid_time_fist_item in mid_time_fist_array:
                            #                 try:
                            #                     if (mid_time_fist_item.replace(second=59) >= out_date_item['attendance_missing_from']) and \
                            #                         (mid_time_fist_item.replace(second=0) <= out_date_item['attendance_missing_to']):
                            #                         delete_item_array.append(mid_time_fist_item)
                            #                 except Exception as ex:
                            #                     print('ex in ra ngoai 1: ', ex)
                                                
                            #         for delete_item in delete_item_array:
                            #             mid_time_fist_array.remove(delete_item)
                                        
                            #         delete_item_array = [] 
                            #         for out_date_item in out_date_array:
                            #             for mid_time_last_item in mid_time_last_array:
                            #                 try:
                            #                     if (mid_time_last_item.replace(second=59) >= out_date_item['attendance_missing_from']) and \
                            #                         (mid_time_last_item.replace(second=0) <= out_date_item['attendance_missing_to']):
                            #                         delete_item_array.append(mid_time_last_item)
                            #                 except Exception as ex:
                            #                     print('ex in ra ngoai: ', ex)
                                                
                            #         for delete_item in delete_item_array:
                            #             mid_time_last_array.remove(delete_item)
                                        
                            #         # found_late_mid = False
                            #         if len(mid_time_last_array) > 0:
                                      
                            #             late_mid_time = mid_time_last_array[0]
                            #             time_late = (late_mid_time - kid_mod_stage_2_end_datetime).total_seconds()//60.0
                            #             # real_total_work_time = real_total_work_time - time_late
                                        
                            #         if len(mid_time_fist_array) > 0:
                
                            #             late_mid_time = mid_time_fist_array[-1]
                            #             time_late = (kid_mod_stage_1_start_datetime - late_mid_time).total_seconds()//60.0

                            #     # if((leave_item['for_reasons']=='2') or (leave_item['for_reasons']==2)) and self.is_calculate_late_early_leave:
                            #     #     time_late_early_for_work += max(leave_item['minutes'], leave_item['time'])
                        if convert_overtime:
                            actual_work_time = 0
                        
                        calculate_actual_work_time = actual_work_time
                        for index, leave_item in hr_leave_employee_date.iterrows():
                            try:
                                if ('nghỉ bù' in leave_item['holiday_status_name'].strip().lower()):
                                    number_cl_date = number_cl_date + 1
                                    # print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                                    total_compensation_leave_date = max(leave_item['minutes'], leave_item['time'])
                                    # if leave_item['is_leave_probationary']:
                                    #     total_compensation_probationary = total_compensation_probationary + \
                                    #         max(leave_item['minutes'],
                                    #             leave_item['time'])
                                    if (calculate_actual_work_time  + total_compensation_leave_date + total_dimuon_vesom_date != total_shift_work_time):
                                        print('actual_work_time:', actual_work_time)
                                        print("update", f"{calculate_actual_work_time  + total_compensation_leave_date + total_dimuon_vesom_date} <- {total_shift_work_time}")
                                    
                                    if total_compensation_leave_date>0:
                                        print("total ototal_compensation_leave_date: ",total_compensation_leave_date)
                                    update_minute = calculate_actual_work_time + total_compensation_leave_date + total_dimuon_vesom_date - total_shift_work_time
                                    print("update_minute : ", update_minute)
                                    if (update_minute > 0) and (((update_minute<50)  and (update_minute != total_compensation_leave_date)) \
                                                                or (total_dimuon_vesom_date >0)) and (total_compensation_leave_date > update_minute):
                                        update_data = {"minutes": total_compensation_leave_date - update_minute,
                                                    "old_minutes": total_compensation_leave_date}
                                        ids = [leave_item['id']]
                                        logger.info(f"update {update_data} - {ids}")
                                        # try:
                                        #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.leave', 'write', [ids, update_data])
                                        #     total_compensation_leave = total_compensation_leave - update_minute
                                        # except Exception as ex:
                                        #     logger.error(f"hr.leave write error: {ex}")
                                        
                                        # print ('success update minute')
                            except Exception as ex:
                                print(ex)                          
                        
                            try:
                                if ('nghỉ phép năm' in leave_item['holiday_status_name'].strip().lower()):
                                    total_al_date = total_al_date + max(leave_item['minutes'], leave_item['time'])
                                    number_al_date = number_al_date + 1
                                    # print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                                elif ( 'có tính lương' in leave_item['holiday_status_name'].strip().lower()):
                                    total_ncl_date = max(leave_item['minutes'], leave_item['time'])
                                elif ('tăng ca' in leave_item['holiday_status_name'].strip().lower()):
                                    total_tc_date = max(leave_item['minutes'], leave_item['time'])
                            
                            except Exception as ex:
                                print(ex)
                                
                        try:
                            if (row['is_holiday']== True) and (auto_calculate_holiday == True):
                                print("Calculate holiday")
                                total_tc_holiday = total_tc_holiday  + \
                                    int((row['holiday_work_time'] - row['night_holiday_work_time_minute']) * 3 + \
                                    row['night_holiday_work_time_minute'] * night_holiday_wage)
                                if probationary_index:
                                    total_tc_holiday_probationary = total_tc_holiday_probationary + \
                                        int((row['holiday_work_time'] - row['night_holiday_work_time_minute']) * 3 + \
                                        row['night_holiday_work_time_minute'] * night_holiday_wage)
                        except Exception as ex:
                            print('auto_calculate_holiday', ex)   
                            
                        if (time_late>0) or (time_out_date >0):
                            real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row) 
                            try:
                                if (time_out_date >0):
                                    real_total_work_time = max(calculate_result_update_total, real_total_work_time)
                            except Exception as ex:
                                pass
                            actual_work_time = real_total_work_time - time_late
                            actual_work_time = max(actual_work_time - time_out_date, 0)
                                
                            if is_ho:
                                if(round(actual_work_time,0) != row['total_work_time']):
                                    if round(actual_work_time,0) in data_total_work_time_update:
                                        
                                        ids_total_work_time_update[round(actual_work_time, 0)].append(int(row['id']))
                                    else:
                                        data_total_work_time_update.append(round(actual_work_time, 0))
                                        ids_total_work_time_update[round(actual_work_time, 0)] = [int(row['id'])]
                        # if(row['employee_code']=='APG210922003') and (row['date'].day==30):
                        #     print(row)
                        #     print(actual_work_time)
                        #     print(time_out_date)
                        #     print(real_total_work_time)
                        #     exit(1)
                        # actual_work_time += time_late_early_for_work
                        total_late_early_for_work = total_late_early_for_work + time_late_early_for_work
                        shift_name_display = actual_work_time if actual_work_time > 0 else '-'
                        #############   start update hr
                        
                        
                        
                        total_normal_working_minutes = total_normal_working_minutes + actual_work_time
                        if probationary_index:
                            total_normal_working_minutes_probationary = total_normal_working_minutes_probationary + actual_work_time
                    
                        
                        col_index = col_index + 1
                        shift_name = row['shift_name'].strip()
                        if shift_name == 'UP':
                            total_unpaid_leave = total_unpaid_leave + 1
                            if probationary_index:
                                total_unpaid_leave_probationary = total_unpaid_leave_probationary + 1
                        elif ('/UP' in shift_name) or ('UP/' in shift_name):
                            total_unpaid_leave = total_unpaid_leave + 0.5
                            if probationary_index:
                                total_unpaid_leave_probationary = total_unpaid_leave_probationary + 0.5
                        # if shift_name == 'CL':
                        #     total_compensation_leave = total_compensation_leave + 1
                        #     if probationary_index:
                        #         total_compensation_probationary = total_compensation_probationary + 1
                        # elif ('/CL' in shift_name) or ('CL/' in shift_name):
                        #     total_compensation_leave = total_compensation_leave + 0.5
                        #     if probationary_index:
                        #         total_compensation_probationary = total_compensation_probationary + 0.5

                        # if shift_name == 'AL':
                        #     total_al = total_al + 1
                        #     if probationary_index:
                        #         total_al_probationary = total_al_probationary + 1
                        # elif ('/AL' in shift_name) or ('AL/' in shift_name):
                        #     total_al = total_al + 0.5
                        #     if probationary_index:
                        #         total_al_probationary = total_al_probationary + 0.5
                        'Ca gay'
                        if (row['fix_rest_time'] == True) and (row['total_work_time']>0):
                            total_fix_rest_time = total_fix_rest_time + 1
                            if probationary_index:
                                total_fix_rest_time_probationary = total_fix_rest_time_probationary + 1

                        if shift_name == 'OFF':
                            list_off_ids.append({'id' : int(row['id']), 'shift_name': row['shift_name'], 'employee_code': row['employee_code'], 'date': row['date']})
                            total_off = total_off + 1
                        elif ('/OFF' in shift_name) or ('OFF/' in shift_name):
                            total_off = total_off + 0.5

                        all_night_hours = row['night_worktime_minute']
                        total_night = total_night + all_night_hours 
                        if probationary_index:
                            total_night_probationary = total_night_probationary + all_night_hours
                        if row['is_holiday']:
                            total_night_holiday = total_night_holiday + row['night_holiday_work_time_minute']
                            if probationary_index:
                                total_night_holiday_probationary = total_night_holiday_probationary + row['night_holiday_work_time_minute']
                            
                        total_ph_date = 0
                        if (shift_name == 'AL') \
                                and (number_cl_date == 0) and (number_al_date == 0):
                            total_al_date = 480
                            total_al = total_al + 480
                            if probationary_index:
                                total_al_probationary = total_al_probationary + 480

                        if (shift_name == 'CL') \
                                and (number_cl_date == 0) and (number_al_date == 0):
                            total_compensation_leave_date = 480
                            total_compensation_leave = total_compensation_leave + 480
                            if probationary_index:
                                total_compensation_probationary = total_compensation_probationary + 480
                                
                        if row['is_holiday'] == True:
                            if shift_name == 'PH':
                                total_ph = total_ph + 1
                                total_ph_date = 1
                                if probationary_index:
                                    total_ph_probationary = total_ph_probationary + 1
                            elif ('/PH' in shift_name) or ('PH/' in shift_name):
                                total_ph_date = max(0.5, (480 - real_total_work_time)/480)
                                total_ph = total_ph + total_ph_date
                                # total_tc_holiday += (1-total_ph_date) * 480
                                if probationary_index:
                                    # total_tc_holiday_probationary += (1-total_ph_date) * 480
                                    total_ph_probationary = total_ph_probationary + total_ph_date
                                    
                            if ('/PH' in shift_name) or ('PH/' in shift_name):
                                total_compensation_leave = total_compensation_leave + real_total_work_time * row['efficiency_factor']
                                if probationary_index:
                                    total_compensation_probationary = total_compensation_probationary + real_total_work_time * row['efficiency_factor']
                                actual_work_time_date = min (total_shift_work_time, 
                                                total_ph_date * 480 + total_ncl_date + 
                                                total_dimuon_vesom_date +
                                                # total_dimuon_vesom_work_date +
                                                total_compensation_leave_date + total_al_date + actual_work_time)
                            else:
                                actual_work_time_date = total_ph_date * 480 + \
                                            total_dimuon_vesom_date + \
                                            total_ncl_date + total_compensation_leave_date + total_al_date + actual_work_time
                        else:
                            actual_work_time_date = total_ph_date * 480 + \
                                            total_dimuon_vesom_date + \
                                            total_ncl_date + total_compensation_leave_date + total_al_date + actual_work_time
                        # try:
                        # if shift_name.strip == 'PH':
                        #     actual_work_time_date = 480
                        update_data = {}
                        is_need_update = False
                        ids = [int(row['id'])]
                        if(actual_work_time>0):
                            total_actual_work_time_date = total_actual_work_time_date + actual_work_time_date
                            # ra ngoai khong co ly do
                            if row['employee_ho']:
                                out_no_explaination_date = row['total_out_ho']
                            else:
                                out_no_explaination_date = total_shift_work_time - (actual_work_time + row['attendance_late'] + row['leave_early']+ time_out_date)
                            total_out_no_explaination = total_out_no_explaination + out_no_explaination_date
                            
                            # if row['employee_code'] == 'APG230419006':
                            #     print('total_shift_work_time', total_shift_work_time)
                            #     print('actual_work_time', actual_work_time)
                            #     print("attendance_late", row['attendance_late'])
                            #     print("leave_early", row['leave_early'])
                            #     print(out_no_explaination_date)
                            #     print(total_out_no_explaination)
                            #     print(row['date'])
                            #     input(row['name'])
                        if is_ho:
                            if(round(actual_work_time_date, 0) != (row['actual_total_work_time'])):
                                if round(actual_work_time_date, 0) in data_actual_total_work_time_update:
                                    ids_actual_total_work_time_update[round(actual_work_time_date,0)].append(int(row['id']))
                                    # if(row['employee_code']=='APG231002004') and (row['date'].day==23):
                                    #     print(row)
                                    #     print(actual_work_time)
                                    #     print(time_out_date)
                                    #     print(real_total_work_time)
                                    #     print(actual_work_time_date)
                                    #     exit(1)
                                else:
                                    data_actual_total_work_time_update.append(round(actual_work_time_date, 0))
                                    ids_actual_total_work_time_update[round(actual_work_time_date, 0)] = ids
                                # update_data = {'actual_total_work_time': int(actual_work_time_date)}
                                # is_need_update = True
                        # if is_ho:
                            
                            if pd.notnull(standard_working_time) and pd.notna(standard_working_time):
                                if (standard_working_time>0) and (total_off > len(self.date_array) - standard_working_time + len(ids_off_remove_array)) \
                                    and ('OFF' in row['shift_name']) and (len(list_off_ids)>1):
                                    select_item = list_off_ids[0]
                                    for find_item in list_off_ids:
                                        if(find_item['date']) > select_item['date']:
                                            select_item = find_item
                                    item_copy = select_item.copy()
                                    ids_off_remove_array.append(item_copy)
                                    ids_all_off_remove.append(item_copy)
                                    new_list = []
                                    for find_item in list_off_ids:
                                        if find_item['date'] != select_item['date']:
                                            new_list.append(find_item)
                                    list_off_ids = new_list
                                if int(standard_working_time)  != int(row['standard_working_day']):
                                    # update_data['standard_working_day'] = int(standard_working_time)
                                    # is_need_update= True
                                    if int(standard_working_time) in data_standard_working_day_update:
                                        if not (int(row['id']) in ids_standard_working_day_update[int(standard_working_time)]):
                                            ids_standard_working_day_update[int(standard_working_time)].append(int(row['id']))
                                    else:
                                        data_standard_working_day_update.append(int(standard_working_time))
                                        ids_standard_working_day_update[int(standard_working_time)] = ids
                            else:
                                if 0 != row['standard_working_day']:
                                    if 0 in data_standard_working_day_update:
                                        if not (int(row['id']) in ids_standard_working_day_update[0]):
                                            ids_standard_working_day_update[0].append(int(row['id']))
                                    else:
                                        data_standard_working_day_update.append(0)
                                        ids_standard_working_day_update[0] = ids
                        # if is_ho:
                            if pd.notnull (row['probationary_contract_termination_date']) and pd.notna(row['probationary_contract_termination_date']):
                                if row['probationary_contract_termination_date'] != row['probation_completion_wage']:
                                    try:
                                        update_data_probation_completion_wage = row['probationary_contract_termination_date'].strftime('%Y-%m-%d')
                                        # is_need_update= True
                                        if update_data_probation_completion_wage in data_probation_completion_wage_update:
                                            if not (int(row['id']) in ids_probation_completion_wage_update[update_data_probation_completion_wage]):
                                                ids_probation_completion_wage_update[update_data_probation_completion_wage].append(int(row['id']))
                                        else:
                                            data_probation_completion_wage_update.append(update_data_probation_completion_wage)
                                            ids_probation_completion_wage_update[update_data_probation_completion_wage] = ids
                                    except Exception as ex:
                                        print(ex)
                        # if is_ho:
                            if pd.notnull(row['probationary_salary_rate']) and pd.notna(row['probationary_salary_rate']):
                                if  int(row['probationary_salary_rate']) != int(row['probation_wage_rate']):
                                    probationary_salary_rate_date = int(row['probationary_salary_rate'])
                                    # update_data['probation_wage_rate'] = probationary_salary_rate_date
                                    # is_need_update= True
                                    if ( probationary_salary_rate_date in data_probationwage_rate_update):
                                        if not (int(row['id']) in ids_probationwage_rate_update[probationary_salary_rate_date]):
                                            ids_probationwage_rate_update[probationary_salary_rate_date].append(int(row['id']))
                                    else:
                                        data_probationwage_rate_update.append(probationary_salary_rate_date)
                                        ids_probationwage_rate_update[probationary_salary_rate_date] = ids
                            
                        # except Exception as ex:
                        #     print("Exception ex in actual: ", row)
                        minutes_working_reduced = max(0, (row['minutes_working_not_reduced'] - total_shift_work_time))
                        total_minutes_working_reduced = total_minutes_working_reduced + minutes_working_reduced
                        
                        if not is_post_inspection:
                            ws.cell(row=start_row, column=start_col +
                                    date.day).value = shift_name_display
                            
                            # total_leave = total_al_date + total_compensation_leave_date
                            # if row['employee_ho']:
                            #     total_leave = total_leave + row['total_out_ho']
                            last_month_al_remaining_leave_minute_date = 0
                            if len(last_month_al_employee.index)>0:
                                last_month_al_remaining_leave_minute_date = last_month_al_employee.iloc[0]['remaining_leave_minute']
                            total_leave_date = min(total_shift_work_time - actual_work_time,
                                                    last_month_al_remaining_leave_minute_date - total_leave)
                            total_leave_date = max(total_leave_date, 0)
                            
                            total_leave = total_leave_date + total_leave
                            
                            ws.cell(row=start_row + 1, column=start_col + date.day).value = \
                                total_leave_date if total_leave_date>0 else '-'
                                
                            ws.cell(row=start_row + 2, column=start_col + date.day).value = \
                                'X' if row['attendance_late'] >= 5 else 'o' if row['attendance_late'] >0 \
                                    else ''
                                    
                            if row['attendance_late'] >= 5:
                                total_attendance_late_more_than_5 = total_attendance_late_more_than_5 + 1
                            elif row['attendance_late'] >0:
                                total_attendance_late_less_than_5 = total_attendance_late_less_than_5 + 1
                                        
                            ws.cell(row=start_row + 3, column=start_col + date.day).value = \
                                minutes_working_reduced if minutes_working_reduced>0 else '-'
                        #     if pd.notnull(row['probationary_contract_termination_date']):
                        #         ws.cell(row=start_row, column=6).value = row['probationary_contract_termination_date'].strftime(
                        #             '%d-%m-%Y')
                        #     if pd.notnull(row['probationary_salary_rate']):
                        #         ws.cell(
                        #             row=start_row, column=7).value = row['probationary_salary_rate']/100
                        #     if col_index == 0:
                        #         if pd.notnull(row['workingday']):
                        #             ws.cell(row=start_row, column=5).value = row['workingday'].strftime(
                        #                 '%d-%m-%Y')
                        #         if pd.notnull(row['job_title']):
                        #             ws.cell(row=start_row,
                        #                     column=4).value = row['job_title']
                            try:
                                
                                # print(resource_calendar_id)
                                # Tiêu chuẩn 48 giờ/tuần
                                ws.cell(row=start_row,
                                            column=5).value = standard_working_time
                            #     if '44' in resource_calendar_id:
                            #         ws.cell(row=start_row,
                            #                 column=8).value = self.working_time_44
                            #     elif '40' in resource_calendar_id:
                            #         ws.cell(row=start_row,
                            #                 column=8).value = self.working_time_40
                            #     else:
                            #         ws.cell(row=start_row,
                            #                 column=8).value = self.working_time_48
                            except Exception as ex:
                                ws.cell(row=start_row,
                                        column=8).value = self.working_time_48
                                print("gio lam ex: ", ex)


                # summary data
                # cl thu viec con lai bang cl thu viec thang truoc cong tang ca thu viec tru cl thu viec thang nay
                # last_month_al_employee = df_al_report[df_al_report['employee_code']==sub_group_index[1]]
                # last_month_cl_employee = df_cl_report[df_cl_report['employee_code']==sub_group_index[1]]
                
                last_month_cl_remaining_leave_minute_probationary = 0
                last_month_cl_remaining_leave_minute = 0
                current_al_remaining_leave_minute = 0
                last_month_al_remaining_leave_minute = 0
                current_cl_remaining_leave_minute = 0
                current_cl_remaining_leave_minute_probationary = 0
                
                if len(last_month_al_employee.index) > 0:
                    last_month_al_remaining_leave_minute = \
                        last_month_al_employee.iloc[0]['remaining_leave_minute'] + \
                        last_month_al_employee.iloc[0][current_month_str] 
                    
                if len(last_month_cl_employee.index) > 0:
                    last_month_cl_remaining_leave_minute = int(last_month_cl_employee.iloc[0]['remaining_total_minute']) + \
                        last_month_cl_employee.iloc[0][f"used_official_{self.date_array[0].month}"] + \
                        last_month_cl_employee.iloc[0][f"used_probationary_{self.date_array[0].month}"]
                        
                    last_month_cl_remaining_leave_minute_probationary = last_month_cl_employee.iloc[0]['remaining_probationary_minute'] + \
                        last_month_cl_employee.iloc[0][f"used_probationary_{self.date_array[0].month}"]
                try:
                    current_al_remaining_leave_minute = max(0,last_month_al_remaining_leave_minute - total_al)
                except:
                    current_al_remaining_leave_minute = 0
                    last_month_al_remaining_leave_minute = 0
                # cl con lai = cl con lai thang truoc cong tang ca cong tang ca tru di tong cl
                try:
                    current_cl_remaining_leave_minute = max(0, last_month_cl_remaining_leave_minute + total_tc - total_compensation_leave)
                except:
                    current_cl_remaining_leave_minute= 0
                    last_month_cl_remaining_leave_minute = 0
                try:
                    current_cl_remaining_leave_minute_probationary = max(0,last_month_cl_remaining_leave_minute_probationary + \
                                                            total_tc_probationary - total_compensation_probationary)
                except:
                    current_cl_remaining_leave_minute_probationary = 0
                    last_month_cl_remaining_leave_minute_probationary = 0 
                total_compensation_leave_dimuon_vesom = 0
                total_al_dimuon_vesom = 0
                if self.is_calculate_late_early_leave:
                
                    # cl di muon ve som = chuyen doi di muon ve som sang cl: gia tri nho nhat cua cl con lai va di muon ve som
                    total_compensation_leave_dimuon_vesom = min(total_dimuon_vesom, current_cl_remaining_leave_minute)

                    # cl di muon ve som chuyen doi sang cl thu viec bang gia tri nho nhat cl thu viec con lai và di muon ve som thu viec
                    total_compensation_dimuon_vesom_probationary = min(total_dimuon_vesom_probationary, current_cl_remaining_leave_minute_probationary)

                    # phep sau khi di muon ve som gia tri nho nhat so di muon ve som tru di phan chuyen doi sang cl va al con lai
                    total_al_dimuon_vesom = min(total_dimuon_vesom - total_compensation_leave_dimuon_vesom, current_al_remaining_leave_minute)
                used_cl =  total_compensation_leave + total_compensation_leave_dimuon_vesom
                used_cl_probationary = total_compensation_probationary + total_compensation_dimuon_vesom_probationary 
                used_cl_offical = used_cl - used_cl_probationary
                
                al_cl_current_month_item = {'used_cl_probationary': used_cl_probationary, 
                                            'used_cl_offical': used_cl_offical,
                                            'total_tc_probationary': total_tc_probationary,
                                            'total_tc_offical': total_tc - total_tc_probationary,
                                            'phatsinhtang_thuviec': phatsinhtang_thuviec,
                                            'phatsinhtang_chinhthuc': phatsinhtang_chinhthuc,
                                            'total_al': total_al_dimuon_vesom + total_al}
                
                self.al_cl_current_month[f"{sub_group_index[1]}"] = al_cl_current_month_item
                if not is_post_inspection:
                    ws.cell(row=start_row+1, column=51).value = total_attendance_late_more_than_5
                    # quy con lai
                    ws.cell(row=start_row+2, column=51).value = total_attendance_late_less_than_5
                    ws.cell(row=start_row, column=37).value = total_actual_work_time_date / minutes_per_day
                    ws.cell(row=start_row + 1, column=37).value = minutes_per_day
                    ws.cell(row=start_row, column=38).value  = total_minutes_working_reduced / minutes_per_day
                    ws.cell(row=start_row + 1, column=38).value = total_out_no_explaination
                    if len(last_month_al_employee.index) > 0:
                        ws.cell(row=start_row, column=40).value = last_month_al_employee.iloc[0]['remaining_leave_day']
                    else:
                        ws.cell(row=start_row, column=40).value = 0
                    ws.cell(row=start_row, column=40).value = total_leave/480
                    try:
                        ws.cell(row=start_row, column=74).value = last_month_cl_remaining_leave_minute
                        ws.cell(row=start_row, column=75).value = last_month_cl_remaining_leave_minute_probationary
                        ws.cell(row=start_row, column=76).value = last_month_al_remaining_leave_minute 
                    except Exception as ex:
                        print ("last month data err", ex)
                    ws.cell(row=start_row, column=77).value = total_compensation_leave_dimuon_vesom 
                    ws.cell(row=start_row, column=78).value = total_al_dimuon_vesom - total_al_probationary
                    ws.cell(row=start_row, column=79).value = total_dimuon_vesom 
                    ws.cell(row=start_row, column=80).value = current_al_remaining_leave_minute 
                    ws.cell(row=start_row, column=81).value = current_cl_remaining_leave_minute 
                    

                    # ws.cell(row=start_row, column=9).value = total_off
                    # ws.cell(row=start_row, column=10).value = total_unpaid_leave
                    # ws.cell(row=start_row, column=11).value = total_ph_probationary
                    # ws.cell(row=start_row, column=12).value = total_tc_probationary / 60.0
                    # ws.cell(row=start_row, column=13).value = total_tc_holiday_probationary / 60.0

                    # ws.cell(row=start_row, column=14).value = (total_night_probationary - total_night_holiday_probationary) / 60.0
                    # ws.cell(row=start_row, column=15).value = total_night_holiday_probationary / 60.0
                    # ws.cell(row=start_row, column=16).value = total_fix_rest_time_probationary
                    # ws.cell(row=start_row, column=17).value = total_ncl_probationary/480
                    # ws.cell(row=start_row, column=18).value = total_compensation_probationary/480 + total_compensation_dimuon_vesom_probationary/480
                    #         # min(total_compensation_probationary/480 + total_compensation_dimuon_vesom_probationary/480,
                    #         #                                   last_month_cl_remaining_leave_minute_probationary/480 + total_tc_probationary/480)
                        
                    # total_normal_working_minutes = total_normal_working_minutes
                    # ws.cell(row=start_row, column=19).value = total_normal_working_minutes_probationary / (8 * 60)

                    # ws.cell(row=start_row, column=21).value = total_ph - total_ph_probationary
                    # ws.cell(row=start_row, column=22).value = (total_tc - total_tc_probationary) / 60.0
                    # ws.cell(row=start_row, column=23).value = (total_tc_holiday - total_tc_holiday_probationary) / 60.0
                    # ws.cell(row=start_row, column=24).value = ((total_night - total_night_probationary) - 
                    #                         (total_night_holiday - total_night_holiday_probationary) ) / 60.0
                    # # if sub_group_index[1] == 'APG220801006':
                    # #     # input('APG220801006')
                    # #     print(total_night_holiday_probationary)
                    # #     print('total: ', total_night_holiday)
                    # #     input('APG220801006')
                    # ws.cell(row=start_row, column=25).value = (total_night_holiday - total_night_holiday_probationary) / 60.0

                    # ws.cell(row=start_row, column=26).value = total_fix_rest_time - total_fix_rest_time_probationary
                    # ws.cell(row=start_row, column=27).value = (total_al - total_al_probationary)/480 + total_al_dimuon_vesom/480
                    #                         # min((total_al - total_al_probationary)/480 + total_al_dimuon_vesom/480, \
                    #                         # last_month_al_remaining_leave_minute)

                    # ws.cell(row=start_row, column=30).value = (total_ncl - total_ncl_probationary) / 480
                    # ws.cell(row=start_row, column=31).value = (total_compensation_leave - total_compensation_probationary) / 480 + \
                    #     (total_compensation_leave_dimuon_vesom - total_compensation_dimuon_vesom_probationary) / 480

                    #     # min((total_compensation_leave - total_compensation_probationary) / 480 + \
                    #     # (total_compensation_leave_dimuon_vesom - total_compensation_dimuon_vesom_probationary) / 480, 
                    #     # (last_month_cl_remaining_leave_minute + total_tc) - \
                    #     # (last_month_cl_remaining_leave_minute_probationary/480 + total_tc_probationary/480))
                    # ws.cell(row=start_row, column=32).value = (total_normal_working_minutes - total_normal_working_minutes_probationary) / (8 * 60)
                    # total_paid_day = total_ph * 480 + total_ncl + total_compensation_leave + \
                    #         total_compensation_leave_dimuon_vesom + total_normal_working_minutes + total_al
                    # ws.cell(row=start_row, column=79).value = max(total_paid_day - standard_working_time * 480, 0)
                    ws.cell(row=start_row, column=80).value = phatsinhtang_thuviec 
                    ws.cell(row=start_row, column=81).value = phatsinhtang_chinhthuc 
                    # if self.is_update_al_cl:
                    #     try:
                    #         print(total_al_dimuon_vesom)

                    #         ids = [int(last_month_al_employee.iloc[0]['id'])]
                    #         if str(int(last_month_al_employee.iloc[0]['id'])) in self.al_result_update:
                    #             self.al_result_update[str(int(last_month_al_employee.iloc[0]['id']))][current_month_str] = \
                    #                     int(total_al_dimuon_vesom) 
                    #         else:
                    #             update_data_al_current = {}
                    #             if int(last_month_al_employee.iloc[0][current_month_str]) != int(total_al_dimuon_vesom): 
                    #                 # input(f"before - {sub_group_index[1]}")
                    #                 update_data_al_current[current_month_str] = int(total_al_dimuon_vesom) 
                    #                 self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'write', 
                    #                                         [ids, update_data_al_current])
                    #             # input(sub_group_index[1])
                    #     except Exception as ex:
                    #         print(ex)
                    #         logger.error(f"{sub_group_index[1]} - ex in al {ex}")
                            # input("ex in al")
                    
                        # try:
                        #     total_compensation_dimuon_vesom_offical = \
                        #         total_compensation_leave_dimuon_vesom - total_compensation_dimuon_vesom_probationary
                            
                        #     print(total_compensation_dimuon_vesom_offical)
                        #     ids = [int(last_month_cl_employee.iloc[0]['id'])]
                        #     id_cl_str = str(int(last_month_cl_employee.iloc[0]['id']))
                            
                        #     if id_cl_str in self.cl_result_update:
                        #         # if sub_group_index[1] =='':
                        #         #     print(self.cl_result_update[id_cl_str])
                        #         #     input('day roii')
                        #         try:
                        #             if f"used_official_{self.date_array[0].month}" in self.cl_result_update[id_cl_str]:
                        #                 self.cl_result_update[id_cl_str] \
                        #                     [f"used_official_{self.date_array[0].month}"] = \
                        #                     self.cl_result_update[id_cl_str][f"used_official_{self.date_array[0].month}"] \
                        #                     + int(total_compensation_dimuon_vesom_offical) 
                        #             else:
                        #                 self.cl_result_update[id_cl_str] \
                        #                     [f"used_official_{self.date_array[0].month}"] = \
                        #                         int(total_compensation_dimuon_vesom_offical) 
                        #             if f"used_probationary_{self.date_array[0].month}" in self.cl_result_update[id_cl_str]:
                        #                 self.cl_result_update[id_cl_str] \
                        #                     [f"used_probationary_{self.date_array[0].month}"] = \
                        #                     self.cl_result_update[id_cl_str][f"used_probationary_{self.date_array[0].month}"] \
                        #                     + int(total_compensation_dimuon_vesom_probationary) 
                        #             else:
                        #                 self.cl_result_update[id_cl_str] \
                        #                     [f"used_probationary_{self.date_array[0].month}"] = \
                        #                         int(total_compensation_dimuon_vesom_probationary) 
                                                
                        #         except Exception as ex:
                        #             logger.error(f"{sub_group_index[1]} - ex in cl {id_cl_str} stage 1 {ex}")
                        #             # if sub_group_index[1] =='':
                        #             #     print(self.cl_result_update[id_cl_str])
                        #             #     input('day roii')
                        #     else:
                        #         update_data_cl_current = {}
                        #         is_update_al_cl = False
                        #         if int(last_month_cl_employee.iloc[0][f"used_official_{self.date_array[0].month}"]) != int(total_compensation_dimuon_vesom_offical): 
                        #             # input(f"before - {sub_group_index[1]}")
                                    
                        #             is_update_al_cl = True
                        #         update_data_cl_current[f"used_official_{self.date_array[0].month}"] = int(total_compensation_dimuon_vesom_offical) 
                        #         if int(last_month_cl_employee.iloc[0][f"used_probationary_{self.date_array[0].month}"]) != int(total_compensation_dimuon_vesom_probationary): 
                        #             # input(f"before - {sub_group_index[1]}")
                                    
                        #             is_update_al_cl = True
                        #         update_data_cl_current[f"used_probationary_{self.date_array[0].month}"] = int(total_compensation_dimuon_vesom_probationary) 
                        #         if is_update_al_cl:    
                        #             self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'write', 
                        #                                     [ids, update_data_cl_current])
                        # except Exception as ex:
                        #     print(ex)
                        #     logger.error(f"{sub_group_index[1]} - ex in cl {ex}")
                            
                    start_row = start_row + 4
        return df_scheduling_ver
    
    def export_sumary_attendance_report_company(self, company_name, df_scheduling_ver, 
                    df_hr_leave, 
                    df_explanation_data, df_cl_report, df_al_report, ws, 
                    is_post_inspection=False, is_ho = False, 
                    auto_calculate_holiday = False,
                    night_holiday_wage = 3.0):
        start_row = 11
        start_col = 37
        if not is_post_inspection:    
            ws.merge_cells('C2:Z2')
            top_left_cell = ws['C2']
            top_left_cell.value = company_name
            ws.merge_cells('C3:Z3')
            top_left_cell = ws['C3']
            top_left_cell.value = ''
            ws.merge_cells('C4:Z4')
            top_left_cell = ws['C4']
            top_left_cell.value = ''
            # ws['AH10'].value = f'Employee Type: All - Working Time: All - Pay Cycle: Month - Payment Period: {self.date_array[0].strftime("%Y/%m/%d")} - {self.date_array[len(self.date_array)-1].strftime("%Y/%m/%d")}'
            # ws['AH6']  = ws['AH10'].value
            ws.merge_cells('A6:BK6')
            top_left_cell = ws['A6']
            top_left_cell.value = f'Employee Type: All - Working Time: All - Pay Cycle: Month - Payment Period: {self.date_array[0].strftime("%d/%m/%Y")} - {self.date_array[len(self.date_array)-1].strftime("%d/%m/%Y")}'
            for date_item in self.date_array:
                day_of_month = date_item.day
                ws.cell(row=start_row-1, column=start_col +
                        day_of_month).value = date_item
                ws.cell(row=start_row-2, column=start_col +
                        day_of_month).value = date_item.strftime('%A')[:3]
            if date_item.day < 31:
                for add_item in range(date_item.day + 1, 32):
                    ws.cell(row=start_row-1, column=start_col + add_item).value = ''
                    ws.cell(row=start_row-2, column=start_col + add_item).value = ''
                    ws.cell(row=start_row, column=start_col + add_item).value = ''
        data_probationwage_rate_update = []
        data_probation_completion_wage_update = []
        # data_actual_total_work_time_update = []
        # data_standard_working_day_update = []
        data_contract_company_update = []
        # data_total_work_time_update = []
        
        # ids_total_work_time_update = {}
        ids_contract_company_update = {}
        ids_probationwage_rate_update = {}
        ids_probation_completion_wage_update = {}
        # ids_actual_total_work_time_update = {}
        # ids_standard_working_day_update = {}
        ids_all_off_remove = []
        month_array_str = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 
                                'october', 'november', 'december']
        current_month_str = month_array_str[self.date_array[0].month-1]
        try:
            df_scheduling_ver[['night_holiday_work_time_minute','holiday_work_time', 'night_worktime_minute']] = \
                    df_scheduling_ver.apply(lambda row: self.calculate_night_holiday_work_time(row),  axis=1, result_type='expand')
            print("sucesss")
        except Exception as ex:
            df_scheduling_ver['night_holiday_work_time_minute'] =0
            df_scheduling_ver['holiday_work_time'] = 0
            df_scheduling_ver['night_worktime_minute']= 0
            print("calculate night: ", ex) 
        for g, data in df_scheduling_ver[(df_scheduling_ver['department_name'].notnull())
                                              & (df_scheduling_ver['department_name'] != False)
                                              & (df_scheduling_ver['workingday'].notnull())
                                              # & (self.df_scheduling_ver['employee_code']=='APG230316014')
                                              & (df_scheduling_ver['workingday'] != False)].groupby('department_name'):
            start_row = start_row + 1
            count = 1
            print(g)
            if not is_post_inspection:
                ws.cell(row=start_row, column=1).value = g
                ws.cell(row=start_row, column=17).value = ''
                ws.cell(row=start_row, column=31).value = ''
                ws.cell(row=start_row, column=32).value = ''
                start_row = start_row + 1

            for sub_group_index, sub_group_data in data.groupby(['name', 'employee_code']):
                if (sub_group_data.iloc[0]['severance_day'] !=False) and (pd.notnull (sub_group_data.iloc[0]['severance_day'])) and \
                        sub_group_data.iloc[0]['severance_day'] < self.date_array[0]:
                    continue
                # if sub_group_index[1] == 'APG220801006':
                    # input('APG220801006')
                    # print(sub_group_data.iloc[0]['night_holiday_work_time_minute'])
                    # print('total: ', total_night_holiday)
                list_off_ids = []
                ids_off_remove_array = []
                # standard_working_time = sub_group_data['standard_working_time'].mean()
                standard_working_time =  sub_group_data[sub_group_data['standard_working_day']>0]['standard_working_day'].mean()
                total_ncl = 0
                total_ncl_probationary = 0
                total_off = 0

                total_ph = 0
                total_al = 0
                total_al_probationary = 0

                total_al_dimuon_vesom = 0

                total_dimuon_vesom = 0
                total_dimuon_vesom_probationary = 0

                total_tc = 0
                total_tc_probationary = 0
                phatsinhtang_thuviec = 0
                phatsinhtang_chinhthuc = 0
                total_tc_holiday = 0
                total_tc_holiday_probationary = 0
                total_normal_working_minutes = 0
                total_normal_working_minutes_probationary = 0

                total_ph_probationary = 0
                total_unpaid_leave = 0
                total_unpaid_leave_probationary = 0

                total_compensation_leave = 0
                total_compensation_probationary = 0

                total_compensation_leave_dimuon_vesom = 0
                total_compensation_dimuon_vesom_probationary = 0

                total_night = 0
                total_night_probationary = 0

                total_night_holiday = 0
                total_night_holiday_probationary = 0

                total_fix_rest_time = 0
                total_fix_rest_time_probationary = 0

                minute_worked_day_holiday = 0
                minute_worked_day_holiday_probationary = 0

                total_late_early_for_work = 0
                if not is_post_inspection:
                    ws.cell(row=start_row, column=2).value = sub_group_index[1]
                    ws.cell(row=start_row, column=3).value = sub_group_index[0]
                col_index = 0
                total_out = 0
                

                hr_leave_employee = df_hr_leave[df_hr_leave['employee_code']
                                                     == sub_group_index[1]]

                for index, leave_item in hr_leave_employee.iterrows():
                    # if sub_group_index[1] == 'APG230610001':
                    #     print(leave_item)
                    #     leave_item['employee_code']
                    #     print(total_tc_holiday)
                    #     print(hr_leave_employee)
                    #     for index2, testitem in hr_leave_employee.iterrows():
                    #         try:
                    #             print(testitem['holiday_status_name'])
                    #         except:
                    #             print(testitem['holiday_status_name'][1])
                    #         if ('tăng ca' in testitem['holiday_status_name'].strip().lower()):
                    #             exit(1)
                    # if shift_name == 'CL':
                    #     total_compensation_leave = total_compensation_leave + 1
                    #     if probationary_index:
                    #         total_compensation_probationary = total_compensation_probationary + 1
                    # elif ('/CL' in shift_name) or ('CL/' in shift_name):
                    #     total_compensation_leave = total_compensation_leave + 0.5
                    #     if probationary_index:
                    #         total_compensation_probationary = total_compensation_probationary + 0.5

                    # if shift_name == 'AL':
                    #     total_al = total_al + 1
                    #     if probationary_index:
                    #         total_al_probationary = total_al_probationary + 1
                    # elif ('/AL' in shift_name) or ('AL/' in shift_name):
                    #     total_al = total_al + 0.5
                    #     if probationary_index:
                    #         total_al_probationary = total_al_probationary + 0.5
                    # if (leave_item['request_date_from'] <= self.date_array[len(self.date_array)-1]):
                    # Đơn xin nghỉ có tính lương
                    # print("--------------------",
                    #       leave_item['holiday_status_name'])
                    # print("--------------------",
                    #       leave_item['holiday_status_name'])
                    try:
                        if ('nghỉ bù' in leave_item['holiday_status_name'].lower()):
                            print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                            total_compensation_leave = total_compensation_leave + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_compensation_probationary = total_compensation_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)
                    try:
                        if ( 'nghỉ phép năm' in leave_item['holiday_status_name'].strip().lower()):
                            print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                            total_al = total_al + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_al_probationary = total_al_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)             
                    try:
                        if ( 'có tính lương' in leave_item['holiday_status_name'].strip().lower()):
                            print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                            total_ncl = total_ncl + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_ncl_probationary = total_ncl_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)
                    # try:
                    #     if ('tăng ca' in leave_item['holiday_status_name'].strip().lower()):
                    #         print("don tang ca")
                    #         if leave_item['is_holiday']==True:
                    #             total_tc_holiday = total_tc_holiday + \
                    #                 max(leave_item['minutes'],
                    #                     leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                    #             if leave_item['is_leave_probationary']:
                    #                 total_tc_holiday_probationary = total_tc_holiday_probationary + \
                    #                     max(leave_item['minutes'],
                    #                         leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                    #         else:
                    #             total_tc = total_tc + \
                    #                 max(leave_item['minutes'],
                    #                     leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                    #             if leave_item['is_leave_probationary']:
                    #                 total_tc_probationary = total_tc_probationary + \
                    #                     max(leave_item['minutes'],
                    #                         leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                                        
                    #         if 'phát sinh tăng' in leave_item['reasons'].strip().lower():
                    #             if leave_item['is_leave_probationary']:
                    #                 phatsinhtang_thuviec = phatsinhtang_thuviec + \
                    #                     max(leave_item['minutes'],
                    #                         leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                    #             else:
                    #                 phatsinhtang_chinhthuc = phatsinhtang_chinhthuc + \
                    #                     max(leave_item['minutes'],
                    #                         leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                        
                    #         # if sub_group_index[1] == 'APG230610001':
                    #         #     print(leave_item)
                    #         #     print(total_tc_holiday)
                    #         #     print(real_total_work_time)
                    #         #     exit(1)
                    # except Exception as ex:
                    #     print(ex)    
                            
                    try:    
                        if (('đi muộn' in leave_item['holiday_status_name'].strip().lower()) or ('về sớm' in leave_item['holiday_status_name'].strip().lower())) and \
                                ((leave_item['for_reasons']=='1') or (leave_item['for_reasons']==1)) and self.is_calculate_late_early_leave:
                            print("don di muon ve som duoc duyet voi ly do ca nhan")
                            total_dimuon_vesom = total_dimuon_vesom + \
                                max(leave_item['minutes'], leave_item['time'])
                            if leave_item['is_leave_probationary']:
                                total_dimuon_vesom_probationary = total_dimuon_vesom_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time'])
                    except Exception as ex:
                        print(ex)   
                    try:                
                        if ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()) and \
                                    leave_item['for_reasons']=='1':
                            duration = leave_item['attendance_missing_to'] - leave_item['attendance_missing_from']
                            seconds = duration.total_seconds()
                            hours = seconds // 3600
                            minutes = (seconds % 3600) // 60.0

                            total_out = total_out + minutes + hours*60
                            print("don ra ngoài cá nhân")
                    except Exception as ex:
                        print(ex)
                # print('hr_learve_employee: ', hr_leave_employee)
                # hr_leave_employee.to_excel('hr_leave_employee.xlsx')
                if not is_post_inspection:
                    ws.cell(row=start_row, column=9).value = 0
                    ws.cell(row=start_row, column=10).value = 0
                    ws.cell(row=start_row, column=11).value = 0
                    ws.cell(row=start_row, column=12).value = 0
                    ws.cell(row=start_row, column=13).value = 0

                    ws.cell(row=start_row, column=14).value = 0
                    ws.cell(row=start_row, column=15).value = 0
                    ws.cell(row=start_row, column=16).value = 0
                    ws.cell(row=start_row, column=17).value = 0
                    ws.cell(row=start_row, column=18).value = 0
                    ws.cell(row=start_row, column=19).value = 0

                    ws.cell(row=start_row, column=21).value = 0
                    ws.cell(row=start_row, column=22).value = 0
                    ws.cell(row=start_row, column=23).value = 0
                    ws.cell(row=start_row, column=24).value = 0
                    ws.cell(row=start_row, column=25).value = 0

                    ws.cell(row=start_row, column=26).value = 0
                    ws.cell(row=start_row, column=27).value = 0
                    ws.cell(row=start_row, column=31).value = 0
                    ws.cell(row=start_row, column=32).value = 0

                # for index, row in sub_group_data.iterrows():
                for probationary_index, probationary_data in sub_group_data.groupby('is_probationary'):   
                    for index, row in probationary_data.iterrows():
                        last_out_date= row['last_attendance_attempt']
                        out_date_array = []
                        hr_leave_employee_date = df_hr_leave[(df_hr_leave['employee_code']
                                                     == sub_group_index[1]) & \
                                                    (df_hr_leave['date_str'] == row['date_str']) ]
                        if pd.notnull(row['contract_company']):
                            if row['company'] != row['contract_company']:
                                try:
                                    # is_need_update= True
                                    if f"{row['contract_company']}" in data_contract_company_update:
                                        ids_contract_company_update[f"{row['contract_company']}"].append({"id":int(row['id']), "employee_code": row['employee_code'], 
                                                                                                        "employee_name": row['employee_name'], 
                                                                                                        'company': row["company"], 'contract_company': row["contract_company"]})
                                    else:
                                        data_contract_company_update.append(f"{row['contract_company']}")
                                        ids_contract_company_update[f"{row['contract_company']}"] = [{"id":int(row['id']), "employee_code": row['employee_code'], 
                                                                                                        'company': row["company"], 'contract_company': row["contract_company"]}]
                                except Exception as ex:
                                    print(ex)
                    # if (sub_group_index[1]=='APG220215002'):
                    #     hr_leave_employee_date.to_excel(os.path.join(output_report_folder, f'{ row["date_str"]}-APG220215002.xlsx'))
                        if len(hr_leave_employee_date.index) >0:
                            print("check date: ",  row['date'].strftime('%Y-%m-%d'))
                        time_late_early_for_work = 0
                        time_bussiness_trip = 0
                        is_update_total_worktime = False
                        real_total_work_time = round(row['total_work_time'],0)
                        start_date_work_kid_mode = 0
                        end_date_work_kid_mode = 0
                        kid_mode_time = 0
                        # if actual_work_time > 0 and row['total_shift_work_time'] > 0:
                        # for index, leave_item in hr_leave_employee_date.iterrows():
                            
                        #     if ('con nhỏ' in leave_item['holiday_status_name'].strip().lower()):
                               
                        #         real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row) 
                                
                        #         kid_mod_stage_1_start_datetime, kid_mod_stage_1_end_datetime, \
                        #             kid_mod_stage_2_start_datetime, kid_mod_stage_2_end_datetime = \
                        #             self.process_kid_mode(row, leave_item)
                        #         try:
                        #             if((leave_item['kid_mod_stage_1_start']!=11) and 
                        #                 (leave_item['kid_mod_stage_1_start']!=11.5) and
                        #                 (leave_item['kid_mod_stage_1_start']!=13)  ):
                        #                 kid_mode_time = 60
                        #                 try:    
                        #                     if(leave_item['kid_mod_stage_2_end'] == 9):    
                        #                         start_date_work_kid_mode = (kid_mod_stage_2_end_datetime.replace(second=0) - row['attendance_attempt_1'].replace(second=0)) \
                        #                                     .total_seconds()//60.0
                        #                         start_date_work_kid_mode =min(60, max(0, start_date_work_kid_mode))
                                                
                        #                     elif(leave_item['kid_mod_stage_1_end'] == 8.5):       
                        #                         start_date_work_kid_mode = (kid_mod_stage_1_end_datetime.replace(second=0) - row['attendance_attempt_1'].replace(second=0)) \
                        #                                     .total_seconds()//60.0
                        #                         start_date_work_kid_mode =min(30, max(0, start_date_work_kid_mode))
                                                
                        #                     print("(leave_item['kid_mod_stage_1_start']: ", leave_item['kid_mod_stage_1_start'])
                        #                     print("(leave_item['kid_mod_stage_1_end']: ", leave_item['kid_mod_stage_1_end'])
                        #                     print("(leave_item['kid_mod_stage_2_start']: ", leave_item['kid_mod_stage_2_start'])
                        #                     print("(leave_item['kid_mod_stage_2_end']: ", leave_item['kid_mod_stage_2_end'])
                        #                     print("row['shift_start_datetime']: ", row['shift_start_datetime'])
                        #                     print("row['attendance_attempt_1']: ", row['attendance_attempt_1'])
                        #                     print('kid_mod_stage_1_start_datetime: ', kid_mod_stage_1_start_datetime)
                                            
                        #                 except:
                        #                     print('start_date_work_kid_mode')
                        #                     start_date_work_kid_mode = 0
                        #                 try:
                        #                     if(leave_item['kid_mod_stage_1_start'] == 16):    
                        #                         end_date_work_kid_mode = (row['last_attendance_attempt'].replace(second=0) - kid_mod_stage_1_start_datetime.replace(second=0)) \
                        #                                 .total_seconds()//60.0
                        #                         end_date_work_kid_mode =min(60, max(0, end_date_work_kid_mode))
                                                
                        #                     elif(leave_item['kid_mod_stage_2_start'] == 16.5):    
                        #                         end_date_work_kid_mode = (row['last_attendance_attempt'].replace(second=0) - kid_mod_stage_2_start_datetime.replace(second=0)) \
                        #                                 .total_seconds()//60.0
                        #                         end_date_work_kid_mode =min(30, max(0, end_date_work_kid_mode))
                        #                 except:
                        #                     print('start_date_work_kid_mode')
                        #                     end_date_work_kid_mode = 0
                                        
                        #                 print('before: real_total_work_time : ',real_total_work_time)
                        #                 real_total_work_time = real_total_work_time + kid_mode_time - \
                        #                             (start_date_work_kid_mode + end_date_work_kid_mode)  

                        #                 print(f'{row["company"]}--{row["name"]}')
                        #                 print('kid_mode_time', kid_mode_time)

                        #                 print('start_date_work_kid_mode', start_date_work_kid_mode)
                        #                 print('real_total_work_time', real_total_work_time)
                        #                 print('end_date_work_kid_mode', end_date_work_kid_mode)
                        #                 # input(f"{row['employee_code']}----{kid_mod_stage_1_start_datetime.replace(second=0)}")
                        #         except Exception as ex:
                        #             print(ex)
                        #             # input("con nho 3")
                        #         # if((leave_item['for_reasons']=='2') or (leave_item['for_reasons']==2)) and self.is_calculate_late_early_leave:
                        #         #     time_late_early_for_work += max(leave_item['minutes'], leave_item['time'])
                            
                        if len(hr_leave_employee_date.index) >0:
                            for index, leave_item in hr_leave_employee_date.iterrows():
                                
                                # print("--------------------",
                                #     leave_item['holiday_status_name'])
                                
                                try:
                                # if ('có tính lương' in leave_item['holiday_status_name'].strip().lower()):
                                #     print("aaaaaa'có tính lương'aaaaaaa")
                                    if ('công tác' in leave_item['holiday_status_name'].strip().lower()):
                                        time_bussiness_trip = time_bussiness_trip + max(leave_item['minutes'], leave_item['time'])
                                except:
                                    print('cong tac')
                                try:
                                    if ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()):
                                        if pd.notnull(leave_item['request_date_from']):
                                            out_date_array.append(leave_item)
                                            real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row)
                                            real_total_work_time = real_total_work_time + kid_mode_time - \
                                                (start_date_work_kid_mode + end_date_work_kid_mode)  
                                            if last_out_date > leave_item['request_date_from']:
                                                last_out_date = leave_item['request_date_from']
                                            if (leave_item['for_reasons']=='1'):                         
                                                try:
                                                    calculate_result_update = self.calculate_actual_work_time_couple(row, row[f'attendance_attempt_1'], leave_item['attendance_missing_to'])
                                                    real_total_work_time = max(calculate_result_update['total_work_time'], real_total_work_time)
                                                except Exception as ex:
                                                    pass
                                                calculate_result = self.calculate_actual_work_time_couple(row, leave_item['attendance_missing_from'], leave_item['attendance_missing_to'])
                                                
                                                print("time out date xfiiiiiiiiiiiiiiiiiistx", calculate_result)
                                                time_out_date = calculate_result['total_work_time']
                                                real_total_work_time = real_total_work_time - time_out_date
                                            
                                except:
                                    print('ra ngoai')
                                # elif (('đi muộn' in leave_item['holiday_status_name'].strip().lower()) or ('về sớm' in leave_item['holiday_status_name'].strip().lower())):
                                #     if row['employee_ho']== False:
                                #         real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row) 
                                #         if row['employee_code'] == 'APG220801006':
                                #             print(row)
                                #             print()
                                #             print(real_total_work_time)
                                #             # exit(1)
                                
                                #     if is_ho:
                                #         if(real_total_work_time != row['total_work_time']):
                                #             if int(real_total_work_time) in data_total_work_time_update:
                                #                 ids_total_work_time_update[int(real_total_work_time)].append(int(row['id']))
                                #             else:
                                #                 data_total_work_time_update.append(int(real_total_work_time))
                                #                 ids_total_work_time_update[int(real_total_work_time)] = [int(row['id'])]

                                #     if((leave_item['for_reasons']=='2') or (leave_item['for_reasons']==2)) and self.is_calculate_late_early_leave:
                                #         time_late_early_for_work += max(leave_item['minutes'], leave_item['time'])
                                # except Exception as ex:
                                #     print(ex)
                            
                        # print(f"{row}")
                        date = row['date']
                        shift_name = row['shift_name'] if row['shift_name'] else ''
                        resource_calendar_id = row['resource_calendar_id']
                        if pd.isnull(resource_calendar_id) or resource_calendar_id == False:
                            contract_employee_filter = self.df_contract[(self.df_contract['date_start']<=date)&
                                        (self.df_contract['date_end']>=date)&
                                        (self.df_contract['employee_code']==sub_group_index[1])].sort_values("date_end", ascending=False)
                            if len(contract_employee_filter.index)>0:
                                resource_calendar_id = contract_employee_filter.iloc[0]['resource_calendar_id']
                            
                            
                        if row['fix_rest_time'] == False:
                            total_shift_work_time = round(row['total_shift_work_time'] * 60, -1)
                        else: 
                            total_shift_work_time = 480

                        # if sub_group_index[1] == 'APG220601007':
                        actual_work_time =  max (time_bussiness_trip, min (real_total_work_time + time_bussiness_trip, total_shift_work_time))

                        # else:    
                        #     actual_work_time =  min (row['total_work_time'] + row['kid_time'], total_shift_work_time)
                        # actual_work_time = actual_work_time - total_out
                        
                        
                        # if resource_calendar_id == False:
                        #     continue
                        # Find explan nation
                        # giai trinh di muon ve som ly do cong viec
                        # thieu cham cong va khong dung gio cho full luon
                        explanation_filter = df_explanation_data[
                            df_explanation_data['date_str'].isin([row['date_str']]) &
                            ((df_explanation_data['invalid_type']==3) | (df_explanation_data['invalid_type']=='3') |
                            (df_explanation_data['invalid_type']==4) | (df_explanation_data['invalid_type']=='4')) &
                            df_explanation_data['employee_code'].str.contains(sub_group_index[1]).fillna(False) &
                            ((df_explanation_data['reason']==2) | (df_explanation_data['reason']== '2')) &
                            ((df_explanation_data['validated']==2) | (df_explanation_data['validated'] == '2'))]
                        
                        if len(explanation_filter.index) > 0 and actual_work_time < row['total_shift_work_time']*60:
                            if ('/OFF' in row['shift_name']) or ('OFF/' in row['shift_name']):
                                actual_work_time = 240
                            elif row['fix_rest_time'] == True:
                                actual_work_time = 480
                            else:
                                actual_work_time = round(row['total_shift_work_time']*60, -1)

                            print('fix: ------------------- buy explanation', explanation_filter.iloc[0])

                        # if len(hr_leave_filter.index)>0 and actual_work_time < 480:
                        #     # try:
                        #     print('fix by cl: -------------------', hr_leave_filter)
                        #     if pd.notnull(hr_leave_filter.iloc[0]['minutes']):

                        #         minute = int(hr_leave_filter.iloc[0]['minutes'])
                        #         actual_work_time = min(row['total_shift_work_time']*60, actual_work_time + minute)
                        #     # except:
                        #     #     pass

                        if (row['missing_checkin_break'] == True) and (not ('/OFF' in shift_name)) and (not ('OFF/' in shift_name)):
                            actual_work_time = min(actual_work_time, 240)
                                # actual_work_time, row['total_shift_work_time']*30)
                            # print(f"Thieu cham cong giua {date}")
                            if not is_post_inspection:
                                redFill = PatternFill(
                                    patternType='solid', fgColor=colors.Color(rgb='00FDF000'))
                                ws.cell(row=start_row, column=start_col +
                                        date.day).fill = redFill
                        else:
                            # print(f"khong thieu cham cong {row['attendance_attempt_3']} - hop dong {resource_calendar_id}" )
                            if not is_post_inspection:
                                blueFill = PatternFill(
                                    patternType='solid', fgColor=colors.Color(rgb='0000FF00'))
                                ws.cell(row=start_row, column=start_col +
                                        date.day).fill = blueFill
                        if not is_post_inspection:
                            ft = Font(color="0000FF")
                            if shift_name == '-' or shift_name == 'OFF' or row['total_shift_work_time'] == 0:
                                ft = Font(color="0000FF")
                            elif (pd.isnull(row['attendance_attempt_1']) or row['attendance_attempt_1'] == False) \
                                    or (pd.isnull(row['last_attendance_attempt']) or row['last_attendance_attempt'] == False):
                                ft = Font(color="FF0000")
                            elif real_total_work_time < 480:
                                ft = Font(color="FC1000")
                            else:
                                ft = Font(color="0000FF")

                            ws.cell(row=start_row, column=start_col +
                                    date.day).font = ft
                        time_out_date = 0
                        total_tc_date =0 
                        total_ncl_date =0 
                        total_compensation_leave_date = 0 
                        number_cl_date = 0
                        total_dimuon_vesom_date = 0
                        total_dimuon_vesom_work_date = 0
                        total_al_date = 0 
                        number_al_date = 0
                        late_mid_time = 0
                        time_late =0 
                        calculate_result_update_total = 0
                        start_date_work_kid_mode = 0
                        end_date_work_kid_mode = 0
                        kid_mode_time = 0
                        total_tc = total_tc + row['ot_normal'] 
                        
                        total_tc_holiday = total_tc_holiday + row['ot_holiday']
                        if(probationary_index):
                            total_tc_probationary = total_tc_probationary + row['ot_normal'] 
                            total_tc_holiday_probationary = total_tc_holiday_probationary + row['ot_holiday']
                        # if actual_work_time > 0 and row['total_shift_work_time'] > 0:
                        # for index, leave_item in hr_leave_employee_date.iterrows():
                            
                        #     if ('con nhỏ' in leave_item['holiday_status_name'].strip().lower()):
                        #         real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row) 
                                
                        #         kid_mod_stage_1_start_datetime, kid_mod_stage_1_end_datetime, \
                        #             kid_mod_stage_2_start_datetime, kid_mod_stage_2_end_datetime = \
                        #             self.process_kid_mode(row, leave_item)
                        #         try:
                        #             if((leave_item['kid_mod_stage_1_start']!=11) and 
                        #                 (leave_item['kid_mod_stage_1_start']!=11.5) and
                        #                 (leave_item['kid_mod_stage_1_start']!=13)):
                        #                 kid_mode_time = 60
                        #                 try:    
                        #                     if(leave_item['kid_mod_stage_2_end'] == 9):    
                        #                         start_date_work_kid_mode = (kid_mod_stage_2_end_datetime.replace(second=0) - row['attendance_attempt_1'].replace(second=0)) \
                        #                                     .total_seconds()//60.0
                        #                         start_date_work_kid_mode =min(60, max(0, start_date_work_kid_mode))
                                                
                        #                     elif(leave_item['kid_mod_stage_1_end'] == 8.5):       
                        #                         start_date_work_kid_mode = (kid_mod_stage_1_end_datetime.replace(second=0) - row['attendance_attempt_1'].replace(second=0)) \
                        #                                     .total_seconds()//60.0
                        #                         start_date_work_kid_mode =min(30, max(0, start_date_work_kid_mode))
                                                
                        #                     print("(leave_item['kid_mod_stage_1_start']: ", leave_item['kid_mod_stage_1_start'])
                        #                     print("(leave_item['kid_mod_stage_1_end']: ", leave_item['kid_mod_stage_1_end'])
                        #                     print("(leave_item['kid_mod_stage_2_start']: ", leave_item['kid_mod_stage_2_start'])
                        #                     print("(leave_item['kid_mod_stage_2_end']: ", leave_item['kid_mod_stage_2_end'])
                        #                     print("row['shift_start_datetime']: ", row['shift_start_datetime'])
                        #                     print("row['attendance_attempt_1']: ", row['attendance_attempt_1'])
                        #                     print('kid_mod_stage_1_start_datetime: ', kid_mod_stage_1_start_datetime)
                                            
                        #                 except:
                        #                     print('start_date_work_kid_mode')
                        #                     start_date_work_kid_mode = 0
                        #                 try:
                        #                     if(leave_item['kid_mod_stage_1_start'] == 16):    
                        #                         end_date_work_kid_mode = (row['last_attendance_attempt'].replace(second=0) - kid_mod_stage_1_start_datetime.replace(second=0)) \
                        #                                 .total_seconds()//60.0
                        #                         end_date_work_kid_mode =min(60, max(0, end_date_work_kid_mode))
                                                
                        #                     elif(leave_item['kid_mod_stage_2_start'] == 16.5):    
                        #                         end_date_work_kid_mode = (row['last_attendance_attempt'].replace(second=0) - kid_mod_stage_2_start_datetime.replace(second=0)) \
                        #                                 .total_seconds()//60.0
                        #                         end_date_work_kid_mode =min(30, max(0, end_date_work_kid_mode))
                        #                 except:
                        #                     print('start_date_work_kid_mode')
                        #                     end_date_work_kid_mode = 0
                                        
                        #                 real_total_work_time = real_total_work_time + kid_mode_time - \
                        #                             (start_date_work_kid_mode + end_date_work_kid_mode)  
                        #                 # input(real_total_work_time)
                        #                 # input(end_date_work_kid_mode)
                        #                 # input(f"{leave_item['kid_mod_stage_1_start']}----{kid_mod_stage_1_start_datetime}---{kid_mod_stage_2_start_datetime}")
                        #         except Exception as ex:
                        #             print(ex)
                        #             # input('con nho 2')
                        #         if (row['employee_ho']== False) and \
                        #             ((leave_item['kid_mod_stage_1_start']==11) or 
                        #             (leave_item['kid_mod_stage_1_start']==11.5) or
                        #             (leave_item['kid_mod_stage_1_start']==13)  ) :
                                    
                        #             if kid_mod_stage_1_start_datetime is None:
                        #                 continue
                        #             mid_time_fist_array = []
                        #             for time_item_index in range(1,16) :
                        #                 try:
                        #                     if (row[f"attendance_attempt_{time_item_index}"] < kid_mod_stage_1_start_datetime) and \
                        #                         (row[f"attendance_attempt_{time_item_index}"] <  row['last_attendance_attempt']) and \
                        #                         (row[f"attendance_attempt_{time_item_index}"] >  row['attendance_attempt_1']) and \
                        #                         (row[f"attendance_attempt_{time_item_index}"] > row['shift_start_datetime'])and \
                        #                         (row[f"attendance_attempt_{time_item_index}"] < row['shift_end_datetime']) :
                        #                         mid_time_fist_array.append(row[f"attendance_attempt_{time_item_index}"])
                        #                 except Exception as ex:
                        #                     print("con nho ex mid_time_fist_array", ex)
                        #             mid_time_last_array = []
                        #             for time_item_index in range(1,16): 
                        #                 try:
                        #                     if (row[f"attendance_attempt_{time_item_index}"] > kid_mod_stage_2_end_datetime) and \
                        #                         (row[f"attendance_attempt_{time_item_index}"] <  row['last_attendance_attempt']) and \
                        #                         (row[f"attendance_attempt_{time_item_index}"] > row['attendance_attempt_1']) and \
                        #                         (row[f"attendance_attempt_{time_item_index}"] > row['shift_start_datetime'])and \
                        #                         (row[f"attendance_attempt_{time_item_index}"] < row['shift_end_datetime']):
                        #                         mid_time_last_array.append(row[f"attendance_attempt_{time_item_index}"])
                        #                 except Exception as ex:
                        #                     print("con nho ex mid_time_last_array", ex)
                        #             delete_item_array = []
                        #             for out_date_item in out_date_array:
                        #                 for mid_time_fist_item in mid_time_fist_array:
                        #                     try:
                        #                         if (mid_time_fist_item.replace(second=59) >= out_date_item['attendance_missing_from']) and \
                        #                             (mid_time_fist_item.replace(second=0) <= out_date_item['attendance_missing_to']):
                        #                             delete_item_array.append(mid_time_fist_item)
                        #                             print('mid_time_fist_item',out_date_item)
                        #                     except:
                        #                         print('error ra ngoai 1')
                                                
                        #             for delete_item in delete_item_array:
                        #                 mid_time_fist_array.remove(delete_item)
                                        
                        #             delete_item_array = [] 
                        #             for out_date_item in out_date_array:
                        #                 for mid_time_last_item in mid_time_last_array:
                        #                     try:
                        #                         if (mid_time_last_item.replace(second=59) >= out_date_item['attendance_missing_from']) and \
                        #                             (mid_time_last_item.replace(second=0) <= out_date_item['attendance_missing_to']):
                        #                             delete_item_array.append(mid_time_last_item)
                        #                             print('mid_time_fist_item', out_date_item)
                        #                     except Exception as ex:
                        #                         print(mid_time_last_item)
                        #                         print(out_date_item)
                        #                         print('error ra ngoai 2', ex)
                                                
                        #             for delete_item in delete_item_array:
                        #                 mid_time_last_array.remove(delete_item)
                        #                 # input(actual_work_time)
                                    
                        #             # found_late_mid = False
                        #             if len(mid_time_last_array) > 0:
                                      
                        #                 late_mid_time = mid_time_last_array[0]
                        #                 time_late = (late_mid_time - kid_mod_stage_2_end_datetime).total_seconds()//60.0
                        #                 # real_total_work_time = real_total_work_time - time_late
                        #                 # input(time_late)
                        #             if len(mid_time_fist_array) > 0:
                                        
                        #                 late_mid_time = mid_time_fist_array[-1]
                        #                 time_late = (kid_mod_stage_1_start_datetime - late_mid_time).total_seconds()//60.0
                        #                 # input(time_late)

                        #         # if((leave_item['for_reasons']=='2') or (leave_item['for_reasons']==2)) and self.is_calculate_late_early_leave:
                        #         #     time_late_early_for_work += max(leave_item['minutes'], leave_item['time'])
                            
                        for index, leave_item in hr_leave_employee_date.iterrows():
                            
                            # print("--------------------",
                            #     leave_item['holiday_status_name'])
                            try:
                                if ('có tính lương' in leave_item['holiday_status_name'].strip().lower()):
                                    print("aaaaaa'có tính lương'aaaaaaa")
                                    # total_ncl = total_ncl + \
                                    #     max(leave_item['minutes'], leave_item['time'])
                                    # if leave_item['is_leave_probationary']:
                                    #     total_tc_probationary = total_tc_probationary + \
                                    #         max(leave_item['minutes'],
                                    #             leave_item['time'])
                            except Exception as ex:
                                print(ex)

                            try:
                                if (('đi muộn' in leave_item['holiday_status_name'].strip().lower()) or ('về sớm' in leave_item['holiday_status_name'].strip().lower())):
                                    if ((leave_item['for_reasons']=='1') or (leave_item['for_reasons']==1)) and self.is_calculate_late_early_leave:
                                        print("don di muon ve som duoc duyet voi ly do ca nhan trong ngay")
                                        total_dimuon_vesom_date = total_dimuon_vesom_date + \
                                            max(leave_item['minutes'], leave_item['time'])
                                        real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row)
                                        real_total_work_time = real_total_work_time + kid_mode_time - \
                                                (start_date_work_kid_mode + end_date_work_kid_mode)
                                    elif ((leave_item['for_reasons']=='2') or (leave_item['for_reasons']==2)) and self.is_calculate_late_early_leave:
                                        total_dimuon_vesom_work_date = total_dimuon_vesom_work_date + \
                                            max(leave_item['minutes'], leave_item['time'])
                                        time_late_early_for_work = time_late_early_for_work + max(leave_item['minutes'], leave_item['time'])

                                        # if leave_item['is_leave_probationary']:
                                            #     total_dimuon_vesom_probationary = total_dimuon_vesom_probationary + \
                                            #         max(leave_item['minutes'],
                                            #             leave_item['time'])
                            except Exception as ex:
                                print(ex)
                            try:
                                if ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()) and \
                                    ((leave_item['for_reasons']=='1') or (leave_item['for_reasons']==1)):
                                    # duration = leave_item['attendance_missing_to'] - leave_item['attendance_missing_from']
                                    calculate_result = self.calculate_actual_work_time_couple(row, leave_item['attendance_missing_from'], leave_item['attendance_missing_to'])
                                    print("time out date xxxxxxxxxxxxxxxx", calculate_result)
                                    calculate_result_update = self.calculate_actual_work_time_couple(row, row[f'attendance_attempt_1'], leave_item['attendance_missing_to'])
                                    calculate_result_update_total = calculate_result_update['total_work_time']
                                    time_out_date = calculate_result['total_work_time']
                                    # if(row['employee_code']=='APG210922003') and (row['date'].day==30):
                                    #     print(row)
                                    #     print(actual_work_time)
                                    #     print(time_out_date)
                                    #     print(real_total_work_time)
                                    #     exit(1)
                                    # seconds = duration.total_seconds()
                                    # hours = seconds // 3600
                                    # minutes = (seconds % 3600) // 60.0
                                    # time_out_date = minutes + hours*60
                                    
                            except Exception as ex:
                                print(ex)
                                
                            
                            
                        calculate_actual_work_time = actual_work_time
                        convert_overtime = False
                        for index, leave_item in hr_leave_employee_date.iterrows():
                            try:
                                if ('nghỉ bù' in leave_item['holiday_status_name'].strip().lower()):
                                    number_cl_date = number_cl_date + 1
                                    # print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                                    total_compensation_leave_date = max(leave_item['minutes'], leave_item['time'])
                                    # if leave_item['is_leave_probationary']:
                                    #     total_compensation_probationary = total_compensation_probationary + \
                                    #         max(leave_item['minutes'],
                                    #             leave_item['time'])
                                    if (calculate_actual_work_time  + total_compensation_leave_date + total_dimuon_vesom_date != total_shift_work_time):
                                        print('actual_work_time:', actual_work_time)
                                        print("update", f"{calculate_actual_work_time  + total_compensation_leave_date + total_dimuon_vesom_date} <- {total_shift_work_time}")
                                    
                                    if total_compensation_leave_date>0:
                                        print("total ototal_compensation_leave_date: ",total_compensation_leave_date)
                                    update_minute = calculate_actual_work_time + total_compensation_leave_date + total_dimuon_vesom_date - total_shift_work_time
                                    print("update_minute : ", update_minute)
                                    if (update_minute > 0) and (((update_minute<50)  and (update_minute != total_compensation_leave_date)) \
                                                                or (total_dimuon_vesom_date >0)) and (total_compensation_leave_date > update_minute):
                                        update_data = {"minutes": total_compensation_leave_date - update_minute,
                                                    "old_minutes": total_compensation_leave_date}
                                        ids = [leave_item['id']]
                                        logger.info(f"update {update_data} - {ids}")
                                        # try:
                                        #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.leave', 'write', [ids, update_data])
                                        #     total_compensation_leave = total_compensation_leave - update_minute
                                        # except Exception as ex:
                                        #     logger.error(f"hr.leave write error: {ex}")
                                        
                                        # print ('success update minute')
                            except Exception as ex:
                                print(ex)                          
                        
                            try:
                                if ('nghỉ phép năm' in leave_item['holiday_status_name'].strip().lower()):
                                    total_al_date = total_al_date + max(leave_item['minutes'], leave_item['time'])
                                    number_al_date = number_al_date + 1
                                    # print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                                elif ( 'có tính lương' in leave_item['holiday_status_name'].strip().lower()):
                                    total_ncl_date = max(leave_item['minutes'], leave_item['time'])
                                elif ('tăng ca' in leave_item['holiday_status_name'].strip().lower()):
                                    total_tc_date = max(leave_item['minutes'], leave_item['time'])
                                    if (leave_item['convert_overtime']==True):
                                        convert_overtime = True
                                        
                            
                            except Exception as ex:
                                print(ex)
                                
                        # try:
                        #     if (row['is_holiday']== True) and (auto_calculate_holiday == True):
                        #         print("Calculate holiday")
                        #         total_tc_holiday = total_tc_holiday  + \
                        #             int((row['holiday_work_time'] - row['night_holiday_work_time_minute']) * 3 + \
                        #             row['night_holiday_work_time_minute'] * night_holiday_wage)
                        #         if probationary_index:
                        #             total_tc_holiday_probationary = total_tc_holiday_probationary + \
                        #                 int((row['holiday_work_time'] - row['night_holiday_work_time_minute']) * 3 + \
                        #                 row['night_holiday_work_time_minute'] * night_holiday_wage)
                        # except Exception as ex:
                        #     print('auto_calculate_holiday', ex)   
                            
                        if (time_late>0) or (time_out_date >0):
                            real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row) 
                            real_total_work_time = real_total_work_time + kid_mode_time - \
                                                (start_date_work_kid_mode + end_date_work_kid_mode)
                            try:
                                if (time_out_date >0):
                                    real_total_work_time = max(calculate_result_update_total, real_total_work_time)
                            except Exception as ex:
                                pass
                            actual_work_time = real_total_work_time - time_late
                            actual_work_time = max(actual_work_time - time_out_date, 0)
                                
                            # if is_ho:
                                # if(round(actual_work_time,0) != row['total_work_time']):
                                #     if round(actual_work_time,0) in data_total_work_time_update:
                                        
                                #         ids_total_work_time_update[round(actual_work_time, 0)].append(int(row['id']))
                                #     else:
                                #         data_total_work_time_update.append(round(actual_work_time, 0))
                                #         ids_total_work_time_update[round(actual_work_time, 0)] = [int(row['id'])]
                        if convert_overtime:
                            real_total_work_time = 0

                            actual_work_time = 0
                            
                        # if(row['employee_code']=='APG210922003') and (row['date'].day==30):
                        #     print(row)
                        #     print(actual_work_time)
                        #     print(time_out_date)
                        #     print(real_total_work_time)
                        #     exit(1)
                        # actual_work_time += time_late_early_for_work
                        total_late_early_for_work = total_late_early_for_work + time_late_early_for_work
                        shift_name_display = f"{round(actual_work_time,0)}_{shift_name}" if actual_work_time else shift_name
                        #############   start update hr
                        
                        
                        
                        total_normal_working_minutes = total_normal_working_minutes + actual_work_time
                        if probationary_index:
                            total_normal_working_minutes_probationary = total_normal_working_minutes_probationary + actual_work_time
                    
                        if not is_post_inspection:
                            ws.cell(row=start_row, column=start_col +
                                    date.day).value = shift_name_display
                            if pd.notnull(row['probationary_contract_termination_date']):
                                ws.cell(row=start_row, column=6).value = row['probationary_contract_termination_date'].strftime(
                                    '%d-%m-%Y')
                            if pd.notnull(row['probationary_salary_rate']):
                                ws.cell(
                                    row=start_row, column=7).value = row['probationary_salary_rate']/100
                            if col_index == 0:
                                if pd.notnull(row['workingday']):
                                    ws.cell(row=start_row, column=5).value = row['workingday'].strftime(
                                        '%d-%m-%Y')
                                if pd.notnull(row['job_title']):
                                    ws.cell(row=start_row,
                                            column=4).value = row['job_title']
                                try:
                                    
                                    # print(resource_calendar_id)
                                    # Tiêu chuẩn 48 giờ/tuần
                                    ws.cell(row=start_row,
                                                column=8).value = standard_working_time
                                #     if '44' in resource_calendar_id:
                                #         ws.cell(row=start_row,
                                #                 column=8).value = self.working_time_44
                                #     elif '40' in resource_calendar_id:
                                #         ws.cell(row=start_row,
                                #                 column=8).value = self.working_time_40
                                #     else:
                                #         ws.cell(row=start_row,
                                #                 column=8).value = self.working_time_48
                                except Exception as ex:
                                    ws.cell(row=start_row,
                                            column=8).value = self.working_time_48
                                    print("gio lam ex: ", ex)

                        col_index = col_index + 1
                        shift_name = row['shift_name'].strip()
                        if shift_name == 'UP':
                            total_unpaid_leave = total_unpaid_leave + 1
                            if probationary_index:
                                total_unpaid_leave_probationary = total_unpaid_leave_probationary + 1
                        elif ('/UP' in shift_name) or ('UP/' in shift_name):
                            total_unpaid_leave = total_unpaid_leave + 0.5
                            if probationary_index:
                                total_unpaid_leave_probationary = total_unpaid_leave_probationary + 0.5
                        # if shift_name == 'CL':
                        #     total_compensation_leave = total_compensation_leave + 1
                        #     if probationary_index:
                        #         total_compensation_probationary = total_compensation_probationary + 1
                        # elif ('/CL' in shift_name) or ('CL/' in shift_name):
                        #     total_compensation_leave = total_compensation_leave + 0.5
                        #     if probationary_index:
                        #         total_compensation_probationary = total_compensation_probationary + 0.5

                        # if shift_name == 'AL':
                        #     total_al = total_al + 1
                        #     if probationary_index:
                        #         total_al_probationary = total_al_probationary + 1
                        # elif ('/AL' in shift_name) or ('AL/' in shift_name):
                        #     total_al = total_al + 0.5
                        #     if probationary_index:
                        #         total_al_probationary = total_al_probationary + 0.5
                        # 'Ca gay'
                        if (row['fix_rest_time'] == True) and (row['total_work_time']>0):
                            total_fix_rest_time = total_fix_rest_time + 1
                            if probationary_index:
                                total_fix_rest_time_probationary = total_fix_rest_time_probationary + 1

                        if shift_name == 'OFF':
                            list_off_ids.append({'id' : int(row['id']), 'shift_name': row['shift_name'], 'employee_code': row['employee_code'], 'date': row['date']})
                            total_off = total_off + 1
                        elif ('/OFF' in shift_name) or ('OFF/' in shift_name):
                            total_off = total_off + 0.5

                        all_night_hours = row['night_worktime_minute']
                        total_night = total_night + all_night_hours 
                        if probationary_index:
                            total_night_probationary = total_night_probationary + all_night_hours
                        if row['is_holiday']:
                            total_night_holiday = total_night_holiday + row['night_holiday_work_time_minute']
                            if probationary_index:
                                total_night_holiday_probationary = total_night_holiday_probationary + row['night_holiday_work_time_minute']
                            
                        total_ph_date = 0
                        if (shift_name == 'AL') \
                                and (number_cl_date == 0) and (number_al_date == 0):
                            total_al_date = 480
                            total_al = total_al + 480
                            if probationary_index:
                                total_al_probationary = total_al_probationary + 480

                        if (shift_name == 'CL') \
                                and (number_cl_date == 0) and (number_al_date == 0):
                            total_compensation_leave_date = 480
                            total_compensation_leave = total_compensation_leave + 480
                            if probationary_index:
                                total_compensation_probationary = total_compensation_probationary + 480
                                
                        if row['is_holiday'] == True:
                            if shift_name == 'PH':
                                total_ph = total_ph + 1
                                total_ph_date = 1
                                if probationary_index:
                                    total_ph_probationary = total_ph_probationary + 1
                            elif ('/PH' in shift_name) or ('PH/' in shift_name):
                                total_ph_date = max(0.5, (480 - real_total_work_time)/480)
                                total_ph = total_ph + total_ph_date
                                # total_tc_holiday += (1-total_ph_date) * 480
                                if probationary_index:
                                    # total_tc_holiday_probationary += (1-total_ph_date) * 480
                                    total_ph_probationary = total_ph_probationary + total_ph_date
                                    
                            if ('/PH' in shift_name) or ('PH/' in shift_name):
                                total_compensation_leave = total_compensation_leave + real_total_work_time * row['efficiency_factor']
                                if probationary_index:
                                    total_compensation_probationary = total_compensation_probationary + real_total_work_time * row['efficiency_factor']
                                actual_work_time_date = min (total_shift_work_time, 
                                                total_ph_date * 480 + total_ncl_date + 
                                                total_dimuon_vesom_date +
                                                # total_dimuon_vesom_work_date +
                                                total_compensation_leave_date + total_al_date + actual_work_time)
                            else:
                                actual_work_time_date = total_ph_date * 480 + \
                                            total_dimuon_vesom_date + \
                                            total_ncl_date + total_compensation_leave_date + total_al_date + actual_work_time
                        else:
                            actual_work_time_date = total_ph_date * 480 + \
                                            total_dimuon_vesom_date + \
                                            total_ncl_date + total_compensation_leave_date + total_al_date + actual_work_time
                        # try:
                        # if shift_name.strip == 'PH':
                        #     actual_work_time_date = 480
                        update_data = {}
                        is_need_update = False
                        ids = [int(row['id'])]
                        if is_ho:
                            # if(round(actual_work_time_date, 0) != (row['actual_total_work_time'])):
                            #     if round(actual_work_time_date, 0) in data_actual_total_work_time_update:
                            #         ids_actual_total_work_time_update[round(actual_work_time_date,0)].append(int(row['id']))
                            #         # if(row['employee_code']=='APG231002004') and (row['date'].day==23):
                            #         #     print(row)
                            #         #     print(actual_work_time)
                            #         #     print(time_out_date)
                            #         #     print(real_total_work_time)
                            #         #     print(actual_work_time_date)
                            #         #     exit(1)
                            #     else:
                            #         data_actual_total_work_time_update.append(round(actual_work_time_date, 0))
                            #         ids_actual_total_work_time_update[round(actual_work_time_date, 0)] = ids
                                # update_data = {'actual_total_work_time': int(actual_work_time_date)}
                                # is_need_update = True
                        # if is_ho:
                            
                            if pd.notnull(standard_working_time) and pd.notna(standard_working_time):
                                if (standard_working_time>0) and (total_off > len(self.date_array) - standard_working_time + len(ids_off_remove_array)) \
                                    and ('OFF' in row['shift_name']) and (len(list_off_ids)>1):
                                    select_item = list_off_ids[0]
                                    for find_item in list_off_ids:
                                        if(find_item['date']) > select_item['date']:
                                            select_item = find_item
                                    item_copy = select_item.copy()
                                    ids_off_remove_array.append(item_copy)
                                    ids_all_off_remove.append(item_copy)
                                    new_list = []
                                    for find_item in list_off_ids:
                                        if find_item['date'] != select_item['date']:
                                            new_list.append(find_item)
                                    list_off_ids = new_list
                                # if int(standard_working_time)  != int(row['standard_working_day']):
                                    # update_data['standard_working_day'] = int(standard_working_time)
                                    # is_need_update= True
                                    # if int(standard_working_time) in data_standard_working_day_update:
                                    #     if not (int(row['id']) in ids_standard_working_day_update[int(standard_working_time)]):
                                    #         ids_standard_working_day_update[int(standard_working_time)].append(int(row['id']))
                                    # else:
                                    #     data_standard_working_day_update.append(int(standard_working_time))
                                    #     ids_standard_working_day_update[int(standard_working_time)] = ids
                            # else:
                            #     if 0 != row['standard_working_day']:
                            #         if 0 in data_standard_working_day_update:
                            #             if not (int(row['id']) in ids_standard_working_day_update[0]):
                            #                 ids_standard_working_day_update[0].append(int(row['id']))
                            #         else:
                            #             data_standard_working_day_update.append(0)
                            #             ids_standard_working_day_update[0] = ids
                        # if is_ho:
                            if pd.notnull (row['probationary_contract_termination_date']) and pd.notna(row['probationary_contract_termination_date']):
                                if row['probationary_contract_termination_date'] != row['probation_completion_wage']:
                                    try:
                                        update_data_probation_completion_wage = row['probationary_contract_termination_date'].strftime('%Y-%m-%d')
                                        # is_need_update= True
                                        if update_data_probation_completion_wage in data_probation_completion_wage_update:
                                            if not (int(row['id']) in ids_probation_completion_wage_update[update_data_probation_completion_wage]):
                                                ids_probation_completion_wage_update[update_data_probation_completion_wage].append(int(row['id']))
                                        else:
                                            data_probation_completion_wage_update.append(update_data_probation_completion_wage)
                                            ids_probation_completion_wage_update[update_data_probation_completion_wage] = ids
                                    except Exception as ex:
                                        print(ex)
                        # if is_ho:
                            if pd.notnull(row['probationary_salary_rate']) and pd.notna(row['probationary_salary_rate']):
                                if  int(row['probationary_salary_rate']) != int(row['probation_wage_rate']):
                                    probationary_salary_rate_date = int(row['probationary_salary_rate'])
                                    # update_data['probation_wage_rate'] = probationary_salary_rate_date
                                    # is_need_update= True
                                    if ( probationary_salary_rate_date in data_probationwage_rate_update):
                                        if not (int(row['id']) in ids_probationwage_rate_update[probationary_salary_rate_date]):
                                            ids_probationwage_rate_update[probationary_salary_rate_date].append(int(row['id']))
                                    else:
                                        data_probationwage_rate_update.append(probationary_salary_rate_date)
                                        ids_probationwage_rate_update[probationary_salary_rate_date] = ids
                            
                        # except Exception as ex:
                        #     print("Exception ex in actual: ", row)


                # summary data
                # cl thu viec con lai bang cl thu viec thang truoc cong tang ca thu viec tru cl thu viec thang nay
                last_month_al_employee = df_al_report[df_al_report['employee_code']==sub_group_index[1]]
                last_month_cl_employee = df_cl_report[df_cl_report['employee_code']==sub_group_index[1]]
                
                last_month_cl_remaining_leave_minute_probationary = 0
                last_month_cl_remaining_leave_minute = 0
                current_al_remaining_leave_minute = 0
                last_month_al_remaining_leave_minute = 0
                current_cl_remaining_leave_minute = 0
                current_cl_remaining_leave_minute_probationary = 0
                
                if len(last_month_al_employee.index) > 0:
                    last_month_al_remaining_leave_minute = \
                        last_month_al_employee.iloc[0]['remaining_leave_minute'] + \
                        last_month_al_employee.iloc[0][current_month_str] 
                    
                if len(last_month_cl_employee.index) > 0:
                    last_month_cl_remaining_leave_minute = int(last_month_cl_employee.iloc[0]['remaining_total_minute']) + \
                        last_month_cl_employee.iloc[0][f"used_official_{self.date_array[0].month}"] + \
                        last_month_cl_employee.iloc[0][f"used_probationary_{self.date_array[0].month}"]
                        
                    last_month_cl_remaining_leave_minute_probationary = last_month_cl_employee.iloc[0]['remaining_probationary_minute'] + \
                        last_month_cl_employee.iloc[0][f"used_probationary_{self.date_array[0].month}"]
                try:
                    current_al_remaining_leave_minute = max(0,last_month_al_remaining_leave_minute - total_al)
                except:
                    current_al_remaining_leave_minute = 0
                    last_month_al_remaining_leave_minute = 0
                # cl con lai = cl con lai thang truoc cong tang ca cong tang ca tru di tong cl
                try:
                    current_cl_remaining_leave_minute = max(0, last_month_cl_remaining_leave_minute + total_tc - total_compensation_leave)
                except:
                    current_cl_remaining_leave_minute= 0
                    last_month_cl_remaining_leave_minute = 0
                try:
                    current_cl_remaining_leave_minute_probationary = max(0,last_month_cl_remaining_leave_minute_probationary + \
                                                            total_tc_probationary - total_compensation_probationary)
                except:
                    current_cl_remaining_leave_minute_probationary = 0
                    last_month_cl_remaining_leave_minute_probationary = 0 
                total_compensation_leave_dimuon_vesom = 0
                total_al_dimuon_vesom = 0
                if self.is_calculate_late_early_leave:
                
                    # cl di muon ve som = chuyen doi di muon ve som sang cl: gia tri nho nhat cua cl con lai va di muon ve som
                    total_compensation_leave_dimuon_vesom = min(total_dimuon_vesom, current_cl_remaining_leave_minute)

                    # cl di muon ve som chuyen doi sang cl thu viec bang gia tri nho nhat cl thu viec con lai và di muon ve som thu viec
                    total_compensation_dimuon_vesom_probationary = min(total_dimuon_vesom_probationary, current_cl_remaining_leave_minute_probationary)

                    # phep sau khi di muon ve som gia tri nho nhat so di muon ve som tru di phan chuyen doi sang cl va al con lai
                    total_al_dimuon_vesom = min(total_dimuon_vesom - total_compensation_leave_dimuon_vesom, current_al_remaining_leave_minute)
                used_cl =  total_compensation_leave + total_compensation_leave_dimuon_vesom
                used_cl_probationary = total_compensation_probationary + total_compensation_dimuon_vesom_probationary 
                used_cl_offical = used_cl - used_cl_probationary
                
                al_cl_current_month_item = {'used_cl_probationary': used_cl_probationary, 
                                            'used_cl_offical': used_cl_offical,
                                            'total_tc_probationary': total_tc_probationary,
                                            'total_tc_offical': total_tc - total_tc_probationary,
                                            'phatsinhtang_thuviec': phatsinhtang_thuviec,
                                            'phatsinhtang_chinhthuc': phatsinhtang_chinhthuc,
                                            'total_al': total_al_dimuon_vesom + total_al}
                
                self.al_cl_current_month[f"{sub_group_index[1]}"] = al_cl_current_month_item
                if not is_post_inspection:
                    # quy con lai
                    ws.cell(row=start_row, column=73).value = total_late_early_for_work
                    try:
                        ws.cell(row=start_row, column=74).value = last_month_cl_remaining_leave_minute
                        ws.cell(row=start_row, column=75).value = last_month_cl_remaining_leave_minute_probationary
                        ws.cell(row=start_row, column=76).value = last_month_al_remaining_leave_minute 
                    except Exception as ex:
                        print ("last month data err", ex)
                    ws.cell(row=start_row, column=77).value = total_compensation_leave_dimuon_vesom 
                    ws.cell(row=start_row, column=78).value = total_al_dimuon_vesom - total_al_probationary
                    ws.cell(row=start_row, column=79).value = total_dimuon_vesom 
                    ws.cell(row=start_row, column=80).value = current_al_remaining_leave_minute 
                    ws.cell(row=start_row, column=81).value = current_cl_remaining_leave_minute 


                    ws.cell(row=start_row, column=9).value = total_off
                    ws.cell(row=start_row, column=10).value = total_unpaid_leave
                    ws.cell(row=start_row, column=11).value = total_ph_probationary
                    ws.cell(row=start_row, column=12).value = total_tc_probationary / 60.0
                    ws.cell(row=start_row, column=13).value = total_tc_holiday_probationary / 60.0

                    ws.cell(row=start_row, column=14).value = (total_night_probationary - total_night_holiday_probationary) / 60.0
                    ws.cell(row=start_row, column=15).value = total_night_holiday_probationary / 60.0
                    ws.cell(row=start_row, column=16).value = total_fix_rest_time_probationary
                    ws.cell(row=start_row, column=17).value = total_ncl_probationary/480
                    ws.cell(row=start_row, column=18).value = total_compensation_probationary/480 + total_compensation_dimuon_vesom_probationary/480
                            # min(total_compensation_probationary/480 + total_compensation_dimuon_vesom_probationary/480,
                            #                                   last_month_cl_remaining_leave_minute_probationary/480 + total_tc_probationary/480)
                        
                    total_normal_working_minutes = total_normal_working_minutes
                    ws.cell(row=start_row, column=19).value = total_normal_working_minutes_probationary / (8 * 60)

                    ws.cell(row=start_row, column=21).value = total_ph - total_ph_probationary
                    ws.cell(row=start_row, column=22).value = (total_tc - total_tc_probationary) / 60.0
                    ws.cell(row=start_row, column=23).value = (total_tc_holiday - total_tc_holiday_probationary) / 60.0
                    ws.cell(row=start_row, column=24).value = ((total_night - total_night_probationary) - 
                                            (total_night_holiday - total_night_holiday_probationary) ) / 60.0
                    # if sub_group_index[1] == 'APG220801006':
                    #     # input('APG220801006')
                    #     print(total_night_holiday_probationary)
                    #     print('total: ', total_night_holiday)
                    #     input('APG220801006')
                    ws.cell(row=start_row, column=25).value = (total_night_holiday - total_night_holiday_probationary) / 60.0

                    ws.cell(row=start_row, column=26).value = total_fix_rest_time - total_fix_rest_time_probationary
                    ws.cell(row=start_row, column=27).value = (total_al - total_al_probationary)/480 + total_al_dimuon_vesom/480
                                            # min((total_al - total_al_probationary)/480 + total_al_dimuon_vesom/480, \
                                            # last_month_al_remaining_leave_minute)

                    ws.cell(row=start_row, column=30).value = (total_ncl - total_ncl_probationary) / 480
                    ws.cell(row=start_row, column=31).value = (total_compensation_leave - total_compensation_probationary) / 480 + \
                        (total_compensation_leave_dimuon_vesom - total_compensation_dimuon_vesom_probationary) / 480

                        # min((total_compensation_leave - total_compensation_probationary) / 480 + \
                        # (total_compensation_leave_dimuon_vesom - total_compensation_dimuon_vesom_probationary) / 480, 
                        # (last_month_cl_remaining_leave_minute + total_tc) - \
                        # (last_month_cl_remaining_leave_minute_probationary/480 + total_tc_probationary/480))
                    ws.cell(row=start_row, column=32).value = (total_normal_working_minutes - total_normal_working_minutes_probationary) / (8 * 60)
                    total_paid_day = total_ph * 480 + total_ncl + total_compensation_leave + \
                            total_compensation_leave_dimuon_vesom + total_normal_working_minutes + total_al
                    ws.cell(row=start_row, column=79).value = max(total_paid_day - standard_working_time * 480, 0)
                    ws.cell(row=start_row, column=80).value = phatsinhtang_thuviec 
                    ws.cell(row=start_row, column=81).value = phatsinhtang_chinhthuc 
                    # if self.is_update_al_cl:
                    #     try:
                    #         print(total_al_dimuon_vesom)

                    #         ids = [int(last_month_al_employee.iloc[0]['id'])]
                    #         if str(int(last_month_al_employee.iloc[0]['id'])) in self.al_result_update:
                    #             self.al_result_update[str(int(last_month_al_employee.iloc[0]['id']))][current_month_str] = \
                    #                     int(total_al_dimuon_vesom) 
                    #         else:
                    #             update_data_al_current = {}
                    #             if int(last_month_al_employee.iloc[0][current_month_str]) != int(total_al_dimuon_vesom): 
                    #                 # input(f"before - {sub_group_index[1]}")
                    #                 update_data_al_current[current_month_str] = int(total_al_dimuon_vesom) 
                    #                 self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'write', 
                    #                                         [ids, update_data_al_current])
                    #             # input(sub_group_index[1])
                    #     except Exception as ex:
                    #         print(ex)
                    #         logger.error(f"{sub_group_index[1]} - ex in al {ex}")
                    #         # input("ex in al")
                    
                    #     try:
                    #         total_compensation_dimuon_vesom_offical = \
                    #             total_compensation_leave_dimuon_vesom - total_compensation_dimuon_vesom_probationary
                            
                    #         print(total_compensation_dimuon_vesom_offical)
                    #         ids = [int(last_month_cl_employee.iloc[0]['id'])]
                    #         id_cl_str = str(int(last_month_cl_employee.iloc[0]['id']))
                            
                    #         if id_cl_str in self.cl_result_update:
                    #             # if sub_group_index[1] =='':
                    #             #     print(self.cl_result_update[id_cl_str])
                    #             #     input('day roii')
                    #             try:
                    #                 if f"used_official_{self.date_array[0].month}" in self.cl_result_update[id_cl_str]:
                    #                     self.cl_result_update[id_cl_str] \
                    #                         [f"used_official_{self.date_array[0].month}"] = \
                    #                         self.cl_result_update[id_cl_str][f"used_official_{self.date_array[0].month}"] \
                    #                         + int(total_compensation_dimuon_vesom_offical) 
                    #                 else:
                    #                     self.cl_result_update[id_cl_str] \
                    #                         [f"used_official_{self.date_array[0].month}"] = \
                    #                             int(total_compensation_dimuon_vesom_offical) 
                    #                 if f"used_probationary_{self.date_array[0].month}" in self.cl_result_update[id_cl_str]:
                    #                     self.cl_result_update[id_cl_str] \
                    #                         [f"used_probationary_{self.date_array[0].month}"] = \
                    #                         self.cl_result_update[id_cl_str][f"used_probationary_{self.date_array[0].month}"] \
                    #                         + int(total_compensation_dimuon_vesom_probationary) 
                    #                 else:
                    #                     self.cl_result_update[id_cl_str] \
                    #                         [f"used_probationary_{self.date_array[0].month}"] = \
                    #                             int(total_compensation_dimuon_vesom_probationary) 
                                                
                    #             except Exception as ex:
                    #                 logger.error(f"{sub_group_index[1]} - ex in cl {id_cl_str} stage 1 {ex}")
                    #                 # if sub_group_index[1] =='':
                    #                 #     print(self.cl_result_update[id_cl_str])
                    #                 #     input('day roii')
                    #         else:
                    #             update_data_cl_current = {}
                    #             is_update_al_cl = False
                    #             if int(last_month_cl_employee.iloc[0][f"used_official_{self.date_array[0].month}"]) != int(total_compensation_dimuon_vesom_offical): 
                    #                 # input(f"before - {sub_group_index[1]}")
                                    
                    #                 is_update_al_cl = True
                    #             update_data_cl_current[f"used_official_{self.date_array[0].month}"] = int(total_compensation_dimuon_vesom_offical) 
                    #             if int(last_month_cl_employee.iloc[0][f"used_probationary_{self.date_array[0].month}"]) != int(total_compensation_dimuon_vesom_probationary): 
                    #                 # input(f"before - {sub_group_index[1]}")
                                    
                    #                 is_update_al_cl = True
                    #             update_data_cl_current[f"used_probationary_{self.date_array[0].month}"] = int(total_compensation_dimuon_vesom_probationary) 
                    #             if is_update_al_cl:    
                    #                 self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'write', 
                    #                                         [ids, update_data_cl_current])
                    #     except Exception as ex:
                    #         print(ex)
                    #         logger.error(f"{sub_group_index[1]} - ex in cl {ex}")
                    start_row = start_row + 1
                    
        if is_ho and (datetime.datetime.now() - self.date_array[0]).days < 55:
            try:
                df_contract_update = pd.DataFrame.from_dict(ids_contract_company_update)
                df_contract_update.to_excel(f'df_contract_update_{company_name}.xlsx')
                for value in data_contract_company_update:
                    print(f"ids_contract_company_update ids_contract_company_update: {value} - ids: {ids_contract_company_update[value]}")
                    ids = [item['id'] for item in ids_contract_company_update[value]]
                    update_data = {'company': value}
                    self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                    
            except Exception as ex :
                logger.error(f"{sub_group_index[1]} - ex in cl {ex}")
        # if is_ho:
            for value in data_probationwage_rate_update:
                try:
                    print(f"probationwage_rate value: {value} - ids: {ids_probationwage_rate_update[value]}")
                    ids = ids_probationwage_rate_update[value]
                    update_data = {'probation_wage_rate': int(value)}
                    self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                except Exception as ex :
                    logger.error(f"Ex data_probationwage_rate_update {ex}")

            for value in data_probation_completion_wage_update:
                try:
                    print(f"robation_completion_wage value: {value} - ids: {ids_probation_completion_wage_update[value]}")
                    ids = ids_probation_completion_wage_update[value]
                    update_data = {'probation_completion_wage': value}
                    self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                except Exception as ex :
                    logger.error(f"Ex data_probation_completion_wage_update {ex}")
        # # else:
            # for value in data_actual_total_work_time_update:
            #     try:
            #         print(f"actual_total_work_time value: {value} - ids: {ids_actual_total_work_time_update[value]}")
            #         ids = ids_actual_total_work_time_update[value]
            #         update_data = {'actual_total_work_time': int(value)}
            #         self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
            #     except Exception as ex :
            #         logger.error(f"Ex data_actual_total_work_time_update {ex}")

            # for value in data_total_work_time_update:
            #     try:
            #         print(f"total_work_time value: {value} - ids: {ids_total_work_time_update[value]}")
            #         ids = ids_total_work_time_update[value]
            #         update_data = {'total_work_time': int(value)}
            #         # print(data_total_work_time_update)
            #         # print(ids_total_work_time_update)
            #         # exit(1)
            #         self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
            #     except Exception as ex :
            #         logger.error(f"Ex data_total_work_time_update {ex}")
                # exit(1)
            # for value in data_standard_working_day_update:
            #     try:
            #         print(f"standard_working_day value: {value} - ids: {ids_standard_working_day_update[value]}")
            #         ids = ids_standard_working_day_update[value]
            #         update_data = {'standard_working_day': int(value)}
            #         self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
            #     except Exception as ex :
            #         logger.error(f"Ex data_standard_working_day_update {ex}")
        return df_scheduling_ver
    
    def getdepartment(self, row):
        try:
            return row['department_id']
        except:
            return ''
    def export_feed_report(self):
        min_io_results= []
        for group_index, df_scheduling_ver in self.df_scheduling_ver.groupby('company'):
            print("find: ", group_index)
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
            else:
                print('outp df_company_filter', df_company_filter)
                continue
            mis_id = df_company_filter.iloc[0]['mis_id']
            print("find: ", mis_id)
            print(mis_id)
            output_report_folder = ''
       
            for item in REPORT_LINK:
                if item['mis-id'] == mis_id:
                    output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m")) 
 
                    # df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index )]
                    # # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                    # df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                    print('out: ', output_report_folder)
                    if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                        os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                    isExist = os.path.exists(output_report_folder)
                    if not isExist:

                    # Create a new directory because it does not exist
                        os.makedirs(output_report_folder)
                    print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                    media_report_folder = init_media_subfoder_report( f"{item['user']}", self.date_array[0].strftime("%Y-%m"))
                    ref_workbook = self.load_workbook_from_url(url_feed_report)
                    ws_epl = ref_workbook.worksheets[1]
                    ws_department = ref_workbook.worksheets[0]
                    self.export_feed_report_company( df_scheduling_ver, ws_epl, ws_department)
                    print('out: ', output_report_folder)
                    # ref_workbook.save(os.path.join(
                    #     output_report_folder, "BAO_CAO_TONG_HOP_AN_CA.xlsx"))
                    new_file_name = "BAO_CAO_TONG_HOP_AN_CA.xlsx"
                    old_file_path = os.path.join(
                        output_report_folder, new_file_name)
                    ref_workbook.save(old_file_path)
                    service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
                    print("BAO_CAO_TONG_HOP_AN_CA url: ", service_url)
                    year = self.date_array[0].year
                    month = self.date_array[0].month
                    
                    company_id = df_company_filter.iloc[0]['id']
                    try:
                        min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}',filename='report_feed', temp_file_path=old_file_path)
                        min_io_results.append( {'company_name': group_index, \
                                        'month' : self.date_array[0].month, \
                                        'year' : self.date_array[0].year, \
                                        'company_id': company_id, \
                                        'template': '7', \
                                        'department_name': '', 'url': min_io_result['url'], \
                                        'type': f"Báo cáo ca ăn Cong ty {mis_id}"} )
                    except Exception as ex:
                        print(ex)
                    break
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data(df_miniio)
    def export_kpi_weekly_report_ho(self):
        min_io_results = []
        
        index = 0
        output_report_folder = ''
        ref_workbook = self.load_workbook_from_url(url_kpi_weekly_report)

        output_report_folder = os.path.join(self.output_report_folder, f"{self.username}", self.date_array[0].strftime("%Y-%m")) 
        isExist = os.path.exists(output_report_folder)
        if not isExist:
        # Create a new directory because it does not exist
            os.makedirs(output_report_folder)
        media_report_folder = init_media_subfoder_report( f"{self.username}", self.date_array[0].strftime("%Y-%m"))
        list_sheet_name = []
        source = ref_workbook.active
        for group_index, df_kpi_weekly_report in self.df_kpi_weekly_report.groupby('company_name'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]

            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['id']
                print('company_ids: ', self.company_ids)
                if mis_id in self.company_ids:
         
                    
                    
                    target = ref_workbook.copy_worksheet(source)
                  
                    # ws = ref_workbook.worksheets[len(ref_workbook.worksheets)-2]
                    ws = target
                    if (pd.notnull( df_company_filter.iloc[0]['mis_id'])) and not (df_company_filter.iloc[0]['mis_id'] in list_sheet_name):
                        try:
                            ws.title = df_company_filter.iloc[0]['mis_id']
                            list_sheet_name.append(df_company_filter.iloc[0]['mis_id'])
                        except:
                            if len(group_index) > 20:
                                if not (group_index[19:] in list_sheet_name):
                                    ws.title = group_index[19:]
                                    list_sheet_name.append(group_index[19:])
                    

                    
                    if not os.path.exists(os.path.join(self.output_report_folder, f"{self.username}")):
                        os.makedirs(os.path.join(self.output_report_folder, f"{self.username}"))
                    
                    
                    print("Tong hop cham cong")
                    
                    
                    is_ho = True
                    self.export_kpi_weekly_report_company( group_index, df_kpi_weekly_report, ws)
        # isExist = os.path.exists(output_report_folder)  
        # if not isExist:

        # # Create a new directory because it does not exist
        #     os.makedirs(output_report_folder)
        #     print("The new output_report_folder is created! ", output_report_folder)

        new_file_name = "BAO_CAO_TONG_HOP_BAO CAO TUAN.xlsx"
        old_file_path = os.path.join(
            output_report_folder, new_file_name)
        ref_workbook.save(old_file_path)
        service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
        print("Summary tBAO_CAO_TONG_HOP_BAO CAO TUAN. url: ", service_url)
                        # if upload_to_min_io == True:
        year = self.date_array[0].year
        month = self.date_array[0].month
        
        is_save_file = (datetime.datetime.now() - self.date_array[0]).days < 50
        if is_save_file == False:
            return
        try:
            min_io_result = self.save_excel_to_min_io(document_type='ouput_report', subfoder=f'HO/{year}-{month}',filename='kpi_bao_cao_tuan_ho', temp_file_path=old_file_path)
            min_io_results.append( {'company_name': self.company_info['name'], \
                            'month' : self.date_array[0].month, \
                            'year' : self.date_array[0].year, \
                            'company_id': self.company_id, \
                            'template': '12', \
                            'department_name': '', 'url': min_io_result['url'], \
                            'type': f"KPI bao cao tuan HO"} )
        except Exception as ex:
            print(ex)
        # print(self.df_explanation_data)
        # print(self.df_hr_leave)
        # index = index + 1
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)
            self.load_upload_report_data(df_miniio)
                        
    def export_kpi_weekly_report(self):
        min_io_results= []
        for group_index, df_kpi_weekly_report in self.df_kpi_weekly_report.groupby('company_name'):
            print("find: ", group_index)
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
            else:
                print('outp df_company_filter', df_company_filter)
                continue
            mis_id = df_company_filter.iloc[0]['mis_id']
            print("find: ", mis_id)
            print(mis_id)
            output_report_folder = ''
       
            for item in REPORT_LINK:
                if item['user'] == self.username :
                    continue
                if item['mis-id'] == mis_id:
                    output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m")) 
 
                    # df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index )]
                    # # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                    # df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                    print('out: ', output_report_folder)
                    if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                        os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                    isExist = os.path.exists(output_report_folder)
                    if not isExist:

                    # Create a new directory because it does not exist
                        os.makedirs(output_report_folder)
                    print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                    media_report_folder = init_media_subfoder_report( f"{item['user']}", self.date_array[0].strftime("%Y-%m"))
                    ref_workbook = self.load_workbook_from_url(url_kpi_weekly_report)

                    ws = ref_workbook.worksheets[0]
                    self.export_kpi_weekly_report_company(group_index, df_kpi_weekly_report, ws)
                    
                    new_file_name = "BAO_CAO_TONG_HOP_BAO CAO TUAN.xlsx"
                    old_file_path = os.path.join(
                        output_report_folder, new_file_name)
                    ref_workbook.save(old_file_path)
                    service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
                    print("BAO_CAO_TONG_HOP_BAO CAO TUAN: ", service_url)
                    year = self.date_array[0].year
                    month = self.date_array[0].month
                    
                    company_id = df_company_filter.iloc[0]['id']
                    try:
                        min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}',filename='PI_WEEKLY_REPORT', temp_file_path=old_file_path)
                        min_io_results.append( {'company_name': group_index, \
                                        'month' : self.date_array[0].month, \
                                        'year' : self.date_array[0].year, \
                                        'company_id': company_id, \
                                        'template': '12', \
                                        'department_name': '', 'url': min_io_result['url'], \
                                        'type': f"KPI WEEKLY REPORT {mis_id}"} )
                    except Exception as ex:
                        print(ex)
                    break
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data(df_miniio)     
            
    def export_kpi_weekly_report_company(self, company_name, df_kpi_weekly_report, ws):
        min_io_results= []
        start_row = 6
        group_start_row = 5
        list_weeks_in_month = calendar.monthcalendar(self.date_array[0].year, self.date_array[0].month)
        index = 0
        week_days= []
        ws.cell(row= 2, column= 3).value = datetime.datetime.now() 
        ws.cell(row= 3, column= 3).value = company_name
        weerly_report_index =index
        for weeks in list_weeks_in_month:
            fist = 0
            last = 0
            for item in weeks:
                if (item != 0):
                    last = item
                    if (fist == 0):
                        fist = item

            week_days.append([fist, last])
            sunday = datetime.datetime(self.date_array[0].year, \
                                            self.date_array[0].month, \
                                            last )
            monday = datetime.datetime(self.date_array[0].year, \
                                            self.date_array[0].month, \
                                            fist )
            friday = sunday - datetime.timedelta(days=2)
            if (friday.month == (monday + datetime.timedelta(days=4)).month):
                weerly_report_index = index
                    
            ws.cell(row=start_row-1, column=4 + index * 2).value = monday.strftime("%d-%m-%Y")
            ws.cell(row=start_row-1, column=5 + index * 2).value = sunday.strftime("%d-%m-%Y")
            index = index + 1
            
        for group_index, subdata in df_kpi_weekly_report.groupby('department_name'):
            start_row = start_row + 1
            group_start_row = group_start_row + 1
            ws.cell(row=group_start_row, column=2).value = group_index
            print("find: ", group_index)
            for subg, data in subdata.sort_values(['employee_code', 'employee_level']).iterrows():
                ws.cell(row=start_row, column=2).value = data['employee_code']
                ws.cell(row=start_row, column=3).value = data['employee_name']
                if (pd.notnull(data['book_review_compensation_status'])):
                    ws.cell(row=start_row, column=14).value = data['book_review_compensation_status']                               
                # ws.cell(row=start_row, column=4).value = data['compensation_status_week_1'] \
                #     if (data['compensation_status_week_1'] != False) else ''
                # if 'đúng' in f"{data['compensation_status_week_1']}":
                sunday = datetime.datetime(self.date_array[0].year, \
                                        self.date_array[0].month, \
                                        week_days[0][1] )
                
                monday = sunday - datetime.timedelta(days=6)
                monday = monday.replace(hour=0, minute=0, second=0) -  datetime.timedelta(hours = 7)
                sunday = sunday.replace(hour=23, minute=59, second=59) -  datetime.timedelta(hours = 7)
                df_weekly_report_employee = self.hr_weekly_report[(self.hr_weekly_report['employee_code']== data['employee_code']) & \
                                                                (self.hr_weekly_report['create_date'] <= sunday)&
                                                                (self.hr_weekly_report['create_date'] >= monday)]
                
                
                if len(df_weekly_report_employee.index) > 0:
                    data_weekly = df_weekly_report_employee.iloc[len(df_weekly_report_employee)-1]
                    try: 
                        create_time = data_weekly['create_date'] + datetime.timedelta(hours=7)
                    except:
                        create_time = ''
                      
                    if not ('chưa' in f'{data_weekly["state"]}') and pd.notnull(data_weekly["state"]):
                        ws.cell(row=start_row, column=5).value = create_time
                    ws.cell(row=start_row, column=4).value = data_weekly['state']
                else:
                    ws.cell(row=start_row, column=4).value = 'chưa gửi'
                    
                # ws.cell(row=start_row, column=6).value = data['compensation_status_week_2'] \
                #     if (data['compensation_status_week_2'] != False) else ''
                # if 'đúng' in f"{data['compensation_status_week_2']}":
                sunday = datetime.datetime(self.date_array[0].year, \
                                        self.date_array[0].month, \
                                        week_days[1][1] )
                monday = sunday - datetime.timedelta(days=6)
                monday = monday.replace(hour=0, minute=0, second=0) -  datetime.timedelta(hours = 7)
                sunday = sunday.replace(hour=23, minute=59, second=59) -  datetime.timedelta(hours = 7)
                df_weekly_report_employee = self.hr_weekly_report[(self.hr_weekly_report['employee_code']== data['employee_code']) & \
                                                                (self.hr_weekly_report['create_date'] <= sunday)&
                                                                (self.hr_weekly_report['create_date'] >= monday)]
                
                if len(df_weekly_report_employee.index) > 0:
                    data_weekly = df_weekly_report_employee.iloc[len(df_weekly_report_employee)-1]
                    try: 
                        create_time = data_weekly['create_date'] + datetime.timedelta(hours=7)
                    except:
                        create_time = ''
                    
                    if not ('chưa' in f'{data_weekly["state"]}') and pd.notnull(data_weekly["state"]):
                        ws.cell(row=start_row, column=7).value = create_time
                    ws.cell(row=start_row, column=6).value = data_weekly['state']
                else:
                    ws.cell(row=start_row, column=6).value = 'chưa gửi'
                # ws.cell(row=start_row, column=8).value = data['compensation_status_week_3'] \
                #     if (data['compensation_status_week_3'] != False) else ''
                # if 'đúng' in f"{data['compensation_status_week_3']}":
                sunday = datetime.datetime(self.date_array[0].year, \
                                        self.date_array[0].month, \
                                        week_days[2][1] )
                monday = sunday - datetime.timedelta(days=6)
                monday = monday.replace(hour=0, minute=0, second=0) -  datetime.timedelta(hours = 7)
                sunday = sunday.replace(hour=23, minute=59, second=59) -  datetime.timedelta(hours = 7)
                df_weekly_report_employee = self.hr_weekly_report[(self.hr_weekly_report['employee_code']== data['employee_code']) & \
                                                                (self.hr_weekly_report['create_date'] <= sunday)&
                                                                (self.hr_weekly_report['create_date'] >= monday)]
                
                if len(df_weekly_report_employee.index) > 0:
                    data_weekly = df_weekly_report_employee.iloc[len(df_weekly_report_employee)-1]
                    try: 
                        create_time = data_weekly['create_date'] + datetime.timedelta(hours=7)
                    except:
                        create_time = ''
                    if not ('chưa' in f'{data_weekly["state"]}') and pd.notnull(data_weekly["state"]):
                        ws.cell(row=start_row, column=9).value = create_time
                    ws.cell(row=start_row, column=8).value = data_weekly['state']
                else:
                    ws.cell(row=start_row, column=8).value = 'chưa gửi'
                # ws.cell(row=start_row, column=10).value = data['compensation_status_week_4'] \
                #     if (data['compensation_status_week_4'] != False) else ''
                # if 'đúng' in f"{data['compensation_status_week_4']}":
                sunday = datetime.datetime(self.date_array[0].year, \
                                        self.date_array[0].month, \
                                        week_days[3][1] )
                monday = sunday - datetime.timedelta(days=6)
                monday = monday.replace(hour=0, minute=0, second=0) -  datetime.timedelta(hours = 7)
                friday = sunday - datetime.timedelta(days=2)
                
                sunday = sunday.replace(hour=23, minute=59, second=59) -  datetime.timedelta(hours = 7)
                df_weekly_report_employee = self.hr_weekly_report[(self.hr_weekly_report['employee_code']== data['employee_code']) & \
                                                                (self.hr_weekly_report['create_date'] <= sunday)&
                                                                (self.hr_weekly_report['create_date'] >= monday)]
                
                if len(df_weekly_report_employee.index) > 0:
                    data_weekly = df_weekly_report_employee.iloc[len(df_weekly_report_employee)-1]
                    try: 
                        create_time = data_weekly['create_date'] + datetime.timedelta(hours=7)
                    except:
                        create_time = ''
                    try:
                        if(weerly_report_index == 3):
                            ws.cell(row=start_row, column=14).value =data_weekly["state"]
                    except:
                        ws.cell(row=start_row, column=14).value = '-'   
                    # if(friday.month == monday.month):
                    #     ws.cell(row=start_row, column=14).value =data_weekly["state"]   
                        
                    if not ('chưa' in f'{data_weekly["state"]}') and pd.notnull(data_weekly["state"]):
                        ws.cell(row=start_row, column=11).value = create_time
                    ws.cell(row=start_row, column=10).value = data_weekly['state']
                else:
                    ws.cell(row=start_row, column=10).value = 'chưa gửi'
                # ws.cell(row=start_row, column=12).value = data['compensation_status_week_5'] \
                #     if (data['compensation_status_week_5'] != False) else ''
                if (len(week_days) > 4):
                    # print(weeks)
                    if list_weeks_in_month[4][-1] != 0:
                        sunday = datetime.datetime(self.date_array[0].year, \
                                            self.date_array[0].month, \
                                            week_days[4][1] )
                        monday = sunday - datetime.timedelta(days=6)
                    else:
                        monday = datetime.datetime(self.date_array[0].year, \
                                            self.date_array[0].month, \
                                            week_days[4][0] )
                        sunday = monday + datetime.timedelta(days=6)
                    monday = monday.replace(hour=0, minute=0, second=0) -  datetime.timedelta(hours = 7)
                    sunday = sunday.replace(hour=23, minute=59, second=59) -  datetime.timedelta(hours = 7)
                    friday = sunday - datetime.timedelta(days=2)
                    
                    df_weekly_report_employee = self.hr_weekly_report[(self.hr_weekly_report['employee_code']== data['employee_code']) & \
                                                                    (self.hr_weekly_report['create_date'] <= sunday)&
                                                                    (self.hr_weekly_report['create_date'] >= monday)]
                    
                    
                    if len(df_weekly_report_employee.index) > 0:
                        data_weekly = df_weekly_report_employee.iloc[len(df_weekly_report_employee)-1]
                        try:
                            if(weerly_report_index == 4):
                                ws.cell(row=start_row, column=14).value =data_weekly["state"]
                        except:
                            ws.cell(row=start_row, column=14).value = '-'
                        ws.cell(row=start_row, column=12).value = data_weekly['state']
                        try: 
                            
                            create_time = data_weekly['create_date'] + datetime.timedelta(hours=7)
                        except:
                            create_time = ''
                        if not ('chưa' in f'{data_weekly["state"]}') and pd.notnull(data_weekly["state"]):
                            ws.cell(row=start_row, column=13).value = create_time
                    else:
                        ws.cell(row=start_row, column=12).value = 'chưa gửi'
                start_row = start_row + 1    
                group_start_row = group_start_row + 1
                
    def export_feed_report_company(self,df_scheduling_ver,ws_epl, ws_department):
        print("export An ca")
        ws_epl.merge_cells('A2:I2')
        top_left_cell = ws_epl['A2']
        top_left_cell.value = f'Từ ngày: {self.date_array[0].strftime("%d/%m/%Y")} - Đến ngày:{self.date_array[len(self.date_array)-1].strftime("%d/%m/%Y")}'
        
        ws_department.merge_cells('A2:AA2')
        top_left_cell = ws_department['A2']
        top_left_cell.value = f'Từ ngày: {self.date_array[0].strftime("%d/%m/%Y")} - Đến ngày:{self.date_array[len(self.date_array)-1].strftime("%d/%m/%Y")}'
        start_row = 3
        group_start_row = 4
        for g, subdata in df_scheduling_ver.groupby('department'):
            if g==False:
                continue
            # print(subdata)
            department_data = {}
            for date_item in self.date_array:
                dp_item = {}
                dp_item['total_dp_breakfast'] = 0
                dp_item['total_dp_lunch'] = 0
                dp_item['total_dp_night_eat'] = 0
                dp_item['total_dp_dinner'] = 0 
                department_data[date_item.strftime('%Y-%m-%d')] = dp_item
            start_row = start_row + 1
            group_start_row = group_start_row + 1
            ws_department.cell(row=group_start_row, column=2).value = g

            index = 1
            for subg, data in subdata.sort_values(['employee_code', 'date']).groupby('employee_code'):
                # print(data)
                ws_epl.cell(row=start_row, column=1).value = index
                try:
                    ws_epl.cell(row=start_row, column=3).value = data.iloc[0]['employee_name']
                except:
                    ws_epl.cell(row=start_row, column=3).value = ''
                ws_epl.cell(row=start_row, column=2).value = data.iloc[0]['employee_code']
                ws_epl.cell(row=start_row, column=4).value = g
                
                index = index + 1
                total_breakfast = 0
                total_lunch = 0
                total_night_eat = 0
                total_dinner = 0 
                for sub_group_index, sub_group_data in data.iterrows():
                    print(sub_group_data)
                    if sub_group_data['breakfast'] == True:
                        total_breakfast = total_breakfast + 1
                        department_data[sub_group_data['date_str']]['total_dp_breakfast'] = \
                            department_data[sub_group_data['date_str']]['total_dp_breakfast']  +1
                    if sub_group_data['lunch'] == True:
                        total_lunch = total_lunch + 1
                        department_data[sub_group_data['date_str']]['total_dp_lunch'] = \
                            department_data[sub_group_data['date_str']]['total_dp_lunch'] + 1
                    if sub_group_data['dinner'] == True:
                        total_dinner = total_dinner + 1
                        department_data[sub_group_data['date_str']]['total_dp_dinner'] = \
                            department_data[sub_group_data['date_str']]['total_dp_dinner'] + 1
                    if sub_group_data['night_eat'] == True:
                        total_night_eat = total_night_eat + 1
                        department_data[sub_group_data['date_str']]['total_dp_night_eat'] = \
                            department_data[sub_group_data['date_str']]['total_dp_night_eat'] + 1
                ws_epl.cell(row=start_row, column=5).value = total_breakfast
                ws_epl.cell(row=start_row, column=6).value = total_lunch
                ws_epl.cell(row=start_row, column=7).value = total_dinner
                ws_epl.cell(row=start_row, column=8).value = total_night_eat
                start_row = start_row + 1

                #     ws_epl.cell(row=start_row, column=1).value = index
                #     try:
                #         ws_epl.cell(row=start_row, column=3).value = sub_group_data['employee_name']
                #     except:
                #         ws_epl.cell(row=start_row, column=3).value = ''
                #     ws_epl.cell(row=start_row, column=2).value = sub_group_data['employee_code']
            # save file
            for date_item in self.date_array:
                dp_item = department_data[date_item.strftime('%Y-%m-%d')] 
                ws_department.cell(row=group_start_row, column=date_item.day*4 -1).value = dp_item['total_dp_breakfast']
                ws_department.cell(row=group_start_row, column=date_item.day*4).value = dp_item['total_dp_lunch']
                ws_department.cell(row=group_start_row, column=date_item.day*4 + 1).value = dp_item['total_dp_night_eat']
                ws_department.cell(row=group_start_row, column=date_item.day*4 + 2).value = dp_item['total_dp_dinner'] 
                    
    def export_attendence_scheduling_report(self):
        min_io_results=[]
        for group_index, df_scheduling_ver in self.df_scheduling_ver.groupby('company'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            found = False
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                for item in REPORT_LINK:
                    if item['mis-id'] == mis_id:
                        output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m")) 
                        found= True
                        df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index )]
                        # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                        df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                        print('out: ', output_report_folder)
                        if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                            os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:

                        # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                        print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                        media_report_folder = init_media_subfoder_report( f"{item['user']}", self.date_array[0].strftime("%Y-%m"))
                        ref_workbook = self.load_workbook_from_url(url_attendace_count)
                        ws = ref_workbook.worksheets[0]
                        source = ref_workbook.worksheets[1]
                        self.export_attendence_scheduling_report_company( df_scheduling_ver, df_hr_leave, ws)
                        print("Save file")
                        # ref_workbook.save(os.path.join(
                        #     output_report_folder, "CHAM_CONG_DEM_SO_LAN.xlsx"))
                        new_file_name = "CHAM_CONG_DEM_SO_LAN.xlsx"
                        old_file_path = os.path.join(
                            output_report_folder, new_file_name)
                        ref_workbook.save(old_file_path)
                        service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
                        print("CHAM_CONG_DEM_SO_LAN url: ", service_url)
                        year = self.date_array[0].year
                        month = self.date_array[0].month
                        
                        company_id = df_company_filter.iloc[0]['id']
                        try:
                            min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}',filename='attendance_scheduling', temp_file_path=old_file_path)
                            min_io_results.append( {'company_name': group_index, \
                                            'month' : self.date_array[0].month, \
                                            'year' : self.date_array[0].year, \
                                            'company_id': company_id, \
                                            'template': '2', \
                                            'department_name': '', 'url': min_io_result['url'], \
                                            'type': f"Báo cáo chấm công đếm số lần Cong ty {mis_id}"} )
                        except Exception as ex:
                            print(ex)
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data(df_miniio)    
    def export_attendence_scheduling_report_company(self, df_scheduling_ver, df_hr_leave, ws):
        """
          load data and save to file unity /template/Chấm công đếm số lần.xlsx
        """
        print("export Chấm công đếm số lần")
        
        # df_scheduling_group.to_excel("acx.xlsx",sheet_name = 'Sheet1')
        start_row = 2
        for index in range(1, 11):
            col_name = f'attendance_attempt_{index}'
            df_scheduling_ver[col_name] = pd.to_datetime(
                df_scheduling_ver[col_name], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        col_name = f'last_attendance_attempt'
        df_scheduling_ver[col_name] = pd.to_datetime(
            df_scheduling_ver[col_name], format="%Y-%m-%d %H:%M:%S", errors='coerce')
        week = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
        previous_name = ''

        # self.df_scheduling_ver= self.df_scheduling_ver.dropna(subset='department_name', inplace=False)
        for g, subdata in df_scheduling_ver[(df_scheduling_ver['department_name'].notnull())
                                                 & (df_scheduling_ver['department_name'] != False)
                                                 & (df_scheduling_ver['workingday'].notnull())
                                                 & (df_scheduling_ver['workingday'] != False)].sort_values(['employee_code', 'date']).groupby('department_name'):
            
            count = 1
            # print(g)
            ws.cell(row=start_row, column=1).value = g
            start_row = start_row + 1
            group_start_row = 3
            previous_name = ''
            for subg, data in subdata.sort_values(['employee_code', 'date']).groupby('employee_code'):
                if (data.iloc[0]['severance_day'] !=False) and (pd.notnull (data.iloc[0]['severance_day'])) and \
                        data.iloc[0]['severance_day'] < self.date_array[0]:
                    continue
                for index, row in data.iterrows():
                    # ws.cell(row=start_row + date_item.day -1, column = 1).value = row['time_keeping_code']
                    ws.cell(
                        row=start_row + row['date'].day - 1, column=2).value = row['employee_code']
                    ws.cell(
                        row=start_row + row['date'].day - 1, column=3).value = row['employee_name']
                    # ws.cell(row=start_row + date_item.day -1, column = 4).value = row['CHỨC VỤ TIẾNG VIỆT']
                    ws.cell(
                        row=start_row + row['date'].day - 1, column=4).value = row['department']
                    
                    # try:
                    # # date join
                    #     if not pd.isnull(row['NGÀY NHẬN VIỆC TẠI CÔNG TY']):
                    #         ws.cell(row=start_row + date_item.day - 1, column = 5).value =row['NGÀY NHẬN VIỆC TẠI CÔNG TY'].strftime('%d/%m/%Y')
                    #     else:
                    #         ws.cell(row=start_row + date_item.day - 1, column = 5).value = 'N/A'
                    # except Exception as ex:
                    #     print("ex")
                    #
                    ws.cell(row=start_row + row['date'].day - 1,
                            column=6).value = row['date'].strftime('%d/%m/%Y')
                    ws.cell(row=start_row + row['date'].day - 1,
                            column=7).value = week[row['date'].weekday()]
                    ws.cell(row=start_row +
                            row['date'].day - 1, column=8).value = '-'
                    ws.cell(row=start_row +
                            row['date'].day - 1, column=9).value = '-'

                    # print(
                    #     f"{row['date']}-{row['date'].day + start_row}-{row['employee_name']}-{row['shift_name']}")
                    start_col = 10
                    for index in range(1, 11):
                        if pd.notnull(row[f'attendance_attempt_{index}']) and row[f'attendance_attempt_{index}'] != False:
                            try:
                                ws.cell(row=start_row + row['date'].day - 1, column=start_col).value = \
                                    row[f'attendance_attempt_{index}'].strftime(
                                        '%H:%M')
                            except Exception as ex:
                                print(ex)
                        else:
                            ws.cell(
                                row=start_row + row['date'].day - 1, column=start_col).value = ''
                        start_col = start_col + 1

                    if pd.notnull(row[f'last_attendance_attempt']) and row[f'last_attendance_attempt'] != False:
                        ws.cell(row=start_row + row['date'].day - 1, column=start_col-1).value = \
                            row[f'last_attendance_attempt'].strftime('%H:%M')
                        
                        attendance_late =( row[f'attendance_attempt_1'] - row['shift_start_datetime']).total_seconds() / 60.0
                        ws.cell(row=start_row + row['date'].day - 1, column=start_col+4).value = max(0, attendance_late)

                        leave_early = (row['shift_end_datetime']-row[f'last_attendance_attempt']).total_seconds() / 60.0
                        ws.cell(row=start_row + row['date'].day - 1, column=start_col+5).value = max(0, leave_early)
                            
                    else:
                        ws.cell(
                            row=start_row + row['date'].day - 1, column=start_col-1).value = ''

                    ws.cell(
                        row=start_row + row['date'].day - 1, column=9).value = row['shift_name']
                    ws.cell(
                        row=start_row + row['date'].day - 1, column=8).value = row['total_work_time']

                    ws.cell(
                        row=start_row + row['date'].day - 1, column=30).value = row['total_work_time']   
                    if 'PH' in row['shift_name']:
                        real_total_work_time, diff_to_total = self.calculate_real_total_work_time(row)
                        # real_total_work_time = real_total_work_time + kid_mode_time - \
                        #                         (start_date_work_kid_mode + end_date_work_kid_mode)
                        ws.cell(
                            row=start_row + row['date'].day - 1, column=31).value = real_total_work_time
                        ws.cell(
                            row=start_row + row['date'].day - 1, column=32).value = diff_to_total
                start_row = start_row + len(self.date_array)
        # for g, subdata in df_scheduling_ver[(df_scheduling_ver['department_name'].notnull())
        #                                          & (df_scheduling_ver['department_name'] != False)
        #                                          & (df_scheduling_ver['workingday'].notnull())
        #                                          & (df_scheduling_ver['workingday'] != False)].sort_values(['employee_code', 'date']).groupby('department_name'):

        #     taget = ref_workbook.copy_worksheet(source)
        #     taget.title = g if len(g)<30 else g[:29]
        #     count = 1
        #     # print(g)
        #     group_start_row = 3
        #     previous_name = ''

        #     for subg, data in subdata.sort_values(['employee_code', 'date']).groupby('employee_code'):
        #         start_row = 0
        #         for index, row in data.iterrows():
        #             # ws.cell(row=start_row + date_item.day -1, column = 1).value = row['time_keeping_code']
                 
        #             if start_row == 0:
        #                 taget.cell(
        #                     row=group_start_row , column=2).value = row['employee_code']
        #                 taget.cell(
        #                     row=group_start_row , column=13).value = row['employee_name']
        #                 # ws.cell(row=start_row + date_item.day -1, column = 4).value = row['CHỨC VỤ TIẾNG VIỆT']
        #                 taget.cell(
        #                     row=group_start_row , column=24).value = row['department']
                    
        #             # group Area
        #             taget.cell(
        #                     row=group_start_row + 1, column=  row['date'].day * 12 - 9).value = row['shift_name']
        #             taget.cell(
        #                     row=group_start_row + 2, column=  row['date'].day * 12 - 9).value = row['shift_start']
        #             taget.cell(
        #                     row=group_start_row + 2, column=  row['date'].day * 12 - 7).value = row['shift_end']
        #             if pd.notnull(row[f'last_attendance_attempt']) and row[f'last_attendance_attempt'] != False:
        #                 taget.cell(
        #                     row=group_start_row + 2, column=  row['date'].day * 12).value = row[f'last_attendance_attempt'].strftime('%H:%M')

        #             if pd.notnull(row['attendance_attempt_1']) and row[f'attendance_attempt_1'] != False:     
        #                 taget.cell(
        #                     row=group_start_row + 2, column=row['date'].day * 12 - 2 ).value = row[f'attendance_attempt_1'].strftime('%H:%M')                  
        #             taget.cell(
        #                     row=group_start_row + 1, column=row['date'].day * 12 - 2 ).value = f"{row['leave_early']}"
        #             taget.cell(
        #                     row=group_start_row + 1, column=row['date'].day * 12 ).value =f"{row['attendance_late']}"
        #             taget.cell(
        #                     row=group_start_row + 3, column=row['date'].day * 12 + 1 ).value = row['total_work_time']
        #             taget.cell(
        #                     row=group_start_row + 4, column=row['date'].day * 12 - 8 ).value = f"{row['actual_total_work_time']}"
        #             # print(row['date_str'])
        #             # print(row['employee_code'])
        #             hr_leave_employee_date = df_hr_leave[(df_hr_leave['employee_code']
        #                                              == row['employee_code']) & \
        #                                             (df_hr_leave['date_str'] == row['date_str']) ]
        #             # df_sum_leave = hr_leave_employee_date.groupby(['holiday_status_name','state'])['minutes'].sum()
        #             # df_size_leave = hr_leave_employee_date.groupby(['holiday_status_name','state']).size()
        #             # print(hr_leave_employee_date)
        #             sum_minute_all_cl_leave = 0
        #             count_cl_leave = 0
        #             sum_minute_validate_cl_leave = 0
        #             sum_minute_all_al_leave = 0
        #             count_al_leave = 0
        #             sum_minute_validate_al_leave = 0

        #             sum_minute_all_tc_leave = 0
        #             count_tc_leave = 0
        #             sum_minute_validate_tc_leave = 0
                    
        #             sum_minute_all_m_leave = 0
        #             count_m_leave = 0
        #             sum_minute_validate_m_leave = 0

        #             sum_minute_all_s_leave = 0
        #             count_s_leave = 0
        #             sum_minute_validate_s_leave = 0

        #             check_child = False
                    
        #             for leave_index, leave_data in hr_leave_employee_date.iterrows():
        #                 if ('bù' in leave_data['holiday_status_name']):
        #                     count_cl_leave += 1
        #                     sum_minute_all_cl_leave += leave_data['minutes']
        #                     if leave_data['state'] == 'validate':
        #                         sum_minute_validate_cl_leave += leave_data['minutes']
        #                 elif ('phép năm' in leave_data['holiday_status_name']):
        #                     count_al_leave +=1
        #                     sum_minute_all_al_leave += leave_data['minutes']
        #                     if leave_data['state'] == 'validate':
        #                         sum_minute_validate_al_leave += leave_data['minutes']
        #                 elif ('tăng ca' in leave_data['holiday_status_name'].lower()):
        #                     count_tc_leave +=1
        #                     sum_minute_all_tc_leave += leave_data['minutes']
        #                     if leave_data['state'] == 'validate':
        #                         sum_minute_validate_tc_leave += leave_data['minutes']
        #                 elif ('muộn' in leave_data['holiday_status_name'].lower()):
        #                     count_m_leave +=1
        #                     sum_minute_all_m_leave += leave_data['minutes']
        #                     if leave_data['state'] == 'validate':
        #                         sum_minute_validate_m_leave += leave_data['minutes']
        #                 elif ('sớm' in leave_data['holiday_status_name'].lower()):
        #                     count_s_leave +=1
        #                     sum_minute_all_s_leave += leave_data['minutes']
        #                     if leave_data['state'] == 'validate':
        #                         sum_minute_validate_s_leave += leave_data['minutes']
        #                 elif (('con nhỏ' in leave_data['holiday_status_name'].lower() ) and 
        #                         (leave_data['state'] == 'validate')):
        #                     check_child = True
                        
        #             taget.cell(
        #                     row=group_start_row + 4, column=  row['date'].day * 12 - 3 ).value = sum_minute_all_cl_leave
        #             taget.cell(
        #                     row=group_start_row + 4, column=  row['date'].day * 12 - 4 ).value = sum_minute_validate_cl_leave
        #             taget.cell(
        #                     row=group_start_row + 4, column=  row['date'].day * 12 - 2 ).value = count_cl_leave

        #             taget.cell(
        #                     row=group_start_row + 5, column=  row['date'].day * 12 - 3 ).value = sum_minute_all_al_leave
        #             taget.cell(
        #                     row=group_start_row + 5, column=  row['date'].day * 12 - 4 ).value = sum_minute_validate_al_leave
        #             taget.cell(
        #                     row=group_start_row + 5, column=  row['date'].day * 12 - 2 ).value = count_al_leave
                        
        #             taget.cell(
        #                     row=group_start_row + 6, column=  row['date'].day * 12 - 3 ).value = sum_minute_all_tc_leave
        #             taget.cell(
        #                     row=group_start_row + 6, column=  row['date'].day * 12 - 4 ).value = sum_minute_validate_tc_leave
        #             taget.cell(
        #                     row=group_start_row + 6, column=  row['date'].day * 12 - 2 ).value = count_tc_leave
                    
        #             taget.cell(
        #                     row=group_start_row + 7, column=  row['date'].day * 12 - 3 ).value = sum_minute_all_m_leave
        #             taget.cell(
        #                     row=group_start_row + 7, column=  row['date'].day * 12 - 4 ).value = sum_minute_validate_m_leave
        #             taget.cell(
        #                     row=group_start_row + 7, column=  row['date'].day * 12 - 2 ).value = count_m_leave
                    
        #             taget.cell(
        #                     row=group_start_row + 8, column=  row['date'].day * 12 - 3 ).value = sum_minute_all_s_leave
        #             taget.cell(
        #                     row=group_start_row + 8, column=  row['date'].day * 12 - 4 ).value = sum_minute_validate_s_leave
        #             taget.cell(
        #                     row=group_start_row + 8, column=  row['date'].day * 12 - 2 ).value = count_s_leave
        #             if check_child == True:
        #                 taget.cell(row=group_start_row + 9, column=  row['date'].day * 12 - 1 ).value = 'x'

        #         start_row = start_row + 1
        #         group_start_row =  group_start_row + 14
            # save file
    def calculate_real_total_work_time(self, row):
        result = row['total_work_time']
        if (pd.notnull(row[f'attendance_attempt_1'])) and (pd.notnull(row[f'last_attendance_attempt'])):
            calculate_result = self.calculate_actual_work_time_couple(row, row[f'attendance_attempt_1'], row[f'last_attendance_attempt'])
            result = calculate_result['total_work_time']
        return result, result - row['total_work_time']
    
    def export_al_cl_report_ho(self, al_sheet_name='AL', cl_sheet_name='CL'):
        min_io_results = []
        
        index = 0
        output_report_folder = ''        
        al_sheet_name='AL' 
        cl_sheet_name='CL'
        # continue
        ref_workbook = self.load_workbook_from_url(url_al_cl_tracking_ho)
        # current_month = self.date_array[len(self.date_array)-1].month - 1
        print(ref_workbook.sheetnames)
        index = 0
        cl_index = 0
        al_index = 0
        for sheet_name in ref_workbook.sheetnames:
            print(sheet_name)
            if al_sheet_name == sheet_name:
                al_sheet_name = sheet_name
                al_index = index
            if cl_sheet_name == sheet_name:
                cl_sheet_name = sheet_name
                cl_index = index
            index = index + 1

        source_ws_al = ref_workbook.worksheets[al_index]
        source_ws_cl = ref_workbook.worksheets[cl_index]
        
        for group_index, df_kpi_weekly_report in self.df_kpi_weekly_report.groupby('company_name'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            df_al_report = self.df_al_report[self.df_al_report['company_name']== group_index]
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['id']
                print('company_ids: ', self.company_ids)
                if mis_id in self.company_ids:
                    try:
                        output_report_folder = os.path.join(self.output_report_folder, f"{self.username}", self.date_array[0].strftime("%Y-%m")) 
                        found= True
                        df_hr_leave = self.df_hr_leave[self.df_hr_leave['company_name']==group_index]
                        df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                        df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                        print('out al cl: ', output_report_folder)
                        if not os.path.exists(os.path.join(self.output_report_folder, f"{self.username}")):
                            os.makedirs(os.path.join(self.output_report_folder, f"{self.username}"))
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:

                        # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                        # print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                        # print('al_result', al_result)
                        # try:
                        df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index] 
                        max_calulate_leave = df_cl_report['date_calculate'].max()
                        df_cl_report =  df_cl_report[df_cl_report['date_calculate']==max_calulate_leave]
                        # print('df_cl_report: ', df_cl_report)
                        # print("AL max: ", al_result[('date_calculate_leave','max')][group_index])
                        
                        df_al_report = self.df_al_report[(self.df_al_report['company_name']==group_index )]
                                                        
                        max_calulate_leave =  df_al_report['date_calculate_leave'].max()
                        df_al_report = df_al_report[df_al_report['date_calculate_leave']==max_calulate_leave]
                        
                        df_cl_report.to_excel(os.path.join(output_report_folder, 'df_cl_report.xlsx'))
                        df_al_report.to_excel(os.path.join(output_report_folder, 'df_al_report.xlsx'))
                        print(output_report_folder)
                        print(df_cl_report)
                        media_report_folder = init_media_subfoder_report( f"{self.username}", self.date_array[0].strftime("%Y-%m"))
                        # target = ref_workbook.copy_worksheet(source)
                        ws_al = ref_workbook.copy_worksheet(source_ws_al)
                        ws_cl = ref_workbook.copy_worksheet(source_ws_cl)
                        ws_al.title = f"{df_company_filter.iloc[0]['mis_id']} AL"
                        ws_cl.title = f"{df_company_filter.iloc[0]['mis_id']} CL"
                        self.export_al_cl_report_company_version2(df_al_report = df_al_report, 
                                                                df_cl_report = df_cl_report,
                                                                ws_al=ws_al, ws_cl=ws_cl)
                        
                    except Exception as ex:
                        print(f"export_al_cl_report {mis_id} - {ex}")
        # isExist = os.path.exists(output_report_folder)  
        # if not isExist:

        # # Create a new directory because it does not exist
        #     os.makedirs(output_report_folder)
        #     print("The new output_report_folder is created! ", output_report_folder)
        new_file_name = "Theo doi ngay phep bu.xlsx"
        old_file_path = os.path.join(
            output_report_folder, new_file_name)
        ref_workbook.save(old_file_path)
        service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
        print("Summary PHEPBU. url: ", service_url)
                        # if upload_to_min_io == True:
        year = self.date_array[0].year
        month = self.date_array[0].month
        
        is_save_file = (datetime.datetime.now() - self.date_array[0]).days < 50
        if is_save_file == False:
            return
        try:
            min_io_result = self.save_excel_to_min_io(document_type='ouput_report', subfoder=f'HO/{year}-{month}',filename='AL_CL', temp_file_path=old_file_path)
            min_io_results.append( {'company_name': self.company_info['name'], \
                            'month' : self.date_array[0].month, \
                            'year' : self.date_array[0].year, \
                            'company_id': self.company_id, \
                            'template': '5', \
                            'department_name': '', 'url': min_io_result['url'], \
                            'type': f"BAO_CAO_PHEP_BU_HO"} )
        except Exception as ex:
            print(ex)
        # print(self.df_explanation_data)
        # print(self.df_hr_leave)
        # index = index + 1
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)
            self.load_upload_report_data(df_miniio)
    
    def export_al_cl_report(self, al_sheet_name='AL', cl_sheet_name='CL', process_callback=None):
        min_io_results=[]
        for group_index, df_al_report in self.df_al_report.groupby('company_name'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            
            found = False
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                for item in REPORT_LINK:
                    if item['user'] == self.username :
                        continue
                    if item['mis-id'] == mis_id:
                        try:
                            output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m")) 
                            found= True
                            df_hr_leave = self.df_hr_leave[self.df_hr_leave['company_name']==group_index]
                            df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                            df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                            print('out al cl: ', output_report_folder)
                            if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                                os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                            isExist = os.path.exists(output_report_folder)
                            if not isExist:

                            # Create a new directory because it does not exist
                                os.makedirs(output_report_folder)
                            # print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                            # print('al_result', al_result)
                            # try:
                            df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index] 
                            max_calulate_leave = df_cl_report['date_calculate'].max()
                            df_cl_report =  df_cl_report[df_cl_report['date_calculate']==max_calulate_leave]
                            # print('df_cl_report: ', df_cl_report)
                            # print("AL max: ", al_result[('date_calculate_leave','max')][group_index])
                            
                            df_al_report = self.df_al_report[(self.df_al_report['company_name']==group_index )]
                                                            
                            max_calulate_leave =  df_al_report['date_calculate_leave'].max()
                            df_al_report = df_al_report[df_al_report['date_calculate_leave']==max_calulate_leave]
                            
                            df_cl_report.to_excel(os.path.join(output_report_folder, 'df_cl_report.xlsx'))
                            df_al_report.to_excel(os.path.join(output_report_folder, 'df_al_report.xlsx'))
                            print(output_report_folder)
                            print(df_cl_report)
                            media_report_folder = init_media_subfoder_report( f"{item['user']}", self.date_array[0].strftime("%Y-%m"))
                            
                            al_sheet_name='AL' 
                            cl_sheet_name='CL'
                            # continue
                            ref_workbook = self.load_workbook_from_url(url_al_cl_tracking)
                            current_month = self.date_array[len(self.date_array)-1].month - 1
                            print(ref_workbook.sheetnames)
                            index = 0
                            cl_index = 0
                            al_index = 0
                            for sheet_name in ref_workbook.sheetnames:
                                print(sheet_name)
                                if al_sheet_name == sheet_name:
                                    al_sheet_name = sheet_name
                                    al_index = index
                                if cl_sheet_name == sheet_name:
                                    cl_sheet_name = sheet_name
                                    cl_index = index
                                index = index + 1

                            ws_al = ref_workbook.worksheets[al_index]
                            ws_cl = ref_workbook.worksheets[cl_index]

                            self.export_al_cl_report_company(df_al_report = df_al_report, 
                                                                    df_cl_report = df_cl_report,
                                                                    ws_al=ws_al, ws_cl=ws_cl)
                            

                            new_file_name = "Theo doi ngay phep bu.xlsx"
                            old_file_path = os.path.join(
                                output_report_folder, new_file_name)
                            ref_workbook.save(old_file_path)
                            service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
                            print("Theo doi phep bu url: ", service_url)
                            year = self.date_array[0].year
                            month = self.date_array[0].month
                            
                            company_id = df_company_filter.iloc[0]['id']
                            try:
                                min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}',filename='al_cl', temp_file_path=old_file_path)
                                min_io_results.append( {'company_name': group_index, \
                                                'month' : self.date_array[0].month, \
                                                'year' : self.date_array[0].year, \
                                                'company_id': company_id, \
                                                'template': '5', \
                                                'department_name': '', 'url': min_io_result['url'], \
                                                'type': f"Báo cáo phep bu Cong ty {mis_id}"} )
                            except Exception as ex:
                                print(ex)
                            break
                        except Exception as ex:
                            print(f"export_al_cl_report {mis_id} - {ex}")
            # except Exception as ex:
            #     print(ex)
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data(df_miniio)
            
    def export_al_cl_report_severance(self, al_sheet_name='AL', cl_sheet_name='CL', process_callback=None):
        min_io_results=[]
        for group_index, df_al_report in self.df_al_report.groupby('company_name'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            
            found = False
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                # for item in REPORT_LINK:
                # if item_user== self.username :
                #     continue
                # if item_mis_id == mis_id:
                try:
                    isExist = os.path.exists(os.path.join(self.output_report_folder, f"{mis_id}m"))
                    if not isExist:

                    # Create a new directory because it does not exist
                        os.makedirs(os.path.join(self.output_report_folder, f"{mis_id}m"))
                    
                    output_report_folder = os.path.join(self.output_report_folder, f"{mis_id}m", self.date_array[0].strftime("%Y-%m")) 
                    isExist = os.path.exists(output_report_folder)
                    if not isExist:

                    # Create a new directory because it does not exist
                        os.makedirs(output_report_folder)
                    
                    df_hr_leave = self.df_hr_leave[self.df_hr_leave['company_name']==group_index]
                    df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                    # df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                    print('out al cl: ', output_report_folder)
                    if not os.path.exists(os.path.join(self.output_report_folder, f"{mis_id}m")):
                        os.makedirs(os.path.join(self.output_report_folder, f"{mis_id}m"))
                    isExist = os.path.exists(output_report_folder)
                    if not isExist:

                    # Create a new directory because it does not exist
                        os.makedirs(output_report_folder)
                    # print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                    # print('al_result', al_result)
                    # try:
                    df_cl_report = self.df_cl_report_severance[self.df_cl_report_severance['company_name']==group_index] 
                    max_calulate_leave = df_cl_report['date_calculate'].max()
                    df_cl_report =  df_cl_report[df_cl_report['date_calculate']==max_calulate_leave]
                    # print('df_cl_report: ', df_cl_report)
                    # print("AL max: ", al_result[('date_calculate_leave','max')][group_index])
                    
                    df_al_report = self.df_al_report_severance[(self.df_al_report_severance['company_name']==group_index )]
                                                    
                    max_calulate_leave =  df_al_report['date_calculate_leave'].max()
                    df_al_report = df_al_report[df_al_report['date_calculate_leave']==max_calulate_leave]
                    
                    df_cl_report.to_excel(os.path.join(output_report_folder, 'df_cl_report_severance.xlsx'))
                    df_al_report.to_excel(os.path.join(output_report_folder, 'df_al_report_severance.xlsx'))
                    print(output_report_folder)
                    print(df_cl_report)
                    media_report_folder = init_media_subfoder_report( f"{mis_id}m", self.date_array[0].strftime("%Y-%m"))
                    
                    al_sheet_name='AL' 
                    cl_sheet_name='CL'
                    # continue
                    ref_workbook = self.load_workbook_from_url(url_al_cl_tracking)
                    current_month = self.date_array[len(self.date_array)-1].month - 1
                    print(ref_workbook.sheetnames)
                    index = 0
                    cl_index = 0
                    al_index = 0
                    for sheet_name in ref_workbook.sheetnames:
                        print(sheet_name)
                        if al_sheet_name == sheet_name:
                            al_sheet_name = sheet_name
                            al_index = index
                        if cl_sheet_name == sheet_name:
                            cl_sheet_name = sheet_name
                            cl_index = index
                        index = index + 1

                    ws_al = ref_workbook.worksheets[al_index]
                    ws_cl = ref_workbook.worksheets[cl_index]

                    self.export_al_cl_report_company(df_al_report = df_al_report, 
                                                            df_cl_report = df_cl_report,
                                                            ws_al=ws_al, ws_cl=ws_cl, client_compute=True)
                    

                    new_file_name = "Theo doi ngay phep bu nghi viec.xlsx"
                    isExist = os.path.exists(output_report_folder)
                    if not isExist:

                    # Create a new directory because it does not exist
                        os.makedirs(output_report_folder)
                        print("The new output_report_folder is created! ", output_report_folder)
                    old_file_path = os.path.join(
                        output_report_folder, new_file_name)
                    ref_workbook.save(old_file_path)
                    service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
                    print("Theo doi phep bu url: ", service_url)
                    year = self.date_array[0].year
                    month = self.date_array[0].month
                    
                    company_id = df_company_filter.iloc[0]['id']
                    try:
                        
                        min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}',filename='al_cl_serverance', temp_file_path=old_file_path)
                        min_io_results.append( {'company_name': group_index, \
                                        'month' : self.date_array[0].month, \
                                        'year' : self.date_array[0].year, \
                                        'company_id': company_id, \
                                        'template': '9', \
                                        'department_name': '', 'url': min_io_result['url'], \
                                        'type': f"Báo cáo phep bu nghi viec Cong ty {mis_id}"} )
                    except Exception as ex:
                        print(ex)
                    # break
                except Exception as ex:
                    print(f"export_al_cl_report {mis_id} - {ex}")
            # except Exception as ex:
            #     print(ex)
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data(df_miniio)
            
    def export_al_cl_report_next_month(self, al_sheet_name='AL', cl_sheet_name='CL', process_callback=None):
        min_io_results=[]
        for group_index, df_al_report in self.df_al_report.groupby('company_name'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            
            found = False
            if len(df_company_filter.index) > 0:

                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                for item in REPORT_LINK:
                    if item['mis-id'] == mis_id:
                        df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index) & 
                                                        (self.df_hr_leave['state']=='validate')]

                        output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}",
                                ( self.date_array[-1]+ datetime.timedelta(days=10)).strftime("%Y-%m")) 
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:
                            # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                        found= True
                        df_hr_leave = self.df_hr_leave[self.df_hr_leave['company_name']==group_index]
                        df_scheduling_ver = self.df_scheduling_ver[self.df_scheduling_ver['company'] == group_index]
                        df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                        df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                        print('out al cl: ', output_report_folder)
                        if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                            os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                        isExist = os.path.exists(output_report_folder)
                        if not isExist:

                        # Create a new directory because it does not exist
                            os.makedirs(output_report_folder)
                        # print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                        # print('al_result', al_result)
                        # try:
                        df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index] 
                        max_calulate_leave = df_cl_report['date_calculate'].max()
                        df_cl_report =  df_cl_report[df_cl_report['date_calculate']==max_calulate_leave]
                        # print('df_cl_report: ', df_cl_report)
                        # print("AL max: ", al_result[('date_calculate_leave','max')][group_index])
                        
                        df_al_report = self.df_al_report[(self.df_al_report['company_name']==group_index )]
                                                        
                        max_calulate_leave =  df_al_report['date_calculate_leave'].max()
                        df_al_report = df_al_report[df_al_report['date_calculate_leave']==max_calulate_leave]
                        
                        df_cl_report.to_excel(os.path.join(output_report_folder, 'df_cl_report.xlsx'))
                        df_al_report.to_excel(os.path.join(output_report_folder, 'df_al_report.xlsx'))
                        print(output_report_folder)
                        print(df_cl_report)
                        media_report_folder = init_media_subfoder_report( f"{item['user']}", (self.date_array[-1] +  + datetime.timedelta(days=5)).strftime("%Y-%m"))
                        
                        al_sheet_name='AL' 
                        cl_sheet_name='CL'
                        # continue
                        ref_workbook = self.load_workbook_from_url(url_al_cl_tracking)
                        
                        print(ref_workbook.sheetnames)
                        index = 0
                        cl_index = 0
                        al_index = 0
                        for sheet_name in ref_workbook.sheetnames:
                            print(sheet_name)
                            if al_sheet_name == sheet_name:
                                print('aaaaa', al_sheet_name)
                                al_sheet_name = sheet_name
                                al_index = index
                            if cl_sheet_name == sheet_name:
                                print('bbbbb', cl_sheet_name)
                                cl_sheet_name = sheet_name
                                cl_index = index
                            index = index + 1

                        ws_al = ref_workbook.worksheets[al_index]
                        ws_cl = ref_workbook.worksheets[cl_index]

                        self.export_al_cl_report_next_month_company(df_al_report = df_al_report, 
                                                                df_cl_report = df_cl_report,
                                                                df_hr_leave = df_hr_leave,
                                                                df_scheduling_ver = df_scheduling_ver, 
                                                                ws_al=ws_al, ws_cl=ws_cl)
                        

                        new_file_name = "Theo doi ngay phep bu-nextmonth.xlsx"
                        old_file_path = os.path.join(
                            output_report_folder, new_file_name)
                        ref_workbook.save(old_file_path)
                
                        year = (self.date_array[-1] + datetime.timedelta(days=5)).year
                        month = (self.date_array[-1] + datetime.timedelta(days=5)).month
                        
                        company_id = df_company_filter.iloc[0]['id']
                        try:
                            min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}',filename='al_cl', temp_file_path=old_file_path)
                            min_io_results.append( {'company_name': group_index, \
                                            'month' :  (self.date_array[-1] + datetime.timedelta(days=5)).month, \
                                            'year' :  (self.date_array[-1] + datetime.timedelta(days=5)).year, \
                                            'company_id': company_id, \
                                            'template': '5', \
                                            'url': min_io_result['url'], \
                                            'type': f"Báo cáo phep bu Cong ty {mis_id}"} )
                        except Exception as ex:
                            print(ex)
                        
                        #     break
                        # except Exception as ex:
                        #     print(f"export_al_cl_report {mis_id} - {ex}")
            # except Exception as ex:
            #     print(ex)
        if len(min_io_results)>0:
            print(min_io_results)
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data(df_minio_link_report = df_miniio, is_next_month = True)

    def export_al_cl_report_next_month_company(self, df_al_report, 
                                                    df_cl_report,
                                                    df_hr_leave,
                                                    df_scheduling_ver,
                                                    ws_al, ws_cl, process_callback=None):
        ws_al.cell(row=1, column=3).value = self.date_array[0].year
        ws_al.cell(row=4, column=15).value = f"Phép tồn {self.date_array[0].year -1}"
        try:
            df_scheduling_ver[['night_holiday_work_time_minute','holiday_work_time', 'night_worktime_minute']] = \
                    df_scheduling_ver.apply(lambda row: self.calculate_night_holiday_work_time(row),  axis=1, result_type='expand')
            print("sucesss")
        except Exception as ex:
            df_scheduling_ver['night_holiday_work_time_minute'] =0
            df_scheduling_ver['holiday_work_time'] = 0
            df_scheduling_ver['night_worktime_minute']= 0
            print("calculate night: ", ex)
            print('------------')
            
        try:
            if process_callback:
                process_callback.emit((10, "Khoi tao workbook"))

            if process_callback:
                process_callback.emit(
                    (100, "Khoi tao thanh cong theo doi phep bu"))
        except Exception as ex:
            print('mini: ', ex)
        current_month = self.date_array[-1].month
        # print(self.df_employees['workingday'])
        index = 1
        start_row = 6
        # for g, data in self.df_scheduling_ver[(self.df_scheduling_ver['department_name'].notnull())
        #                                       & (self.df_scheduling_ver['department_name'] != False)
        #                                       & (self.df_scheduling_ver['workingday'].notnull())
        #                                       & (self.df_scheduling_ver['workingday'] != False)].sort_values(['employee_code', 'date']).groupby('department_name'):
        
        for group_index, sub_group_data in df_al_report.iterrows():
            total_ncl = 0
            total_ncl_probationary = 0
            total_off = 0

            total_ph = 0
            total_al = 0
            total_al_probationary = 0

      
            total_dimuon_vesom = 0
            total_dimuon_vesom_probationary = 0

            total_tc = 0
            total_tc_probationary = 0
            total_tc_holiday = 0
            total_tc_holiday_probationary = 0
  

            total_compensation_leave = 0
            total_compensation_probationary = 0

     
            hr_leave_employee = df_hr_leave[df_hr_leave['employee_code']
                                                    == sub_group_data['employee_code']]

            for leave_index, leave_item in hr_leave_employee.iterrows():
                if leave_item['holiday_status_name']:
                    if ('nghỉ bù' in leave_item['holiday_status_name'].strip().lower()):
                        print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                        total_compensation_leave = total_compensation_leave + \
                            max(leave_item['minutes'], leave_item['time'])
                        if leave_item['is_leave_probationary']:
                            total_compensation_probationary = total_compensation_probationary + \
                                max(leave_item['minutes'],
                                    leave_item['time'])
                # except Exception as ex:
                #     print(ex)
                # try:
                    if ( 'nghỉ phép năm' in leave_item['holiday_status_name'].strip().lower()):
                        print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                        total_al = total_al + \
                            max(leave_item['minutes'], leave_item['time'])
                        if leave_item['is_leave_probationary']:
                            total_al_probationary = total_al_probationary + \
                                max(leave_item['minutes'],
                                    leave_item['time'])

                    elif ( 'có tính lương' in leave_item['holiday_status_name'].strip().lower()):
                        print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
                        total_ncl = total_ncl + \
                            max(leave_item['minutes'], leave_item['time'])
                        if leave_item['is_leave_probationary']:
                            total_ncl_probationary = total_ncl_probationary + \
                                max(leave_item['minutes'],
                                    leave_item['time'])
                    elif ('tăng ca' in leave_item['holiday_status_name'].strip().lower()):
                        print("don tang ca")
                        if leave_item['is_holiday']==True:
                            total_tc_holiday = total_tc_holiday + \
                                max(leave_item['minutes'],
                                    leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                            if leave_item['is_leave_probationary']:
                                total_tc_holiday_probationary = total_tc_holiday_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                        else:
                            total_tc = total_tc + \
                                max(leave_item['minutes'],
                                    leave_item['time']) * max(1, leave_item['multiplier_work_time'])
                            if leave_item['is_leave_probationary']:
                                total_tc_probationary = total_tc_probationary + \
                                    max(leave_item['minutes'],
                                        leave_item['time']) * max(1, leave_item['multiplier_work_time'])

                    elif (('đi muộn' in leave_item['holiday_status_name'].strip().lower()) or ('về sớm' in leave_item['holiday_status_name'].strip().lower())) and \
                            ((leave_item['for_reasons']=='1') or (leave_item['for_reasons']==1)) and self.is_calculate_late_early_leave:
                        print("don di muon ve som duoc duyet voi ly do ca nhan")
                        total_dimuon_vesom = total_dimuon_vesom + \
                            max(leave_item['minutes'], leave_item['time'])
                        if leave_item['is_leave_probationary']:
                            total_dimuon_vesom_probationary = total_dimuon_vesom_probationary + \
                                max(leave_item['minutes'],
                                    leave_item['time'])
                                    
                    elif ('ra ngoài' in leave_item['holiday_status_name'].strip().lower()) and \
                                leave_item['for_reasons']=='1':
                        duration = leave_item['attendance_missing_to'] - leave_item['attendance_missing_from']
                        seconds = duration.total_seconds()
                        hours = seconds // 3600
                        minutes = (seconds % 3600) // 60.0

                        total_out = total_out + minutes + hours*60
                        print("don ra ngoài cá nhân")
                # except Exception as ex:
                #     print(ex)
            # print('al: ',sub_group_data['january'])
            if index == 1:
                ws_al.cell(row=2, column=12).value = sub_group_data['date_calculate_leave']
                ws_al.cell(row=2, column=11).value = sub_group_data['date_calculate_leave']
            ws_al.cell(row=start_row, column=1).value = index
            try:
                ws_al.cell(row=start_row, column=3).value = sub_group_data['employee_name']
            except:
                ws_al.cell(row=start_row, column=3).value = ''
            try:
                ws_al.cell(row=start_row, column=4).value = sub_group_data['department_name']
            except:
                ws_al.cell(row=start_row, column=4).value = ''
            ws_al.cell(row=start_row, column=2).value = sub_group_data['employee_code']
            # '','','','','','','',
            #                                             '','','','','december'
            ws_al.cell(row=start_row, column=15).value = sub_group_data['remaining_leave']/60.0 if current_month > 1 else sub_group_data['remaining_leave_minute']/60.0 
            ws_al.cell(row=start_row, column=16).value = sub_group_data['january'] if current_month > 1 else total_al if current_month ==1 else 0
            ws_al.cell(row=start_row, column=17).value = sub_group_data['february']  if current_month > 2 else total_al if current_month ==2 else 0
            ws_al.cell(row=start_row, column=18).value = sub_group_data['march'] if current_month > 3 else total_al if current_month ==3 else 0
            ws_al.cell(row=start_row, column=19).value = sub_group_data['april'] if current_month > 4 else total_al if current_month ==4 else 0
            ws_al.cell(row=start_row, column=20).value = sub_group_data['may'] if current_month > 5 else total_al if current_month ==5 else 0
            ws_al.cell(row=start_row, column=21).value = sub_group_data['june'] if current_month > 6 else total_al if current_month ==6 else 0
            ws_al.cell(row=start_row, column=22).value = sub_group_data['july'] if current_month > 7 else total_al if current_month ==7 else 0
            ws_al.cell(row=start_row, column=23).value = sub_group_data['august'] if current_month > 8 else total_al if current_month ==8 else 0
            ws_al.cell(row=start_row, column=24).value = sub_group_data['september'] if current_month > 9 else total_al if current_month ==9 else 0 
            ws_al.cell(row=start_row, column=25).value = sub_group_data['october']if current_month > 10 else total_al if current_month ==10 else 0
            ws_al.cell(row=start_row, column=26).value = sub_group_data['november'] if current_month > 11 else total_al if current_month ==11 else 0
            ws_al.cell(row=start_row, column=27).value = sub_group_data['december'] if current_month > 12 else total_al if current_month ==12 else 0
            ws_al.cell(row=start_row, column=29).value = sub_group_data['remaining_leave_minute'] - total_al
            ws_al.cell(row=start_row, column=30).value = sub_group_data['remaining_leave_day'] - total_al//480
            
            ws_al.cell(row=start_row,column=5).value = '' if sub_group_data['job_title'] == False else sub_group_data['job_title']
            ws_al.cell(row=start_row,column=6).value = '' if sub_group_data['standard_day'] == False else sub_group_data['standard_day']
            ws_al.cell( row=start_row, column=8).value = '' if sub_group_data['date_sign'] == False else sub_group_data['date_sign']
            ws_al.cell(row=start_row, column=7).value = '' if sub_group_data['workingday'] == False else sub_group_data['workingday']
            ws_al.cell(row=start_row, column=9).value = '' if sub_group_data['date_apply_leave'] == False else sub_group_data['date_apply_leave']
            ws_al.cell( row=start_row, column=10).value = '' if sub_group_data['severance_day'] == False else sub_group_data['severance_day']
            start_row = start_row + 1
            index = index + 1

        index = 1
        start_row = 5
        print ("export CL ", df_cl_report)
        for group_index, sub_group_data in df_cl_report.iterrows():
            print(sub_group_data)
            ws_cl.cell(row=start_row, column=1).value = index
            ws_cl.cell(row=start_row, column=2).value = sub_group_data['employee_code']
            ws_cl.cell(row=start_row, column=4).value = '' if sub_group_data['workingday'] == False else sub_group_data['workingday']
            ws_cl.cell(row=start_row, column=5).value = '' if sub_group_data['tier'] == False else sub_group_data['tier']
            ws_cl.cell(row=start_row, column=6).value = '' if sub_group_data['department_name'] == False else sub_group_data['department_name']
            ws_cl.cell(row=start_row, column=7).value = '' if sub_group_data['job_title'] == False else sub_group_data['job_title']
            ws_cl.cell(row=start_row, column=8).value = '' if sub_group_data['severance_day'] == False else sub_group_data['severance_day']
            ws_cl.cell(row=start_row, column=9).value = '' if sub_group_data['date_sign'] == False else sub_group_data['date_sign']

            try:
                ws_cl.cell(row=start_row, column=3).value = sub_group_data['employee_name']
            except:
                ws_cl.cell(row=start_row, column=3).value = ''
            index = index + 1
            col_index = 0
            # Note: Update case month is 1 -- TBD
            if (self.date_array[0] - relativedelta(months = 1)).year == self.date_array[0].year:
                for month_index in range(0,(self.date_array[0] - relativedelta(months = 1)).month):
                    print(sub_group_data[f'increase_probationary_{month_index+1}'])
                    # print('month_index',month_index)
                    
                    ws_cl.cell(row=start_row,
                                        column=month_index * 6 + 12).value = str(sub_group_data[f'increase_probationary_{month_index+1}']) \
                                                                if sub_group_data[f'increase_probationary_{month_index+1}'] > 0 else ''
                    ws_cl.cell(row=start_row,
                                        column=month_index* 6 + 13).value = str(sub_group_data[f'increase_official_{month_index+1}']) \
                                                                if sub_group_data[f'increase_official_{month_index+1}'] > 0 else ''
                    ws_cl.cell(row=start_row,
                                        column=month_index * 6 + 14).value = str(sub_group_data[f'used_probationary_{month_index+1}']) \
                                                                if sub_group_data[f'used_probationary_{month_index+1}']> 0 else ''
                    ws_cl.cell(row=start_row,
                                        column=month_index* 6 + 15).value = str(sub_group_data[f'used_official_{month_index+1}']) \
                                                                if sub_group_data[f'used_official_{month_index+1}'] > 0 else ''
                    ws_cl.cell(row=start_row,
                                        column=month_index* 6 + 16).value = str(sub_group_data[f'overtime_probationary_{month_index+1}']) \
                                                                if sub_group_data[f'overtime_probationary_{month_index+1}'] > 0 else ''
                    ws_cl.cell(row=start_row,
                                        column=month_index* 6 + 17).value = str(sub_group_data[f'overtime_official_{month_index+1}']) \
                                                                if sub_group_data[f'overtime_official_{month_index+1}'] > 0 else ''
            month_index = current_month -1
            print(sub_group_data[f'increase_probationary_{month_index+1}'])
                # print('month_index',month_index)
                
            ws_cl.cell(row=start_row,
                                column=month_index * 6 + 12).value = str(sub_group_data[f'increase_probationary_{month_index+1}']) \
                                                        if sub_group_data[f'increase_probationary_{month_index+1}'] > 0 else ''
            ws_cl.cell(row=start_row,
                                column=month_index* 6 + 13).value = str(sub_group_data[f'increase_official_{month_index+1}']) \
                                                        if sub_group_data[f'increase_official_{month_index+1}'] > 0 else ''
            ws_cl.cell(row=start_row,
                                column=month_index * 6 + 14).value = str(sub_group_data[f'used_probationary_{month_index+1}']) \
                                                        if sub_group_data[f'used_probationary_{month_index+1}']> 0 else ''
            ws_cl.cell(row=start_row,
                                column=month_index* 6 + 15).value = str(sub_group_data[f'used_official_{month_index+1}']) \
                                                        if sub_group_data[f'used_official_{month_index+1}'] > 0 else ''
            ws_cl.cell(row=start_row,
                                column=month_index* 6 + 16).value = str(sub_group_data[f'overtime_probationary_{month_index+1}']) \
                                                        if sub_group_data[f'overtime_probationary_{month_index+1}'] > 0 else ''
            ws_cl.cell(row=start_row,
                                column=month_index* 6 + 17).value = str(sub_group_data[f'overtime_official_{month_index+1}']) \
                                                        if sub_group_data[f'overtime_official_{month_index+1}'] > 0 else ''

            ws_cl.cell(row=start_row,
                                    column=90).value = sub_group_data[f'remaining_probationary_minute'] \
                                                            if sub_group_data[f'remaining_probationary_minute'] > 0 else ''
            ws_cl.cell(row=start_row,
                                    column=91).value = sub_group_data[f'remaining_official_minute'] \
                                                            if sub_group_data[f'remaining_official_minute'] > 0 else ''

            ws_cl.cell(row=start_row,
                                    column=93).value = sub_group_data[f'remaining_probationary_minute']/480 \
                                                            if sub_group_data[f'remaining_probationary_minute'] > 0 else ''
            ws_cl.cell(row=start_row,
                                    column=94).value = sub_group_data[f'remaining_official_minute']/480 \
                                                            if sub_group_data[f'remaining_official_minute'] > 0 else ''
                                                            
            # Current month
            month_index = month_index+1
            usecl_probationary = total_compensation_probationary + total_dimuon_vesom_probationary
            ws_cl.cell(row=start_row, column=month_index * 6 + 14).value = \
                str(usecl_probationary) if usecl_probationary>0 else ''
            usecl_offical = total_compensation_leave + total_dimuon_vesom - \
                (total_compensation_probationary + total_dimuon_vesom_probationary)      
            
            ws_cl.cell(row=start_row, column=month_index* 6 + 15).value = \
                str(usecl_offical) if usecl_offical > 0 else ''
                
            ws_cl.cell(row=start_row, column=month_index* 6 + 16).value = \
                str(total_tc_probationary) if total_tc_probationary >0 else ''
            
            ws_cl.cell(row=start_row, column=month_index* 6 + 17).value = \
                str(total_tc - total_tc_probationary) if (total_tc - total_tc_probationary) >0 else ''
            start_row = start_row + 1
            

        # print(f'cl inde:{cl_index}- al_index {al_index}')
        
    def export_al_cl_report_company_version2(self, df_al_report, 
                                                    df_cl_report,
                                                    ws_al, ws_cl, process_callback=None):
        month_array_str = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 
                                'october', 'november', 'december']
        try:
            if process_callback:
                process_callback.emit((10, "Khoi tao workbook"))

            if process_callback:
                process_callback.emit(
                    (100, "Khoi tao thanh cong theo doi phep bu"))
        except Exception as ex:
            print('mini: ', ex)
        # print(self.df_employees['workingday'])
        index = 1
        start_row = 6
        # for g, data in self.df_scheduling_ver[(self.df_scheduling_ver['department_name'].notnull())
        #                                       & (self.df_scheduling_ver['department_name'] != False)
        #                                       & (self.df_scheduling_ver['workingday'].notnull())
        #                                       & (self.df_scheduling_ver['workingday'] != False)].sort_values(['employee_code', 'date']).groupby('department_name'):
        ws_al.cell(row=1, column=3).value = self.date_array[0].year
        for group_index, sub_group_data in df_al_report.iterrows():
            # print('al: ',sub_group_data['january'])
            if index == 1:
                ws_al.cell(row=2, column=12).value = sub_group_data['date_calculate_leave']
                ws_al.cell(row=2, column=11).value = sub_group_data['date_calculate_leave']
            ws_al.cell(row=start_row, column=1).value = index
            try:
                ws_al.cell(row=start_row, column=3).value = sub_group_data['employee_name']
            except:
                ws_al.cell(row=start_row, column=3).value = ''
            try:
                ws_al.cell(row=start_row, column=4).value = sub_group_data['department_name']
            except:
                ws_al.cell(row=start_row, column=4).value = ''
            ws_al.cell(row=start_row, column=2).value = sub_group_data['employee_code']
            
            ws_al.cell(row=start_row, column=14).value = sub_group_data['leave_year'] 
            ws_al.cell(row=start_row, column=15).value = sub_group_data['remaining_leave'] 
            ws_al.cell(row=start_row, column=16).value = sub_group_data['january'] 
            ws_al.cell(row=start_row, column=17).value = sub_group_data['february'] 
            ws_al.cell(row=start_row, column=18).value = sub_group_data['march'] 
            ws_al.cell(row=start_row, column=19).value = sub_group_data['april'] 
            ws_al.cell(row=start_row, column=20).value = sub_group_data['may'] 
            ws_al.cell(row=start_row, column=21).value = sub_group_data['june']
            ws_al.cell(row=start_row, column=22).value = sub_group_data['july'] 
            ws_al.cell(row=start_row, column=23).value = sub_group_data['august']
            ws_al.cell(row=start_row, column=24).value = sub_group_data['september'] 
            ws_al.cell(row=start_row, column=25).value = sub_group_data['october']
            ws_al.cell(row=start_row, column=26).value = sub_group_data['november'] 
            ws_al.cell(row=start_row, column=27).value = sub_group_data['december'] 
            
            if f"{sub_group_data['employee_code']}" in self.al_cl_current_month:
                old_value = int(ws_al.cell(row=start_row, column=15 + self.date_array[0].month).value)
                ws_al.cell(row=start_row, column=15 + self.date_array[0].month).value \
                    = self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_al']
                    
            ws_al.cell(row=start_row,column=5).value = '' if sub_group_data['job_title'] == False else sub_group_data['job_title']
            ws_al.cell(row=start_row,column=6).value = '' if sub_group_data['standard_day'] == False else sub_group_data['standard_day']
            ws_al.cell( row=start_row, column=8).value = '' if sub_group_data['date_sign'] == False else sub_group_data['date_sign']
            ws_al.cell(row=start_row, column=7).value = '' if sub_group_data['workingday'] == False else sub_group_data['workingday']
            ws_al.cell(row=start_row, column=9).value = '' if sub_group_data['date_apply_leave'] == False else sub_group_data['date_apply_leave']
            ws_al.cell( row=start_row, column=10).value = '' if sub_group_data['severance_day'] == False else sub_group_data['severance_day']

            try:
                note_al = f"di muon ve som: {sub_group_data['total_dimuon_vesom']}"
                ws_al.cell(row=start_row, column=31).value = note_al
            except Exception as ex:
                print('note al ex: ',ex)
            start_row = start_row + 1
            index = index + 1

        index = 1
        start_row = 5
        print ("export CL ", df_cl_report)
        for group_index, sub_group_data in df_cl_report.iterrows():
            print(sub_group_data)
            
            ws_cl.cell(row=start_row, column=2).value = sub_group_data['employee_code']
            ws_cl.cell(row=start_row, column=4).value = '' if sub_group_data['workingday'] == False else sub_group_data['workingday']
            ws_cl.cell(row=start_row, column=5).value = '' if sub_group_data['tier'] == False else sub_group_data['tier']
            ws_cl.cell(row=start_row, column=6).value = '' if sub_group_data['department_name'] == False else sub_group_data['department_name']
            ws_cl.cell(row=start_row, column=7).value = '' if sub_group_data['job_title'] == False else sub_group_data['job_title']
            ws_cl.cell(row=start_row, column=8).value = '' if sub_group_data['severance_day'] == False else sub_group_data['severance_day']
            ws_cl.cell(row=start_row, column=9).value = '' if sub_group_data['date_sign'] == False else sub_group_data['date_sign']

            try:
                ws_cl.cell(row=start_row, column=3).value = sub_group_data['employee_name']
            except:
                ws_cl.cell(row=start_row, column=3).value = ''
                continue
            index = index + 1
            col_index = 0
            # Note: Update case month is 1 -- TBD
            # for month_index in range(0,(self.date_array[0] - relativedelta(months = 1)).month):
            for month_index in range(0,(self.date_array[0]).month):
                # print('month_index',month_index)
                increase_probationary = sub_group_data[f'increase_probationary_{month_index+1}']
                increase_official = sub_group_data[f'increase_official_{month_index+1}']
                overtime_probationary = sub_group_data[f'overtime_probationary_{month_index+1}']
                overtime_official = sub_group_data[f'overtime_official_{month_index+1}']
                used_probationary = sub_group_data[f'used_probationary_{month_index+1}']
                used_official = sub_group_data[f'used_official_{month_index+1}']
                increase = increase_probationary + increase_official + overtime_probationary + overtime_official
                used = used_probationary + used_official
                ws_cl.cell(row=start_row,
                                    column=month_index * 2 + 11).value = increase if increase > 0 else ''
                ws_cl.cell(row=start_row,
                                    column=month_index* 2 + 12).value = used if used > 0 else ''
               
            remain_last_year = sub_group_data[f'increase_official_1'] + sub_group_data[f'increase_probationary_1']
            ws_cl.cell(row=start_row, column=10).value = remain_last_year if remain_last_year > 0 else ''
            if f"{sub_group_data['employee_code']}" in self.al_cl_current_month: 
                current_month_index = self.date_array[0].month-1
                increase_probationary = sub_group_data[f'increase_probationary_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['phatsinhtang_thuviec']
                increase_official = sub_group_data[f'increase_official_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['phatsinhtang_chinhthuc']
                used_probationary = sub_group_data[f'used_probationary_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['used_cl_probationary']
                used_official = sub_group_data[f'used_official_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_tc_probationary']
                overtime_probationary = sub_group_data[f'overtime_probationary_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_tc_probationary']
                overtime_official = sub_group_data[f'overtime_official_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_tc_offical']
                increase = increase_probationary + increase_official + overtime_probationary + overtime_official
                used = used_probationary + used_official
                ws_cl.cell(row=start_row, column=current_month_index * 2 + 11).value = increase
                ws_cl.cell(row=start_row, column=current_month_index* 2 + 12).value = used
                
            try:
                note_cl = f"di muon ve som: {sub_group_data['total_dimuon_vesom']} \n"
                note_cl = f"{note_cl} thu viec: {sub_group_data['total_dimuon_vesom_probationary']} \n"
                note_cl = f"{note_cl} chinh thuc: {sub_group_data['total_dimuon_vesom_official']}"
                ws_cl.cell(row=start_row, column=96).value = note_cl
            except Exception as ex:
                print('note cl ex: ',ex)
            start_row = start_row + 1

        # print(f'cl inde:{cl_index}- al_index {al_index}')
        
        
    def export_al_cl_report_company(self, df_al_report, 
                                                    df_cl_report,
                                                    ws_al, ws_cl, process_callback=None,
                                                    client_compute = False):
        month_array_str = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 
                                'october', 'november', 'december']
        try:
            if process_callback:
                process_callback.emit((10, "Khoi tao workbook"))

            if process_callback:
                process_callback.emit(
                    (100, "Khoi tao thanh cong theo doi phep bu"))
        except Exception as ex:
            print('mini: ', ex)
        # print(self.df_employees['workingday'])
        index = 1
        start_row = 6
        # for g, data in self.df_scheduling_ver[(self.df_scheduling_ver['department_name'].notnull())
        #                                       & (self.df_scheduling_ver['department_name'] != False)
        #                                       & (self.df_scheduling_ver['workingday'].notnull())
        #                                       & (self.df_scheduling_ver['workingday'] != False)].sort_values(['employee_code', 'date']).groupby('department_name'):
        ws_al.cell(row=1, column=3).value = self.date_array[0].year
        for group_index, sub_group_data in df_al_report.iterrows():
            # print('al: ',sub_group_data['january'])
            if index == 1:
                ws_al.cell(row=2, column=12).value = sub_group_data['date_calculate_leave']
                ws_al.cell(row=2, column=11).value = sub_group_data['date_calculate_leave']
            ws_al.cell(row=start_row, column=1).value = index
            try:
                ws_al.cell(row=start_row, column=3).value = sub_group_data['employee_name']
            except:
                ws_al.cell(row=start_row, column=3).value = ''
            try:
                ws_al.cell(row=start_row, column=4).value = sub_group_data['department_name']
            except:
                ws_al.cell(row=start_row, column=4).value = ''
            ws_al.cell(row=start_row, column=2).value = sub_group_data['employee_code']
            # '','','','','','','',
            #                                             '','','','','december'
            ws_al.cell(row=start_row, column=14).value = sub_group_data['leave_year'] 
            ws_al.cell(row=start_row, column=15).value = sub_group_data['remaining_leave'] 
            ws_al.cell(row=start_row, column=16).value = sub_group_data['january'] 
            ws_al.cell(row=start_row, column=17).value = sub_group_data['february'] 
            ws_al.cell(row=start_row, column=18).value = sub_group_data['march'] 
            ws_al.cell(row=start_row, column=19).value = sub_group_data['april'] 
            ws_al.cell(row=start_row, column=20).value = sub_group_data['may'] 
            ws_al.cell(row=start_row, column=21).value = sub_group_data['june']
            ws_al.cell(row=start_row, column=22).value = sub_group_data['july'] 
            ws_al.cell(row=start_row, column=23).value = sub_group_data['august']
            ws_al.cell(row=start_row, column=24).value = sub_group_data['september'] 
            ws_al.cell(row=start_row, column=25).value = sub_group_data['october']
            ws_al.cell(row=start_row, column=26).value = sub_group_data['november'] 
            ws_al.cell(row=start_row, column=27).value = sub_group_data['december'] 
                
            if f"{sub_group_data['employee_code']}" in self.al_cl_current_month:
                old_value = int(ws_al.cell(row=start_row, column=15 + self.date_array[0].month).value)
                ws_al.cell(row=start_row, column=15 + self.date_array[0].month).value \
                    = self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_al']
            # if not client_compute:
            ws_al.cell(row=start_row, column=29).value = sub_group_data['remaining_leave_minute'] 
            ws_al.cell(row=start_row, column=30).value = sub_group_data['remaining_leave_day'] 
            # else:
            # try:
            #     ids = [int(sub_group_data['id'])]
            #     update_data = {month_array_str[self.date_array[0].month - 1]: int(self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_al'])}
            #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.al.report', 'write', 
            #                                         [ids, update_data])
            # except Exception as ex:
            #     print(ex) 
            #     # new_remain_minute = sub_group_data['remaining_leave_minute'] + old_value \
            #     #    - self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_al']
            #     # ws_al.cell(row=start_row, column=29).value = new_remain_minute
            #     # ws_al.cell(row=start_row, column=30).value = new_remain_minute / 480
                
            #     # input(sub_group_data['employee_code'])
            ws_al.cell(row=start_row,column=5).value = '' if sub_group_data['job_title'] == False else sub_group_data['job_title']
            ws_al.cell(row=start_row,column=6).value = '' if sub_group_data['standard_day'] == False else sub_group_data['standard_day']
            ws_al.cell( row=start_row, column=8).value = '' if sub_group_data['date_sign'] == False else sub_group_data['date_sign']
            ws_al.cell(row=start_row, column=7).value = '' if sub_group_data['workingday'] == False else sub_group_data['workingday']
            ws_al.cell(row=start_row, column=9).value = '' if sub_group_data['date_apply_leave'] == False else sub_group_data['date_apply_leave']
            ws_al.cell( row=start_row, column=10).value = '' if sub_group_data['severance_day'] == False else sub_group_data['severance_day']

            try:
                note_al = f"di muon ve som: {sub_group_data['total_dimuon_vesom']}"
                ws_al.cell(row=start_row, column=31).value = note_al
            except Exception as ex:
                print('note al ex: ',ex)
            start_row = start_row + 1
            index = index + 1

        index = 1
        start_row = 5
        print ("export CL ", df_cl_report)
        for group_index, sub_group_data in df_cl_report.iterrows():
            print(sub_group_data)
            ws_cl.cell(row=start_row, column=1).value = index
            ws_cl.cell(row=start_row, column=2).value = sub_group_data['employee_code']
            ws_cl.cell(row=start_row, column=4).value = '' if sub_group_data['workingday'] == False else sub_group_data['workingday']
            ws_cl.cell(row=start_row, column=5).value = '' if sub_group_data['tier'] == False else sub_group_data['tier']
            ws_cl.cell(row=start_row, column=6).value = '' if sub_group_data['department_name'] == False else sub_group_data['department_name']
            ws_cl.cell(row=start_row, column=7).value = '' if sub_group_data['job_title'] == False else sub_group_data['job_title']
            ws_cl.cell(row=start_row, column=8).value = '' if sub_group_data['severance_day'] == False else sub_group_data['severance_day']
            ws_cl.cell(row=start_row, column=9).value = '' if sub_group_data['date_sign'] == False else sub_group_data['date_sign']

            try:
                ws_cl.cell(row=start_row, column=3).value = sub_group_data['employee_name']
            except:
                ws_cl.cell(row=start_row, column=3).value = ''
            index = index + 1
            col_index = 0
            # Note: Update case month is 1 -- TBD
            # for month_index in range(0,(self.date_array[0] - relativedelta(months = 1)).month):
            for month_index in range(0,(self.date_array[0]).month):
                # print('month_index',month_index)
                
                ws_cl.cell(row=start_row,
                                    column=month_index * 6 + 12).value = str(sub_group_data[f'increase_probationary_{month_index+1}']) \
                                                            if sub_group_data[f'increase_probationary_{month_index+1}'] > 0 else ''
                ws_cl.cell(row=start_row,
                                    column=month_index* 6 + 13).value = str(sub_group_data[f'increase_official_{month_index+1}']) \
                                                            if sub_group_data[f'increase_official_{month_index+1}'] > 0 else ''
                ws_cl.cell(row=start_row,
                                    column=month_index * 6 + 14).value = str(sub_group_data[f'used_probationary_{month_index+1}']) \
                                                            if sub_group_data[f'used_probationary_{month_index+1}']> 0 else ''
                ws_cl.cell(row=start_row,
                                    column=month_index* 6 + 15).value = str(sub_group_data[f'used_official_{month_index+1}']) \
                                                            if sub_group_data[f'used_official_{month_index+1}'] > 0 else ''
                ws_cl.cell(row=start_row,
                                    column=month_index* 6 + 16).value = str(sub_group_data[f'overtime_probationary_{month_index+1}']) \
                                                            if sub_group_data[f'overtime_probationary_{month_index+1}'] > 0 else ''
                ws_cl.cell(row=start_row,
                                    column=month_index* 6 + 17).value = str(sub_group_data[f'overtime_official_{month_index+1}']) \
                                                            if sub_group_data[f'overtime_official_{month_index+1}'] > 0 else ''
            if f"{sub_group_data['employee_code']}" in self.al_cl_current_month: 
                current_month_index = self.date_array[0].month-1
                increase_probationary = sub_group_data[f'increase_probationary_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['phatsinhtang_thuviec']
                increase_official = sub_group_data[f'increase_official_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['phatsinhtang_chinhthuc']
                used_probationary = sub_group_data[f'used_probationary_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['used_cl_probationary']
                used_official = sub_group_data[f'used_official_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_tc_probationary']
                overtime_probationary = sub_group_data[f'overtime_probationary_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_tc_probationary']
                overtime_official = sub_group_data[f'overtime_official_{current_month_index+1}'] \
                        + self.al_cl_current_month[f"{sub_group_data['employee_code']}"]['total_tc_offical']
                ws_cl.cell(row=start_row,
                    column=current_month_index * 6 + 12).value = increase_probationary
                ws_cl.cell(row=start_row,
                                    column=current_month_index* 6 + 13).value = increase_official
                ws_cl.cell(row=start_row,
                                    column=current_month_index * 6 + 14).value = used_probationary
                ws_cl.cell(row=start_row, column=current_month_index* 6 + 16).value = used_official
                                    
                ws_cl.cell(row=start_row,
                                    column=current_month_index* 6 + 16).value = overtime_probationary
                ws_cl.cell(row=start_row,
                                    column=current_month_index* 6 + 17).value = overtime_official
                # if client_compute:
                # try:
                #     ids = [int(sub_group_data['id'])]
                #     update_data = {f'increase_probationary_{current_month_index+1}': int(increase_probationary), 
                #                     f'used_probationary_{current_month_index+1}': int(used_probationary),
                #                     f'used_official_{current_month_index+1}': int(used_official), 
                #                     f'increase_official_{current_month_index+1}': int(increase_official), 
                #                     f'overtime_probationary_{current_month_index+1}': int(overtime_probationary), 
                #                     f'overtime_official_{current_month_index+1}': int(overtime_official)}
                #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.cl.report', 'write', 
                #                                         [ids, update_data])
                # except Exception as ex:
                #     print(ex)
                    
            ws_cl.cell(row=start_row,
                                    column=90).value = sub_group_data[f'remaining_probationary_minute'] \
                                                            if sub_group_data[f'remaining_probationary_minute'] > 0 else ''
            ws_cl.cell(row=start_row,
                                    column=91).value = sub_group_data[f'remaining_official_minute'] \
                                                            if sub_group_data[f'remaining_official_minute'] > 0 else ''

            ws_cl.cell(row=start_row,
                                    column=93).value = sub_group_data[f'remaining_probationary_minute']/480 \
                                                            if sub_group_data[f'remaining_probationary_minute'] > 0 else ''
            ws_cl.cell(row=start_row,
                                    column=94).value = sub_group_data[f'remaining_official_minute']/480 \
                                                            if sub_group_data[f'remaining_official_minute'] > 0 else ''

            try:
                note_cl = f"di muon ve som: {sub_group_data['total_dimuon_vesom']} \n"
                note_cl = f"{note_cl} thu viec: {sub_group_data['total_dimuon_vesom_probationary']} \n"
                note_cl = f"{note_cl} chinh thuc: {sub_group_data['total_dimuon_vesom_official']}"
                ws_cl.cell(row=start_row, column=96).value = note_cl
            except Exception as ex:
                print('note cl ex: ',ex)
            start_row = start_row + 1

        # print(f'cl inde:{cl_index}- al_index {al_index}')
        

    def export_late_in_5_miniutes_report(self):
        min_io_results=[]
        for group_index, df_scheduling_ver in self.df_scheduling_ver.groupby('company'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            found = False
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                for item in REPORT_LINK:
                    if item['mis-id'] == mis_id:
                        try:
                            output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m")) 
                            found= True
                            df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index ) &
                                            (self.df_hr_leave['state']=='validate')]
                            # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                            df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                            print('out: ', output_report_folder)
                            if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                                os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                            isExist = os.path.exists(output_report_folder)
                            if not isExist:

                            # Create a new directory because it does not exist
                                os.makedirs(output_report_folder)
                            print("kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk: ",group_index)
                            # df_hr_leave = self.df_hr_leave[self.df_hr_leave['company_name']==group_index]
                            # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                            # df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index]
                            # df_al_report = self.df_al_report[self.df_al_report['company_name']==group_index]
                            df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                            media_report_folder = init_media_subfoder_report( f"{item['user']}", self.date_array[0].strftime("%Y-%m"))
    
                            ref_workbook = self.load_workbook_from_url(url_late_5_minute)
                            ws = ref_workbook.worksheets[0]
                            self.export_late_in_5_miniutes_report_company( df_scheduling_ver, 
                                                                    df_hr_leave, df_explanation_data, ws)
                            # ref_workbook.save(os.path.join(
        #     output_report_folder, "BAO_CAO_DI_MUON_TREN_5P.xlsx"))
                            new_file_name = "BAO_CAO_DI_MUON_TREN_5P.xlsx"
                            old_file_path = os.path.join(
                                output_report_folder, new_file_name)
                            ref_workbook.save(old_file_path)
                            service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
                            print("BAO_CAO_DI_MUON_TREN_5P: ", service_url)
                            year = self.date_array[0].year
                            month = self.date_array[0].month
                            
                            company_id = df_company_filter.iloc[0]['id']
                            try:
                                min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}',filename='5m_late', temp_file_path=old_file_path)
                                min_io_results.append( {'company_name': group_index, \
                                                'month' : self.date_array[0].month, \
                                                'year' : self.date_array[0].year, \
                                                'company_id': company_id, \
                                                'template': '3', \
                                                'department_name': '', 'url': min_io_result['url'], \
                                                'type': f"Báo cáo di muon 5ph Cong ty {mis_id}"} )
                            except Exception as ex:
                                print(ex)
                            break
                        except Exception as ex:
                            print('5minute ex: ', ex)
        if len(min_io_results)>0:
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data(df_miniio)

    def export_late_in_5_miniutes_report_company(self, df_scheduling_ver, 
                                                    df_hr_leave, df_explanation_data, ws):
        

        start_row = 4
        count_more_than = 0
        more_than = 0
        # explanation_results = self.collect_late_in_5_minute()
        # df_explanation_results = pd.DataFrame.from_dict(explanation_results)
        # df_late= self.df_scheduling_ver[self.df_scheduling_ver['is_late_in'] & (self.df_scheduling_ver['is_late_in'] > 0)].groupby('code')
        # result = []
        for g, data in df_scheduling_ver[(df_scheduling_ver['department_name'].notnull())
                                              & (df_scheduling_ver['department_name'] != False)
                                              & (df_scheduling_ver['workingday'].notnull())
                                              & (df_scheduling_ver['workingday'] != False)].sort_values(['employee_code', 'date']).groupby('department_name'):
            print('aaa')
            for sub_group_index, sub_group_data in data.groupby(['employee_name', 'employee_code']):
                if (sub_group_data.iloc[0]['severance_day'] !=False) and (pd.notnull (sub_group_data.iloc[0]['severance_day'])) and \
                        sub_group_data.iloc[0]['severance_day'] < self.date_array[0]:
                    continue
                count_less_than_equal = 0
                count_more_than = 0
                total_number_learve = 0
                total_number_explanation = 0
                hr_leave_employee = df_hr_leave[df_hr_leave['employee_code']
                                                     == sub_group_index[1]]
                for index, item in sub_group_data.iterrows():
                    number_explanation = 0
                    
                    found_leave = False
                    for _, leave_item in hr_leave_employee.iterrows():
                        if item['date'] >= leave_item['request_date_from'] \
                            and item ['date'] <= leave_item['request_date_to']:
                            # print('holiday_status_name: ', leave_item['holiday_status_name'])
                            # if ('đi muộn' in  leave_item['holiday_status_name']):
                            if ('bù' in leave_item['holiday_status_name']) or \
                                ('phép năm' in  leave_item['holiday_status_name']) or \
                                ('về sớm' in  leave_item['holiday_status_name']) or \
                                ('đi muộn' in  leave_item['holiday_status_name']):
                                total_number_learve = total_number_learve + 1
                                found_leave = True

                    explanation_filter = df_explanation_data[
                        df_explanation_data['date_str'].str.contains(item['date_str']).fillna(False) &
                        df_explanation_data['employee_code'].str.contains(sub_group_index[1]).fillna(False) &
                        # self.df_explanation_data['reason'].str.contains('2') &
                        ((df_explanation_data['validated']==2) | (df_explanation_data['validated']=='2'))]
                    
                    # hr_leave_filter = self.df_hr_leave[self.df_hr_leave['date_from_str'].str.contains(row['date_str']) &
                    #             self.df_hr_leave['employee_code'].str.contains(sub_group_index[1])
                    #             ]

                    # if g in self.df_explanation_data.index and actual_work_time < 480:
                    if len(explanation_filter.index) > 0 :
                        number_explanation = number_explanation + 1
                        total_number_explanation = number_explanation + total_number_explanation
                    if number_explanation > 0 or found_leave:
                        print('has leave', found_leave)
                        continue           
                    try:

                        if int(item['attendance_late']) > 5:
                            count_more_than = count_more_than + 1
                        elif int(item['attendance_late']) > 0:
                            count_less_than_equal = count_less_than_equal + 1
                    except Exception as ex:
                        print('8', ex)
                if count_less_than_equal > 0 or count_more_than > 0:
                    ws.cell(row=start_row, column=3).value = sub_group_index[0]
                    ws.cell(row=start_row, column=2).value = sub_group_index[1]
                    ws.cell(row=start_row, column=4).value = count_less_than_equal
                    ws.cell(row=start_row, column=5).value = count_more_than
                    ws.cell(row=start_row, column=11).value = total_number_learve
                    ws.cell(row=start_row, column=10).value = total_number_explanation
                    start_row = start_row + 1
                
        
    def export_night_shift_report(self):
        min_io_results = []
        index = 0
        for group_index, df_scheduling_ver in self.df_scheduling_ver.groupby('company'):
            df_company_filter = self.df_all_company[self.df_all_company['name']==group_index]
            found = False
            if len(df_company_filter.index) > 0:
                print(df_company_filter)
                mis_id = df_company_filter.iloc[0]['mis_id']
                for item in REPORT_LINK:
                    if item['mis-id'] == mis_id:
                        try:
                            output_report_folder = os.path.join(self.output_report_folder, f"{item['user']}", self.date_array[0].strftime("%Y-%m")) 
                            found= True
                            df_hr_leave = self.df_hr_leave[(self.df_hr_leave['company_name']==group_index ) &
                                            (self.df_hr_leave['state']=='validate')]
                            # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                            df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                            print('out: ', output_report_folder)
                            if not os.path.exists(os.path.join(self.output_report_folder, f"{item['user']}")):
                                os.makedirs(os.path.join(self.output_report_folder, f"{item['user']}"))
                            isExist = os.path.exists(output_report_folder)
                            if not isExist:

                            # Create a new directory because it does not exist
                                os.makedirs(output_report_folder)
                            df_hr_leave = self.df_hr_leave[self.df_hr_leave['company_name']==group_index]
                            # df_hr_leave.to_excel(os.path.join(output_report_folder,'out_df_hr_leave.xlsx'))
                            # df_cl_report = self.df_cl_report[self.df_cl_report['company_name']==group_index]
                            # df_al_report = self.df_al_report[self.df_al_report['company_name']==group_index]
                            # df_explanation_data = self.df_explanation_data[self.df_explanation_data['company_name']==group_index]
                            media_report_folder = init_media_subfoder_report( f"{item['user']}", self.date_array[0].strftime("%Y-%m"))
                            ref_workbook = self.load_workbook_from_url(url_night_split_shift)
                            ws_night_shift = ref_workbook['Night Shift']
                            ws_split_shift = ref_workbook['Split Shift']
                            self.export_night_shift_report_company( group_index, df_scheduling_ver, ws_night_shift, ws_split_shift)
                            
                            new_file_name = "CA_DEM_CA_GAY.xlsx"
                            old_file_path = os.path.join(
                                output_report_folder, new_file_name)
                            ref_workbook.save(old_file_path)
                            service_url = copy_to_default_storage(old_file_path, new_file_name, media_report_folder)
                            print("CA_DEM_CA_GAY url: ", service_url)
                            year = self.date_array[0].year
                            month = self.date_array[0].month
                            
                            company_id = df_company_filter.iloc[0]['id']
                            try:
                                min_io_result = self.save_excel_to_min_io(document_type='output_report', subfoder=f'{mis_id}/{year}-{month}',filename='night_split', temp_file_path=old_file_path)
                                min_io_results.append( {'company_name': group_index, \
                                                'month' : self.date_array[0].month, \
                                                'year' : self.date_array[0].year, \
                                                'company_id': company_id, \
                                                'template': '4', \
                                                'department_name': '', 'url': min_io_result['url'], \
                                                'type': f"Báo cáo ca dem ca gay Cong ty {mis_id}"} )
                            except Exception as ex:
                                print(ex)
                            break
                        except Exception as ex:
                            print('Ex export_night_shift_report: ', ex)
        if len(min_io_results)>0:                
            df_miniio = pd.DataFrame.from_dict(min_io_results)        
            self.load_upload_report_data(df_miniio)

    def export_night_shift_report_company(self, company_name, df_scheduling_ver, 
                    ws_night_shift, ws_split_shift):
        # df_scheduling_ver.to_excel(os.path.join(self.output_report_folder,"convitthu5.xlsx"), sheet_name='Sheet1')
        
        df_explanation_results = df_scheduling_ver[(df_scheduling_ver['department_name'].notnull())
                                                        & (df_scheduling_ver['department_name'] != False)
                                                        & (df_scheduling_ver['workingday'].notnull())
                                                        & (df_scheduling_ver['workingday'] != False)
                                                        & df_scheduling_ver['night']].sort_values(by=['date'], ascending=False).groupby('department_name')
        ws = ws_night_shift
        ws.merge_cells('C2:K2')
        top_left_cell = ws['C2']
        top_left_cell.value = company_name
        ws.merge_cells('B8:T8')
        top_left_cell = ws['B8']
        top_left_cell.value = f'Từ ngày: {self.date_array[0].strftime("%d/%m/%Y")} - Đến ngày:{self.date_array[len(self.date_array)-1].strftime("%d/%m/%Y")}'

        ws.merge_cells('C3:E3')
        top_left_cell = ws['C3']
        top_left_cell.value = ''
        ws.merge_cells('C4:K4')
        top_left_cell = ws['C4']
        top_left_cell.value = ''
        start_row = 11
        # each item in filter data
        for g, data in df_explanation_results:
            ws.cell(row=start_row, column=4).value = g
            start_row = start_row + 1
            count = 1
            # groups.append(g)
            for index, row in data.iterrows():
                try:
                    ws.cell(row=start_row, column=3).value = count
                    ws.cell(row=start_row,
                            column=4).value = row['employee_code']
                    ws.cell(row=start_row, column=5).value = str(
                        row['employee_name'])
                    ws.cell(row=start_row, column=6).value = row['job_title']
                    if pd.notnull(row['date']):
                        ws.merge_cells(f'G{start_row}:I{start_row}')
                        top_left_cell = ws[f'G{start_row}']
                        top_left_cell.value = row['date'].strftime(
                            "%d/%m/%Y")

                    # shift code
                    ws.merge_cells(f'J{start_row}:L{start_row}')
                    top_left_cell = ws[f'J{start_row}']
                    top_left_cell.value = row['shift_name']

                    # print("row['start_work_date_time']: ", row['start_work_date_time'])
                    if pd.notnull(row['shift_start']):
                        h, m, _ = float_to_hours(row['shift_start'])
                        ws.merge_cells(f'M{start_row}:N{start_row}')
                        top_left_cell = ws[f'M{start_row}']
                        top_left_cell.value = '%02d:%02d' % (h, m)
                        # ws.cell(row=start_row,
                        #         column=12).value = '%02d:%02d' % (h, m)
                        # ws.cell(row=start_row, column=13).value = ''
                    if pd.notnull(row['shift_end']):
                        h, m, _ = float_to_hours(row['shift_end'])
                        ws.merge_cells(f'O{start_row}:P{start_row}')
                        top_left_cell = ws[f'O{start_row}']
                        top_left_cell.value = '%02d:%02d' % (h, m)
                        # ws.cell(row=start_row,
                        #         column=14).value = '%02d:%02d' % (h, m)
                        # ws.cell(row=start_row, column=15).value = ''
                    total_work_hour_h, total_work_hour_m, _ = float_to_hours(
                        row['total_shift_work_time'])
                    ws.cell(row=start_row, column=18).value = '%02d:%02d' % (
                        total_work_hour_h, total_work_hour_m)
                    # ws.cell(row=start_row, column=25).value = row['note']
                    count = count + 1

                    start_row = start_row + 1
                except Exception as ex:
                    print("Night shift: ", ex)
        # SPLIT SHIFT SHEET
        ws = ws_split_shift
        ws.merge_cells('C2:K2')
        top_left_cell = ws['C2']
        top_left_cell.value = company_name
        ws.merge_cells('B8:X8')
        top_left_cell = ws['B8']
        top_left_cell.value = f'Từ ngày: {self.date_array[0].strftime("%d/%m/%Y")} - Đến ngày:{self.date_array[len(self.date_array)-1].strftime("%d/%m/%Y")}'
        ws.merge_cells('C3:E3')
        top_left_cell = ws['C3']
        top_left_cell.value = ''
        ws.merge_cells('C4:K4')
        top_left_cell = ws['C4']
        top_left_cell.value = ''
        
        df_explanation_results = df_scheduling_ver[(df_scheduling_ver['department_name'].notnull())
                                                        & (df_scheduling_ver['department_name'] != False)
                                                        & (df_scheduling_ver['workingday'].notnull())
                                                        & (df_scheduling_ver['workingday'] != False)
                                                        & df_scheduling_ver['fix_rest_time']] \
            .sort_values(by=['date'], ascending=False).groupby('department_name')

        start_row = 11
        # explanation_results = self.collect_explanation()
        # df_explanation_results = pd.DataFrame.from_dict(explanation_results)
        # print(df_explanation_results)
        for g, data in df_explanation_results:
            ws.cell(row=start_row, column=4).value = g
            start_row = start_row + 1
            count = 1
            # groups.append(g)
            for index, row in data.iterrows():
                try:
                    if (row['total_work_time'] <= 0):
                        continue
                    ws.cell(row=start_row, column=3).value = count
                    ws.cell(row=start_row,
                            column=4).value = row['employee_code']
                    ws.cell(row=start_row, column=5).value = str(
                        row['employee_name'])
                    ws.cell(row=start_row, column=6).value = row['job_title']
                    if pd.notnull(row['date']):
                        ws.merge_cells(f'G{start_row}:I{start_row}')
                        top_left_cell = ws[f'G{start_row}']
                        top_left_cell.value  = row['date'].strftime(
                            "%d/%m/%Y")
                        # ws.cell(row=start_row, column=9).value = row['type']

                    # shift code
                    ws.merge_cells(f'J{start_row}:L{start_row}')
                    top_left_cell = ws[f'J{start_row}']
                    top_left_cell.value =  row['shift_name']
                    if pd.notnull(row['shift_start']):
                        h, m, _ = float_to_hours(row['shift_start'])
                        # ws.cell(row=start_row,
                        #         column=12).value = '%02d:%02d' % (h, m)
                        # ws.cell(row=start_row, column=13).value = ''
                        ws.merge_cells(f'M{start_row}:N{start_row}')
                        top_left_cell = ws[f'M{start_row}']
                        top_left_cell.value = '%02d:%02d' % (h, m)
                    if pd.notnull(row['shift_end']):
                        h, m, _ = float_to_hours(row['shift_end'])
                        ws.merge_cells(f'O{start_row}:P{start_row}')
                        top_left_cell = ws[f'O{start_row}']
                        top_left_cell.value = '%02d:%02d' % (h, m)
                        # ws.cell(row=start_row,
                        #         column=14).value = '%02d:%02d' % (h, m)
                        # ws.cell(row=start_row, column=15).value = ''
                    total_work_hour_h, total_work_hour_m, _ = float_to_hours(
                        row['total_shift_work_time'])
                    ws.cell(row=start_row, column=18).value = '%02d:%02d' % (
                        total_work_hour_h, total_work_hour_m)
                    # ws.cell(row=start_row, column=25).value = row['note']
                    count = count + 1

                    start_row = start_row + 1
                except Exception as ex:
                    print("SPlit: ", ex)
        # SPLIT SHIFT SHEET
        # ws = ref_workbook['Night Shift']

        # ref_workbook.save(os.path.join(
        #     output_report_folder, "CA_DEM_CA_GAY.xlsx"))
        # self.df_scheduling_ver.to_excel(os.path.join(self.output_report_folder,"convitthu2.xlsx"), sheet_name='Sheet1')
        
    def export_log(self, output_file):
        self.df_employees.to_excel(output_file)

    def analyst_log(self, output_file):
        df = pd.read_excel(output_file)
        code_series = df["code"]

        duplicates = code_series[code_series.duplicated(keep=False)]

        print("Các giá trị bị trùng lặp:")
        print(duplicates)

        df_without_duplicates = df.drop_duplicates(
            subset=["code"], keep="first")

        duplicates = df[~df.index.isin(df_without_duplicates.index)]

        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
            duplicates.to_excel(writer, sheet_name="Duplicates", index=False)

        # In ra thông báo khi hoàn thành
        print("Các giá trị bị trùng lặp đã được đưa vào sheet 'Duplicates'")

    def update_dumny_shift(self):
        result = False
        self.df_employees[self.df_employees['employee_ho']==True].to_excel("ho.xlsx")
        # exit()
        # for g, data in self.df_scheduling_ver.groupby('employee_code'):
        for index, item in self.df_employees[self.df_employees['employee_ho']==True].iterrows():

            # date_str_data_array = []
            # shift_error = []
            # shift_null = []
            # for index, item in data.iterrows():
            #     if not item['date'] in date_str_data_array:
            #         date_str_data_array.append(item['date'])
            #     if item['shift_name'] == '' or item['shift_name'] == False:
            #         shift_error.append(item['date'])
            #     if pd.isnull(item['id_y'] ):
            #         shift_null.append((item['date'],  item['shift_name']))
            # date_str_array = [x for x in self.date_array if not x.strftime('%Y-%m-%d') in date_str_data_array]        
            # print(f"{item['code']}")
            # print(f"{item['part_time_department_name']}")
            employees_scheduling = self.df_scheduling_ver[self.df_scheduling_ver['employee_code']==item['code']]

            for date_item in self.date_array:
                if (date_item in employees_scheduling['date'].values):
                    continue
                    # print ('date exist', date_item)
                    # print(f"{item['code']}")
                    # print(item)
                    # select_item_array = employees_scheduling[employees_scheduling['date']==date_item]
                    # select_item = select_item_array.iloc[0]
                    # print('this is id', select_item['id'])
                    # ids = [select_item['id'].item()]
                    # update_data = {'employee_code': item['code'], 'employee_name':item['name'],'date': date_item.strftime('%Y-%m-%d') ,'shift_name':'-'}
                    # update_data['department'] = item['department_name']
                    # update_data['company'] = self.company_info['name']
                    # id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                    # print('update thanh cong', update_data)

                    # update_data = {'employee_code': g, 'employee_name':data.iloc[0]['name'],'date': date_blank.strftime('%Y-%m-%d') ,'shift_name':'-'}
                    # update_data['department'] = False if not data.iloc[0]['department_id'] else data.iloc[0]['department_id'][1]
                    # update_data['company'] = self.company_info['name']
                    # id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'create', [update_data])
                    # print('Them thanh cong', update_data)
                # else:
                #     update_data = {'employee_code': item['code'], 'employee_name':item['name'],'date': date_item.strftime('%Y-%m-%d') ,'shift_name':'-'}
                #     update_data['department'] = item['department_name']
                #     update_data['company'] = self.company_info['name']
                #     try:
                #         id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'create', [update_data])
                    
                #         print('Them thanh cong', update_data)
                #     except Exception as ex:
                #         logger.error(f"create attendance report: {ex}")
        
        return result

    def import_roster_data(self,input_file, sheet_name='01-20.08', progress_callback = None):
        self.list_error_shift = []
        self.list_error_employee = []
        # models.execute_kw(self.db, uid, password, 'res.partner', 'check_access_rights', ['read'], {'raise_exception': False})
        # models.execute_kw(self.db, uid, password, 'res.partner', 'search', [[['is_company', '=', True]]])
        # department_ids = models.execute_kw(self.db, uid, password, 'hr.department','search', [[]])
        # list_departments  = models.execute_kw(self.db, uid, password, 'hr.department', 'read', [department_ids], {'fields': ['name', 'total_employee', 'company_id', 'member_ids']})

        # ids = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'search', [[]], {})
        # self.list_shifts  = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'read', [ids], {'fields': ['id', 'name', 'start_work_time', 'end_work_time']})

        # self.download_employee_info()

        ref_workbook= openpyxl.load_workbook(input_file)
        ws = ref_workbook[sheet_name]
        header_start_row = 1
        print(ws)
        for i in range(1,50):
            for j in range(1, 50):
                try:
                    if i>=5 and i<=7:
                        print( ws.cell(row=i, column=j).value)
                    if '"CN","T2","T3","T4","T5","T6","T7"' in f'{ws.cell(row=i, column=j).value}':
                        header_start_row= i
                        break
                except Exception as ex:
                    pass
                    # print(ex)
        print(header_start_row)
        header_start_row = header_start_row - 1
        df_dict = pd.read_excel(input_file, header=[header_start_row-2, header_start_row-1, header_start_row], 
                                sheet_name=sheet_name)
        
        self.group_range_date = []
        self.msnv_column_header = ''
        self.name_column_header = ''
        for item in list(df_dict.columns.values):
            try:
                print(item)
                if item[2]== 'T2'or item[2]== 'T3' or item[2]== 'T4' or item[2]== 'T5'or item[2]== 'T6'or item[2]== 'T7' or item[2]== 'CN':
                    self.group_range_date.append(item)
                elif (item[0].strip().upper() == 'MSNV' )or (item[1].strip() == 'Mã NV' ):
                    self.msnv_column_header = item
                elif (item[0] == 'HỌ VÀ TÊN') or (item[1].strip() == 'Họ & Tên'):
                    self.name_column_header = item
                # print('check:',item)
            except:
                print(item)
        # df_dict = df_dict[df_dict[self.msnv_column_header]=='APS210120002']
        print('aaaaaaaaâ', self.msnv_column_header)
        df_employees = self.df_employees[self.df_employees['employee_ho'] == True]

        df_dict = df_dict.merge(df_employees[['id','code', 'department_id', 'name','job_title','time_keeping_code', 'company_name','part_time_department_name']], \
                left_on=[self.msnv_column_header], right_on = ['code'], how='left', suffixes=( '' ,'_employee' ))
        df_dict.to_excel('df_dict.xlsx')
        df_dict[df_dict['name'].notnull()].apply(lambda row: self.process_content_dropship(row), axis=1, result_type='expand') 

    def process_content_dropship(self, row):
        msnv = row[self.msnv_column_header]
        # find id nhan vien
        id_nv = row['id']
        
        
        if id_nv =='':
            print(f"Nhan vien: {msnv} Khong ton tai")
            self.list_error_employee.append(msnv)
            return
        # group = self.group_range_date
            
        start_date = self.group_range_date[0][1].strftime("%Y-%m-%d")
        end_date = self.group_range_date[len(self.group_range_date)-1][1].strftime("%Y-%m-%d")
        
       
        for group_item in self.group_range_date:
            try:
                shift_name = row[group_item].strip()
            except:
                print("ex: ", row[group_item])
                self.list_error_shift.append(row[group_item])
                continue
            date_str = group_item[1].strftime("%Y-%m-%d")
                # find shift value in list shift
            # shift_value = ''
            # for shift in self.list_shifts:
            #     if shift['name'] == shift_name:
            #         shift_value = shift['id']
            #         break
            try:
                # print(f'1 except {shift_name} -{row["company_name"]} - {self.df_shift["name"]}')
                
                filter_shift = self.df_shift[(self.df_shift['name'] == shift_name.strip() )
                                & (self.df_shift['company_name']==row['company_name'])]
            except:
                # print(f'1 except)
                filter_shift = self.df_shift[self.df_shift['name'] == f'{shift_name}'.strip()] 
                filter_shift  =  filter_shift[filter_shift['company_name']== row['company_name']]
                
            employees_scheduling = self.df_scheduling_ver[(self.df_scheduling_ver['employee_code']==row['code'] )& \
                                                            (self.df_scheduling_ver['date_str']==  date_str) &
                                                            (self.df_scheduling_ver['company'] ==  self.company_info['name'])]

            update_data = {'employee_name':row['name'], 'date': date_str, 
                    'attendance_attempt_1': False, 'attendance_attempt_2': False, 'attendance_attempt_3': False,
                    'attendance_attempt_4': False, 'attendance_attempt_5': False, 'attendance_attempt_6': False, 
                    'attendance_attempt_7': False, 'attendance_attempt_8': False, 'attendance_attempt_9': False, 
                    'attendance_attempt_10': False, 'attendance_attempt_11': False, 'attendance_attempt_12': False,
                    'attendance_attempt_13': False, 'attendance_attempt_14': False, 'attendance_attempt_15': False,
                    'last_attendance_attempt': False,'attendance_late':0, 'leave_early':0}
            try:
                resource_calendar_id = row['resource_calendar_id']
                update_data['resource_calendar'] = resource_calendar_id
            except:
                update_data['resource_calendar'] = '-'
                # update_data['missing_checkin_break'] = False
            update_data['department'] = row['part_time_department_name'] 
            update_data['company'] = self.company_info['name']
            update_data['additional_company'] = row['company_name']
            print('found shift',filter_shift )
            if len(filter_shift.index) >0:
                shift_value = filter_shift.iloc[0]['id']
                update_data['shift_name'] = shift_name
                update_data['shift_start'] = filter_shift.iloc[0]['start_work_time'].item()
                update_data['shift_end'] = filter_shift.iloc[0]['end_work_time'].item()
                update_data['rest_start'] = filter_shift.iloc[0]['start_rest_time'].item()
                update_data['rest_end'] = filter_shift.iloc[0]['end_rest_time'].item()
                update_data['rest_shift'] = filter_shift.iloc[0]['rest_shifts'].item()
                update_data['total_shift_work_time'] = filter_shift.iloc[0]['total_work_time'].item()
                update_data['night_shift'] = filter_shift.iloc[0]['night'].item()
                update_data['free_shift_time'] = filter_shift.iloc[0]['fix_rest_time'].item()
                # update_data[''] = filter_shift.iloc[0]['fix_rest_time']
                if self.progress_callback:
                    self.progress_callback.emit((10,f'{date_str}-{group_item[1]}-{shift_value} - {shift_name}'))
                else:
                    print(f'{date_str}-{group_item[1]}-{shift_value} - {shift_name}')
                # ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [["&",("employee_code", "=", msnv),("date", "=", date_str)]], {})

                if len(employees_scheduling.index) > 0:
                    count_scheduling = 0
                    for index_s, select_item in employees_scheduling.iterrows():
                        ids = [select_item['id']]
                        try:
                            if count_scheduling == 0:
                                print('found')
                                
                                print(f'{group_item[1]}-{shift_value} - {shift_name}-found and update-{ids}')
                                
                                self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                            else:
                                self.models.execute_kw(
                                        self.db, self.uid, self.password, 'hr.apec.attendance.report', 'unlink', [ids])
                        except Exception as ex:
                            logger.error(f"update attendance report: {ex}")
                        count_scheduling = count_scheduling + 1
                            
                    
                # else:
                #     try:
                #         id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'create', [update_data])
                        
                #         print("thanhcong CREATE",id_scheduling)
                #     except Exception as ex:
                #         logger.error(f"create attendance report else case: {ex}")
            else:
                if pd.notna(shift_name) and (not shift_name in self.list_error_shift) :
                    self.list_error_shift.append(shift_name)
                # remove
                update_data['shift_name'] = False
                if len(employees_scheduling.index) > 0:
                    count_scheduling = 0
                    for index_s, select_item in employees_scheduling.iterrows():
                        ids = [select_item['id']]
                        try:
                            if count_scheduling == 0:
                                print('found')
                                # select_item = scheduling_data
                                
                                print(f'{group_item[1]} -REMOVE SHIFT-{ids}')
                                self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                            else:
                                self.models.execute_kw(
                                        self.db, self.uid, self.password, 'hr.apec.attendance.report', 'unlink', [ids])
                        except Exception as ex:
                            logger.error(f"unlink or create attendance report: {ex}")
                        count_scheduling = count_scheduling + 1
                # else:
                #     try:
                #         print(f'{group_item[1]} -REMOVE SHIFT-create blank')
                #         # created blank
                #         id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'create', [update_data])
                #     except Exception as ex:
                #         logger.error(f"create blank attendance report: {ex}")
                
        try:
            print('{}-{}'.format(start_date, end_date))
            # id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'employees_scheduling', 'create', [{'start': start_date, 'stop': end_date ,'shift_monday':shift_monday,
            #         'shift_tuesday':shift_tuesday,'shift_wednesday':shift_wednesday, 'shift_thursday':shift_thursday, 
            #         'shift_friday': shift_friday,'shift_saturday': shift_saturday, 'shift_sunday': shift_sunday,'employee_ids':[id_nv]}])
            # print("thanhcong ",id_scheduling)
        except Exception as ex:
            print(ex)

                    
                        
        # print(start_date.strftime("%Y-%m-%d"))
        print(row[self.name_column_header])
    def copy_shift_ho(self, from_date, to_todate, filter_company_array=None):
        if filter_company_array != None:
            hr_ho_employee = self.df_employees[(self.df_employees['company_sid'].isin(filter_company_array)) & 
                            (self.df_employees['employee_ho']==True)]
            scheduling_data = {}
            for index, item in hr_ho_employee.iterrows():
                print(item)
                input_shift_array = self.df_scheduling_ver[(self.df_scheduling_ver['date']<= to_todate)&
                    (self.df_scheduling_ver['date']>= from_date) &
                    (self.df_scheduling_ver['employee_code']== item['code'])]
                input_shift_array['day_of_week'] = input_shift_array['date'].apply(lambda x: x.weekday())
                for date_item in self.date_array:
                    input_shift_filter = input_shift_array[input_shift_array['day_of_week']==date_item.weekday()]
                    shift_name = (input_shift_filter.iloc[0]['shift_name']).strip()
                    scheduler_item = {'employee_code': item['code'], 'shift_code': shift_name, 
                                    'company_name': item['company_name']}
                    scheduling_data.append(scheduler_item)
            df_shift = pd.DataFrame.from_dict(scheduling_data)
            df_shift = df_shift.merge(self.df_shift[['name', 'company_name']], left_on= ['Mã ca', 'Công ty'], right_on=['name', 'company_name'], how='left')
                
    def import_shift_ho(self, filepath):
        df_shift = pd.read_excel(filepath, header=[0,])
        df_shift = df_shift.merge(self.df_shift[['name', 'company_name']], left_on= ['Mã ca', 'Công ty'], right_on=['name', 'company_name'], how='left')
        # df_shift = df_shift[df_shift['name'].isna()]
        df_shift = df_shift.merge(self.df_all_company, left_on='Công ty', right_on='name', how='left', suffixes=( '' ,'_company' ))
        df_shift = df_shift[~df_shift['id'].isna()]
        df_shift.to_excel('df_shift_ho.xlsx')
        for index, row in df_shift.iterrows():
            print(row)
            start_work_time = self.hour_to_float(row['Giờ bắt đầu làm:'])
        
            end_work_time = self.hour_to_float(row['Giờ kết thúc làm:'])

            start_rest_time =self.hour_to_float(row['Giờ bắt đầu nghỉ:'])
            end_rest_time= self.hour_to_float(row['Giờ kết thúc nghỉ:'])
            rest_shifts = row['Ca nghỉ']
            fix_rest_time = row['Ca gãy']
            
            update_data = {'name': row['Mã ca'], 'company_id': row['id'], 'start_work_time': start_work_time,
                'end_work_time': end_work_time,'start_rest_time':start_rest_time,'end_rest_time': end_rest_time,
                'rest_shifts': rest_shifts, 'fix_rest_time': fix_rest_time }
            print(update_data)
            # print(row)
            ids = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'search', [["&",("name", "=", row['Mã ca']),("company_id", "=",  row['id'])]], {})
            # if len(ids) > 0:
            # #     print('Số lần chấm công: ',row['Số lần chấm công'])
            # #     if row['Số lần chấm công'] > 0:
            #     try:
            #         update_data={'start_rest_time':start_rest_time,'end_rest_time': end_rest_time}
            #         self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'write', [ids, update_data])
            #         print('updated')
            #     except Exception as ex:
            #         logger.error(f"update shift : {ex}")
            # else:
            #     try:
            #         id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'create', [update_data])
                
            #         print("thanhcong CREATE",id_scheduling)
            #     except Exception as ex:
            #         logger.error(f"create shift: {ex}")

    def hour_to_float(self, hour_str):
        hour_array = hour_str.split(':') 
        hour = int(hour_array[0])
        minute = int(hour_array[1])
        minute = round(minute/60.0, 2)
        return hour+minute
    def merge_calendar_leaves(self):

        print('***************CALENDAR************************************')
        # try:
        #     self.df_resource_calendar_leaves[['date_from', 'date_to']]= self.df_resource_calendar_leaves.apply(lambda row: self.refact_date_from(row), axis=1, result_type='expand')
        # except:
        #     self.df_resource_calendar_leaves = pd.DataFrame(columns=['date_from', 'date_to'])
        # self.df_resource_calendar_leaves['date_to']=self.df_resource_calendar_leaves.apply(lambda row: row['date_to'].replace(hour=23,minute=59,second=59), axis=1)
        print(self.df_resource_calendar_leaves)
        # print(list_scheduling_ver)
        self.df_normal_attendances[['is_holiday', 'holiday_from', 'holiday_to', 'holiday_name']] = \
            self.df_normal_attendances.apply(lambda row: self.merge_holiday(row), axis=1, result_type='expand')
    # find nearest ho
    def find_nearest_ho(self, row):
        # print(self.df_normal_attendances['employee_id'])
        # print(row['employee_id'])
        is_holiday = False
        holiday_from = None
        holiday_to = None
        holiday_name = ''
        shift = 0
        shift_id = 0
        normal_time = 0
        rest_shifts = False
        fix_rest_time = False
        night = False
        label = ''
        scheduling = -1
        total_shift_work_time = 0
        shift_start = 12
        shift_end = 12
        rest_start = 12
        rest_end = 12
        report_sid = -1
        in_out= False
        # try: 
        df_compare = self.df_normal_attendances[self.df_normal_attendances['time_keeping_code']==int(row['ID'])].sort_values(by=['time'])
        if len(df_compare['time'])>0:
            dates = df_compare['time'].to_list()
            # print(dates)
            normal_shift_row = df_compare.iloc[0]
            for index, time_item in df_compare.iterrows():
                try:
                    if time_item['time'].date() == row["Giờ"].date():
                        normal_shift_row = time_item
                except Exception as ex:
                    print(f'errror nè {time_item["time"]}-{row["Giờ"]}-{type(time_item["time"])}-{type(row["Giờ"])}-{ex}')
                
            # print('min: ', min_time)
            # normal_shift_row = df_compare.iloc[df_compare['time'].get_loc(row['Giờ'],method='nearest')]
            # print(df_compare['time'].searchsorted(row['Giờ']))
            # normal_shift_row = df_compare[df_compare['time']==min_time].iloc[0]
            scheduling = normal_shift_row['scheduling']
            report_sid = normal_shift_row['report_sid']
            shift = normal_shift_row['shift_name']
            # shift_id = normal_shift_row['shift']
            normal_time = normal_shift_row['time']
            
            rest_shifts = normal_shift_row['rest_shifts']
            if pd.notnull(normal_shift_row['total_work_time']):
                total_shift_work_time = normal_shift_row['total_work_time']
            if pd.notnull(normal_shift_row['start_work_time']):
                shift_start = normal_shift_row['start_work_time']
            if pd.notnull(normal_shift_row['end_work_time']):
                shift_end = normal_shift_row['end_work_time']
            if pd.notnull(normal_shift_row['start_rest_time']):
                rest_start = normal_shift_row['start_rest_time']
            if pd.notnull(normal_shift_row['end_rest_time']):
                rest_end = normal_shift_row['end_rest_time']

            fix_rest_time = normal_shift_row['fix_rest_time']
            night = normal_shift_row['night']
            if fix_rest_time:
                print(f"fix_rest_time: {shift}")
            label = normal_shift_row['label']
            is_holiday = normal_shift_row['is_holiday']
            holiday_from = normal_shift_row['holiday_from']
            holiday_to = normal_shift_row['holiday_to']
            holiday_name = normal_shift_row['holiday_name']
            in_out = row['in_out']

            # return shift, shift_id, normal_time, scheduling, rest_shifts, fix_rest_time, night ,label, is_holiday, holiday_from, holiday_to, holiday_name
        # except Exception as ex:
        #     print('Find nearest errrrrrrrrr: ', ex)
            
            
        return shift, normal_time, scheduling, rest_shifts, fix_rest_time, night ,\
                label, is_holiday, holiday_from, holiday_to, holiday_name, total_shift_work_time, \
                shift_start, shift_end, rest_start, rest_end, report_sid, in_out

    ########################### 'resource.calendar.leaves
    def merge_holiday(self, row):
        is_holiday = False
        holiday_from = None
        holiday_to = None
        holiday_name = ''
        try:
            shift_time = row['time']
            df_compare = self.df_resource_calendar_leaves[(self.df_resource_calendar_leaves['date_from']<=shift_time) & \
                                            (self.df_resource_calendar_leaves['date_to'] > shift_time) & 
                                            (self.df_resource_calendar_leaves['company_name']==row['additional_company'])]
            
            if len(df_compare.index) > 0:
                fist_row = df_compare.iloc[0]
                is_holiday = True
                holiday_from = fist_row['date_from']
                holiday_to = fist_row['date_to']
                holiday_name = fist_row['name']

        except Exception as ex:
            is_holiday = False

            print('merge error: ', ex)
            
        return is_holiday, holiday_from, holiday_to, holiday_name
    def process_shift_attendance(self):
        # self.join_shift_info()
    
        
        if self.is_ho:
            df_scheduling_ver = self.df_scheduling_ver[self.df_scheduling_ver['company']==self.company_info['name']]
            # df_scheduling_ver.to_excel('df_scheduling_ver_viet.xlsx')
            self.process_raw_data(df_scheduling_ver)
            self.merge_calendar_leaves()
            # self.df_old[['shift_name', 'normal_time', 'scheduling', 'rest_shifts', 'fix_rest_time', 'night' , 'label', \
            #     'is_holiday', 'holiday_from', 'holiday_to', 'holiday_name', 'total_shift_work_time', \
            #     'shift_start', 'shift_end', 'rest_start', 'rest_end', 'report_sid','in_out']] = self.df_old.apply(lambda row: self.find_nearest_ho(row), axis=1, result_type='expand') 

            self.df_old['ID'] = self.df_old['ID'].apply(lambda row: self.convert_code_to_int(row))
            self.df_employees['time_keeping_code'] = self.df_employees['time_keeping_code'].apply \
                        (lambda row: self.convert_code_to_int(row))
            self.df_old = self.df_old.merge(self.df_employees[self.df_employees['employee_ho']==True][['time_keeping_code','company_name','code']],
                    left_on=['ID'], right_on=['time_keeping_code'], how='left')
            # self.df_old['shift_name']=self.df_old['shift_name'].str.strip()
            # self.df_old = self.df_old.merge(self.df_shift[['id','name', 'total_work_time','company_name']], \
            #         left_on=['shift_name','company_name'], right_on = ['name','company_name'], how='left', suffixes=( '' ,'_y')).drop(['name'], axis=1)
            self.df_old = self.df_old.drop_duplicates(subset=["ID", "Giờ"], keep="last") \
                                .sort_values(['Giờ'])  
            df_scheduling_ver['time_keeping_code'] = df_scheduling_ver['time_keeping_code'].apply \
                        (lambda row: self.convert_code_to_int(row))
            # df_scheduling_ver = self.assign_time_to_shift(df_scheduling_ver)
            df_scheduling_ver[['has_attendance' ,'min_time_out' ,'max_time_in'] ] = \
                df_scheduling_ver.apply(lambda row: self.get_scheduling_time(row), axis=1, result_type='expand')
        else:
            df_scheduling_ver = self.df_scheduling_ver[self.df_scheduling_ver['company']==self.company_info['name']]
            self.df_old[['shift_name', 'normal_time', 'scheduling', 'rest_shifts', 'fix_rest_time', 'night' , 'label', \
                'is_holiday', 'holiday_from', 'holiday_to', 'holiday_name', 'total_shift_work_time', \
                'shift_start', 'shift_end', 'rest_start', 'rest_end', 'report_sid']] = self.df_old.apply(lambda row: self.find_nearest(row), axis=1, result_type='expand') 
            
        
            self.assign_time_to_shift(df_scheduling_ver)
        try:
            self.df_scheduling_ver.to_excel("rawdata_report.xlsx", sheet_name='Sheet1') 
        except Exception as ex:
            print(ex)
        try:
            self.df_old.to_excel("olddata_report.xlsx", sheet_name='Sheet1') 
        except Exception as ex:
            print(ex)
    def convert_code_to_int(self, row):
        try:
            return int(row)
        except:
            return 0
    def get_scheduling_time(self, row):
        has_attendance = False
        min_time_out = None
        max_time_in = None
        # df_attendence_group = 
        # df_scheduling_ver['has_attendance'] = False
        data = self.df_old[(self.df_old['Ngày']==row['date_str']) & (self.df_old['time_keeping_code']==row['time_keeping_code'])]
        update_data = {'attendance_attempt_1': False, 'attendance_attempt_2': False, 'attendance_attempt_3': False,
                    'attendance_attempt_4': False, 'attendance_attempt_5': False, 'attendance_attempt_6': False, 
                    'attendance_attempt_7': False, 'attendance_attempt_8': False, 'attendance_attempt_9': False, 
                    'attendance_attempt_10': False, 'attendance_attempt_11': False, 'attendance_attempt_12': False,
                    'attendance_attempt_13': False, 'attendance_attempt_14': False, 'attendance_attempt_15': False,
                    'attendance_inout_1': False, 'attendance_inout_2': False, 'attendance_inout_3': False,
                    'attendance_inout_4': False, 'attendance_inout_5': False, 'attendance_inout_6': False, 
                    'attendance_inout_7': False, 'attendance_inout_8': False, 'attendance_inout_9': False, 
                    'attendance_inout_10': False, 'attendance_inout_11': False, 'attendance_inout_12': False,
                    'attendance_inout_13': False, 'attendance_inout_14': False, 'attendance_inout_15': False,
                    'last_attendance_attempt': False, 
                    'night_shift': False}
        t_fist = None
        t_last = None
        t_mid_array = []
        inout_mid_array = []
        
        len_df = len(data['Giờ'])
        if (len_df > 1):
            t_last = data.iloc[len_df -1]['Giờ']

        index_col = 1
        if len_df>0:
            t_fist = data.iloc[0]['Giờ']
            for index, item in data.iterrows():
                update_data[f'attendance_attempt_{index_col}'] = item['Giờ'] 
                update_data[f'attendance_inout_{index_col}'] = item['in_out']
                if item['Giờ'] != t_fist and item['Giờ'] != t_last:
                    t_mid_array.append (item['Giờ'])
                    inout_mid_array.append(item['in_out'])

                update_data['last_attendance_attempt'] = item['Giờ']
                index_col = min(15, index_col+1)

        max_time_in = t_fist
        update_data['couple'] = self.find_couple(update_data)
        update_data['fix_rest_time'] = row['fix_rest_time'] if pd.notnull(row['fix_rest_time']) else False
        
        try:
            if max_time_in != None and  max_time_in.replace(second=0) <  row['shift_start_datetime']:
                max_time_in = max([a for a in data['Giờ'] if a.replace(second=0) <= row['shift_start_datetime']])

        except Exception as ex:
            # self.df_scheduling_ver.to_excel('nho ti ti.xlsx')
            print("nho ti ti: ", ex)
        min_time_out = t_last
        try:
            if min_time_out != None and  min_time_out.replace(second=0) > row['shift_end_datetime']:
                min_time_out = min([a for a in data['Giờ'] if a.replace(second=0) >= row['shift_end_datetime']])
        except Exception as ex:
            # self.df_old.to_excel('Hoi kho nhi.xlsx')
            print("Hoi kho nhi: ", ex)

        # df_scheduling_ver.at[g, 'min_time_out'] = min_time_out
        # df_scheduling_ver.at[g, 'max_time_in'] = max_time_in
        keys = update_data.keys()
        for key in keys:
            print(key, update_data[key], type(update_data[key]))

        update_data['id'] =  row['id']
        update_data['shift_name'] =  row['shift_name']
        update_data['rest_shift'] =  row['rest_shift']
        update_data['total_shift_work_time'] = row['total_shift_work_time']
        update_data['shift_start_datetime'] = row['shift_start_datetime']
        update_data['shift_end_datetime'] = row['shift_end_datetime']
        update_data['rest_start_datetime'] = row['rest_start_datetime']
        update_data['rest_end_datetime'] = row['rest_end_datetime']
        # update_data['shift_start_datetime']
        # if (self.s_shift == None) or (self.s_shift == update_data['shift_name']):
        self.calculate_actual_work_time_ho(update_data)
        return has_attendance ,min_time_out ,max_time_in
        

    def assign_time_to_shift(self, df_scheduling_ver):
        self.df_old['ID'] = self.df_old['ID'].apply(lambda row: self.convert_code_to_int(row))
        self.df_employees['time_keeping_code'] = self.df_employees['time_keeping_code'].apply \
                    (lambda row: self.convert_code_to_int(row))
        self.df_old = self.df_old.merge(self.df_employees[self.df_employees['employee_ho']==True][['time_keeping_code','company_name','code']],
                left_on=['ID'], right_on=['time_keeping_code'], how='left')
        self.df_old['shift_name']=self.df_old['shift_name'].str.strip()
        self.df_old = self.df_old.merge(self.df_shift[['id','name', 'total_work_time','company_name']], \
                left_on=['shift_name','company_name'], right_on = ['name','company_name'], how='left', suffixes=( '' ,'_y')).drop(['name'], axis=1)
        # find max time in and min timeout
        # df_attendence_group = self.df_old[~self.df_old['is_from_explanation']].sort_values(['Giờ']).groupby('scheduling')
        df_attendence_group = self.df_old.drop_duplicates(subset=["ID", "Giờ"], keep="last") \
                                .sort_values(['Giờ']).groupby('scheduling')
        df_scheduling_ver['has_attendance'] = False
        result = []
        for g, data in df_attendence_group:
            update_data = {'attendance_attempt_1': False, 'attendance_attempt_2': False, 'attendance_attempt_3': False,
                    'attendance_attempt_4': False, 'attendance_attempt_5': False, 'attendance_attempt_6': False, 
                    'attendance_attempt_7': False, 'attendance_attempt_8': False, 'attendance_attempt_9': False, 
                    'attendance_attempt_10': False, 'attendance_attempt_11': False, 'attendance_attempt_12': False,
                    'attendance_attempt_13': False, 'attendance_attempt_14': False, 'attendance_attempt_15': False,
                    'attendance_inout_1': False, 'attendance_inout_2': False, 'attendance_inout_3': False,
                    'attendance_inout_4': False, 'attendance_inout_5': False, 'attendance_inout_6': False, 
                    'attendance_inout_7': False, 'attendance_inout_8': False, 'attendance_inout_9': False, 
                    'attendance_inout_10': False, 'attendance_inout_11': False, 'attendance_inout_12': False,
                    'attendance_inout_13': False, 'attendance_inout_14': False, 'attendance_inout_15': False,
                    'last_attendance_attempt': False, 
                    'night_shift': False}

            t_fist = None
            t_last = None
            t_mid_array = []
            inout_mid_array = []
            t_fist = data.iloc[0]['Giờ']
            len_df = len(data['Giờ'])
            if (len_df > 1):
                t_last = data.iloc[len_df -1]['Giờ']
                has_attendance = True
            index_col = 1
            for index, item in data.iterrows():
                update_data[f'attendance_attempt_{index_col}'] = item['Giờ'] 
                update_data[f'attendance_inout_{index_col}'] = item['in_out']
                if item['Giờ'] != t_fist and item['Giờ'] != t_last:
                    t_mid_array.append (item['Giờ'])
                    inout_mid_array.append(item['in_out'])

                update_data['last_attendance_attempt'] = item['Giờ']
                index_col = min(15, index_col+1)
            
            update_data['couple'] = self.find_couple(update_data)

            # result_item = {'scheduling': g, 't_fist': t_fist, 't_last': t_last, 't_mid_array': t_mid_array}
            # result.append(result_item)
            if g>=0:
                df_scheduling_ver.at[g, 'has_attendance'] = True

                update_data['fix_rest_time'] = df_scheduling_ver.at[g, 'fix_rest_time']
                max_time_in = t_fist
                try:
                    if max_time_in != None and  max_time_in.replace(second=0) < df_scheduling_ver['shift_start_datetime'][g]:
                        max_time_in = max([a for a in data['Giờ'] if a.replace(second=0) <= df_scheduling_ver['shift_start_datetime'][g]])

                except Exception as ex:
                    # self.df_scheduling_ver.to_excel('nho ti ti.xlsx')
                    print("nho ti ti: ", ex)
                min_time_out = t_last
                try:
                    if min_time_out != None and  min_time_out.replace(second=0) > df_scheduling_ver['shift_end_datetime'][g]:
                        min_time_out = min([a for a in data['Giờ'] if a.replace(second=0) >= df_scheduling_ver['shift_end_datetime'][g]])
                except Exception as ex:
                    # self.df_old.to_excel('Hoi kho nhi.xlsx')
                    print("Hoi kho nhi: ", ex)

                df_scheduling_ver.at[g, 'min_time_out'] = min_time_out
                df_scheduling_ver.at[g, 'max_time_in'] = max_time_in

                
                # result_item = {'scheduling': g, 't_fist': t_fist, 't_last': t_last, 't_mid_array': t_mid_array}
                
                # result.append(self.df_scheduling_ver.at[g, 'min_time_out'])
                # Update to server
                # ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [["&",("employee_code", "=", msnv),("date", "=", date_str)]], {})
                # if len(ids) > 0:
                #     print(f'{group_item[1]}-{shift_value} - {shift_name}-found and update-{ids}')
                # update_data = {'attendance_attempt_1': False, 'attendance_attempt_2': False, 'attendance_attempt_3': False,
                #     'attendance_attempt_4': False, 'attendance_attempt_5': False, 'attendance_attempt_6': False, 
                #     'attendance_attempt_7': False, 'attendance_attempt_8': False, 'attendance_attempt_9': False, 
                #     'attendance_attempt_10': False, 'attendance_attempt_11': False, 'attendance_attempt_12': False,
                #     'attendance_attempt_13': False, 'attendance_attempt_14': False, 'attendance_attempt_15': False,
                #     'last_attendance_attempt': False, 
                #     # 'shift_start': data.iloc[0]['shift_start'].item(), 
                #     # 'shift_end': data.iloc[0]['shift_end'].item(), 
                #     # 'rest_start': data.iloc[0]['rest_start'].item(), 
                #     # 'rest_end': data.iloc[0]['rest_end'].item()
                #     'night_shift': False,}
                # for index_col in range (1,16):
                #     update_data[f'attendance_inout_{index_col}'] = False
                # for g, data in df_attendence_group:

                # update_data['attendance_inout_last'] = False
                # if pd.notnull(data.iloc[0]['shift_name']):   
                #     if data.iloc[0]['shift_name']:
                #         update_data['shift_name'] = f"{data.iloc[0]['shift_name']}"
                # if pd.notnull( data.iloc[0]['fix_rest_time']):
                #     update_data['split_shift'] = True if data.iloc[0]['fix_rest_time'] else False
                # update_data['company'] = self.company_info['name']
                # # department_name = df_scheduling_ver.at[g, 'department_name']
                # # update_data['department'] = department_name
                # # print('update department: ', update_data['department'])
                # if pd.notnull(data.iloc[0]['ID']):
                #     update_data['time_keeping_code'] = "{:06d}".format(int(data.iloc[0]['ID']))
                # if pd.notnull(data.iloc[0]['rest_shifts']):
                #     update_data ['rest_shift'] = True if data.iloc[0]['rest_shifts'] else False
                # if pd.notnull(data.iloc[0]['night']):
                #     update_data ['night_shift'] = True if data.iloc[0]['night'] else False

                # try:
                #     late_time = (max_time_in - df_scheduling_ver['shift_start_datetime'][g]).total_seconds()/60
                #     late_time = max(0, late_time)
                #     update_data['attendance_late'] = late_time
                #     early_time = (df_scheduling_ver['shift_end_datetime'][g] - min_time_out).total_seconds()/60
                #     early_time = max(0, early_time)
                #     update_data['leave_early'] = early_time
                # except:
                #     print('calculate late err')
                # try:
                #     if pd.notnull(data.iloc[0]['total_work_time']):
                #         update_data['total_shift_work_time'] = data.iloc[0]['total_work_time'].item()
                #     elif pd.notnull(df_scheduling_ver['total_work_time'][g]):
                #         update_data['total_shift_work_time'] = df_scheduling_ver['total_work_time'][g].item()
                #     else:
                #         update_data['total_shift_work_time'] = 0
                # except:
                #     print("update_data['total_shift_work_time'] error ")
                #     update_data['total_shift_work_time'] = 0
                    
                # if t_fist != None:
                    
                #     t_fist_str = t_fist.strftime('%Y-%m-%d %H:%M:%S')
                #     update_data['attendance_attempt_1'] = t_fist_str
                #     update_data['attendance_inout_1'] = data.iloc[0]['in_out']
                #     t_index = 2
                # else:
                #     t_index = 1
                # inout_index = 0
                # for t_itme in t_mid_array:
                #     t_fist_str = t_itme.strftime('%Y-%m-%d %H:%M:%S')
                #     update_data[f'attendance_attempt_{t_index}'] = t_fist_str
                #     update_data[f'attendance_inout_{t_index}'] = inout_mid_array[inout_index]
                #     t_index = t_index + 1
                #     inout_index = inout_index + 1
                #     if t_index > 15:
                #         break
                # if (update_data['attendance_inout_1'] == False )and (update_data['attendance_attempt_1'] !=False):
                #     update_data['attendance_inout_1'] = 'I'
                # if t_last != None:
                #     t_fist_str = t_last.strftime('%Y-%m-%d %H:%M:%S')
                #     update_data['last_attendance_attempt'] = t_fist_str
                #     update_data['attendance_inout_last']= data.iloc[len_df -1]['in_out']
                #     # if t_index <=15:
                #     #     update_data[f'attendance_attempt_{t_index}'] = t_fist_str
                #     #     update_data[f'attendance_inout_{t_index}'] = update_data['attendance_inout_last']
                # elif (t_fist != None) and (t_index == 2):
                #     update_data['last_attendance_attempt'] = update_data['attendance_attempt_1'] 
                #     update_data['attendance_inout_last'] = update_data['attendance_inout_1']
                # if (update_data['attendance_inout_last'] == False )and (update_data['last_attendance_attempt'] !=False):
                #     update_data['attendance_inout_last'] = 'O'
                # for t_index in range(1,16):

                #     if (update_data[f'attendance_attempt_{t_index}'] == update_data['last_attendance_attempt']) and \
                #             (update_data[f'attendance_inout_{t_index}'] == False ) and \
                #             (update_data[f'attendance_attempt_{t_index}']  !=False):
                #         update_data[f'attendance_inout_{t_index}'] = update_data['attendance_inout_last']
                # try:
                #     resource_calendar_id = df_scheduling_ver.at[g,'resource_calendar_id'][1]
                #     update_data['resource_calendar'] = resource_calendar_id
                #     shift_name = update_data['shift_name']
                #     if  (not '/' in shift_name ) and (update_data['total_shift_work_time']>0) and \
                #         ('44' in resource_calendar_id) and (update_data['attendance_attempt_3']) == False:
                        
                #         update_data['missing_checkin_break'] = True
                #         print('Thieu cham cong')
                #     else: 
                #         update_data['missing_checkin_break'] = False
                # except Exception as ex:
                #     update_data['resource_calendar'] = '-'
                #     update_data['missing_checkin_break'] = False
                #     print("calculate thieu cham cong err: ", ex)
                # try:
                keys = update_data.keys()
                for key in keys:
                    print(key, update_data[key], type(update_data[key]))
                ids = [int(data.iloc[0]['report_sid'])]
                print(f'update: {update_data} - {ids}')
                update_data['id'] =  df_scheduling_ver.at[g, 'id']
                update_data['shift_name'] =  df_scheduling_ver.at[g, 'shift_name']
                update_data['rest_shift'] =  df_scheduling_ver.at[g, 'rest_shift']
                update_data['total_shift_work_time'] = df_scheduling_ver.at[g, 'total_shift_work_time']
                update_data['shift_start_datetime'] = df_scheduling_ver.at[g, 'shift_start_datetime']
                update_data['shift_end_datetime'] = df_scheduling_ver.at[g, 'shift_end_datetime']
                update_data['rest_start_datetime'] = df_scheduling_ver.at[g, 'rest_start_datetime']
                update_data['rest_end_datetime'] = df_scheduling_ver.at[g, 'rest_end_datetime']
                # update_data['shift_start_datetime']
                # if (self.s_shift == None) or (self.s_shift == update_data['shift_name']):
                self.calculate_actual_work_time_ho(update_data)
                
                # except Exception as ex:
                #     print('data upload error: ', ex)
                # for cliend -side
                update_data['scheduling'] = g
                result.append(update_data)
        df_result_shift = self.df_attendence_hor= pd.DataFrame.from_dict(result)
        df_scheduling_ver['backup'] = df_scheduling_ver.index
        df_scheduling_ver = df_scheduling_ver.merge(df_result_shift, left_index=True, right_on = ['scheduling'], how='left', suffixes=( '' ,'_y'))
        # df_scheduling_ver.to_excel('self_df_scheduling_ver_v2.xlsx')
        return df_scheduling_ver
        # try:
        #     self.df_attendence_hor= pd.DataFrame.from_dict(result)
        #     self.df_scheduling_ver = self.df_scheduling_ver.merge(self.df_attendence_hor, left_index=True, right_on = ['scheduling'], how='left')
        # except Exception as ex:
        #     print('Merge err: ', ex)
        # self.df_scheduling_ver[~self.df_scheduling_ver['has_attendance']].apply(
        #         lambda row: self.remove_attendance(row), axis=1)

    def process_raw_data(self, df_scheduling_ver):
        df_scheduling_ver[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                df_scheduling_ver.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
        normal_attendances = []
        for index, row in df_scheduling_ver.iterrows():
            # print("index day nhe: ", self.df_scheduling_ver['id'][index])
            start_data = {"employee_code": row['employee_code'], 'time_keeping_code': int(row['time_keeping_code']), \
                'shift_name':row['shift_name'], 'time':row['shift_start_datetime'], 'scheduling': index, 'report_sid':row['id'],
                'rest_shifts':row['rest_shifts'],'fix_rest_time':row['fix_rest_time'], 'night':row['night'],'label': 'In', 
                'total_work_time': row['total_work_time'], 'start_work_time': row['shift_start'], 
                'end_work_time': row['shift_end'], 'start_rest_time': row['rest_start'],
                'end_rest_time': row['rest_end']}

            end_data = {"employee_code": row['employee_code'], 'time_keeping_code': int(row['time_keeping_code']), \
                'shift_name':row['shift_name'], 'time':row['shift_end_datetime'], 'scheduling': index, 'report_sid':row['id'],\
                'rest_shifts':row['rest_shifts'],'fix_rest_time':row['fix_rest_time'], 'night':row['night'],'label': 'Out', \
                'total_work_time': row['total_work_time'], 'start_work_time': row['shift_start'], 
                'end_work_time': row['shift_end'], 'start_rest_time': row['rest_start'],
                'end_rest_time': row['rest_end']}
            if self.is_ho:
                start_data['additional_company'] = row['additional_company']
                end_data['additional_company'] = row['additional_company']
            normal_attendances.append(start_data)
            normal_attendances.append(end_data)
        self.df_normal_attendances = pd.DataFrame.from_dict(normal_attendances)
        try:
            self.df_normal_attendances.to_excel("temp_df_normal_attendances.xlsx")
        except Exception as ex:
            print(ex)
    def process_kid_mode(self, row, hr_leave):
        
        kid_mod_stage_1_start_datetime = None
        kid_mod_stage_1_end_datetime = None
        kid_mod_stage_2_start_datetime = None
        kid_mod_stage_2_end_datetime = None
        is_normal_rest = True
        distance_from_start = (row['rest_start_datetime'] - row['shift_start_datetime']).total_seconds()//60
        if (distance_from_start < 360) or (distance_from_start >540):
            is_normal_rest = False
            
        start_work_time = row['start_work_time']
        end_work_time = row['end_work_time']
        start_rest_time =row['start_rest_time']
        if pd.isnull(start_rest_time):
            start_rest_time = row['rest_start']

        end_rest_time = row['end_rest_time']
        if pd.isnull(end_rest_time):
            end_rest_time = row['rest_end']
        if pd.isnull(start_work_time):
            start_work_time = row['shift_start']
        if pd.isnull(end_work_time):
            end_work_time = row['shift_end']
        rest_start_datetime = row['rest_start_datetime'] if is_normal_rest or pd.isnull(row['rest_start_datetime']) \
            else (row['shift_start_datetime'] + datetime.timedelta(hours=4))
        rest_end_datetime = (row['rest_end_datetime'] if is_normal_rest else (row['shift_start_datetime'] + datetime.timedelta(hours=5)))
        try:
            if (hr_leave['kid_mod_stage_1_start'] == 11) and (hr_leave['kid_mod_stage_2_end'] == 12):
                kid_mod_stage_1_start_datetime = rest_start_datetime - datetime.timedelta(hours=1)
                kid_mod_stage_1_end_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_2_start_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_2_end_datetime = kid_mod_stage_2_start_datetime + datetime.timedelta(minutes=30)
            
            elif (hr_leave['kid_mod_stage_1_start'] == 11.5) and (hr_leave['kid_mod_stage_2_end'] == 13.5):
                kid_mod_stage_1_start_datetime = rest_start_datetime - datetime.timedelta(minutes=30)
                kid_mod_stage_1_end_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_1_start_datetime = rest_end_datetime
                kid_mod_stage_2_end_datetime = kid_mod_stage_2_start_datetime + datetime.timedelta(minutes=30)
                
            elif (hr_leave['kid_mod_stage_1_start'] == 13) and (hr_leave['kid_mod_stage_2_end'] == 14.0):
                kid_mod_stage_1_start_datetime = rest_end_datetime
                kid_mod_stage_1_end_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_2_start_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_2_end_datetime = kid_mod_stage_2_start_datetime + datetime.timedelta(minutes=30)
            
            elif (hr_leave['kid_mod_stage_1_start'] == 16) and (hr_leave['kid_mod_stage_2_end'] == 17):
                kid_mod_stage_1_start_datetime = row['shift_end_datetime'] - datetime.timedelta(minutes=60)
                kid_mod_stage_1_end_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_2_start_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_2_end_datetime = kid_mod_stage_2_start_datetime + datetime.timedelta(minutes=30)
            
            elif (hr_leave['kid_mod_stage_1_start'] == 8) and (hr_leave['kid_mod_stage_2_end'] == 9):
                kid_mod_stage_1_start_datetime = row['shift_start_datetime']
                kid_mod_stage_1_end_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_2_start_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_2_end_datetime = kid_mod_stage_2_start_datetime + datetime.timedelta(minutes=30)
                
            else:    
                kid_mod_stage_1_start_datetime = row['shift_start_datetime']
                kid_mod_stage_1_end_datetime = kid_mod_stage_1_start_datetime + datetime.timedelta(minutes=30)
                kid_mod_stage_2_start_datetime = row['shift_end_datetime'] - datetime.timedelta(minutes=30)
                kid_mod_stage_2_end_datetime = kid_mod_stage_2_start_datetime + datetime.timedelta(minutes=30)
        except Exception as ex:
            print(ex)
            print("row['shift_start_datetime']: ", row['shift_start_datetime'])
            print("row['shift_end_datetime']: ", row['shift_end_datetime'])
            print("row['rest_end_datetime'] : ", row['rest_end_datetime'] )
            print("row['rest_start_datetime'] : ", row['rest_start_datetime'] )
            # input('process kid mode')
            
                
        return kid_mod_stage_1_start_datetime, kid_mod_stage_1_end_datetime, \
                    kid_mod_stage_2_start_datetime, kid_mod_stage_2_end_datetime
    def process_report_raw(self, row):
        # print(row)
        shift_start_datetime = None 
        shift_end_datetime = None
        rest_start_datetime = None
        rest_end_datetime = None
        try:
            start_work_time = row['start_work_time']
            if pd.isnull(start_work_time):
                start_work_time = row['shift_start']

            end_work_time = row['end_work_time']
            if pd.isnull(end_work_time):
                end_work_time = row['shift_end']
            start_rest_time =row['start_rest_time']
            if pd.isnull(start_rest_time):
                start_rest_time = row['rest_start']

            end_rest_time = row['end_rest_time']
            if pd.isnull(end_rest_time):
                end_rest_time = row['rest_end']
            h, m, s = float_to_hours(start_work_time)

            shift_start_datetime = row['date'].replace(hour=h, minute=m, second=s)

            h, m, s = float_to_hours(end_work_time)
            shift_end_datetime = row['date'].replace(hour=h, minute=m, second=s)
            shift_end_datetime = shift_end_datetime if end_work_time >= start_work_time else shift_end_datetime + datetime.timedelta(days=1)
            h, m, s = float_to_hours(start_rest_time)
            rest_start_datetime = row['date'].replace(hour=h, minute=m, second=s)
            rest_start_datetime = rest_start_datetime if start_rest_time >= start_work_time else rest_start_datetime + datetime.timedelta(days=1)
            h, m, s = float_to_hours(end_rest_time)
            rest_end_datetime = row['date'].replace(hour=h, minute=m, second=s)
            rest_end_datetime = rest_end_datetime if end_rest_time >= start_work_time else rest_end_datetime + datetime.timedelta(days=1)
        except Exception as ex:
            print(ex)
        return shift_start_datetime, shift_end_datetime, rest_start_datetime, rest_end_datetime 
    


    ################################# CALCULATE  VALUE *********************************
    def find_couple(self, row, list_out_leave_for_work):
        result = [[]]
        time_stack = []
        flags = []
        for index in range(1,16):
            flags.append('')
        for leave_item in list_out_leave_for_work:
            min_delete_index = 16
            max_delete_index = -1
            for index in range(15,0,-1):
                if (row[f'attendance_attempt_{index}'] >= leave_item['attendance_missing_from']) and \
                    (row[f'attendance_attempt_{index}'] <= leave_item['attendance_missing_to']):
                        flags[index] = 'x' # delete
                if min_delete_index < index:
                    min_delete_index = index
                if max_delete_index < index:
                    max_delete_index = index
            
            # for index in range(min_delete_index,0,-1):
                    
                    
                        
        for index in range(1,16):
            attendance_attempt = row[f'attendance_attempt_{index}']
            attendance_inout = row[f'attendance_inout_{index}']
            
            if attendance_inout == 'I':
                print(f'{attendance_inout} --- {attendance_attempt}')
                time_stack.append(index)
            elif attendance_inout =='O':
                if len(time_stack)>0:
                    result[0].append((time_stack[len(time_stack)-1], index))
                time_stack = []
        # print(f'{row["employee_code"]}',result)
        return result
    
    def find_couple_out_in(self, row):
        result = [[]]
        time_stack = []
        flags = []
        total_out_minute = 0
        for index in range(1,16):
            flags.append('')
                    
                        
        for index in range(1,16):
            attendance_attempt = row[f'attendance_attempt_{index}']
            attendance_inout = row[f'attendance_inout_{index}']
            
            if attendance_inout == 'O':
                print(f'{attendance_inout} --- {attendance_attempt}')
                time_stack.append(index)
            elif attendance_inout =='I':
                if len(time_stack)>0:
                    result[0].append((time_stack[len(time_stack)-1], index))
                    
                    in_index = time_stack[len(time_stack)-1]
                    out_index = index
                    real_time_in = row[f'attendance_attempt_{in_index}']
                    real_time_out = row[f'attendance_attempt_{out_index}']
                    update_data_item = self.calculate_actual_work_time_couple(row, real_time_in, real_time_out)
                    # update_data['total_work_time'] = update_data['total_work_time'] + update_data_item['total_work_time']
                    # update_data['night_hours_normal'] = update_data['night_hours_normal'] + update_data_item['night_hours_normal']
                    total_out_minute = total_out_minute + update_data_item['total_work_time']
                time_stack = []
        print(f'{row["employee_code"]} --- {result} - {total_out_minute}')
        # if (row['employee_code'] == 'APG230419006'):
        #     input('APG230419006')
        return result, total_out_minute
    
    def find_couple_in_out(self, df_scheduling_ver):
        df_scheduling_ver[['couple']] = df_scheduling_ver.apply(lambda row: self.find_couple(row), axis=1, result_type='expand')
        try:
            df_scheduling_ver.to_excel('couple.xlsx')
        except:
            print('save cople ex')

        return df_scheduling_ver

        

    def calculate_work_hour_values(self):
        # self.download_attendence_report()
        df_scheduling_ver = self.df_scheduling_ver[self.df_scheduling_ver['company']==self.company_info['name']]
        df_scheduling_ver.to_excel('df_scheduling_ver_viet.xlsx')
        # self.df_scheduling_ver[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
        #         self.df_scheduling_ver.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
        print("Now! calculate total")
        # self.df_scheduling_ver.to_excel("rawdata_report.xlsx", sheet_name='Sheet1') 
        if self.is_ho:
            print(df_scheduling_ver)
            # df_scheduling_ver = self.find_couple_in_out(df_scheduling_ver)
            df_scheduling_ver[['couple']] = df_scheduling_ver.apply(lambda row: self.find_couple(row), axis=1, result_type='expand')
            # calculate_actual_work_time_ho
            df_scheduling_ver.apply(lambda row: self.calculate_actual_work_time_ho(row), axis=1, result_type='expand')
        else:
            df_scheduling_ver[['total_work_time', 'night_hours_normal']] = df_scheduling_ver.apply(lambda row: self.calculate_actual_work_time(row), axis=1, result_type='expand')
        df_scheduling_ver.to_excel("self_rawdata_report_v2.xlsx", sheet_name='Sheet1') 
        return df_scheduling_ver
    
    def calculate_late_early_values(self):
        # self.download_attendence_report()
        
            self.df_scheduling_ver[['couple_out_in', 'total_out_ho']] = self.df_scheduling_ver.apply(lambda row: self.find_couple_out_in(row), axis=1, result_type='expand')
            # calculate_actual_work_time_ho
            # self.df_scheduling_ver[['total_out_ho']] = self.df_scheduling_ver.apply(lambda row: self.calculate_late_early_time_ho(row), axis=1, result_type='expand')
    
    def calculate_late_early_time_ho(self, row):
        result = 0
        # update_data = {'total_work_time':0,'night_hours_normal':0}
        print(row['couple_out_in'])
        if len(row['couple_out_in'])>0:
            for couple in row['couple_out_in']:
                if len(couple)<2:
                    continue
                in_index = couple[0]
                out_index = couple[1]
                real_time_in = row[f'attendance_attempt_{in_index}']
                real_time_out = row[f'attendance_attempt_{out_index}']
                update_data_item = self.calculate_actual_work_time_couple(row, real_time_in, real_time_out)
                # update_data['total_work_time'] = update_data['total_work_time'] + update_data_item['total_work_time']
                # update_data['night_hours_normal'] = update_data['night_hours_normal'] + update_data_item['night_hours_normal']
                result = result + update_data_item['total_work_time']
                
        return result
    
    def calculate_actual_work_time_ho(self, row):
        result = 0
        update_data = {'total_work_time':0,'night_hours_normal':0}
        for couple in row['couple'][0]:
            in_index = couple[0]
            out_index = couple[1]
            real_time_in = row[f'attendance_attempt_{in_index}']
            real_time_out = row[f'attendance_attempt_{out_index}']
            update_data_item = self.calculate_actual_work_time_couple(row, real_time_in, real_time_out)
            update_data['total_work_time'] = update_data['total_work_time'] + update_data_item['total_work_time']
            update_data['night_hours_normal'] = update_data['night_hours_normal'] + update_data_item['night_hours_normal']
            
        if row['id']>0:
            ids = [int(row['id'])]
            #         self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
            #         print(f'sucess: {update_data}')
            for col_index in range(1,16):
                try:
                    update_data[f'attendance_attempt_{col_index}'] = row[f'attendance_attempt_{col_index}'].strftime("%Y-%m-%d %H:%M:%S")
                    if row[f'attendance_inout_{col_index}'] in ['I', 'O']:
                        update_data[f'attendance_inout_{col_index}'] = row[f'attendance_inout_{col_index}']
                    else:
                        update_data[f'attendance_inout_{col_index}'] = False
                except Exception as ex:
                    print (f"update_data[f'attendance_attempt_{col_index}] {ex}")
                    update_data[f'attendance_attempt_{col_index}'] = False
                    update_data[f'attendance_inout_{col_index}'] = False
            try:
                self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
            except Exception as ex:
                logger.error(f"update attendance report when calculate: {ex}")
            list_scheduling_ver  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'read', [ids], 
                                                {'fields': ['id','employee_name', 'date', 'shift_name', 'employee_code', 
                                                'shift_start', 'shift_end', 'rest_start', 'rest_end', 'rest_shift', 'last_attendance_attempt','attendance_attempt_1']})
            print("ket qua")
            print(list_scheduling_ver)
    def calculate_actual_work_time_couple(self, row, real_time_in, real_time_out):
        """
        calculate actual work time with real time in, time out , restime in out


        """

        # print('aaaaaa')
        result = row['total_work_time']
        night_work_time = 0
        fix_rest_time = row['fix_rest_time']
        update_data = {'total_work_time': result, 'night_hours_normal': night_work_time}
        try:
            # try:
            #     rest_shifts = row['rest_shifts']
            # except:
            #     rest_shifts = row['rest_shift']
            # total_work_time = row['total_shift_work_time']
            # # print(f"{rest_shifts}-{total_work_time}")
            # if rest_shifts or total_work_time == 0:
            #     # print(f"{rest_shifts}-{total_work_time}")
            #     result = 0
            # else:
            # try:
                # real_time_in = row['attendance_attempt_1'].replace(second=0)
                # real_time_out = row['last_attendance_attempt'].replace(second=0)
                # print(f"{real_time_in}-{real_time_out}")
            if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
                # print(f"{real_time_in}-{real_time_out}")
                result = 0
            else:
                start_work_date_time = row['shift_start_datetime'].replace(second=0)
                end_work_date_time = row['shift_end_datetime'].replace(second=0)

                start_rest_date_time = row['rest_start_datetime'].replace(second=0)
                end_rest_date_time = row['rest_end_datetime'].replace(second=0)
                
                current_program = min(real_time_out.replace(second=0), start_rest_date_time) \
                                            - max(real_time_in.replace(second=0), start_work_date_time)
                stage_fist = max(0, current_program.total_seconds()//60.0)
                
                current_program = min(real_time_out.replace(second=0), end_work_date_time) \
                                            - max(real_time_in.replace(second=0), end_rest_date_time)
                stage_second = max(0, current_program.total_seconds()//60.0)

                result = int(stage_fist + stage_second)
                # print(start_work_date_time)
                # print(stage_fist)
                # print(end_work_date_time)
                # print(start_rest_date_time)
                # print(end_rest_date_time)
                # print(stage_second)
                # print(real_time_in)
                # print(real_time_out)
                if fix_rest_time:
                    result = min(result, 480)
                # night_work_time = self.calculate_night_work_time(row)
                # if ('/OFF' in row['shift_name']) or ('OFF/' in row['shift_name']):
                #     result = min(result, 240)
                # night_work_time = min(night_work_time, result)
        except Exception as ex:
            print(real_time_out)
            print(real_time_in)
            print("In cal actual", ex)
        update_data = {'total_work_time': result, 'night_hours_normal': night_work_time}
        # if result> 0:
        #     try:
        #         resource_calendar_id = row['resource_calendar_id']
        #         update_data['resource_calendar'] = resource_calendar_id
        #         shift_name = row['shift_name']
        #         if  (not '/' in shift_name ) and (row['total_shift_work_time']>0) and \
        #             '44' in resource_calendar_id and row['attendance_attempt_3'] == False:
                    
        #             update_data['missing_checkin_break'] = True
        #             print('Thieu cham cong')
        #         else: 
        #             update_data['missing_checkin_break'] = False
        #     except Exception as ex:
        #         update_data['resource_calendar'] = ''
        #         update_data['missing_checkin_break'] = False
        #         print("calculate thieu cham cong err: ", ex)
        # else:
        #     update_data['missing_checkin_break'] = False

        # print(f'night_work_time {night_work_time}')
        # print(f'total_work_time {result}-{real_time_in}-{real_time_out}-{start_rest_date_time}-{end_rest_date_time}--{stage_fist} + {stage_second}')
        
        
        return update_data

        # except Exception as ex:
        #     print("calcuate actual time err", ex)
        #     return 0

    def calculate_actual_work_time(self, row):
        """
        calculate actual work time with real time in, time out , restime in out


        """
        # print('aaaaaa')
        result = 0
        night_work_time = 0
        fix_rest_time = row['fix_rest_time']
        try:
            try:
                rest_shifts = row['rest_shifts']
            except:
                rest_shifts = row['rest_shift']
            total_work_time = row['total_shift_work_time']
            # print(f"{rest_shifts}-{total_work_time}")
            if rest_shifts or total_work_time == 0:
                # print(f"{rest_shifts}-{total_work_time}")
                result = 0
            else:
                try:
                    real_time_in = row['attendance_attempt_1'].replace(second=0)
                    real_time_out = row['last_attendance_attempt'].replace(second=0)
                    # print(f"{real_time_in}-{real_time_out}")
                    if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
                        # print(f"{real_time_in}-{real_time_out}")
                        result = 0
                    else:
                        start_work_date_time = row['shift_start_datetime'].replace(second=0)
                        end_work_date_time = row['shift_end_datetime'].replace(second=0)

                        start_rest_date_time = row['rest_start_datetime'].replace(second=0)
                        end_rest_date_time = row['rest_end_datetime'].replace(second=0)
                        
                        current_program = min(real_time_out, start_rest_date_time) - max(real_time_in, start_work_date_time)
                        stage_fist = max(0, current_program.total_seconds()//60.0)
                        
                        current_program = min(real_time_out, end_work_date_time) - max(real_time_in, end_rest_date_time)
                        stage_second = max(0, current_program.total_seconds()//60.0)

                        result = int(stage_fist + stage_second)
                        if fix_rest_time:
                            result = min(result, 480)
                        night_work_time = self.calculate_night_work_time(row)
                        
                        if ('/OFF' in row['shift_name'])or ('OFF/' in row['shift_name']):
                            result = min(result, 240)
                        night_work_time = min(night_work_time, result)
                except Exception as ex:
                    print("In cal actual", ex)
            update_data = {'total_work_time': result, 'night_hours_normal': night_work_time}
            if result> 0:
                try:
                    resource_calendar_id = row['resource_calendar_id']
                    update_data['resource_calendar'] = resource_calendar_id
                    shift_name = row['shift_name'].strip()
                    if  (not '/' in shift_name ) and (row['total_shift_work_time']>0) and \
                        '44' in resource_calendar_id and row['attendance_attempt_3'] == False:
                        
                        # update_data['missing_checkin_break'] = True
                        print('Thieu cham cong')
                    else: 
                        # update_data['missing_checkin_break'] = False
                        print('abc')
                except Exception as ex:
                    update_data['resource_calendar'] = ''
                    # update_data['missing_checkin_break'] = False
                    print("calculate thieu cham cong err: ", ex)
            else:
                # update_data['missing_checkin_break'] = False
                print('abc')

            # print(f'night_work_time {night_work_time}')
            # print(f'total_work_time {result}-{real_time_in}-{real_time_out}-{start_rest_date_time}-{end_rest_date_time}--{stage_fist} + {stage_second}')
            
            # if ((self.s_shift == None) or (self.s_shift == row['shift_name'])) and (row['id']>0):
            #     ids = [int(row['id'])]
            #     self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
            #     print(f'sucess: {update_data}')
            return result, night_work_time

        except Exception as ex:
            print("calcuate actual time err", ex)
            return 0, 0
    def calculate_night_work_time(self, row):
        """
        calculate night holiday work time with real time in, time out , restime in out
        cobine night holiday time
        """
        try:
            rest_shifts = row['rest_shifts']
            total_work_time = row['total_work_time']
            if rest_shifts or total_work_time == 0:
                return 0
            real_time_in = row['attendance_attempt_1'].replace(second=0)
            real_time_out = row['last_attendance_attempt'].replace(second=0)
            if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
                return 0
            start_work_date_time = row['shift_start_datetime'].replace(second=0)
            end_work_date_time = row['shift_end_datetime'].replace(second=0)

            start_rest_date_time = row['rest_start_datetime'].replace(second=0)
            end_rest_date_time = row['rest_end_datetime'].replace(second=0)
            
            

            # start night datetime from 22h to 6h:00
            start_night_date_time = start_work_date_time.replace(hour=22, minute=0, second=0)
            end_night_date_time = (start_night_date_time + datetime.timedelta(days=1)).replace(hour=6)
            
            
            current_program = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time)
            stage_fist = max(0, current_program.total_seconds()/60.0)
            
            current_program = min(real_time_out, end_work_date_time, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time)
            stage_second = max(0, current_program.total_seconds()/60.0)
            # print('{}: night after now: {} - {} - {} -night - {}- {} - real - {} - {} '.format(row['name_employee'] ,row['name'], \
            #                     stage_fist, stage_second, start_night_date_time, end_night_date_time, row['start_work_date_time'], row['end_work_date_time']))
            stage_night =  int(stage_fist + stage_second)

            end_night_date_time = start_work_date_time.replace(hour=6, minute=0, second=0)
            start_night_date_time = (end_night_date_time+ datetime.timedelta(days=-1)).replace(hour=22, minute=0, second=0)
           
            current_program = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time)
            stage_fist = max(0, current_program.total_seconds()/60.0)
            
            current_program = min(real_time_out, end_work_date_time, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time)
            stage_second = max(0, current_program.total_seconds()/60.0)
            # print('{}: night before now: {} - {} - {} -night - {}- {} - real - {} - {} '.format(row['name_employee'] ,row['name'], \
            #                     stage_fist, stage_second, start_night_date_time, end_night_date_time, row['start_work_date_time'], row['end_work_date_time']))
            
            stage_morning = int(stage_fist + stage_second)
            return stage_night + stage_morning

        except Exception as ex:
            # print('{}: night - holiday: {} - {} - {} -night - {}- {} - real - {} - {}'.format(row['name_employee'] ,row['name'], 0, 0, 0, 0, \
            #                                                                         row['start_work_date_time'], row['end_work_date_time']))
            print("calcuate night", ex)
            return 0

    def overide_time_keeping(self, trans_file_path):
        wb = load_workbook(filename = trans_file_path)
        for sheetname in wb.sheetnames:
            # try:
            print("file upload: ", sheetname)
            df_dict = pd.read_excel(trans_file_path, header=[0,], sheet_name= sheetname)
            id_col, hour_col, in_out_col = self.find_in_out_col(df_dict)
            if (in_out_col == None ) or (hour_col == None) or (id_col == None):
                print ('not found in out col: {} -{} -{}'.format(id_col, hour_col, in_out_col))
                continue
            else:
                print('found inout col:{} -{} -{}'.format(id_col, hour_col, in_out_col))
            df_dict.rename(columns={id_col: "ID"})
            id_col = 'ID'
            df_dict[hour_col] = df_dict[hour_col].apply(lambda x: self.convert(x))
            df_dict['Gio'] = df_dict[hour_col].apply(lambda x: self.time_to_string(x))
            # start_time = df_dict['Gio_raw'].min().strftime('%Y-%m-%d')
            # end_time = df_dict['Gio_raw'].max().strftime('%Y-%m-%d')
            df_dict = df_dict[(df_dict[hour_col]>=self.date_array[0]) & \
                        (df_dict[hour_col]<=self.date_array[len(self.date_array)-1])]
            # if in_out_col != None:
            #     df_dict[in_out_col'] = df_dict['Gio_raw'].apply(lambda x: self.time_to_string(x))

            # self.re_download_attandance(start_time, end_time)
            
            # df_dict['ID'] = df_dict[id_col].astype(str)
            df_dict[id_col]  = df_dict[id_col].apply(lambda row: self.convert_code_to_int(row))
            # index_notDigit = df_dict[~df_dict[id_col].apply(lambda x: x.isdigit())].index
            # df_dict.drop(index_notDigit, inplace=True)
            df_dict.drop_duplicates(subset=[id_col, hour_col], inplace=True) 
            # df_dict[hour_col] = all_data['Order Day new'].dt.strftime('%Y-%m-%d')
            # df = pd.merge(df_dict, self.df_old , on=['ID',hour_col], how="outer", indicator=True
            #       ).query('_merge=="left_only"')
            # df = pd.merge(df_dict, self.df_old , on=['ID','Gio'], how="outer", indicator=True
            #     ).query('_merge=="left_only"')
            df = df_dict.merge(self.df_old, left_on=[id_col, hour_col], right_on = ['ID','Giờ'], how='left', suffixes=( '' ,'_raw' ))
            df.to_excel(f'kq_test_may cham cong_{sheetname}.xlsx')
            print(df)
            df.apply(lambda row:self.update_attendance_trans_row(row, id_col=id_col, hour_col = hour_col, in_out_col= in_out_col), axis=1)
            for group_index, group_data in df.groupby([id_col, hour_col]):
                count = 0
                for index, row in group_data.iterrows():
                    if (count == len(group_data.index)-1) :
                        if (pd.isnull(row['id']) or pd.isnal(row['id'])):
                            print('CREATE')
                            self.update_attendance_trans_row(row, id_col=id_col, hour_col = hour_col, in_out_col= in_out_col)
                        else:
                            print('CREATE')
                            self.update_attendance_trans_row(row, id_col=id_col, hour_col = hour_col, 
                                    in_out_col= in_out_col, is_create=False)
                    else:
                        try:
                            print('REMOVE')
                            ids = [row['id']]
                            id_att_tran = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'unlink', 
                                    [ids])
                            print(f"remove trans: , {id_att_tran}")
                        except Exception as ex:
                            logger.error(f"Unlink trans attendance: {ex}")
                    count = count + 1
            # except Exception as ex:
            #     print('Inpimport ex:', ex)

    def update_attendance_trans_row(self, row, id_col = 'ID', hour_col = 'Giờ', location_col = None, in_out_col = None, is_create =True):
        if pd.isnull(row[id_col]) or pd.isna(row[id_col]):
            print('id null ')
            return
        try:
            finger_name = row[id_col]
            finger_name_int = finger_name
            print(f"{hour_col} - {row[hour_col]} - {row[hour_col]} - {finger_name_int}")
            try:
                time = row[hour_col].strftime("%Y-%m-%d %H:%M:%S")
            except:
                time = row[hour_col]
            try:
                day = row[hour_col].strftime("%Y-%m-%d")
            except:
                day = row[hour_col].split(' ')[0].strip()
            
            print(time)
            try:
                if location_col == None:
                    update_data = {'name': finger_name, 'time': time ,'day':day}
                    # id_att_tran = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'create', 
                    #     [update_data])
                    # print(f"trans: , {id_att_tran} - {finger_name} -{time} - {day}")
                else:
                    update_data = {'name': finger_name, 'time': time ,'day':day, 'location': f'{row[location_col]}'}
                    location = f'{row[location_col]}'
                    # id_att_tran = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'create', 
                    #     [update_data])
                    # print(f"trans: , {id_att_tran} - {finger_name} -{time} - {day}- {location}")
            except Exception as ex:
                print("location ex: ", ex)
                update_data = {'name': finger_name, 'time': time ,'day':day}
            if in_out_col != None:
                if row[in_out_col] == 'I':
                    update_data['in_out'] = 'I'
                elif row[in_out_col] == 'O':
                    update_data['in_out'] = 'O'
            try:
                if 'in' in f"{row[in_out_col]}".lower():
                    print(f"in  {row[hour_col]} - {row[id_col]}")
                    update_data['in_out'] = 'I'
                if 'out' in f"{row[in_out_col]}".lower():
                    print(f"out  {row[hour_col]} - {row[id_col]}")
                    update_data['in_out'] = 'O'
            except Exception as ex:
                print("Điểm sự kiện ex: ", ex)

            print("data up: ", update_data)
            try:
                if is_create:
                    print('CREATE')
                    id_att_tran = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'create', 
                            [update_data])
                    print(f"trans: , {id_att_tran} - {finger_name} -{time} - {day}")
                else:
                    ids = [row['id']]
                    id_att_tran = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'write', 
                            [ids, update_data])
                    print(f"update trans: , {id_att_tran} - {finger_name} -{time} - {day}")
            except Exception as ex:
                logger.error(f"create update trans attendance: {ex}")
        except Exception as ex:
            print('hour ', ex)

    def find_in_out_col(self, df_dict):
        result_ID = None
        result_inout = None 
        result_time = None
        for item in list(df_dict.columns.values):
            print(item)
            find_input_inout = df_dict[(df_dict[item]=='I') | (df_dict[item]=='O') | 
                        (df_dict[item].astype(str).str.contains("check in")).fillna(False) | (df_dict[item].astype(str).str.contains("check out").fillna(False))]
            if ("ID" in f'{item}') or ("Mã nhân viên" in f'{item}'):
                result_ID = item
            elif len(find_input_inout.index) > 10:
                result_inout = item
            elif ("Thời gian" in f'{item}') or ("Giờ" in f'{item}'):
                result_time = item
        return result_ID, result_time, result_inout
    def convert(self, x):
        try:
            if isinstance(x, datetime.datetime): 
                return x
            elif '.' in x:
                return datetime.datetime.strptime(x, "%d/%m/%Y %H:%M:%S.%f")
            else:
                return datetime.datetime.strptime(x, "%d/%m/%Y %H:%M:%S")
                
        except Exception as ex:
            print(f'time convert: {x}----{ex}')
            return x
    def time_to_string(self,x):
        try:
            return x.strftime('%Y-%m-%d %H:%M:%S')
        except Exception as ex:
            print('time_to_string', ex)
            return x
    def remove_duplicates_df_old(self):
        df_without_duplicates = self.df_old.drop_duplicates(
            subset=["ID", "Giờ"], keep="last")

        duplicates = self.df_old[~self.df_old.index.isin(df_without_duplicates.index)]
        duplicates.apply(lambda row: self.remove_duplicate_row_server(row), axis=1)

    def remove_duplicate_row_server(self, row):
        ids = [row['id']]
        self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'unlink', [ids])


    def update_department_ho(self, file_path, is_create_department=True):
        df_master_file = pd.read_excel(file_path, index_col=0) 
        if is_create_department and self.is_ho:
            for department_name in df_master_file['BỘ PHẬN TIẾNG VIỆT'].unique():
                print(department_name)
                department_ids = self.models.execute_kw(
                    self.db, self.uid, password, 'hr.department', 'search', [["&", ('company_id', '=', self.company_id), ('name','=',department_name)]])
                if len(department_ids)==0:
                    try:
                        id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.department', 'create', 
                            [{'name': department_name, 'company_id': self.company_id }])
                        print('created: ',id_deparment)
                    except Exception as ex:
                        logger.error(f"create deparment: {ex}")
        # get list department
        department_ids = self.models.execute_kw(
            self.db, self.uid, password, 'hr.department', 'search', [["&", ('company_id', '=', self.company_id), ('active', '=', True)]])
        list_departments = self.models.execute_kw(self.db, self.uid, password, 'hr.department', 'read', [department_ids], {
                                                  'fields': ['id', 'name', 'total_employee', 'company_id', 'member_ids', 'time_keeping_count']})
        df_departments = pd.DataFrame.from_dict(list_departments)

        df_employees = self.df_employees[self.df_employees['part_time_company_id'].astype(str).str.contains(self.company_info['name']).fillna(False)]

        df_master_file = df_master_file.merge(df_employees, left_on=['MÃ NV MIS'], right_on=['code'])

        df_master_file = df_master_file.merge(df_departments, left_on = ['BỘ PHẬN TIẾNG VIỆT'], right_on=['name'], suffixes=['','_department'])

        df_master_file[df_master_file['part_time_department_id']!=df_master_file['id_department']].apply(lambda row: self.update_part_time_deparment(row), axis =1)

        df_scheduling_ver = self.df_scheduling_ver[self.df_scheduling_ver['company'] == self.company_info['name']]

        df_scheduling_ver = df_scheduling_ver.merge(df_master_file, left_on=['employee_code'], right_on=['code'], suffixes=('','_master'))

        df_scheduling_ver[df_scheduling_ver['department'] != df_scheduling_ver['part_time_department_name']].apply(lambda row: self.update_department_scheduling_ver(row), axis =1)


    def update_part_time_deparment(self, row):
        try:
            update_data = {'part_time_department_id':row['id_department']}
            ids = [row['id']]
            self.models.execute_kw(self.db, self.uid, self.password, 'hr.employee', 'write', [ids, update_data])
            print('complete')
        except Exception as ex:
            logger.error(f"write employee: {ex}")

    def update_department_scheduling_ver(self, row):
        try:
            update_data = {'department':row['part_time_department_name']}
            ids = [row['id']]
            self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
            print('complete department', ids)
        except Exception as ex:
            logger.error(f"calculate create update trans attendance: {ex}")
        

