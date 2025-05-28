import sys, requests, json
import datetime
import openpyxl
import os
# from dotenv import load_dotenv
# load_dotenv()
from openpyxl.utils import get_column_letter
import re
import pandas as pd
import numpy as np
from collections import Counter
import xmlrpc.client
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors
from unities.unities import float_to_hours
import glob
from openpyxl import load_workbook
from pathlib import Path
import logging
import requests
# logging.basicConfig(filename='my_log.txt', format='%(asctime)s - %(levelname)s - %(name)s - %(message)s')
# logger = logging.getLogger('my_program')
# logger.info('This is an info message')
logging.basicConfig(filename='my_log.txt',
                    filemode='a',
                    format='%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S',
                    level=logging.DEBUG)

logging.info("Running HRMS")
logger = logging.getLogger('attendence_report')

url = 'https://hrm.mandalahotel.com.vn'
db = 'apechrm_product_v3'
username = 'admin'
password = 'admin'
class AttendenceReportModel():
    def __init__(self, start_date_str, end_date_str, server_url = None, server_db= None, server_username = None, 
        emplyee_code = None, 
        server_password= None, progress_callback = None, s_shift = None, is_load_common_info = False):

        super(AttendenceReportModel, self).__init__()
        self.employee_code = emplyee_code
        self.s_shift = s_shift
        if server_url:
            self.url = server_url
        else:
            self.url = url
        if server_db:
            self.db = server_db
        else:
            self.db = db
        if server_username:
            self.username = server_username
        else:
            self.username = username
        if server_password:
            self.password = server_password
        else:
            self.password = password
        self.date_array = pd.date_range(start=start_date_str, end=end_date_str)
        self.progress_callback = progress_callback
        
        common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(self.url))
        self.models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(self.url))
        self.uid = common.authenticate(self.db, self.username, self.password, {})
        self.user_profile = self.models.execute_kw(self.db, self.uid, self.password, 'res.users', 'read', [[self.uid] ], {'fields': ['id', 'name', 'company_id']})[0]
        self.company_id = self.user_profile['company_id'][0]
        # read shift infomation
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'search', [[("company_id","=",self.company_id)]], {})
        if self.progress_callback:
            self.progress_callback.emit((5, 'Start import'))

        if is_load_common_info:
            self.list_shifts  = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'read', [ids], {'fields': ['id', 'name', 'start_work_time', 
                                                'end_work_time','total_work_time','start_rest_time','end_rest_time',
                                                'rest_shifts', 'fix_rest_time', 'night']})
            if self.progress_callback:
                self.progress_callback.emit((10, 'Load shift completed'))

            self.list_shifts.append({'id':-1, 'name': False, 'start_work_time':12.00, 'end_work_time': 12.00, \
                    'total_work_time':0.00,'start_rest_time':12.00,'end_rest_time':12.00,'rest_shifts':True, 'fix_rest_time':False, 'night':False})
            # print (self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'fields_get', [], {'attributes': ['string', 'type']}))
            self.df_shift = pd.DataFrame.from_dict(self.list_shifts)
            try:
                self.df_shift.to_excel("raw_df_shift_test.xlsx")
            except Exception as ex:
                print(ex)
            print(self.df_shift['total_work_time']*60)

            print("(***************)")
            self.list_error_employee = []
            self.list_error_shift = []
        
                
            # self.sync_data_scheduling(None)
            self.selected_range_index = 0
            self.split_range_date()

            # Check 'resource.calendar.leaves
            resource_calendar_leaves_ids = self.models.execute_kw(self.db, self.uid, self.password, 'resource.calendar.leaves','search', [[]])
            resource_calendar_leaves_list  = self.models.execute_kw(self.db, self.uid, self.password, 'resource.calendar.leaves', 'read', [resource_calendar_leaves_ids], \
                                                            {'fields': ['id','name', 'company_id', 'calendar_id', 'date_from', 'date_to', 'resource_id', 'time_type']})
            self.df_resource_calendar_leaves= pd.DataFrame.from_dict(resource_calendar_leaves_list)

            department_ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.department','search', [[]])
            list_departments  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.department', 'read', [department_ids], {'fields': ['name', 'total_employee', 'company_id', 'member_ids', 'mis_id']})

            self.df_departments = pd.DataFrame.from_dict(list_departments)

            
            self.company_info  = self.models.execute_kw(self.db, self.uid, self.password, 'res.company', 'read', [[self.company_id]], \
                                                            {'fields': ['id','name','is_ho']})[0]
            self.is_ho = self.company_info['is_ho']
            if self.progress_callback:
                self.progress_callback.emit((100, 'Initialize Attendance model completed'))
    def merge_download_attendance (self):
        start_str = (self.date_array[0] + datetime.timedelta(days=-1)).strftime('%Y-%m-%d')
        end_str = (self.date_array[len(self.date_array)-1]+ datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        # self.df_old = pd.read_excel(self.input_attendence_file , index_col=None, header=[0,] ,sheet_name='Sheet1')
        # self.df_old['is_from_explanation'] = False
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'search', [["&",("day", ">=", start_str),("day", "<=", end_str)]], {'offset': 0, 'limit': 1000})
        list_attendance_trans  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'read', [ids], {'fields': ['id', 'name', 'day', 'time','in_out']})
        self.df_old = pd.DataFrame.from_dict(list_attendance_trans)
        # self.df_old['shift_name'] = self.df_old['shift_name'].str.strip()
        index = 1
        while len(ids) == 1000:
            ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'search', [["&",("day", ">=", start_str),("day", "<=", end_str)]], {'offset': index * len(ids), 'limit': 1000})
            list_attendance_trans  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'read', [ids], {'fields': ['id', 'name', 'day', 'time','in_out']})
            self.df_old = pd.concat([self.df_old, pd.DataFrame.from_dict(list_attendance_trans)], ignore_index=True)

            index = index + 1
            print(f'load {index}- {len(ids)}')
        # self.df_old = self.df_old[self.df_old['time'].apply(lambda x: type(x) == datetime.datetime)]
        self.df_old ['Gio'] = self.df_old ['time']
        self.df_old ['time'] = pd.to_datetime(self.df_old ['time'], format="%Y-%m-%d %H:%M:%S",errors='coerce')
        # self.df_old ['time'] += pd.Timedelta(hours=7)
        self.df_old = self.df_old.dropna(subset=['time'])
        self.df_old.rename(columns = {'name':'ID', 'time':'Giờ', 'day':'Ngày'}, inplace = True)
        try:
            self.df_old.to_excel('old.xlsx', sheet_name='Sheet1')
        except Exception as ex:
            print(ex)

    def update_attendance_trans_row(self, row, id_col = 'ID', hour_col = 'Giờ', location_col = None, in_out_col = None):
        try:
            finger_name = row[id_col]
            finger_name_int = finger_name
            print(f"{hour_col} - {row[hour_col]} - {row['Gio']} - {finger_name_int}")
            try:
                time = row[hour_col].strftime("%Y-%m-%d %H:%M:%S")
            except:
                time = row[hour_col]
            try:
                day = row[hour_col].strftime("%Y-%m-%d")
            except:
                day = row[hour_col].split(' ')[0].strip()
            
            print(time)
            if location_col == None:
                update_data = {'name': finger_name, 'time': time ,'day':day}
                # id_att_tran = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'create', 
                #     [update_data])
                # print(f"trans: , {id_att_tran} - {finger_name} -{time} - {day}")
            else:
                update_data = {'name': finger_name, 'time': time ,'day':day, 'location': f'{row[location_col]}'}
                
                # id_att_tran = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'create', 
                #     [update_data])
                # print(f"trans: , {id_att_tran} - {finger_name} -{time} - {day}- {location}")
            if (row['thang_cuon']):
                update_data['in_out'] = 'O'
            else:
                if in_out_col != None:
                    if row[in_out_col] == 'I':
                        update_data['in_out'] = 'I'
                    elif row[in_out_col] == 'O':
                        update_data['in_out'] = 'O'
                try:
                    if 'in' in f"{row['Điểm sự kiện']}".lower():
                        print(f"in  {row[hour_col]} - {row[id_col]}")
                        update_data['in_out'] = 'I'
                    if 'out' in f"{row['Điểm sự kiện']}".lower():
                        print(f"out  {row[hour_col]} - {row[id_col]}")
                        update_data['in_out'] = 'O'
                except Exception as ex:
                    print("Điểm sự kiện ex: ", ex)

            print("data up: ", update_data)
            id_att_tran = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'create', 
                    [update_data])
            print(f"trans: , {id_att_tran} - {finger_name} -{time} - {day}")
        except Exception as ex:
            print('hour ', ex)
            # print(day)
    def re_download_attandance(self, start_str, end_str):
        # start_str = (self.date_array[0] + datetime.timedelta(days=-1)).strftime('%Y-%m-%d')
        # end_str = (self.date_array[len(self.date_array)-1]+ datetime.timedelta(days=1)).strftime('%Y-%m-%d')
        # self.df_old = pd.read_excel(self.input_attendence_file , index_col=None, header=[0,] ,sheet_name='Sheet1')
        # self.df_old['is_from_explanation'] = False
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'search', [["&",("day", ">=", start_str),("day", "<=", end_str)]], {'offset': 0, 'limit': 1000})
        list_attendance_trans  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'read', [ids], {'fields': ['id', 'name', 'day', 'time']})
        self.df_old = pd.DataFrame.from_dict(list_attendance_trans)
        # self.df_old['shift_name'] = self.df_old['shift_name'].str.strip()
        index = 1
        while len(ids) == 1000:
            ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'search', [["&",("day", ">=", start_str),("day", "<=", end_str)]], {'offset': index * len(ids), 'limit': 1000})
            list_attendance_trans  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'read', [ids], {'fields': ['id', 'name', 'day', 'time']})
            self.df_old = pd.concat([self.df_old, pd.DataFrame.from_dict(list_attendance_trans)], ignore_index=True)

            index = index + 1
            print(f'load {index}- {len(ids)}')
        # self.df_old = self.df_old[self.df_old['time'].apply(lambda x: type(x) == datetime.datetime)]
        self.df_old ['Gio'] = self.df_old ['time']
        self.df_old ['time'] = pd.to_datetime(self.df_old['time'], format="%Y-%m-%d %H:%M:%S",errors='coerce')
        # self.df_old ['time'] += pd.Timedelta(hours=7)
        self.df_old = self.df_old.dropna(subset=['time'])
        self.df_old.rename(columns = {'name':'ID', 'time':'Giờ', 'day':'Ngày'}, inplace = True)
        try:
            self.df_old.to_excel('old_remake.xlsx', sheet_name='Sheet1')
        except Exception as ex:
            print(ex)

    def update_attendance(self, trans_file_path, in_out_col=None):
        wb = load_workbook(filename = trans_file_path)
        for sheetname in wb.sheetnames:
            try:
                print("file upload: ", sheetname)
                df_dict = pd.read_excel(trans_file_path, header=[0,], sheet_name= sheetname)
                df_dict['Gio_raw'] = df_dict['Giờ'].apply(lambda x: self.convert(x))
                df_dict['Gio'] = df_dict['Gio_raw'].apply(lambda x: self.time_to_string(x))
                start_time = df_dict['Gio_raw'].min().strftime('%Y-%m-%d')
                end_time = df_dict['Gio_raw'].max().strftime('%Y-%m-%d')

                # if in_out_col != None:
                #     df_dict[in_out_col'] = df_dict['Gio_raw'].apply(lambda x: self.time_to_string(x))

                self.re_download_attandance(start_time, end_time)

                df_dict['ID'] = df_dict['ID'].astype(str)
                df_dict['ID']  = df_dict['ID'].apply(lambda x: x.replace('.00','').replace('.0',''))
                index_notDigit = df_dict[~df_dict['ID'].apply(lambda x: x.isdigit())].index
                df_dict.drop(index_notDigit, inplace=True)
                df_dict.drop_duplicates(subset=["ID", "Gio_raw"], inplace=True) 
                # df_dict['Giờ'] = all_data['Order Day new'].dt.strftime('%Y-%m-%d')
                # df = pd.merge(df_dict, self.df_old , on=['ID','Giờ'], how="outer", indicator=True
                #       ).query('_merge=="left_only"')
                df = pd.merge(df_dict, self.df_old , on=['ID','Gio'], how="outer", indicator=True
                    ).query('_merge=="left_only"')
                df.to_excel(f'kq_test_may cham cong_{sheetname}.xlsx')
                print(df)
                df['thang_cuon'] = False
                df.apply(lambda row:self.update_attendance_trans_row(row,id_col='ID', hour_col = 'Gio_raw', in_out_col= in_out_col), axis=1)
            except Exception as ex:
                print('Inpimport ex:', ex)
    def datetime_from_logline(self, line):
        x = None
        ip = ''
        port = ''
        try:
            line_array = line.split(':')
            # print(line_array[0].split(' ')[0])
            time_data = line_array[0].split(' ')[0].split('/')
            day = int(time_data[1])
            month = int(time_data[0])
            year = int(time_data[2]) 
            x = datetime.datetime(year, month, day)
            if len(line_array) == 4: # case Khong thanh cong
                ip_port_array = line_array[3].split(',PORT')
                ip = ip_port_array[0].strip()
                port = ip_port_array[1].strip()
            elif len(line_array) == 3: # case thanh cong
                ip_port_array = line_array[2].split('Port')
                ip = ip_port_array[0].split('Kết nối thành công IP')[1].strip()
                port = ip_port_array[1].split(',')[0].strip()
        except Exception as ex:
            print(ex)
        return x, ip, port

    def read_log_file(self, location, app_name):
        result = []
        result_data = {}
        # try:
        folder_path = os.path.join(location, app_name)
        file_path = os.path.join(folder_path, 'LogFile.txt')
        print("read file path log: ", file_path)
        lines = []
        try:
            file = open(file_path, 'r',  encoding="utf8")
            while True:
                line = file.readline()
                if not line:
                    break
                lines.append(line)
        except Exception as ex:
            logger.error(f"File path {file_path} Create new  recorder: {ex}")   
            
        for line in lines:
            if 'Kết nối thành công IP ' in line:
                date, ip, port = self.datetime_from_logline(line)
                print(f"thanh cong: {os.path.basename(location)} - {app_name} - {date} - {ip} - {port} - {line}")
                result_item = {'location': os.path.basename(location), 'app_name': app_name, 'date': date, 'ip': ip, 'port': port, 'status': True}
                result.append(result_item)
            elif 'Kết nối không thành công IP: ' in line:
                date, ip, port  = self.datetime_from_logline(line)
                print(f"Khong thanh cong: {os.path.basename(location)} - {app_name} - {date} - {ip} - {port} - {line}")
                result_item = {'location': os.path.basename(location), 'app_name': app_name, 'date': date, 'ip': ip, 'port': port, 'status': False}
                result.append(result_item)
        # except Exception as ex:
        #     print("Ex in read_log_file: ", ex)
        return result
    def load_timerecorder_state(self):
        ids_recorder = self.models.execute_kw(
            self.db, self.uid, self.password, "hr.time.recorder", "search_read", [[]], {'fields': ['id','time_recorder_id', 'address_ip','connection', 
                                                                            'com_port','read_type', 'description','work_address']}
        )
        self.df_recorder = pd.DataFrame.from_dict(ids_recorder)

        ids_recorder_state = self.models.execute_kw(
            self.db, self.uid, self.password, "time.recorder.state.v2", "search_read", [[]], {'fields': ['id','state_id', 'time_recorder','date', 
                                                                            'state','address_ip', 'connection','work_address']}
        )
        self.df_recorder_state= pd.DataFrame.from_dict(ids_recorder_state)
    def update_attendance_from_ho_timekeeper(self):
        url = 'http://10.0.0.33:5984/event_data/_find'
        headers = {'Content-type': 'application/json', 'Accept': 'text/plain', "Authorization": "Basic YWRtaW46MTIzNDU2"}
        data = {
            "selector": {
                "time": {
                    "$gt": (self.date_array[0]).strftime(
                        "%Y-%m-%d"
                    ),
                    "$lt": (
                        datetime.datetime.now() + datetime.timedelta(days=3)
                    ).strftime("%Y-%m-%d"),
                }
            },
            "limit": 200000,
            "sort": [
                    {"time": "desc"}
                ]
        }
        # try:
       
        r = requests.post(url, data=json.dumps(data), headers=headers)
      
        list_result = r.json()['docs']
        print("data export: ", r.json())
        print("data now di dau")
        self.df_attendence_download_ho_timekeeper = pd.DataFrame.from_dict(list_result)
        self.df_attendence_download_ho_timekeeper.to_excel('df_attendence_download_ho_timekeeper.xlsx')
        self.df_attendence_download_ho_timekeeper['ID'] = \
                self.df_attendence_download_ho_timekeeper['cardId'].astype(str)

        self.df_attendence_download_ho_timekeeper['ID']  = \
            self.df_attendence_download_ho_timekeeper['ID'].apply(lambda x: x.replace('.00','').replace('.0',''))
        self.df_attendence_download_ho_timekeeper['Gio_raw'] = \
            self.df_attendence_download_ho_timekeeper['time'].apply(lambda x: self.convert(x))
        self.df_attendence_download_ho_timekeeper['Gio'] = \
            self.df_attendence_download_ho_timekeeper['Gio_raw'].apply(lambda x: self.time_to_string(x))
        self.df_attendence_download_ho_timekeeper['location'] = \
            self.df_attendence_download_ho_timekeeper['deviceId'] + " door:" + \
            self.df_attendence_download_ho_timekeeper['door']
            
        self.df_attendence_download_ho_timekeeper['in_out'] = self.df_attendence_download_ho_timekeeper['door'] \
            .apply(lambda x: 'I' if x in ['1','2',1,2] else 'O')
            
        index_notDigit = self.df_attendence_download_ho_timekeeper[~self.df_attendence_download_ho_timekeeper['ID']
                    .apply(lambda x: x.isdigit())].index
        self.df_attendence_download_ho_timekeeper.drop(index_notDigit, inplace=True)
        self.df_attendence_download_ho_timekeeper.drop_duplicates(subset=["ID", "Gio_raw"], inplace=True) 
        try:
            start_time = self.df_attendence_download_ho_timekeeper['Gio_raw'].min().strftime('%Y-%m-%d')
        except:
            print( self.df_attendence_download_ho_timekeeper['Gio_raw'].min())

        end_time = self.df_attendence_download_ho_timekeeper['Gio_raw'].max().strftime('%Y-%m-%d')

        # if in_out_col != None:
        #     df_dict[in_out_col'] = df_dict['Gio_raw'].apply(lambda x: self.time_to_string(x))

        self.re_download_attandance(start_time, end_time)
        # df_dict['Giờ'] = all_data['Order Day new'].dt.strftime('%Y-%m-%d')
        # df = pd.merge(df_dict, self.df_old , on=['ID','Giờ'], how="outer", indicator=True
        #       ).query('_merge=="left_only"')
        df = pd.merge(self.df_attendence_download_ho_timekeeper, self.df_old , on=['ID','Gio'], how="outer", indicator=True
            ).query('_merge=="left_only"')
        df.to_excel(f'kq_test_may cham cong_ho.xlsx')
        in_out_col = 'in_out'
        df['thang_cuon'] = False
        df.apply(lambda row:self.update_attendance_trans_row(row,id_col='ID', hour_col = 'Gio_raw',
                            location_col='location',
                            in_out_col= in_out_col), axis=1)
        # except Exception as ex:
        #     logger.error(f"collect error {ex}")
        
    def update_attendance_ho_from_excel(self, input_folder, df_time_recorder = None):
        self.load_timerecorder_state()
        parent_path = os.path.abspath(os.path.join(input_folder, os.pardir))
        print(parent_path)
        count = 0
        count_status = 0
        result = []
        # list_sub_folder = [x[0] for x in os.walk(parent_path) if os.path.exists(os.path.join('MCC APEC/Release'))]
        app_folder =  os.path.join(parent_path,"MCC APEC/Release")
        all_files = os.listdir(app_folder)
        print("all_files laaaa",all_files)
        df_async = False
        for file_path in all_files:
            row = os.path.basename(file_path)
            if not ('.xlsx' in file_path):
                continue
            try:
                # print(row)
                name_array = row.split('_')
                year = int(name_array[0])
                month = int(name_array[1])
                day = int(name_array[2])
                ip_address = (name_array[3])
                port = (name_array[4]).replace('.xlsx','')
                excel_file_path = os.path.join(app_folder, row)
                

                # print('month: infile ', month_in_file)
                date_in_file = datetime.datetime(year, month, day)
                if (date_in_file < self.date_array[0]) or (date_in_file > self.date_array[len(self.date_array)-1]):
                # if month_in_file < datetime.datetime.now().month -1:
                    print(f'Out of date: {date_in_file} - {row}')
                    continue
                csv_file_path = os.path.join(app_folder, row)
                if count == 0:
                    df_async = pd.read_excel(excel_file_path)
                    df_async['ip_address']= ip_address
                    if (ip_address in [ "10.0.180.173", "10.0.180.174"]):
                        df_async["in_out"]="O"
                    elif(ip_address in [ "10.0.180.171", "10.0.180.172"]):
                        df_async["in_out"]="I"
                    else:
                        df_async["in_out"]=""
                    df_async["port"]= port
                else:
                    df_item = pd.read_excel(excel_file_path)
                    df_item["port"]= port
                    df_item['ip_address']= ip_address
                    if (ip_address in [ "10.0.180.173", "10.0.180.174"]):
                        df_item["in_out"]="O"
                    elif(ip_address in [ "10.0.180.171", "10.0.180.172"]):
                        df_item["in_out"]="I"
                    else: 
                        df_item["in_out"]=""
                    df_async = pd.concat([df_async,df_item], ignore_index=True)
                    
                count = count+ 1
            except Exception as ex:
                print('file path', ex)
        print("df async", df_async)
        # self.df_attendence_download_ho_timekeeper = pd.DataFrame.from_dict(list_result)
        self.df_attendence_download_ho_timekeeper = df_async
        self.df_attendence_download_ho_timekeeper.to_excel('df_attendence_download_ho_timekeeper_excel.xlsx')
        self.df_attendence_download_ho_timekeeper['ID'] = \
                self.df_attendence_download_ho_timekeeper['User ID'].astype(str)

        self.df_attendence_download_ho_timekeeper['ID']  = \
            self.df_attendence_download_ho_timekeeper['ID'].apply(lambda x: x.replace('.000','').replace('.00','').replace('.0',''))
        self.df_attendence_download_ho_timekeeper['Gio_raw'] = \
            self.df_attendence_download_ho_timekeeper['Verify Date']
        self.df_attendence_download_ho_timekeeper['Gio'] = \
            self.df_attendence_download_ho_timekeeper['Gio_raw'].apply(lambda x: self.time_to_string(x))
        self.df_attendence_download_ho_timekeeper['location'] = \
            self.df_attendence_download_ho_timekeeper['ip_address'] + \
            self.df_attendence_download_ho_timekeeper['port'] 
            
            
        index_notDigit = self.df_attendence_download_ho_timekeeper[~self.df_attendence_download_ho_timekeeper['ID']
                    .apply(lambda x: x.isdigit())].index
        self.df_attendence_download_ho_timekeeper.drop(index_notDigit, inplace=True)
        self.df_attendence_download_ho_timekeeper.drop_duplicates(subset=["ID", "Gio_raw"], inplace=True) 
        try:
            start_time = self.df_attendence_download_ho_timekeeper['Gio_raw'].min().strftime('%Y-%m-%d')
        except:
            print( self.df_attendence_download_ho_timekeeper['Gio_raw'].min())

        end_time = self.df_attendence_download_ho_timekeeper['Gio_raw'].max().strftime('%Y-%m-%d')

        # if in_out_col != None:
        #     df_dict[in_out_col'] = df_dict['Gio_raw'].apply(lambda x: self.time_to_string(x))

        self.re_download_attandance(start_time, end_time)
        # df_dict['Giờ'] = all_data['Order Day new'].dt.strftime('%Y-%m-%d')
        # df = pd.merge(df_dict, self.df_old , on=['ID','Giờ'], how="outer", indicator=True
        #       ).query('_merge=="left_only"')
        df = pd.merge(self.df_attendence_download_ho_timekeeper, self.df_old , on=['ID','Gio'], how="outer", indicator=True
            ).query('_merge=="left_only"')
        df.to_excel(f'kq_test_may cham cong_ho_excel.xlsx')
        in_out_col = 'in_out'
        df['thang_cuon'] = False
        df.apply(lambda row:self.update_attendance_trans_row(row,id_col='ID', hour_col = 'Gio_raw',
                            location_col='location',
                            in_out_col= in_out_col), axis=1)
       
        
    def update_attendance_from_csv(self, input_folder, df_time_recorder = None):
        self.load_timerecorder_state()
        parent_path = os.path.abspath(os.path.join(input_folder, os.pardir))
        print(parent_path)
        app_names = ['App', 'App1','App2', 'App3','App4']
        count = 0
        count_status = 0
        result = []
        for app_name in app_names:
            list_sub_folder = [x[0] for x in os.walk(parent_path) if os.path.exists(os.path.join(x[0],app_name))]
            print(list_sub_folder)
            
            for item in list_sub_folder:
                print (item)
                item_base = os.path.basename(item)
                app_folder = os.path.join(item, app_name)
                result_log_file = self.read_log_file(item, app_name)
                if len(result_log_file)>0:
                    if count_status == 0:
                        df_result_log_file = pd.DataFrame.from_dict(result_log_file)
                    else:
                        df_result_log_file = pd.concat([df_result_log_file, pd.DataFrame.from_dict(result_log_file)], ignore_index=True)
                count_status += len(result_log_file)
                print('app_folder:', app_folder)
                # all_files = glob.glob(os.path.join(app_folder, "/*.csv")  )
                all_files = os.listdir(app_folder)
                
                # all_files = [x[0] for x in os.walk(app_folder)]
                # check_async.csv contain list file
                # check_async_path = os.path.join(app_folder,'check_async.csv')
                
                # if Path(check_async_path).is_file():
                    # print('names: ', all_files)
                
                for file_path in all_files:
                    row = os.path.basename(file_path)
                    if not ('.csv' in file_path):
                        continue
                    try:
                        # print(row)
                        name_array = row.split('_')
                        month_in_file = int(name_array[1])
                        day_in_file = int(name_array[0])
                        # print('month: infile ', month_in_file)
                        date_in_file = datetime.datetime(int(name_array[2]), month_in_file, day_in_file)
                        if (date_in_file < self.date_array[0]) or (date_in_file > self.date_array[len(self.date_array)-1]):
                        # if month_in_file < datetime.datetime.now().month -1:
                            print(f'Out of date: {date_in_file} - {row}')
                            continue
                        csv_file_path = os.path.join(app_folder, row)
                        if count == 0:
                            df_async = pd.read_csv(csv_file_path, header=None)
                            df_async[4]= item_base
                        else:
                            df_item = pd.read_csv(csv_file_path, header=None)
                            df_item[4]= item_base
                            df_async = pd.concat([df_async,df_item], ignore_index=True)
                        
                        count = count+ 1
                        select_item_result = {}
                        found = False
                        for item_result in result:
                            if (item_result['file'] == f"{name_array[3]}.{name_array[4]}.{name_array[5]}.{name_array[6]}" ) \
                                and (item_result['item_base'] == item_base) and (item_result['app_name'] == app_name):
                                found = True
                                select_item_result = item_result
                                break
                        select_item_result['file'] = f"{name_array[3]}.{name_array[4]}.{name_array[5]}.{name_array[6]}"
                        select_item_result[f"{name_array[0]}.{name_array[1]}.{name_array[2]}"] = True
                        select_item_result['folder'] = app_folder
                        select_item_result['item_base'] = item_base
                        select_item_result['app_name'] = app_name
                        if not found:
                            result.append(select_item_result)
                        # print(select_item_result)
                    except Exception as ex:
                        print('file path', ex)
        print (df_async)
        df_async['ID'] = df_async[0].astype(str)
        df_async['ID']  = df_async['ID'].apply(lambda x: x.replace('.00','').replace('.0',''))
        df_async['Gio_raw'] = df_async[1].apply(lambda x: self.convert(x))
        df_async['Gio'] = df_async['Gio_raw'].apply(lambda x: self.time_to_string(x))
        df_async['location'] = df_async[3]
        df_async['thang_cuon'] = df_async.apply(lambda x: (x[3] == '10.0.160.78') and ('HOI SO' in x[4]),  axis=1, result_type='expand')
        index_notDigit = df_async[~df_async['ID'].apply(lambda x: x.isdigit())].index
        df_async.drop(index_notDigit, inplace=True)
        last_day_lastmonth = (self.date_array[0] + datetime.timedelta(days=-1)).replace(hour=23, minute=59, second=59)
        fist_day_nextmonth = (self.date_array[len(self.date_array)-1] + datetime.timedelta(days=1)).replace(hour=0, minute=0, second=0)
        df_async = df_async[(df_async['Gio_raw'] < fist_day_nextmonth) & (df_async['Gio_raw'] > last_day_lastmonth)]
        start_time = df_async['Gio_raw'].min().strftime('%Y-%m-%d')
        end_time = df_async['Gio_raw'].max().strftime('%Y-%m-%d')
        print(f"download data from {start_time} - {end_time}")
        # sorting by first name 
        df_async.sort_values("ID", inplace=True) 
        
        # dropping ALL duplicate values 
        df_async.drop_duplicates(subset=["ID", "Gio_raw"], inplace=True) 
        df_async.to_csv('abc.csv')
        self.re_download_attandance(start_time, end_time)

        # df_dict['Giờ'] = all_data['Order Day new'].dt.strftime('%Y-%m-%d')
        df = pd.merge(df_async, self.df_old , on=['ID','Gio'], how="outer", indicator=True
              ).query('_merge=="left_only"')
        
        # try:
        #     df.to_excel('kq_test_may cham cong.xlsx')
        # except Exception as ex:
        #     print('kq_test_may cham cong.xlsx',ex)
        df.apply(lambda row:self.update_attendance_trans_row(row, 'ID','Gio_raw','location'), axis=1)

        df_result_sync = pd.DataFrame.from_dict(result)
        df_result_sync.to_excel('df_result_sync.xlsx')

        # update recorder
        if not(df_time_recorder is None):
            print('recorder_filter')
            print(self.df_recorder)
            df_time_recorder['address_ip'] = df_time_recorder.apply(lambda x: f"{x['IP']}:{x['PORT']}", axis=1, result_type='expand')
            df_time_recorder['connection'] = df_time_recorder.apply(lambda x: f"{x['RELATIVE-PATH']}:{x['APP-LOCATION']}", axis=1, result_type='expand')
            
            df_time_recorder_avail= self.df_recorder.merge(df_time_recorder, on=['connection', 'address_ip'])
            # self.df_recorder = self.df_recorder.merge(df_time_recorder, on=['connection', 'address_ip'],
            #         how='inner')
            for index, row in df_time_recorder_avail[df_time_recorder_avail['id'].notnull()].iterrows():
                print(row)
                ids = [row['id']]
                update_data={}
                is_need_update = False
                print('df_time_recorder')
                print(df_time_recorder_avail)
                
               
                if pd.notnull(row['NAME']):
                    if row['time_recorder_id'] != row['NAME']:
                        is_need_update = True
                        update_data['time_recorder_id'] = row['NAME']
                if pd.notnull(row['COMPANY']):
                    if row['work_address'] != row['COMPANY']:
                        is_need_update = True
                        update_data['work_address'] = row['COMPANY']
                if is_need_update== True:
                    
                    self.models.execute_kw(self.db, self.uid, self.password, 'hr.time.recorder', 'write', [ids, update_data]) 
                    
        df_recorder_in_log = df_result_log_file.drop_duplicates(['location','app_name','ip','port'])
        for index, recorder in df_recorder_in_log.iterrows():
            print('check: ', recorder)
            recorder_filter = self.df_recorder[(self.df_recorder['connection']==f"{recorder['location']}:{recorder['app_name']}") & \
                                                (self.df_recorder['address_ip']==f"{recorder['ip']}:{recorder['port']}")]
            if len(recorder_filter.index) == 0:
                update_data = {
                    'connection': f"{recorder['location']}:{recorder['app_name']}",
                    'address_ip': f"{recorder['ip']}:{recorder['port']}",
                    "com_port": f"{recorder['port']}"
                    }
                try:
                    id_recorder = self.models.execute_kw(self.db, self.uid, self.password, 'hr.time.recorder', 'create',
                                                    [update_data])
                    print('create new recorder: ', id_recorder)
                except Exception as ex:
                    logger.error(f"Create new recorder: {ex}")   
            
        print(self.df_recorder)
        df_recorder_state_in_log = df_result_log_file.drop_duplicates(['location','app_name','ip','port','date','status'], keep="last")

        # self.df_recorder_state
        try:
            self.df_recorder_state.to_excel('kq_test_may cham cong.xlsx')
        except Exception as ex:
            print('kq_test_may cham cong.xlsx',ex)
        for index, recorder in df_recorder_state_in_log.iterrows():
            try:
                print('check: ', recorder)
                recorder_state_filter = self.df_recorder_state[(self.df_recorder_state['connection']==f"{recorder['location']}:{recorder['app_name']}") & \
                                                    (self.df_recorder_state['address_ip']==f"{recorder['ip']}:{recorder['port']}") & \
                                                    ( self.df_recorder_state['date'] == recorder['date'].strftime("%Y-%m-%d 00:00:00"))]
                
                recorder_filter = self.df_recorder[(self.df_recorder['connection']==f"{recorder['location']}:{recorder['app_name']}") & \
                                                    (self.df_recorder['address_ip']==f"{recorder['ip']}:{recorder['port']}")]
                if (len(recorder_filter.index) > 0 ) and pd.notna(recorder['date']):
                    id_recorder = recorder_filter.iloc[0]['id']
                    update_data = {
                            'time_recorder': int(id_recorder),
                            'date': recorder['date'].strftime("%Y-%m-%d 00:00:00"),
                            'state': 'T' if recorder['status'] == True else 'F'}
                    if len(recorder_state_filter.index) == 0:
                        
                        try:
                            ids_recorder_state = self.models.execute_kw(self.db, self.uid, self.password, 'time.recorder.state.v2', 'create',
                                                            [update_data])
                            print('create new recorder state: ', ids_recorder_state)
                        except Exception as ex:
                            logger.error(f"Create new recorder: {ex}")  
            except Exception as ex:
                logger.error(f"Create new recorder : {recorder} : {ex}")  
    def time_to_string(self,x):
        try:
            return x.strftime('%Y-%m-%d %H:%M:%S')
        except Exception as ex:
            # print('time_to_string', ex)
            return x

    def convert(self, x):
        try:
            if isinstance(x, datetime.datetime): 
                return x
            
            x = x.strip()
            if '.' in x:
                return datetime.datetime.strptime(x, "%d/%m/%Y %H:%M:%S.%f")
            elif '-' in x:
                if 'T' in x:
                    if (x.index('-') == 4):
                        return datetime.datetime.strptime(x, "%Y-%m-%dT%H:%M:%S")
                    return datetime.datetime.strptime(x, "%d-%m-%YT%H:%M:%S")
                else:
                    if (x.index('-') == 4):
                        return datetime.datetime.strptime(x, "%Y-%m-%d %H:%M:%S")
                    return datetime.datetime.strptime(x, "%d-%m-%Y %H:%M:%S")
            else:
                return datetime.datetime.strptime(x, "%d/%m/%Y %H:%M:%S")
                
        except Exception as ex:
            print(f'time convert: {x}----{ex}--{x.index("-")}')
            return x

    def process_content_dropship(self, row):
        msnv = row[self.msnv_column_header]
        # find id nhan vien
        id_nv = row['id']
        
        
        if id_nv =='':
            print(f"Nhan vien: {msnv} Khong ton tai")
            self.list_error_employee.append(msnv)
            return
        # group = self.group_range_date
            
        start_date = self.group_range_date[0][1].strftime("%Y-%m-%d")
        end_date = self.group_range_date[len(self.group_range_date)-1][1].strftime("%Y-%m-%d")
        
       
        for group_item in self.group_range_date:
            try:
                shift_name = row[group_item].strip()
            except:
                print("ex: ", row[group_item])
                self.list_error_shift.append(row[group_item])
                continue
            date_str = group_item[1].strftime("%Y-%m-%d")
                # find shift value in list shift
            # shift_value = ''
            # for shift in self.list_shifts:
            #     if shift['name'] == shift_name:
            #         shift_value = shift['id']
            #         break
            try:
                filter_shift = self.df_shift[self.df_shift['name'] == shift_name.strip()]
            except:
                filter_shift = self.df_shift[self.df_shift['name'] == f'{shift_name}'.strip()]
            if len(filter_shift.index) >0:
                shift_value = filter_shift.iloc[0]['id']
                if self.progress_callback:
                    self.progress_callback.emit((10,f'{date_str}-{group_item[1]}-{shift_value} - {shift_name}'))
                
                ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [["&",("employee_code", "=", msnv),("date", "=", date_str)]], {})
                if len(ids) > 0:
                    print(f'{group_item[1]}-{shift_value} - {shift_name}-found and update-{ids}')
                    update_data = {'employee_name':row['name'], 'shift_name': shift_name}
                    update_data['department'] = False if not row['department_id'] else row['department_id'][1]
                    update_data['company'] = self.company_info['name']
                    self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                    
                else:
                    update_data = {'employee_code': msnv, 'employee_name':row['name'],'date': date_str ,'shift_name':shift_name}
                    update_data['department'] = False if not row['department_id'] else row['department_id'][1]
                    update_data['company'] = self.company_info['name']
                    id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'create', [update_data])
                    
                    print("thanhcong CREATE",id_scheduling)
            else:
                if pd.notna(shift_name) and (not shift_name in self.list_error_shift) :
                    self.list_error_shift.append(shift_name)
                # remove
                ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [["&",("employee_code", "=", msnv),("date", "=", date_str)]], {})
                if len(ids) > 0:
                    print(f'{group_item[1]} -REMOVE SHIFT-{ids}')
                    update_data = {'employee_name':row['name'], 'shift_name': False}
                    update_data['department'] = False if not row['department_id'] else row['department_id'][1]
                    update_data['company'] = self.company_info['name']
                    self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])

                else:
                    print(f'{group_item[1]} -REMOVE SHIFT-create blank')
                    # created blank
                    update_data = {'employee_code': msnv, 'employee_name':row['name'],'date': date_str ,'shift_name':False}
                    update_data['department'] = False if not row['department_id'] else row['department_id'][1]
                    update_data['company'] = self.company_info['name']
                    id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'create', [update_data])
                
        try:
            print('{}-{}'.format(start_date, end_date))
            # id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'employees_scheduling', 'create', [{'start': start_date, 'stop': end_date ,'shift_monday':shift_monday,
            #         'shift_tuesday':shift_tuesday,'shift_wednesday':shift_wednesday, 'shift_thursday':shift_thursday, 
            #         'shift_friday': shift_friday,'shift_saturday': shift_saturday, 'shift_sunday': shift_sunday,'employee_ids':[id_nv]}])
            # print("thanhcong ",id_scheduling)
        except Exception as ex:
            print(ex)

                    
                        
        # print(start_date.strftime("%Y-%m-%d"))
        print(row[self.name_column_header])

    def import_data(self, sheet_name='Sheet1', progress_callback = None):
        
        # models.execute_kw(db, uid, password, 'res.partner', 'check_access_rights', ['read'], {'raise_exception': False})
        # models.execute_kw(db, uid, password, 'res.partner', 'search', [[['is_company', '=', True]]])
        # department_ids = models.execute_kw(db, uid, password, 'hr.department','search', [[]])
        # list_departments  = models.execute_kw(db, uid, password, 'hr.department', 'read', [department_ids], {'fields': ['name', 'total_employee', 'company_id', 'member_ids']})

        # ids = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'search', [[]], {})
        # self.list_shifts  = self.models.execute_kw(self.db, self.uid, self.password, 'shifts', 'read', [ids], {'fields': ['id', 'name', 'start_work_time', 'end_work_time']})

        self.download_employee_info()

        ref_workbook= openpyxl.load_workbook(self.input_file)
        ws = ref_workbook[sheet_name]
        header_start_row = 1
        print(ws)
        for i in range(1,50):
            for j in range(1, 50):
                try:
                    if i>=5 and i<=7:
                        print( ws.cell(row=i, column=j).value)
                    if '"CN","T2","T3","T4","T5","T6","T7"' in f'{ws.cell(row=i, column=j).value}':
                        header_start_row= i
                        break
                except Exception as ex:
                    pass
                    # print(ex)
        print(header_start_row)
        header_start_row = header_start_row - 1
        df_dict = pd.read_excel(self.input_file, header=[header_start_row-2, header_start_row-1, header_start_row], 
                                sheet_name=sheet_name)
        self.group_range_date = []
        self.msnv_column_header = ''
        self.name_column_header = ''
        for item in list(df_dict.columns.values):
            try:
                print(item)
                if item[2]== 'T2'or item[2]== 'T3' or item[2]== 'T4' or item[2]== 'T5'or item[2]== 'T6'or item[2]== 'T7' or item[2]== 'CN':
                    self.group_range_date.append(item)
                elif (item[0].strip().upper() == 'MSNV' )or (item[1].strip() == 'Mã NV' ):
                    self.msnv_column_header = item
                elif (item[0] == 'HỌ VÀ TÊN') or (item[1].strip() == 'Họ & Tên'):
                    self.name_column_header = item
                print('check:',item)
            except:
                print(item)
        print('aaaaaaaaâ', self.msnv_column_header)
        df_dict = df_dict.merge(self.df_employees[['id','code', 'department_id', 'name','job_title','time_keeping_code']], \
                left_on=[self.msnv_column_header], right_on = ['code'], how='left', suffixes=( '' ,'_employee' ))
        df_dict[df_dict['name'].notnull()].apply(lambda row: self.process_content_dropship(row), axis=1, result_type='expand') 

    def download_attendence_report(self):
        miss_data = True
        while miss_data:
            self.download_employee_info()
            start_str = self.date_array[0].strftime('%Y-%m-%d')
            end_str = self.date_array[len(self.date_array)-1].strftime('%Y-%m-%d')
            # self.df_old = pd.read_excel(self.input_attendence_file , index_col=None, header=[0,] ,sheet_name='Sheet1')
            # self.df_old['is_from_explanation'] = False
            # ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.attendance.trans', 'search', [["&",("date", ">=", start_str),("date", "<=", end_str)]], {})
            ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [["&",("date", ">=", start_str),("date", "<=", end_str)]], {})
            list_scheduling_ver  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'read', [ids], 
                                                        {'fields': ['id','employee_name', 'date', 'shift_name', 'employee_code', 
                                                            'shift_start', 'shift_end', 'rest_start', 'rest_end', 'rest_shift',
                                                            'total_shift_work_time',
                                                            'department', 'attendance_attempt_1', 'attendance_attempt_2',
                                                            'attendance_attempt_3', 'attendance_attempt_4', 'attendance_attempt_5',
                                                            'attendance_attempt_6', 'attendance_attempt_7', 'attendance_attempt_8',
                                                            'attendance_attempt_9', 'attendance_attempt_10', 'attendance_attempt_11',
                                                            'attendance_attempt_12', 'attendance_attempt_13', 'attendance_attempt_14',
                                                            'attendance_attempt_15', 'last_attendance_attempt','attendance_inout_1',
                                                            'attendance_inout_2','attendance_inout_3','attendance_inout_4',
                                                            'attendance_inout_5','attendance_inout_6','attendance_inout_7',
                                                            'attendance_inout_8','attendance_inout_9','attendance_inout_10',
                                                            'attendance_inout_11','attendance_inout_12','attendance_inout_13',
                                                            'attendance_inout_14','attendance_inout_15','attendance_inout_last']})
            
            self.df_scheduling_ver = pd.DataFrame.from_dict(list_scheduling_ver)
            # self.df_scheduling_ver ['attendance_attempt_1'] = pd.to_datetime(self.df_scheduling_ver ['attendance_attempt_1'], format="%Y-%m-%d %H:%M:%S",errors='coerce')
            # self.df_scheduling_ver ['attendance_attempt_2'] = pd.to_datetime(self.df_scheduling_ver ['attendance_attempt_2'], format="%Y-%m-%d %H:%M:%S",errors='coerce')
            for index in range(1,16):
                attendance_attempt = f'attendance_attempt_{index}'
                self.df_scheduling_ver [attendance_attempt] = pd.to_datetime(self.df_scheduling_ver [attendance_attempt], format="%Y-%m-%d %H:%M:%S",errors='coerce')
            self.df_scheduling_ver ['last_attendance_attempt'] = pd.to_datetime(self.df_scheduling_ver ['last_attendance_attempt'], format="%Y-%m-%d %H:%M:%S",errors='coerce')
            print(self.df_shift)
            self.df_scheduling_ver['shift_name'] = self.df_scheduling_ver['shift_name'].str.strip()
            self.df_shift['name'] = self.df_shift['name'].str.strip()
            self.df_scheduling_ver =  self.df_scheduling_ver.merge(self.df_shift[['id','name', 'total_work_time', \
                    'start_work_time','end_work_time','rest_shifts', \
                    'start_rest_time','end_rest_time', 'fix_rest_time', 'night']], \
                    left_on=['shift_name'], right_on = ['name'], how='left', suffixes=( '' ,'_y'))
            self.df_scheduling_ver.rename(columns = {'name':'name_shift'}, inplace = True)
            try:
                self.df_scheduling_ver.to_excel("temp_rawdata_report.xlsx", sheet_name='Sheet1') 
            except Exception as ex:
                print(ex)
                
            self.merge_employee()
            miss_data = self.update_miss_shift()

    def update_miss_shift(self):
        result = False
        for g, data in self.df_scheduling_ver.groupby('employee_code'):
            date_str_data_array = []
            shift_error = []
            shift_null = []
            for index, item in data.iterrows():
                if not item['date'] in date_str_data_array:
                    date_str_data_array.append(item['date'])
                if item['shift_name'] == '' or item['shift_name'] == False:
                    shift_error.append(item['date'])
                if pd.isnull(item['id_y'] ):
                    shift_null.append((item['date'],  item['shift_name']))
            date_str_array = [x for x in self.date_array if not x.strftime('%Y-%m-%d') in date_str_data_array]        

                
            if len (date_str_array) >0:
                result = True
                print(g)
                print ("day la danh sach trong ", date_str_array)
                for date_blank  in date_str_array:
                    update_data = {'employee_code': g, 'employee_name':data.iloc[0]['name'],'date': date_blank.strftime('%Y-%m-%d') ,'shift_name':'-'}
                    update_data['department'] = False if not data.iloc[0]['department_id'] else data.iloc[0]['department_id'][1]
                    update_data['company'] = self.company_info['name']
                    id_scheduling = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'create', [update_data])
                    print('Them thanh cong', update_data)
            if len (shift_error) >0:
                print(g)
                print('day la item loi', shift_error)
            if len (shift_null) >0:
                print(g)
                print('day la shift null', shift_null)
        return result
    def get_data_from_server(self):
        self.download_attendence_report()
        self.process_shift_attendance()

    def process_shift_attendance(self):
        self.process_raw_data()
        self.merge_calendar_leaves()
        if self.is_ho:
            self.df_old[['shift_name', 'normal_time', 'scheduling', 'rest_shifts', 'fix_rest_time', 'night' , 'label', \
                'is_holiday', 'holiday_from', 'holiday_to', 'holiday_name', 'total_shift_work_time', \
                'shift_start', 'shift_end', 'rest_start', 'rest_end', 'report_sid','in_out']] = self.df_old.apply(lambda row: self.find_nearest_ho(row), axis=1, result_type='expand') 
        else:
            self.df_old[['shift_name', 'normal_time', 'scheduling', 'rest_shifts', 'fix_rest_time', 'night' , 'label', \
                'is_holiday', 'holiday_from', 'holiday_to', 'holiday_name', 'total_shift_work_time', \
                'shift_start', 'shift_end', 'rest_start', 'rest_end', 'report_sid']] = self.df_old.apply(lambda row: self.find_nearest(row), axis=1, result_type='expand') 
            
        
        self.assign_time_to_shift()
        try:
            self.df_scheduling_ver.to_excel("rawdata_report.xlsx", sheet_name='Sheet1') 
        except Exception as ex:
            print(ex)
        try:
            self.df_old.to_excel("olddata_report.xlsx", sheet_name='Sheet1') 
        except Exception as ex:
            print(ex)
    def assign_time_to_shift(self):
        self.df_old['shift_name']=self.df_old['shift_name'].str.strip()
        self.df_old = self.df_old.merge(self.df_shift[['id','name', 'total_work_time']], \
                left_on=['shift_name'], right_on = ['name'], how='left', suffixes=( '' ,'_y')).drop(['name'], axis=1)
        # find max time in and min timeout
        # df_attendence_group = self.df_old[~self.df_old['is_from_explanation']].sort_values(['Giờ']).groupby('scheduling')
        df_attendence_group = self.df_old.sort_values(['Giờ']).groupby('scheduling')
        self.df_scheduling_ver['has_attendance'] = False
        result = []
        for g, data in df_attendence_group:
            t_fist = None
            t_last = None
            t_mid_array = []
            inout_mid_array = []
            t_fist = data.iloc[0]['Giờ']
            len_df = len(data['Giờ'])
            if (len_df > 1):
                t_last = data.iloc[len_df -1]['Giờ']
            for index, item in data.iterrows():
                if item['Giờ'] != t_fist and item['Giờ'] != t_last:
                    t_mid_array.append (item['Giờ'])
                    inout_mid_array.append(item['in_out'])

            # result_item = {'scheduling': g, 't_fist': t_fist, 't_last': t_last, 't_mid_array': t_mid_array}
            # result.append(result_item)
            if g>=0:
                self.df_scheduling_ver.at[g, 'has_attendance'] = True
                max_time_in = t_fist
                try:
                    if max_time_in != None and  max_time_in.replace(second=0) < self.df_scheduling_ver['shift_start_datetime'][g]:
                        max_time_in = max([a for a in data['Giờ'] if a.replace(second=0) <= self.df_scheduling_ver['shift_start_datetime'][g]])

                except Exception as ex:
                    # self.df_scheduling_ver.to_excel('nho ti ti.xlsx')
                    print("nho ti ti: ", ex)
                min_time_out = t_last
                try:
                    if min_time_out != None and  min_time_out.replace(second=0) > self.df_scheduling_ver['shift_end_datetime'][g]:
                        min_time_out = min([a for a in data['Giờ'] if a.replace(second=0) >= self.df_scheduling_ver['shift_end_datetime'][g]])
                except Exception as ex:
                    # self.df_old.to_excel('Hoi kho nhi.xlsx')
                    print("Hoi kho nhi: ", ex)

                self.df_scheduling_ver.at[g, 'min_time_out'] = min_time_out
                self.df_scheduling_ver.at[g, 'max_time_in'] = max_time_in
                # result_item = {'scheduling': g, 't_fist': t_fist, 't_last': t_last, 't_mid_array': t_mid_array}
                
                result.append(self.df_scheduling_ver.at[g, 'min_time_out'])
                # Update to server
                # ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'search', [["&",("employee_code", "=", msnv),("date", "=", date_str)]], {})
                # if len(ids) > 0:
                #     print(f'{group_item[1]}-{shift_value} - {shift_name}-found and update-{ids}')
                update_data = {'attendance_attempt_1': False, 'attendance_attempt_2': False, 'attendance_attempt_3': False,
                    'attendance_attempt_4': False, 'attendance_attempt_5': False, 'attendance_attempt_6': False, 
                    'attendance_attempt_7': False, 'attendance_attempt_8': False, 'attendance_attempt_9': False, 
                    'attendance_attempt_10': False, 'attendance_attempt_11': False, 'attendance_attempt_12': False,
                    'attendance_attempt_13': False, 'attendance_attempt_14': False, 'attendance_attempt_15': False,
                    'last_attendance_attempt': False, 'night_shift': False,
                    'shift_start': data.iloc[0]['shift_start'].item(), 
                    'shift_end': data.iloc[0]['shift_end'].item(), 'rest_start': data.iloc[0]['rest_start'].item(), 
                    'rest_end': data.iloc[0]['rest_end'].item(),'rest_end': data.iloc[0]['rest_end'].item()}
                for index_col in range (1,16):
                    update_data[f'attendance_inout_{index_col}'] = False
                update_data['attendance_inout_last'] = False
                if pd.notnull(data.iloc[0]['shift_name']):   
                    if data.iloc[0]['shift_name']:
                        update_data['shift_name'] = f"{data.iloc[0]['shift_name']}"
                if pd.notnull( data.iloc[0]['fix_rest_time']):
                    update_data['split_shift'] = True if data.iloc[0]['fix_rest_time'] else False
                update_data['company'] = self.company_info['name']
                department_id = self.df_scheduling_ver.at[g, 'department_id']
                update_data['department'] = False if not department_id else department_id[1]
                print('update department: ', update_data['department'])
                if pd.notnull(data.iloc[0]['ID']):
                    update_data['time_keeping_code'] = "{:06d}".format(int(data.iloc[0]['ID']))
                if pd.notnull(data.iloc[0]['rest_shifts']):
                    update_data ['rest_shift'] = True if data.iloc[0]['rest_shifts'] else False
                if pd.notnull(data.iloc[0]['night']):
                    update_data ['night_shift'] = True if data.iloc[0]['night'] else False

                try:
                    late_time = (max_time_in - self.df_scheduling_ver['shift_start_datetime'][g]).total_seconds()/60
                    late_time = max(0, late_time)
                    update_data['attendance_late'] = late_time
                    early_time = (self.df_scheduling_ver['shift_end_datetime'][g] - min_time_out).total_seconds()/60
                    early_time = max(0, early_time)
                    update_data['leave_early'] = early_time
                except:
                    print('calculate late err')
                try:
                    if pd.notnull(data.iloc[0]['total_work_time']):
                        update_data['total_shift_work_time'] = data.iloc[0]['total_work_time'].item()
                    elif pd.notnull(self.df_scheduling_ver['total_work_time'][g]):
                        update_data['total_shift_work_time'] = self.df_scheduling_ver['total_work_time'][g].item()
                    else:
                        update_data['total_shift_work_time'] = 0
                except:
                    print("update_data['total_shift_work_time'] error ")
                    update_data['total_shift_work_time'] = 0
                    
                if t_fist != None:
                    
                    t_fist_str = t_fist.strftime('%Y-%m-%d %H:%M:%S')
                    update_data['attendance_attempt_1'] = t_fist_str
                    update_data['attendance_inout_1'] = data.iloc[0]['in_out']
                    t_index = 2
                else:
                    t_index = 1
                inout_index = 0
                for t_itme in t_mid_array:
                    t_fist_str = t_itme.strftime('%Y-%m-%d %H:%M:%S')
                    update_data[f'attendance_attempt_{t_index}'] = t_fist_str
                    update_data[f'attendance_inout_{t_index}'] = inout_mid_array[inout_index]
                    t_index = t_index + 1
                    inout_index = inout_index + 1
                    if t_index > 15:
                        break
                if (update_data['attendance_inout_1'] == False )and (update_data['attendance_attempt_1'] !=False):
                    update_data['attendance_inout_1'] = 'I'
                if t_last != None:
                    t_fist_str = t_last.strftime('%Y-%m-%d %H:%M:%S')
                    update_data['last_attendance_attempt'] = t_fist_str
                    update_data['attendance_inout_last']= data.iloc[len_df -1]['in_out']
                    if t_index <=15:
                        update_data[f'attendance_attempt_{t_index}'] = t_fist_str
                        update_data[f'attendance_inout_{t_index}'] = update_data['attendance_inout_last']
                elif (t_fist != None) and (t_index == 2):
                    update_data['last_attendance_attempt'] = update_data['attendance_attempt_1'] 
                    update_data['attendance_inout_last'] = update_data['attendance_inout_1']
                if (update_data['attendance_inout_last'] == False )and (update_data['last_attendance_attempt'] !=False):
                    update_data['attendance_inout_last'] = 'O'
                for t_index in range(1,16):

                    if (update_data[f'attendance_attempt_{t_index}'] == update_data['last_attendance_attempt']) and \
                            (update_data[f'attendance_inout_{t_index}'] == False ) and \
                            (update_data[f'attendance_attempt_{t_index}']  !=False):
                        update_data[f'attendance_inout_{t_index}'] = update_data['attendance_inout_last']
                try:
                    resource_calendar_id = self.df_scheduling_ver.at[g,'resource_calendar_id'][1]
                    update_data['resource_calendar'] = resource_calendar_id
                    shift_name = update_data['shift_name']
                    if  (not '/' in shift_name ) and (update_data['total_shift_work_time']>0) and \
                        ('44' in resource_calendar_id) and (update_data['attendance_attempt_3']) == False:
                        
                        update_data['missing_checkin_break'] = True
                        print('Thieu cham cong')
                    else: 
                        update_data['missing_checkin_break'] = False
                except Exception as ex:
                    update_data['resource_calendar'] = '-'
                    update_data['missing_checkin_break'] = False
                    print("calculate thieu cham cong err: ", ex)
                # try:
                keys = update_data.keys()
                for key in keys:
                    print(key, update_data[key], type(update_data[key]))
                ids = [int(data.iloc[0]['report_sid'])]
                print(f'update: {update_data} - {ids}')
                if (self.s_shift == None) or (self.s_shift == update_data['shift_name']):
                    self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                    list_scheduling_ver  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'read', [ids], 
                                                        {'fields': ['id','employee_name', 'date', 'shift_name', 'employee_code', 
                                                        'shift_start', 'shift_end', 'rest_start', 'rest_end', 'rest_shift', 'last_attendance_attempt','attendance_attempt_1']})
                    print("ket qua")
                    print(list_scheduling_ver)
                # except Exception as ex:
                #     print('data upload error: ', ex)

       
        try:
            self.df_attendence_hor= pd.DataFrame.from_dict(result)
            self.df_scheduling_ver = self.df_scheduling_ver.merge(self.df_attendence_hor, left_index=True, right_on = ['scheduling'], how='left')
        except Exception as ex:
            print('Merge err: ', ex)
        self.df_scheduling_ver[~self.df_scheduling_ver['has_attendance']].apply(
                lambda row: self.remove_attendance(row), axis=1)
    def remove_attendance(self, row):
        update_data = {'attendance_attempt_1': False, 'attendance_attempt_2': False, 'attendance_attempt_3': False,
                    'attendance_attempt_4': False, 'attendance_attempt_5': False, 'attendance_attempt_6': False, 
                    'attendance_attempt_7': False, 'attendance_attempt_8': False, 'attendance_attempt_9': False, 
                    'attendance_attempt_10': False, 'attendance_attempt_11': False, 'attendance_attempt_12': False,
                    'attendance_attempt_13': False, 'attendance_attempt_14': False, 'attendance_attempt_15': False,
                    'last_attendance_attempt': False,'attendance_late':0, 'leave_early':0}
        update_data['company'] = self.company_info['name']
        try:
            resource_calendar_id = row['resource_calendar_id'][1]
            update_data['resource_calendar'] = resource_calendar_id
        except:
            update_data['resource_calendar'] = '-'
            update_data['missing_checkin_break'] = False
        if pd.notnull(row['start_work_time']):
            update_data['shift_start'] = row['start_work_time']
        else:
            update_data['shift_start'] = row['shift_start']
        if pd.notnull(row['end_work_time']):
            update_data['shift_end'] = row['end_work_time']
        else:
            update_data['shift_end'] = row['shift_end'] 
        if pd.notnull(row['start_rest_time']):
            update_data['rest_start'] = row['start_rest_time']
        else:
            update_data['rest_start'] = row['rest_start'] 
        if pd.notnull(row['end_rest_time']):
            update_data['rest_end'] = row['end_rest_time']
        else:
            update_data['rest_end'] = row['rest_end'] 
            
        department_id = row['department_id']
        update_data['department'] = False if not department_id else department_id[1]
        ids = [row['id']]
        print(f'update blank shift: {update_data} - {ids}')
        self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
    def find_nearest_ho(self, row):
        # print(self.df_normal_attendances['employee_id'])
        # print(row['employee_id'])
        is_holiday = False
        holiday_from = None
        holiday_to = None
        holiday_name = ''
        shift = 0
        shift_id = 0
        normal_time = 0
        rest_shifts = False
        fix_rest_time = False
        night = False
        label = ''
        scheduling = -1
        total_shift_work_time = 0
        shift_start = 12
        shift_end = 12
        rest_start = 12
        rest_end = 12
        report_sid = -1
        in_out= False
        # try: 
        df_compare = self.df_normal_attendances[self.df_normal_attendances['time_keeping_code']==int(row['ID'])].sort_values(by=['time'])
        if len(df_compare['time'])>0:
            dates = df_compare['time'].to_list()
            # print(dates)
            normal_shift_row = df_compare.iloc[0]
            for index, time_item in df_compare.iterrows():
                try:
                    if time_item['time'].date() == row["Giờ"].date():
                        normal_shift_row = time_item
                except Exception as ex:
                    print(f'errror nè {time_item["time"]}-{row["Giờ"]}-{type(time_item["time"])}-{type(row["Giờ"])}-{ex}')
                
            # print('min: ', min_time)
            # normal_shift_row = df_compare.iloc[df_compare['time'].get_loc(row['Giờ'],method='nearest')]
            # print(df_compare['time'].searchsorted(row['Giờ']))
            # normal_shift_row = df_compare[df_compare['time']==min_time].iloc[0]
            scheduling = normal_shift_row['scheduling']
            report_sid = normal_shift_row['report_sid']
            shift = normal_shift_row['shift_name']
            # shift_id = normal_shift_row['shift']
            normal_time = normal_shift_row['time']
            
            rest_shifts = normal_shift_row['rest_shifts']
            if pd.notnull(normal_shift_row['total_work_time']):
                total_shift_work_time = normal_shift_row['total_work_time']
            if pd.notnull(normal_shift_row['start_work_time']):
                shift_start = normal_shift_row['start_work_time']
            if pd.notnull(normal_shift_row['end_work_time']):
                shift_end = normal_shift_row['end_work_time']
            if pd.notnull(normal_shift_row['start_rest_time']):
                rest_start = normal_shift_row['start_rest_time']
            if pd.notnull(normal_shift_row['end_rest_time']):
                rest_end = normal_shift_row['end_rest_time']

            fix_rest_time = normal_shift_row['fix_rest_time']
            night = normal_shift_row['night']
            if fix_rest_time:
                print(f"fix_rest_time: {shift}")
            label = normal_shift_row['label']
            is_holiday = normal_shift_row['is_holiday']
            holiday_from = normal_shift_row['holiday_from']
            holiday_to = normal_shift_row['holiday_to']
            holiday_name = normal_shift_row['holiday_name']
            in_out = row['in_out']

            # return shift, shift_id, normal_time, scheduling, rest_shifts, fix_rest_time, night ,label, is_holiday, holiday_from, holiday_to, holiday_name
        # except Exception as ex:
        #     print('Find nearest errrrrrrrrr: ', ex)
            
            
        return shift, normal_time, scheduling, rest_shifts, fix_rest_time, night ,\
                label, is_holiday, holiday_from, holiday_to, holiday_name, total_shift_work_time, \
                shift_start, shift_end, rest_start, rest_end, report_sid, in_out
    def find_nearest(self, row):
        # print(self.df_normal_attendances['employee_id'])
        # print(row['employee_id'])
        is_holiday = False
        holiday_from = None
        holiday_to = None
        holiday_name = ''
        shift = 0
        shift_id = 0
        normal_time = 0
        rest_shifts = False
        fix_rest_time = False
        night = False
        label = ''
        scheduling = -1
        total_shift_work_time = 0
        shift_start = 12
        shift_end = 12
        rest_start = 12
        rest_end = 12
        report_sid = -1
        # try: 
        df_compare = self.df_normal_attendances[self.df_normal_attendances['time_keeping_code']==int(row['ID'])].sort_values(by=['time'])
        if len(df_compare['time'])>0:
            dates = df_compare['time'].to_list()
            # print(dates)
            normal_shift_row = df_compare.iloc[0]
            for index, time_item in df_compare.iterrows():
                try:
                    
                    delta_time_item = abs((time_item['time']  - row['Giờ']))
                    if time_item['time'].date() != row['Giờ'].date():
                        delta_time_item = delta_time_item + datetime.timedelta(hours=2)
                    if time_item['label'] == 'Out':    
                        delta_time_item = delta_time_item + datetime.timedelta(hours=-4)

                    delta_normal_time = abs (normal_shift_row['time']- row['Giờ'])
                    if normal_shift_row['label'] == 'Out':    
                        delta_normal_time = delta_normal_time + datetime.timedelta(hours=-4)
                    if normal_shift_row['time'].date() != row['Giờ'].date():
                        delta_normal_time = delta_normal_time + datetime.timedelta(hours=2)
                    

                    if  delta_time_item < delta_normal_time:
                        normal_shift_row = time_item
                except Exception as ex:
                    print(f'errror nè {time_item["time"]}-{row["Giờ"]}-{type(time_item["time"])}-{type(row["Giờ"])}-{ex}')
                
            # print('min: ', min_time)
            # normal_shift_row = df_compare.iloc[df_compare['time'].get_loc(row['Giờ'],method='nearest')]
            # print(df_compare['time'].searchsorted(row['Giờ']))
            # normal_shift_row = df_compare[df_compare['time']==min_time].iloc[0]
            scheduling = normal_shift_row['scheduling']
            report_sid = normal_shift_row['report_sid']
            shift = normal_shift_row['shift_name']
            # shift_id = normal_shift_row['shift']
            normal_time = normal_shift_row['time']
            
            rest_shifts = normal_shift_row['rest_shifts']
            if pd.notnull(normal_shift_row['total_work_time']):
                total_shift_work_time = normal_shift_row['total_work_time']
            if pd.notnull(normal_shift_row['start_work_time']):
                shift_start = normal_shift_row['start_work_time']
            if pd.notnull(normal_shift_row['end_work_time']):
                shift_end = normal_shift_row['end_work_time']
            if pd.notnull(normal_shift_row['start_rest_time']):
                rest_start = normal_shift_row['start_rest_time']
            if pd.notnull(normal_shift_row['end_rest_time']):
                rest_end = normal_shift_row['end_rest_time']

            fix_rest_time = normal_shift_row['fix_rest_time']
            night = normal_shift_row['night']
            if fix_rest_time:
                print(f"fix_rest_time: {shift}")
            label = normal_shift_row['label']
            is_holiday = normal_shift_row['is_holiday']
            holiday_from = normal_shift_row['holiday_from']
            holiday_to = normal_shift_row['holiday_to']
            holiday_name = normal_shift_row['holiday_name']

            # return shift, shift_id, normal_time, scheduling, rest_shifts, fix_rest_time, night ,label, is_holiday, holiday_from, holiday_to, holiday_name
        # except Exception as ex:
        #     print('Find nearest errrrrrrrrr: ', ex)
            
            
        return shift, normal_time, scheduling, rest_shifts, fix_rest_time, night ,\
                label, is_holiday, holiday_from, holiday_to, holiday_name, total_shift_work_time, \
                shift_start, shift_end, rest_start, rest_end, report_sid


    def refact_date_from(self, row):
        date_from = row['date_from']
        print('date {} from {}'.format (date_from, type(date_from)))
        try:
            date_from =  datetime.datetime.strptime(row['date_from'],'%Y-%m-%d %H:%M:%S').replace(hour=0,minute=0,second=0)
        except Exception as ex:
            print (ex)
        date_to = row['date_to']
        try:
            date_to =  datetime.datetime.strptime(row['date_to'], '%Y-%m-%d %H:%M:%S').replace(hour=0,minute=0,second=0)
            date_to = date_to + datetime.timedelta(days=1)
        except Exception as ex:
            print ('date_to', ex)
        return date_from, date_to

    def merge_calendar_leaves(self):

        print('***************CALENDAR************************************')
        try:
            self.df_resource_calendar_leaves[['date_from', 'date_to']]= self.df_resource_calendar_leaves.apply(lambda row: self.refact_date_from(row), axis=1, result_type='expand')
        except:
            self.df_resource_calendar_leaves = pd.DataFrame(columns=['date_from', 'date_to'])
        # self.df_resource_calendar_leaves['date_to']=self.df_resource_calendar_leaves.apply(lambda row: row['date_to'].replace(hour=23,minute=59,second=59), axis=1)
        print(self.df_resource_calendar_leaves)
        # print(list_scheduling_ver)
        self.df_normal_attendances[['is_holiday', 'holiday_from', 'holiday_to', 'holiday_name']] = \
            self.df_normal_attendances.apply(lambda row: self.merge_holiday(row), axis=1, result_type='expand')

    def download_employee_info(self):
        
        if self.employee_code == None:
            employee_Sids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.employee', 'search', [[("company_id", "=", self.company_id)]])
            list_employees  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.employee', 'read', [employee_Sids], {'fields': ['id', 'name', 'user_id', \
                                                    'company_id', 'code', 'department_id', 'time_keeping_code', 'job_title','resource_calendar_id']})
        else:
            employee_Sids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.employee', 'search', [[("code", "=", self.employee_code)]])
            list_employees  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.employee', 'read', [employee_Sids], {'fields': ['id', 'name', 'user_id', \
                                                    'company_id', 'code', 'department_id', 'time_keeping_code', 'job_title','resource_calendar_id']})
        
        # The above code is iterating through a list of employees and checking if their
        # 'time_keeping_code' attribute is not None. If it is not None, it removes the '.0' string
        # from the end of the attribute value using the replace() method.
        for employee in list_employees:
            try:
                if employee['time_keeping_code'] != None:
                    employee['time_keeping_code'] = int(employee['time_keeping_code'].replace('.00', '').replace('.0', ''))
            except Exception as ex:
                print(ex)
        self.df_employees = pd.DataFrame.from_dict(list_employees)
        # print(self.df_employees)
        
    def merge_employee(self):
        self.df_scheduling_ver =  self.df_scheduling_ver.merge(self.df_employees[['id','code', 'department_id', 'name', \
                'job_title','time_keeping_code', 'resource_calendar_id']], \
                left_on=['employee_code'], right_on = ['code'], how='inner', suffixes=( '' ,'_employee' ))
        # self.df_scheduling_ver.rename(columns = {'id':'employee_sid'}, inplace = True)
        # self.df_scheduling_ver  = self.df_scheduling_ver.dropna(subset=['code'])
        
    def process_raw_data(self):
        self.df_scheduling_ver[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                self.df_scheduling_ver.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
        normal_attendances = []
        for index, row in self.df_scheduling_ver.iterrows():
            # print("index day nhe: ", self.df_scheduling_ver['id'][index])
            start_data = {"employee_code": row['employee_code'], 'time_keeping_code': row['time_keeping_code'], \
                'shift_name':row['shift_name'], 'time':row['shift_start_datetime'], 'scheduling': index, 'report_sid':row['id'],
                'rest_shifts':row['rest_shifts'],'fix_rest_time':row['fix_rest_time'], 'night':row['night'],'label': 'In', 
                'total_work_time': row['total_work_time'], 'start_work_time': row['start_work_time'], 
                'end_work_time': row['end_work_time'], 'start_rest_time': row['start_rest_time'],
                'end_rest_time': row['end_rest_time']}

            end_data = {"employee_code": row['employee_code'], 'time_keeping_code': row['time_keeping_code'], \
                'shift_name':row['shift_name'], 'time':row['shift_end_datetime'], 'scheduling': index, 'report_sid':row['id'],\
                'rest_shifts':row['rest_shifts'],'fix_rest_time':row['fix_rest_time'], 'night':row['night'],'label': 'Out', \
                'total_work_time': row['total_work_time'], 'start_work_time': row['start_work_time'], 
                'end_work_time': row['end_work_time'], 'start_rest_time': row['start_rest_time'],
                'end_rest_time': row['end_rest_time']}

            normal_attendances.append(start_data)
            normal_attendances.append(end_data)
        self.df_normal_attendances = pd.DataFrame.from_dict(normal_attendances)
        try:
            self.df_normal_attendances.to_excel("temp_df_normal_attendances.xlsx")
        except Exception as ex:
            print(ex)

    ########################### 'resource.calendar.leaves
    def merge_holiday(self, row):
        is_holiday = False
        holiday_from = None
        holiday_to = None
        holiday_name = ''
        try:
            shift_time = row['time']
            df_compare = self.df_resource_calendar_leaves[(self.df_resource_calendar_leaves['date_from']<=shift_time) & \
                                            (self.df_resource_calendar_leaves['date_to']>=shift_time)]
            
            if len(df_compare.index) > 0:
                fist_row = df_compare.iloc[0]
                is_holiday = True
                holiday_from = fist_row['date_from']
                holiday_to = fist_row['date_to']
                holiday_name = fist_row['name']

        except Exception as ex:
            is_holiday = False

            print('merge error: ', ex)
            
        return is_holiday, holiday_from, holiday_to, holiday_name
    
    def process_report_raw(self, row):
        # print(row)
        start_work_time = row['start_work_time']
        if pd.isnull(start_work_time):
            start_work_time = row['shift_start']

        end_work_time = row['end_work_time']
        if pd.isnull(end_work_time):
            end_work_time = row['shift_end']
        start_rest_time =row['start_rest_time']
        if pd.isnull(start_rest_time):
            start_rest_time = row['rest_start']

        end_rest_time = row['end_rest_time']
        if pd.isnull(end_rest_time):
            end_rest_time = row['rest_end']
        h, m, s = float_to_hours(start_work_time)

        shift_start_datetime = datetime.datetime.strptime(row['date'],"%Y-%m-%d").replace(hour=h, minute=m, second=s)

        h, m, s = float_to_hours(end_work_time)
        shift_end_datetime = datetime.datetime.strptime(row['date'],"%Y-%m-%d").replace(hour=h, minute=m, second=s)
        shift_end_datetime = shift_end_datetime if end_work_time >= start_work_time else shift_end_datetime + datetime.timedelta(days=1)
        h, m, s = float_to_hours(start_rest_time)
        rest_start_datetime = datetime.datetime.strptime(row['date'],"%Y-%m-%d").replace(hour=h, minute=m, second=s)
        rest_start_datetime = rest_start_datetime if start_rest_time >= start_work_time else rest_start_datetime + datetime.timedelta(days=1)
        h, m, s = float_to_hours(end_rest_time)
        rest_end_datetime = datetime.datetime.strptime(row['date'],"%Y-%m-%d").replace(hour=h, minute=m, second=s)
        rest_end_datetime = rest_end_datetime if end_rest_time >= start_work_time else rest_end_datetime + datetime.timedelta(days=1)
        return shift_start_datetime, shift_end_datetime, rest_start_datetime, rest_end_datetime 
    
    def assign_shift(self, shift_name, selected_employee_code, day):
        print(f'assign {day}')
        assign_date = datetime.datetime.strptime(day, '%d-%m-%Y')
        if not self.df_shift['name'].str.contains(shift_name).any():
            return False
        # selected_range = self.group_range_date[self.selected_range_index]
        # assign_date = selected_range[day]
        print('assign new shift: {} - {} -{}'.format(selected_employee_code, shift_name, assign_date))
        # self.report_shift_ver[self.report_shift_ver['code']==selected_employee_code, \
        #     self.report_shift_ver['date']==assign_date]['new-shift'] = shift_name
        # for idex, item in self.df_scheduling_ver[self.df_scheduling_ver['employee_code']==selected_employee_code].iterrows():
        #     if item['date'] == assign_date:
        #         # scheduling = item['scheduling']
        #         self.df_scheduling_ver.at[idex,'new-shift'] = shift_name
        #         self.df_scheduling_ver.at[idex,'is-modified'] = True
        #         break

    # *********************** INVALID TIME SHEET AREA **************
    def split_range_date(self):
        self.group_range_date = []
        range_date = []
        for item in self.date_array:
            print(item)
            if item.dayofweek == 0:
                range_date = []

            range_date.append(item)

            if item.dayofweek == 6:
                self.group_range_date.append(range_date)

        if len(range_date) <7:
            print("ton tai < 7")
            self.group_range_date.append(range_date)
        else:
            print("ko ton tai" )

    def initialize_table(self, tableInvalidView):
        self.get_data_from_server()
        self.date_array = pd.date_range(start='07/01/2023', end='07/30/2023')
        # self.sync_data_scheduling(None)
        self.selected_range_index = 0
        self.split_range_date()
        selected_range = self.group_range_date[self.selected_range_index]
        tableInvalidView.setRowCount(300) 
        #Column count
        tableInvalidView.setColumnCount(36)  
        count = 0
        start_row = 0
        col = 2
        for date_item in self.date_array:
            newItem = QTableWidgetItem(date_item.strftime('%d-%m-%Y'))
            newItem.setForeground(QBrush(QColor(250, 125, 0)))
            tableInvalidView.setItem(start_row,col, newItem) 
            col = col + 1
        start_row = start_row + 1
        self.df_scheduling_ver['weekday'] = self.df_scheduling_ver.apply(lambda row:datetime.datetime.strptime(row['date'],'%Y-%m-%d'), axis=1)
        for g, data in self.df_scheduling_ver[self.df_scheduling_ver['weekday'].dt.month==5 ].groupby('department'):
            newItem = QTableWidgetItem(f'{g}')
            newItem.setForeground(QBrush(QColor(150, 125, 0)))
            tableInvalidView.setItem(start_row,1, newItem) 
            start_row = start_row + 1
            for sub_g, sub_data in data.groupby('employee_code'):
                employee_code = sub_g
                name_employee = sub_data.iloc[0]['employee_name']
                # invalid_type = row['invalid_type']
                tableInvalidView.setItem(start_row,0, QTableWidgetItem(f'{employee_code}'))
                tableInvalidView.setItem(start_row,1, QTableWidgetItem(f'{name_employee}')) 
                for dayofweek, day_data in sub_data.groupby('weekday'):
                    col = 1 + int(dayofweek.day)
                    shift_name = '-' if day_data.iloc[0]['shift_name'] == False else  f"{day_data.iloc[0]['shift_name']}"
                    newItem = QTableWidgetItem(f'{shift_name}')
                    newItem.setForeground(QBrush(QColor(0, 255, 0)))
                    tableInvalidView.setItem(start_row,col, newItem) 
                # tableInvalidView.setItem(count,2, QTableWidgetItem(f'{reason}'))
                # tableInvalidView.setItem(count,3, QTableWidgetItem(f'{invalid_type}'))
                

                start_row = start_row + 1

    ################################# CALCULATE  VALUE ***********************
    def find_couple(self, row):
        result = [[]]
        time_stack = []
        for index in range(1,16):
            attendance_attempt = row[f'attendance_attempt_{index}']
            attendance_inout = row[f'attendance_inout_{index}']
            if attendance_inout == 'I':
                time_stack.append(index)
            elif attendance_inout =='O':
                if len(time_stack)>0:
                    result[0].append((time_stack[len(time_stack)-1], index))
                time_stack = []

        return result
    def find_couple_in_out(self):
        self.df_scheduling_ver[['couple']] = self.df_scheduling_ver.apply(lambda row: self.find_couple(row), axis=1, result_type='expand')
        try:
            self.df_scheduling_ver.to_excel('couple.xlsx')
        except:
            print('save cople ex')

    def calculate_work_hour_values(self):
        self.download_attendence_report()
        
        
        self.df_scheduling_ver[['shift_start_datetime','shift_end_datetime', 'rest_start_datetime', 'rest_end_datetime']]  = \
                self.df_scheduling_ver.apply(lambda row: self.process_report_raw(row), axis=1, result_type='expand')
        print("Now! calculate total")
        # self.df_scheduling_ver.to_excel("rawdata_report.xlsx", sheet_name='Sheet1') 
        if self.is_ho:
            self.find_couple_in_out()
            # calculate_actual_work_time_ho
            self.df_scheduling_ver.apply(lambda row: self.calculate_actual_work_time_ho(row), axis=1, result_type='expand')
        else:
            self.df_scheduling_ver.apply(lambda row: self.calculate_actual_work_time(row), axis=1, result_type='expand')
        # self.df_scheduling_ver.to_excel("rawdata_report2.xlsx", sheet_name='Sheet1') 
    def calculate_actual_work_time_ho(self, row):
        result = 0
        update_data = {'total_work_time':0,'night_hours_normal':0}
        for couple in row['couple']:
            in_index = couple[0]
            out_index = couple[1]
            real_time_in = row[f'attendance_attempt_{in_index}']
            real_time_out = row[f'attendance_attempt_{out_index}']
            update_data_item = self.calculate_actual_work_time_couple(row, real_time_in, real_time_out)
            update_data['total_work_time'] = update_data['total_work_time'] + update_data_item['total_work_time']
            update_data['night_hours_normal'] = update_data['night_hours_normal'] + update_data_item['night_hours_normal']
        if row['id']>0:
                ids = [int(row['id'])]
                self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                print(f'sucess: {update_data}')
    def calculate_actual_work_time_couple(self, row, real_time_in, real_time_out):
        """
        calculate actual work time with real time in, time out , restime in out


        """
        # print('aaaaaa')
        result = 0
        night_work_time = 0
        try:
            rest_shifts = row['rest_shift']
            total_work_time = row['total_shift_work_time']
            # print(f"{rest_shifts}-{total_work_time}")
            if rest_shifts or total_work_time == 0:
                # print(f"{rest_shifts}-{total_work_time}")
                result = 0
            else:
                try:
                    # real_time_in = row['attendance_attempt_1'].replace(second=0)
                    # real_time_out = row['last_attendance_attempt'].replace(second=0)
                    # print(f"{real_time_in}-{real_time_out}")
                    if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
                        # print(f"{real_time_in}-{real_time_out}")
                        result = 0
                    else:
                        start_work_date_time = row['shift_start_datetime'].replace(second=0)
                        end_work_date_time = row['shift_end_datetime'].replace(second=0)

                        start_rest_date_time = row['rest_start_datetime'].replace(second=0)
                        end_rest_date_time = row['rest_end_datetime'].replace(second=0)
                        
                        current_program = min(real_time_out, start_rest_date_time) - max(real_time_in, start_work_date_time)
                        stage_fist = max(0, current_program.total_seconds()//60.0)
                        
                        current_program = min(real_time_out, end_work_date_time) - max(real_time_in, end_rest_date_time)
                        stage_second = max(0, current_program.total_seconds()//60.0)

                        result = int(stage_fist + stage_second)
                        result = min(result, 480)
                        night_work_time = self.calculate_night_work_time(row)
                        if '/OFF' in row['shift_name']:
                            result = min(result, 240)
                        night_work_time = min(night_work_time, result)
                except Exception as ex:
                    print("In cal actual", ex)
            update_data = {'total_work_time': result, 'night_hours_normal': night_work_time}
            if result> 0:
                try:
                    resource_calendar_id = row['resource_calendar_id'][1]
                    update_data['resource_calendar'] = resource_calendar_id
                    shift_name = row['shift_name']
                    if  (not '/' in shift_name ) and (row['total_shift_work_time']>0) and \
                        '44' in resource_calendar_id and row['attendance_attempt_3'] == False:
                        
                        update_data['missing_checkin_break'] = True
                        print('Thieu cham cong')
                    else: 
                        update_data['missing_checkin_break'] = False
                except Exception as ex:
                    update_data['resource_calendar'] = ''
                    update_data['missing_checkin_break'] = False
                    print("calculate thieu cham cong err: ", ex)
            else:
                update_data['missing_checkin_break'] = False

            # print(f'night_work_time {night_work_time}')
            # print(f'total_work_time {result}-{real_time_in}-{real_time_out}-{start_rest_date_time}-{end_rest_date_time}--{stage_fist} + {stage_second}')
            
            
            return update_data

        except Exception as ex:
            print("calcuate actual time err", ex)
            return 0

    def calculate_actual_work_time(self, row):
        """
        calculate actual work time with real time in, time out , restime in out


        """
        # print('aaaaaa')
        result = 0
        night_work_time = 0
        try:
            rest_shifts = row['rest_shift']
            total_work_time = row['total_shift_work_time']
            # print(f"{rest_shifts}-{total_work_time}")
            if rest_shifts or total_work_time == 0:
                # print(f"{rest_shifts}-{total_work_time}")
                result = 0
            else:
                try:
                    real_time_in = row['attendance_attempt_1'].replace(second=0)
                    real_time_out = row['last_attendance_attempt'].replace(second=0)
                    # print(f"{real_time_in}-{real_time_out}")
                    if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
                        # print(f"{real_time_in}-{real_time_out}")
                        result = 0
                    else:
                        start_work_date_time = row['shift_start_datetime'].replace(second=0)
                        end_work_date_time = row['shift_end_datetime'].replace(second=0)

                        start_rest_date_time = row['rest_start_datetime'].replace(second=0)
                        end_rest_date_time = row['rest_end_datetime'].replace(second=0)
                        
                        current_program = min(real_time_out, start_rest_date_time) - max(real_time_in, start_work_date_time)
                        stage_fist = max(0, current_program.total_seconds()//60.0)
                        
                        current_program = min(real_time_out, end_work_date_time) - max(real_time_in, end_rest_date_time)
                        stage_second = max(0, current_program.total_seconds()//60.0)

                        result = int(stage_fist + stage_second)
                        result = min(result, 480)
                        night_work_time = self.calculate_night_work_time(row)
                        if '/OFF' in row['shift_name']:
                            result = min(result, 240)
                        night_work_time = min(night_work_time, result)
                except Exception as ex:
                    print("In cal actual", ex)
            update_data = {'total_work_time': result, 'night_hours_normal': night_work_time}
            if result> 0:
                try:
                    resource_calendar_id = row['resource_calendar_id'][1]
                    update_data['resource_calendar'] = resource_calendar_id
                    shift_name = row['shift_name']
                    if  (not '/' in shift_name ) and (row['total_shift_work_time']>0) and \
                        '44' in resource_calendar_id and row['attendance_attempt_3'] == False:
                        
                        update_data['missing_checkin_break'] = True
                        print('Thieu cham cong')
                    else: 
                        update_data['missing_checkin_break'] = False
                except Exception as ex:
                    update_data['resource_calendar'] = ''
                    update_data['missing_checkin_break'] = False
                    print("calculate thieu cham cong err: ", ex)
            else:
                update_data['missing_checkin_break'] = False

            # print(f'night_work_time {night_work_time}')
            # print(f'total_work_time {result}-{real_time_in}-{real_time_out}-{start_rest_date_time}-{end_rest_date_time}--{stage_fist} + {stage_second}')
            
            if ((self.s_shift == None) or (self.s_shift == row['shift_name'])) and (row['id']>0):
                ids = [int(row['id'])]
                self.models.execute_kw(self.db, self.uid, self.password, 'hr.apec.attendance.report', 'write', [ids, update_data])
                print(f'sucess: {update_data}')
            return result

        except Exception as ex:
            print("calcuate actual time err", ex)
            return 0
    def calculate_night_work_time(self, row):
        """
        calculate night holiday work time with real time in, time out , restime in out
        cobine night holiday time
        """
        try:
            rest_shifts = row['rest_shifts']
            total_work_time = row['total_work_time']
            if rest_shifts or total_work_time == 0:
                return 0
            real_time_in = row['attendance_attempt_1'].replace(second=0)
            real_time_out = row['last_attendance_attempt'].replace(second=0)
            if (real_time_in == None) or (real_time_in == '') or (real_time_out == None) or (real_time_out == ''):
                return 0
            start_work_date_time = row['shift_start_datetime'].replace(second=0)
            end_work_date_time = row['shift_end_datetime'].replace(second=0)

            start_rest_date_time = row['rest_start_datetime'].replace(second=0)
            end_rest_date_time = row['rest_end_datetime'].replace(second=0)
            
            

            # start night datetime from 22h to 6h:00
            start_night_date_time = start_work_date_time.replace(hour=22, minute=0, second=0)
            end_night_date_time = (start_night_date_time + datetime.timedelta(days=1)).replace(hour=6)
            
            
            current_program = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time)
            stage_fist = max(0, current_program.total_seconds()/60.0)
            
            current_program = min(real_time_out, end_work_date_time, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time)
            stage_second = max(0, current_program.total_seconds()/60.0)

            stage_night =  int(stage_fist + stage_second)

            end_night_date_time = start_work_date_time.replace(hour=6, minute=0, second=0)
            start_night_date_time = (end_night_date_time+ datetime.timedelta(days=-1)).replace(hour=22, minute=0, second=0)
           
            current_program = min(real_time_out, start_rest_date_time, end_night_date_time) - \
                    max(real_time_in, start_work_date_time, start_night_date_time)
            stage_fist = max(0, current_program.total_seconds()/60.0)
            
            current_program = min(real_time_out, end_work_date_time, end_night_date_time) - \
                    max(real_time_in, end_rest_date_time, start_night_date_time)
            stage_second = max(0, current_program.total_seconds()/60.0)
            stage_morning = int(stage_fist + stage_second)
            return stage_night + stage_morning

        except Exception as ex:
            # print('{}: night - holiday: {} - {} - {} -night - {}- {} - real - {} - {}'.format(row['name_employee'] ,row['name'], 0, 0, 0, 0, \
            #                                                                         row['start_work_date_time'], row['end_work_date_time']))
            print("calcuate night", ex)
            return 0

    # **************************** EMPLOYEE AREA
    def syn_company_employees(self, company_mis_id):
        url = 'https://mis.apec.com.vn/session/login'
        data = {"username": "quyeths","password": "123123"}
        headers = {'Content-type': 'application/json', 'Accept': 'text/plain', "Authorization": "Basic YWRtaW46MTIzNDU2"}
        r = requests.post(url, data=json.dumps(data), headers=headers)

        r.status_code
        print(r.json())
        SessionKey   = r.json()['result']['SessionKey']

        # mis id of phu yen: "MHS-PY"
        data = {
                "values": [],
                "conditions": [
                    {
                    "ConditionID": "D16",
                    "Operator": "001",
                    "Value": company_mis_id,
                    "ID": 1
                    }
                ],
                "data": {
                    "pageSize": 1000
                },
                "moduleInfo": {
                    "ModuleID": "03266",
                    "SubModule": "MMN"
                },
                "sessionId": "fd4831ccd764e0dae3943440908014c"
            }
        headers = {'Content-type': 'application/json', 'Accept': 'text/plain', 'Authorization': f'Base {SessionKey}'}

        url = 'https://mis.apec.com.vn//search/executeSearch'

        r = requests.post(url, data=json.dumps(data), headers=headers)
        print(r.json())
        # get list company
        ids = self.models.execute_kw(self.db, self.uid, self.password, 'res.company', 'search', [[("mis_id", "=", company_mis_id)]], {})
        list_company  = self.models.execute_kw(self.db, self.uid, self.password, 'res.company', 'read', [ids], {'fields': ['id', 'name']})
        print("list company: ",list_company)
        if len(list_company) == 0:
            return
        company_id = list_company[0]['id']
        # Check deparment
        department_ids = self.models.execute_kw(self.db, self.uid, self.password, 'hr.department','search', [[("company_id", "=", company_id)]])
        list_departments  = self.models.execute_kw(self.db, self.uid, self.password, 'hr.department', 'read', [department_ids], {'fields': ['id','name', 'total_employee', 'company_id', 'member_ids', 'mis_id']})
        
        
        # with open('data-page-phuyen.json', 'w') as f:
        #     json.dump(data, f)
        for item in data['result']['data']:
            STAFF_ID = item['STAFF_ID'] # ma so nhan vien
            FING_ID = item['FING_ID'] # ma cham cong
            DEPARTMENT_NAME = item['DEPARTMENT_NAME']
            MANAGER = item['MANAGER']
            PHONE = item['PHONE']
            NOTE = item['NOTE'] # "gmail: kimchi.dhpy@gmail.com",
            CONTACT_TYPE = item['CONTACT_TYPE']
            IDENTIFICATION = item['IDENTIFICATION']
            EMAIL = item['EMAIL']
            COUNTRY = item['COUNTRY'] # dia chi
            DATEOFBIRTH = item['DATEOFBIRTH']
            FINISH_DATE  = item['FINISH_DATE']
            START_DATE = item['START_DATE']
            STAFF_NAME = item['STAFF_NAME']
            found_department = [ dp for dp in list_departments if dp['mis_id'] ==  DEPARTMENT_NAME]
            if len(found_department) == 0:
                # Created new deparment
                id_deparment = self.models.execute_kw(self.db, self.uid, self.password, 'hr.department', 'create', 
                    [{'name': DEPARTMENT_NAME, 'company_id': company_id , 'mis_id': DEPARTMENT_NAME}])
                print('list created', id_deparment)
                # id_deparment = id_deparments[0]
            else:
                id_deparment = found_department[0]['id']

            # create employee
            self.download_employee_info()
            emloyee_list = self.df_employees[self.df_employees['code'] == STAFF_ID]
            if len (emloyee_list) == 0:
                id_deparments = self.models.execute_kw(self.db, self.uid, self.password, 'hr.employee', 'create', 
                    [{'name': STAFF_NAME, 'department_id': id_deparment , 'code': STAFF_ID, 'time_keeping_code': f'{int(FING_ID) if FING_ID else False}',
                        'workingday': START_DATE}])
                print('emoloyee created', id_deparments)
            
            else:
                id = emloyee_list.iloc[0]['id'].item()
                ids = [id]
                update_data = {'company_id': company_id}
                self.models.execute_kw(self.db, self.uid, self.password, 'hr.employee', 'write', [ids, update_data])
                print(f"nhan vien ton tai {STAFF_ID}- {STAFF_NAME}")


        

    def add_new_employee(self, employee_code):
        self.download_employee_info()
        url = 'https://mis.apec.com.vn/session/login'
        data = {"username": "quyeths","password": "123123"}
        headers = {'Content-type': 'application/json', 'Accept': 'text/plain'}
        r = requests.post(url, data=json.dumps(data), headers=headers)

        r.status_code
        print(r.json())
        SessionKey   = r.json()['result']['SessionKey']

        data = {
                "values": [],
                "conditions": [
                    {
                    "ConditionID": "D04",
                    "Operator": "003",
                    "Value":  employee_code, #"APG230316002",
                    "ID": 1
                    }
                ],
                "data": {
                    "pageSize": 100
                },
                "moduleInfo": {
                    "ModuleID": "03266",
                    "SubModule": "MMN"
                },
                "sessionId": "fd4831ccd764e0dae3943440908014c"
            }

        headers = {'Content-type': 'application/json', 'Accept': 'text/plain', 'Authorization': f'Base {SessionKey}'}

        url = 'https://mis.apec.com.vn//search/executeSearch'

        r = requests.post(url, data=json.dumps(data), headers=headers)
        print(r.status_code)
        data = r.json()
        mis_department = data['result']['data'][0]['DEPARTMENT_NAME']
        mis_name = data['result']['data'][0]['STAFF_NAME']
        mis_start_date = data['result']['data'][0]['START_DATE'].replace("T07:00:00", "")
        mis_birthday =  data['result']['data'][0]['DATEOFBIRTH']
        mis_POSITION = data['result']['data'][0]['POSITION']
        mis_email = data['result']['data'][0]['EMAIL']
        mis_fing_id = f"{data['result']['data'][0]['FING_ID']}".replace('.0','')
        
        print("Mis Department: ", mis_department)
        dpm = self.df_departments[self.df_departments['mis_id']== mis_department]

        deparment = dpm.iloc[0]
        department_id = deparment['id'].item()
        department_company_id = deparment['company_id'][0]
        id = self.models.execute_kw(self.db, self.uid, self.password, 'hr.employee', 'create', 
            [{'code': employee_code, 'time_keeping_code': mis_fing_id ,
                'department_id':department_id, 'workingday':mis_start_date, 'name': mis_name}])
        print(id)
        # with open('data-page-alinhdan.json', 'w') as f:
        #     json.dump(data, f)
