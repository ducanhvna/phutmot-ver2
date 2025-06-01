import os
import tempfile
import pandas as pd
import openpyxl
import pytest
from app.utils.report_exporters import export_summary_attendance_report

TEST_DATA_PATH = os.path.join(os.path.dirname(__file__), 'data', 'transform_data_20250601.xlsx')

class DummyMinioClient:
    def __init__(self, test_result_dir):
        self.test_result_dir = test_result_dir
        os.makedirs(test_result_dir, exist_ok=True)
    def fput_object(self, bucket, object_name, file_path):
        # Thay thế upload bằng copy file sang thư mục test_result
        dest = os.path.join(self.test_result_dir, object_name)
        from shutil import copyfile
        copyfile(file_path, dest)
    def presigned_get_object(self, bucket, object_name):
        # Trả về đường dẫn file local
        return os.path.join(self.test_result_dir, object_name)

@pytest.mark.parametrize("sheet_name", ["apec_attendance_report"])
def test_export_summary_attendance_report(sheet_name):
    # Đọc dữ liệu test từ file Excel
    df = pd.read_excel(TEST_DATA_PATH, sheet_name=sheet_name)
    data = {sheet_name: df}
    with tempfile.TemporaryDirectory() as tmpdir:
        test_result_dir = os.path.join(tmpdir, "test_result")
        minio_client = DummyMinioClient(test_result_dir)
        # Gọi hàm với mock minio_client
        output_files = export_summary_attendance_report(
            data,
            tmpdir,
            data_key=sheet_name,
            minio_client=minio_client,
            minio_bucket="test-bucket",
            minio_prefix=None,
            remove_local_file=True
        )
        # Kiểm tra có file xuất ra (metadata trả về)
        assert output_files, "No output file generated!"
        for item in output_files:
            assert isinstance(item, dict) and 'url' in item and 'minio_key' in item, "Output item missing url/minio_key!"
            file_path = item['url']
            assert os.path.exists(file_path), f"File {file_path} does not exist!"
            # Kiểm tra file Excel có sheet đầu tiên và dữ liệu tên/mã nhân viên
            wb = openpyxl.load_workbook(file_path)
            ws = wb.worksheets[0]
            # Kiểm tra tên công ty ở ô C2
            assert ws['C2'].value is not None and ws['C2'].value != '', "Company name missing in C2!"
            # Kiểm tra có ít nhất 1 dòng dữ liệu nhân viên (từ dòng 11 trở đi)
            found_employee = False
            for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=2, max_col=3):
                code, name = row[0].value, row[1].value
                if code and name:
                    found_employee = True
                    break
            assert found_employee, "No employee data found in output file!"
