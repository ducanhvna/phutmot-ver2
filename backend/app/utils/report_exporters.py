import pandas as pd
import os
from openpyxl import Workbook
from app.models.file_metadata import save_file_metadata_list
import openpyxl
from shutil import copyfile

def export_al_cl_report_department(data, output_dir, data_key='al_report'):
    """
    Xuất báo cáo AL/CL theo từng phòng ban, mỗi công ty là 1 file, file tổng sẽ ghép từ các file con.
    """
    df = data.get(data_key)
    if (
        not isinstance(df, pd.DataFrame)
        or df.empty
        or 'company_name' not in df.columns
        or 'department_name' not in df.columns
    ):
        # Nếu không có dữ liệu hoặc thiếu cột, tạo file tổng No Data
        file_path = os.path.join(output_dir, 'al_cl_report_department.xlsx')
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        return file_path
    # Tạo file cho từng công ty
    company_files = []
    for company, group_company in df.groupby('company_name'):
        safe_company = str(company)[:20].replace('/', '_').replace('\\', '_')
        file_path = os.path.join(output_dir, f"al_cl_report_department_{safe_company}.xlsx")
        with pd.ExcelWriter(file_path) as writer:
            for dept, group_dept in group_company.groupby('department_name'):
                if not group_dept.empty:
                    sheet_name = str(dept)[:31].replace('/', '_').replace('\\', '_')
                    group_dept.to_excel(writer, sheet_name=sheet_name, index=False)
        company_files.append(file_path)
    # Ghép các file con thành file tổng (nhiều sheet, mỗi sheet là 1 công ty)
    import openpyxl
    file_path = os.path.join(output_dir, 'al_cl_report_department.xlsx')
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for company, group_company in df.groupby('company_name'):
            # Gộp toàn bộ dữ liệu công ty vào 1 sheet
            sheet_name = str(company)[:31].replace('/', '_').replace('\\', '_')
            group_company.to_excel(writer, sheet_name=sheet_name, index=False)
    return file_path

def export_sumary_attendance_report_department(data, output_dir, data_key='attendance'):
    """
    Xuất báo cáo tổng hợp chấm công theo phòng ban, mỗi sheet là 1 phòng ban của 1 công ty.
    """
    file_path = os.path.join(output_dir, 'sumary_attendance_report_department.xlsx')
    df = data.get(data_key)
    with pd.ExcelWriter(file_path) as writer:
        if not isinstance(df, pd.DataFrame) or df.empty:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        else:
            # Nếu có dữ liệu, group theo company, department (hoặc trường phù hợp)
            if 'company' in df.columns and 'department' in df.columns:
                group_cols = ['company', 'department']
            elif 'company_name' in df.columns and 'department_name' in df.columns:
                group_cols = ['company_name', 'department_name']
            else:
                df.to_excel(writer, sheet_name='No Data', index=False)
                return file_path
            has_data = False
            grouped = df.groupby(group_cols)
            for (company, dept), group in grouped:
                if not group.empty:
                    sheet_name = f"{str(company)[:15]}_{str(dept)[:10]}"
                    sheet_name = sheet_name.replace('/', '_').replace('\\', '_')[:31]
                    group.to_excel(writer, sheet_name=sheet_name, index=False)
                    has_data = True
            if not has_data:
                pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
    return file_path

def export_late_in_5_miniutes_report(data, output_dir):
    """
    Xuất báo cáo đi muộn dưới 5 phút, mỗi công ty là 1 file riêng biệt.
    """
    # Hỗ trợ cả trường hợp data là dict chứa DataFrame hoặc trực tiếp là DataFrame
    df = data.get('late_in_5_miniutes') if isinstance(data, dict) else data
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        file_path = os.path.join(output_dir, 'late_in_5_miniutes_report_No_Data.xlsx')
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        return [file_path]
    # Xác định cột group theo công ty
    if 'company' in df.columns:
        group_col = 'company'
    elif 'company_name' in df.columns:
        group_col = 'company_name'
    else:
        file_path = os.path.join(output_dir, 'late_in_5_miniutes_report_No_Group.xlsx')
        df.to_excel(file_path, index=False)
        return [file_path]
    file_paths = []
    for company, group in df.groupby(group_col):
        if not group.empty:
            safe_company = str(company)[:20].replace('/', '_').replace('\\', '_')
            file_path = os.path.join(output_dir, f"late_in_5_miniutes_report_{safe_company}.xlsx")
            with pd.ExcelWriter(file_path) as writer:
                group.to_excel(writer, sheet_name='Data', index=False)
            file_paths.append(file_path)
    if not file_paths:
        file_path = os.path.join(output_dir, 'late_in_5_miniutes_report_No_Data.xlsx')
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        return [file_path]
    return file_paths

def export_feed_report(data, output_dir):
    """
    Xuất báo cáo tổng hợp ăn ca, mỗi sheet là 1 công ty.
    """
    file_path = os.path.join(output_dir, 'feed_report.xlsx')
    df = data.get('feed_report')
    if not isinstance(df, pd.DataFrame) or df.empty:
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        return file_path

    # Group theo company hoặc company_name
    if 'company' in df.columns:
        group_col = 'company'
    elif 'company_name' in df.columns:
        group_col = 'company_name'
    else:
        # Nếu không có cột group, xuất 1 sheet
        df.to_excel(file_path, index=False)
        return file_path

    with pd.ExcelWriter(file_path) as writer:
        for company, group in df.groupby(group_col):
            sheet_name = str(company)[:31].replace('/', '_').replace('\\', '_')
            group.to_excel(writer, sheet_name=sheet_name, index=False)
    return file_path

def export_kpi_weekly_report_ho(data, output_dir):
    """
    Xuất báo cáo KPI weekly cho khối HO, mỗi sheet là 1 công ty.
    """
    file_path = os.path.join(output_dir, 'kpi_weekly_report_ho.xlsx')
    df = data.get('kpi_weekly_report_summary')
    if not isinstance(df, pd.DataFrame) or df.empty:
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        return file_path

    # Group theo company hoặc company_name
    if 'company' in df.columns:
        group_col = 'company'
    elif 'company_name' in df.columns:
        group_col = 'company_name'
    else:
        # Nếu không có cột group, xuất 1 sheet
        df.to_excel(file_path, index=False)
        return file_path

    with pd.ExcelWriter(file_path) as writer:
        for company, group in df.groupby(group_col):
            sheet_name = str(company)[:31].replace('/', '_').replace('\\', '_')
            group.to_excel(writer, sheet_name=sheet_name, index=False)
    return file_path

def export_kpi_weekly_report(data, output_dir):
    file_path = os.path.join(output_dir, 'kpi_weekly_report.xlsx')
    df = data.get('kpi_weekly_report_summary')
    if not isinstance(df, pd.DataFrame) or df.empty:
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        return file_path
    df.to_excel(file_path, index=False)
    return file_path

def export_al_cl_report_ho(data, output_dir):
    file_path = os.path.join(output_dir, 'al_cl_report_ho.xlsx')
    df = data.get('al_report')
    if not isinstance(df, pd.DataFrame) or df.empty:
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        return file_path
    df.to_excel(file_path, index=False)
    return file_path

def export_al_cl_report(data, output_dir):
    file_path = os.path.join(output_dir, 'al_cl_report.xlsx')
    df = data.get('al_report')
    if not isinstance(df, pd.DataFrame) or df.empty:
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        return file_path
    df.to_excel(file_path, index=False)
    return file_path

def export_al_cl_report_severance(data, output_dir):
    file_path = os.path.join(output_dir, 'al_cl_report_severance.xlsx')
    df = data.get('al_report')
    if not isinstance(df, pd.DataFrame) or df.empty:
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        return file_path
    df.to_excel(file_path, index=False)
    return file_path

def export_json_report(data: dict, output_dir: str, file_name: str):
    """
    Lưu dict chuẩn hóa ra file json với tên file_name trong output_dir.
    Tự động chuyển các kiểu datetime, Timestamp, numpy types về string hoặc native cho JSON.
    """
    import json
    import numpy as np
    import pandas as pd
    from datetime import datetime, date
    def default_serializer(obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, (np.integer, np.floating)):
            return obj.item()
        if isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if hasattr(obj, '__iter__') and not isinstance(obj, (str, bytes, dict)):
            return list(obj)
        return str(obj)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, file_name)
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=default_serializer)
    return file_path

def export_summary_attendance_report(data, output_dir, data_key=None, template_path=None, date_array=None, minio_client=None, minio_bucket=None, minio_prefix=None, remove_local_file=True):
    """
    Xuất báo cáo tổng hợp chấm công từng công ty, sử dụng template Excel chuẩn.
    Mỗi công ty là 1 file, sheet đầu tiên sẽ được ghi dữ liệu tổng hợp.
    Nếu truyền minio_client và minio_bucket thì upload từng file lên MinIO ngay sau khi tạo, trả về danh sách metadata hoặc url.
    """
    import pandas as pd
    import openpyxl
    from shutil import copyfile
    import os
    from datetime import datetime, timedelta
    import calendar
    # Tự động chọn data_key nếu không truyền vào
    if data_key is None:
        if 'apec_attendance_report' in data:
            data_key = 'apec_attendance_report'
        elif 'attendance' in data:
            data_key = 'attendance'
        else:
            raise ValueError("Không tìm thấy key dữ liệu phù hợp trong dict data!")
    if template_path is None:
        template_path = os.path.join(os.path.dirname(__file__), 'template', 'Summary time attendant OK.xlsx')
    df = data.get(data_key)
    result_list = []
    if not isinstance(df, pd.DataFrame) or df.empty:
        file_path = os.path.join(output_dir, 'summary_attendance_report_No_Data.xlsx')
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        if minio_client and minio_bucket:
            minio_key = os.path.join(minio_prefix or '', os.path.basename(file_path))
            minio_client.fput_object(minio_bucket, minio_key, file_path)
            url = minio_client.presigned_get_object(minio_bucket, minio_key)
            if remove_local_file:
                os.remove(file_path)
            result_list.append({'company': None, 'minio_key': minio_key, 'url': url})
            return result_list
        return [file_path]
    # Xác định cột group theo công ty
    group_col = 'mis_id' if 'mis_id' in df.columns else 'company_name' if 'company_name' in df.columns else None
    if not group_col:
        file_path = os.path.join(output_dir, 'summary_attendance_report_No_Group.xlsx')
        df.to_excel(file_path, index=False)
        if minio_client and minio_bucket:
            minio_key = os.path.join(minio_prefix or '', os.path.basename(file_path))
            minio_client.fput_object(minio_bucket, minio_key, file_path)
            url = minio_client.presigned_get_object(minio_bucket, minio_key)
            if remove_local_file:
                os.remove(file_path)
            result_list.append({'company': None, 'minio_key': minio_key, 'url': url})
            return result_list
        return [file_path]
    # Nếu date_array không truyền vào thì lấy từ ngày đầu tháng đến cuối tháng hiện tại
    if date_array is None:
        now = datetime.now()
        first_day = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day = now.replace(day=calendar.monthrange(now.year, now.month)[1], hour=0, minute=0, second=0, microsecond=0)
        date_array = [first_day + timedelta(days=i) for i in range((last_day - first_day).days + 1)]
    for company, group in df.groupby(group_col):
        if not group.empty:
            safe_company = str(company)[:20].replace('/', '_').replace('\\', '_')
            file_path = os.path.join(output_dir, f"1__summary_attendance_report_{safe_company}.xlsx")
            copyfile(template_path, file_path)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.worksheets[0]
            # Merge và ghi thông tin công ty, ngày tháng lên đầu file
            ws.merge_cells('C2:Z2')
            ws['C2'].value = company
            ws.merge_cells('C3:Z3')
            ws['C3'].value = ''
            ws.merge_cells('C4:Z4')
            ws['C4'].value = ''
            # Chỉ merge dòng 6 cho tiêu đề tổng, KHÔNG merge các dòng 9, 10, 11
            ws.merge_cells('A6:BK6')
            ws['A6'].value = f'Employee Type: All - Working Time: All - Pay Cycle: Month - Payment Period: {date_array[0].strftime("%d/%m/%Y")} - {date_array[-1].strftime("%d/%m/%Y")}'
            # Nếu không có cột department_name thì bỏ qua phần ghi tên phòng ban
            if 'department_name' in group.columns:
                start_row = 11
                start_col = 33
                # --- Bỏ merge các ô ở dòng 9, 10, 11, cột từ 33 trở đi (nếu có) ---
                unmerge_rows = [start_row-2, start_row-1, start_row]
                for row in unmerge_rows:
                    for col in range(start_col+1, start_col+32):
                        cell = ws.cell(row=row, column=col)
                        for merged_range in list(ws.merged_cells.ranges):
                            if cell.coordinate in merged_range:
                                ws.unmerge_cells(str(merged_range))
                # Ghi thứ/ngày vào dòng 9, 10 (không merge các dòng này)
                for date_item in date_array:
                    day_of_month = date_item.day
                    ws.cell(row=start_row-2, column=start_col + day_of_month).value = date_item.strftime('%A')[:3]
                    ws.cell(row=start_row-1, column=start_col + day_of_month).value = date_item
                if date_array and date_array[-1].day < 31:
                    for add_item in range(date_array[-1].day + 1, 32):
                        ws.cell(row=start_row-2, column=start_col + add_item).value = ''
                        ws.cell(row=start_row-1, column=start_col + add_item).value = ''
                        ws.cell(row=start_row, column=start_col + add_item).value = ''
                for dept, dept_group in group[group['department_name'].notnull() & (group['department_name'] != False)].groupby('department_name'):
                    # In tên phòng ban
                    ws.cell(row=start_row, column=1).value = dept
                    start_row += 1    
                    # Ghi dữ liệu từng nhân viên: name vào cột 3, employee_code vào cột 2
                    name_col = 'name' if 'name' in group.columns else 'employee_name' if 'employee_name' in group.columns else None
                    code_col = 'employee_code' if 'employee_code' in group.columns else 'code' if 'code' in group.columns else None
                    if not name_col or not code_col:
                        raise KeyError("Không tìm thấy cột tên hoặc mã nhân viên trong dữ liệu!")
                    for sub_group_index, sub_group_data in dept_group.groupby([name_col, code_col]):
                        ws.cell(row=start_row, column=2).value = sub_group_index[1]
                        ws.cell(row=start_row, column=3).value = sub_group_index[0]
                        # ...
                        max_row_by_a_date = 0
                        for date, date_data in sub_group_data.groupby('date'):
                            date_value = datetime.strptime(date, '%Y-%m-%d') if isinstance(date, str) else date
                            subrow_index = 0
                            for row in date_data.iterrows():
                                day_of_month = date_value.day
                                col_index = start_col + day_of_month
                                # Ghi dữ liệu vào ô tương ứng
                                ws.cell(row=start_row + subrow_index, column=col_index).value = date_data['shift_name'].values[0] if 'shift_name' in date_data.columns else ''
                                subrow_index += 1
                            max_row_by_a_date = max(max_row_by_a_date, subrow_index)    
                        start_row += max_row_by_a_date
            wb.save(file_path)
            # Nếu có minio_client thì upload lên MinIO và trả về url/metadata
            if minio_client and minio_bucket:
                minio_key = os.path.join(minio_prefix or '', os.path.basename(file_path))
                minio_client.fput_object(minio_bucket, minio_key, file_path)
                url = minio_client.presigned_get_object(minio_bucket, minio_key)
                if remove_local_file:
                    os.remove(file_path)
                result_list.append({'company': company, 'minio_key': minio_key, 'url': url})
            else:
                result_list.append(file_path)
    if not result_list:
        file_path = os.path.join(output_dir, 'summary_attendance_report_No_Data.xlsx')
        with pd.ExcelWriter(file_path) as writer:
            pd.DataFrame().to_excel(writer, sheet_name='No Data', index=False)
        if minio_client and minio_bucket:
            minio_key = os.path.join(minio_prefix or '', os.path.basename(file_path))
            minio_client.fput_object(minio_bucket, minio_key, file_path)
            url = minio_client.presigned_get_object(minio_bucket, minio_key)
            if remove_local_file:
                os.remove(file_path)
            result_list.append({'company': None, 'minio_key': minio_key, 'url': url})
            return result_list
        return [file_path]
    return result_list
